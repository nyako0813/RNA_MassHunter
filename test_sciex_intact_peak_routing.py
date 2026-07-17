from hashlib import sha256
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock

import numpy as np
import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

import main as main_module
from rna_masshunter.audit_policy import AuditPolicy
from rna_masshunter.config import load_config, resolve_paths, validate_config
from rna_masshunter.excel_report import (
    SCIEX_INTACT_DIAGNOSTIC_SHEET,
    SCIEX_INTACT_OPTIONAL_RESULT_KEY,
    SCIEX_INTACT_PEAK_SHEET,
    write_excel_report,
)
from rna_masshunter.models import RunConfig


PARSER_DIAGNOSTIC_SHEET = "SCIEX_Profile_Diagnostics"
PARSER_INPUT_SHEET = "SCIEX_Profile_Input"
FORMAL_COLUMNS = (
    "SCIEX_Intact_Peak_Detection_Applied_To_Formal_Score",
    "SCIEX_Intact_Peak_Detection_Applied_To_Ranking",
    "SCIEX_Intact_Peak_Detection_Applied_To_Candidate_Filtering",
)


def write_neutral_profile(path, *, flat=False):
    masses = np.arange(1000.0, 1100.5, 0.5)
    if flat:
        intensities = np.ones_like(masses)
    else:
        intensities = 1.0 + 20.0 * np.exp(-0.5 * ((masses - 1050.0) / 2.0) ** 2)
    lines = ["Mass\tIntensity"]
    lines.extend(f"{mass}\t{intensity}" for mass, intensity in zip(masses, intensities))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_mz_profile(path):
    path.write_text(
        "Mass/Charge\tIntensity\n100.0\t10\n100.1\t11\n100.2\t9\n",
        encoding="utf-8",
    )
    return path


def route_config(path=None, *, enabled=True, detection=True):
    return RunConfig(
        sciex_profile={
            "enabled": enabled,
            "path": path,
            "intact_peak_detection": {"enabled": detection},
        }
    )


def route(config, level="full", warnings=None):
    return main_module.build_sciex_profile_optional_results(
        config,
        AuditPolicy.from_level(level),
        [] if warnings is None else warnings,
    )


def writer_config(report_limit=1000):
    return SimpleNamespace(
        analysis={"mode": "full"},
        project={"name": "sciex-routing-test"},
        input={}, organism={}, sequence={}, experiment={}, instrument={},
        sciex_profile={}, reconstruction={"enabled": False}, digestion={"enabled": False},
        alkaline_phosphatase={}, fragment_mapping={}, modification_search={},
        peak_filtering={}, p1_annotation={}, ms2_annotation={},
        modification_evidence_ranking={}, biological_context={}, performance={},
        reporting={
            "max_excel_rows_per_sheet": report_limit,
            "truncate_large_sheets": True,
        },
    )


def write_report(tmp_path, level, optional_results, label):
    return write_excel_report(
        output_dir=tmp_path / label,
        config=writer_config(),
        diagnostics={}, intact_results=[], charge_state_peaks=[], warnings=[],
        modifications=[], rule_set={}, pathways=[], theoretical_fragments=[],
        fragment_ms1_matches=[], known_modification_candidates=[],
        known_modification_summary=[], optional_results=optional_results,
        audit_policy=AuditPolicy.from_level(level),
    )


def sheet_names(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def read_sheet(path, name):
    return pd.read_excel(path, sheet_name=name, header=2, dtype=object)


def test_config_defaults_are_backward_compatible_and_unknown_keys_are_retained(tmp_path):
    config_path = tmp_path / "config.yaml"
    config_path.write_text("project:\n  name: old-config\n", encoding="utf-8")
    warnings = []
    config = load_config(config_path, warnings=warnings)
    validate_config(config, warnings=warnings)
    assert not any("sciex_profile" in row["Message"] for row in warnings)
    assert config.sciex_profile == {
        "enabled": False,
        "path": None,
        "intact_peak_detection": {"enabled": True},
        "intact_mass_comparison": {
            "enabled": True,
            "strict_tolerance_da": 1.0,
            "broad_tolerance_da": 5.0,
        },
        "delta_mass_cluster_audit": {
            "enabled": True,
            "cluster_tolerance_da": 0.5,
            "duplicate_apex_tolerance_da": 0.25,
            "isotope_spacing_da": 1.003355,
            "isotope_spacing_tolerance_da": 0.15,
            "integer_spacing_tolerance_da": 0.15,
            "minimum_cluster_size": 2,
            "max_pair_spacing_da": 200.0,
            "max_pair_rows": 20000,
        },
    }

    config_path.write_text(
        yaml.safe_dump({
            "sciex_profile": {
                "enabled": False,
                "path": None,
                "intact_peak_detection": {"enabled": True},
                "future_option": "preserved",
            }
        }),
        encoding="utf-8",
    )
    config = load_config(config_path)
    assert config.sciex_profile["future_option"] == "preserved"


def test_relative_sciex_path_uses_existing_project_root_rule(tmp_path):
    config = route_config("profiles/sample-full.txt")
    resolve_paths(config, tmp_path)
    assert Path(config.sciex_profile["path"]) == tmp_path / "profiles/sample-full.txt"


@pytest.mark.parametrize(
    "path_value,message",
    [
        (None, "path is required"),
        ("", "path must not be empty"),
    ],
)
def test_enabled_profile_requires_nonempty_path(path_value, message):
    config = route_config(path_value)
    with pytest.raises(ValueError, match=message):
        validate_config(config)


def test_disabled_routing_does_not_call_parser_or_detector(monkeypatch):
    monkeypatch.setattr(
        main_module, "parse_sciex_profile",
        lambda _path: pytest.fail("parser must not run"),
    )
    monkeypatch.setattr(
        main_module, "detect_sciex_intact_peaks",
        lambda *_args, **_kwargs: pytest.fail("detector must not run"),
    )
    assert route(route_config(None, enabled=False)) == {}


def test_detection_disabled_parses_profile_without_running_detector(tmp_path, monkeypatch):
    profile = write_neutral_profile(tmp_path / "sample-full.txt")
    monkeypatch.setattr(
        main_module, "detect_sciex_intact_peaks",
        lambda *_args, **_kwargs: pytest.fail("detector must not run"),
    )
    results = route(route_config(profile, detection=False), "full")
    assert results[PARSER_DIAGNOSTIC_SHEET][0]["Profile_Type"] == "NEUTRAL_MASS_PROFILE"
    assert PARSER_INPUT_SHEET in results
    assert SCIEX_INTACT_OPTIONAL_RESULT_KEY not in results


def test_neutral_profile_routes_parser_and_detector(tmp_path, monkeypatch):
    profile = write_neutral_profile(tmp_path / "sample-full.txt")
    calls = []
    original = main_module.detect_sciex_intact_peaks

    def recording_detector(masses, intensities, **metadata):
        calls.append((tuple(masses), tuple(intensities), dict(metadata)))
        return original(masses, intensities, **metadata)

    monkeypatch.setattr(main_module, "detect_sciex_intact_peaks", recording_detector)
    results = route(route_config(profile), "full")
    diagnostic = results[PARSER_DIAGNOSTIC_SHEET][0]
    assert diagnostic["Profile_Type"] == "NEUTRAL_MASS_PROFILE"
    assert diagnostic["Input_Status"] == "SUPPORTED_INPUT"
    assert diagnostic["Eligible_For_Neutral_Mass_Analysis"] is True
    assert len(calls) == 1
    assert calls[0][2] == {
        "profile_type": "NEUTRAL_MASS_PROFILE",
        "input_status": "SUPPORTED_INPUT",
        "eligible_for_neutral_mass_analysis": True,
    }
    routed = results[SCIEX_INTACT_OPTIONAL_RESULT_KEY]
    assert routed["source_file"] == profile
    assert routed["result"].diagnostics["Detection_Status"].startswith("DETECTION_COMPLETED")


def test_mz_profile_is_not_converted_or_sent_to_detector(tmp_path, monkeypatch):
    profile = write_mz_profile(tmp_path / "sample-T1.txt")
    monkeypatch.setattr(
        main_module, "detect_sciex_intact_peaks",
        lambda *_args, **_kwargs: pytest.fail("MZ_PROFILE must not reach detector"),
    )
    results = route(route_config(profile), "full")
    diagnostic = results[PARSER_DIAGNOSTIC_SHEET][0]
    assert diagnostic["Profile_Type"] == "MZ_PROFILE"
    assert diagnostic["Input_Status"] == "UNSUPPORTED_PROFILE_TYPE"
    assert diagnostic["Eligible_For_Neutral_Mass_Analysis"] is False
    assert SCIEX_INTACT_OPTIONAL_RESULT_KEY not in results
    assert results[PARSER_INPUT_SHEET][0]["MZ"] == 100.0
    assert results[PARSER_INPUT_SHEET][0]["Neutral_Mass"] == ""


@pytest.mark.parametrize("kind", ["missing", "directory"])
def test_invalid_profile_paths_raise_clear_errors(tmp_path, kind):
    profile = tmp_path / "input"
    if kind == "directory":
        profile.mkdir()
        expected = IsADirectoryError
    else:
        expected = FileNotFoundError
    with pytest.raises(expected, match="SCIEX profile"):
        route(route_config(profile))


@pytest.mark.parametrize(
    "name,text,status",
    [
        ("empty.txt", "", "EMPTY_INPUT"),
        ("malformed.txt", "Time\tSignal\n1\t2\n", "UNRECOGNIZED_COLUMNS"),
        ("header-full.txt", "Mass\tIntensity\n", "EMPTY_INPUT"),
    ],
)
def test_empty_malformed_and_header_only_profiles_are_parser_only(tmp_path, name, text, status):
    profile = tmp_path / name
    profile.write_text(text, encoding="utf-8")
    results = route(route_config(profile), "audit")
    assert results[PARSER_DIAGNOSTIC_SHEET][0]["Input_Status"] == status
    assert SCIEX_INTACT_OPTIONAL_RESULT_KEY not in results


def test_zero_peak_neutral_profile_still_routes_diagnostic_result(tmp_path):
    profile = write_neutral_profile(tmp_path / "flat-full.txt", flat=True)
    results = route(route_config(profile), "audit")
    detection = results[SCIEX_INTACT_OPTIONAL_RESULT_KEY]["result"]
    assert detection.diagnostics["Detected_Sensitive_Peak_Count"] == 0


def test_detector_failure_retains_parser_diagnostics_and_adds_warning(tmp_path, monkeypatch):
    profile = write_neutral_profile(tmp_path / "failure-full.txt")
    warnings = []

    def fail(*_args, **_kwargs):
        raise RuntimeError("synthetic detector failure")

    monkeypatch.setattr(main_module, "detect_sciex_intact_peaks", fail)
    results = route(route_config(profile), "audit", warnings)
    assert PARSER_DIAGNOSTIC_SHEET in results
    assert SCIEX_INTACT_OPTIONAL_RESULT_KEY not in results
    assert any(
        row["Source"] == "sciex_intact_peak_detection"
        and "synthetic detector failure" in str(row["Context"])
        for row in warnings
    )


@pytest.mark.parametrize("level,expect_intact", [("standard", False), ("audit", True), ("full", True)])
def test_routed_neutral_profile_follows_workbook_policy(tmp_path, level, expect_intact):
    profile = write_neutral_profile(tmp_path / f"{level}-full.txt")
    optional = route(route_config(profile), level)
    report = write_report(tmp_path, level, optional, f"workbook-{level}")
    names = set(sheet_names(report))
    assert (SCIEX_INTACT_DIAGNOSTIC_SHEET in names) is expect_intact
    assert (SCIEX_INTACT_PEAK_SHEET in names) is expect_intact
    if not expect_intact:
        return
    diagnostics = read_sheet(report, SCIEX_INTACT_DIAGNOSTIC_SHEET)
    peaks = read_sheet(report, SCIEX_INTACT_PEAK_SHEET)
    detection = optional[SCIEX_INTACT_OPTIONAL_RESULT_KEY]["result"]
    assert len(diagnostics) == 1
    assert len(peaks) == detection.diagnostics["Detected_Sensitive_Peak_Count"]
    for column in FORMAL_COLUMNS:
        assert not diagnostics[column].map(bool).any()
        assert not peaks[column].map(bool).any()
    assert not peaks["Molecular_Identity_Assigned"].map(bool).any()


def test_mz_profile_workbook_keeps_existing_parser_policy(tmp_path):
    profile = write_mz_profile(tmp_path / "policy-T1.txt")
    for level, expected in {
        "standard": set(),
        "audit": {PARSER_DIAGNOSTIC_SHEET},
        "full": {PARSER_DIAGNOSTIC_SHEET, PARSER_INPUT_SHEET},
    }.items():
        optional = route(route_config(profile), level)
        report = write_report(tmp_path, level, optional, f"mz-{level}")
        names = set(sheet_names(report))
        assert names & {PARSER_DIAGNOSTIC_SHEET, PARSER_INPUT_SHEET} == expected
        assert SCIEX_INTACT_DIAGNOSTIC_SHEET not in names
        assert SCIEX_INTACT_PEAK_SHEET not in names


def test_main_places_routed_result_in_excel_optional_results(tmp_path, monkeypatch):
    profile = write_neutral_profile(tmp_path / "main-full.txt")
    config_path = tmp_path / "main.yaml"
    output_dir = tmp_path / "output"
    config_path.write_text(
        yaml.safe_dump({
            "analysis": {"mode": "intact_only"},
            "project": {
                "name": "main-routing-test",
                "output_dir": str(output_dir),
                "log_dir": str(tmp_path / "logs"),
                "cache_dir": str(tmp_path / "cache"),
            },
            "input": {"raw_path": "", "mzml_path": "", "msconvert_path": ""},
            "sequence": {"sequence": ""},
            "reconstruction": {"enabled": True},
            "digestion": {"enabled": False},
            "sciex_profile": {
                "enabled": True,
                "path": str(profile),
                "intact_peak_detection": {"enabled": True},
            },
        }),
        encoding="utf-8",
    )
    captured = {}

    def capture_writer(**kwargs):
        captured.update(kwargs)
        return tmp_path / "captured.xlsx"

    monkeypatch.setattr(main_module, "setup_logger", lambda _path: Mock())
    monkeypatch.setattr(main_module, "run_startup_check", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(main_module, "write_excel_report", capture_writer)
    main_module.main(["--config", str(config_path), "--audit-level", "audit"])
    optional = captured["optional_results"]
    assert SCIEX_INTACT_OPTIONAL_RESULT_KEY in optional
    assert PARSER_DIAGNOSTIC_SHEET in optional
    assert optional[SCIEX_INTACT_OPTIONAL_RESULT_KEY]["result"].diagnostics[
        "Detection_Status"
    ].startswith("DETECTION_COMPLETED")


def test_sciex_enabled_ab_preserves_formal_workbook_values(tmp_path):
    profile = write_neutral_profile(tmp_path / "ab-full.txt")
    formal_optional = {
        "Modification_Evidence_Ranking": [{
            "Candidate_Key": "candidate-1", "Final_Score": 7.5,
            "Final_Confidence": "High", "Final_Rank": 1,
        }],
        "P1_Summary": [{"Observed_Peak_Count": 3}],
        "MS2_Summary": [{"Spectrum_Count": 2}],
    }
    disabled_report = write_report(tmp_path, "full", formal_optional, "ab-disabled")
    enabled_optional = dict(formal_optional)
    enabled_optional.update(route(route_config(profile), "full"))
    enabled_report = write_report(tmp_path, "full", enabled_optional, "ab-enabled")

    for sheet in (
        "Known_Modification_Candidates", "Known_Modification_Summary",
        "Modification_Evidence_Ranking", "P1_Summary", "MS2_Summary",
    ):
        pd.testing.assert_frame_equal(
            read_sheet(disabled_report, sheet),
            read_sheet(enabled_report, sheet),
        )

    disabled_summary = read_sheet(disabled_report, "Run_summary")
    enabled_summary = read_sheet(enabled_report, "Run_summary")
    disabled_values = dict(zip(disabled_summary["Item"], disabled_summary["Value"]))
    enabled_values = dict(zip(enabled_summary["Item"], enabled_summary["Value"]))
    for key in set(disabled_values) - {"Generated"}:
        if pd.isna(disabled_values[key]) and pd.isna(enabled_values[key]):
            continue
        assert enabled_values[key] == disabled_values[key]


def test_routing_is_deterministic_nonmutating_and_preserves_input(tmp_path):
    profile = write_neutral_profile(tmp_path / "deterministic-full.txt")
    before_hash = sha256(profile.read_bytes()).hexdigest()
    config = route_config(profile)
    settings_before = {
        "enabled": config.sciex_profile["enabled"],
        "path": config.sciex_profile["path"],
        "intact_peak_detection": dict(config.sciex_profile["intact_peak_detection"]),
    }
    first = route(config, "full")
    second = route(config, "full")
    first_result = first[SCIEX_INTACT_OPTIONAL_RESULT_KEY]["result"]
    second_result = second[SCIEX_INTACT_OPTIONAL_RESULT_KEY]["result"]
    assert first_result.diagnostics_row() == second_result.diagnostics_row()
    assert first_result.peak_rows() == second_result.peak_rows()
    assert first_result.provenance_rows() == second_result.provenance_rows()
    assert config.sciex_profile == settings_before
    assert sha256(profile.read_bytes()).hexdigest() == before_hash
