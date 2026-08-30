from enum import Enum
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from rna_masshunter.audit_policy import AuditPolicy, included_sheet_names
from rna_masshunter.excel_report import (
    SCIEX_INTACT_DIAGNOSTIC_SHEET,
    SCIEX_INTACT_PEAK_SHEET,
    SCIEX_MASS_COMPARISON_DETAIL_SHEET,
    SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY,
    SCIEX_MASS_COMPARISON_SUMMARY_SHEET,
    _sciex_mass_comparison_excel_sheets,
    write_excel_report,
)
from rna_masshunter.intact_rna_mass import (
    FivePrimeState,
    IntactRnaMassParameters,
    ThreePrimeState,
    calculate_intact_rna_mass,
)
from rna_masshunter.sciex_intact_mass_comparison import (
    DETAIL_COLUMNS,
    SUMMARY_COLUMNS,
    compare_sciex_intact_masses,
)

SEQUENCE = "ACGUCCA"
FORMAL_FLAGS = (
    "SCIEX_Intact_Mass_Matching_Applied_To_Formal_Score",
    "SCIEX_Intact_Mass_Matching_Applied_To_Ranking",
    "SCIEX_Intact_Mass_Matching_Applied_To_Candidate_Filtering",
    "SCIEX_Intact_Mass_Matching_Applied_To_Final_Consensus",
)
FALSE_CERTAINTY_FLAGS = (
    "Structure_Identity_Assigned",
    "Position_Assigned",
    "Molecular_Identity_Assigned",
    "Modification_Assigned",
)


class Detection:
    def __init__(self, peaks):
        self._peaks = [dict(row) for row in peaks]

    def peak_rows(self):
        return [dict(row) for row in self._peaks]

    def diagnostics_row(self):
        return {"Detection_Status":"DETECTION_COMPLETED", "Input_Status":"SUPPORTED_INPUT"}


def reference_mass():
    return calculate_intact_rna_mass(
        SEQUENCE,
        parameters=IntactRnaMassParameters(
            five_prime_state=FivePrimeState.MONOPHOSPHATE,
            three_prime_state=ThreePrimeState.OH,
        ),
    ).monoisotopic_neutral_mass


def synthetic_peak(peak_id, mass, **extra):
    return {
        "Peak_ID":peak_id,
        "Apex_Mass":mass,
        "Centroid_Mass":mass + 0.2,
        "Apex_Intensity_Raw":100.0,
        "Detection_Tier":"STRICT",
        "Strict_Threshold_Passed":True,
        "Peak_Area_Complete":True,
        "Centroid_Complete":True,
        "Broad_Peak_Flag":False,
        "Possible_Shoulder":False,
        "Edge_Peak_Flag":False,
        **extra,
    }


def comparison(identity_status="CONFIRMED", peaks=None):
    mass = reference_mass()
    if peaks is None:
        peaks = [
            synthetic_peak("exact", mass),
            synthetic_peak("exploratory", mass + 5.0),
            synthetic_peak("broad", mass, Broad_Peak_Flag=True),
            synthetic_peak(
                "edge", mass, Edge_Peak_Flag=True, Peak_Area_Complete=False,
            ),
        ]
    return compare_sciex_intact_masses(
        Detection(peaks),
        999.0,
        source_file="/synthetic/input.txt",
        sequence=SEQUENCE,
        sequence_source="synthetic-excel-test",
        identity_status=identity_status,
    )


def writer_config(limit=1000):
    return SimpleNamespace(
        analysis={"mode":"full"},
        project={"name":"comparison-excel"},
        input={}, organism={}, sequence={}, experiment={}, instrument={}, sciex_profile={},
        reconstruction={"enabled":False}, digestion={"enabled":False},
        alkaline_phosphatase={}, fragment_mapping={}, modification_search={},
        peak_filtering={}, p1_annotation={}, ms2_annotation={},
        modification_evidence_ranking={}, biological_context={}, performance={},
        reporting={"max_excel_rows_per_sheet":limit, "truncate_large_sheets":True},
    )


def write_comparison(tmp_path, result, level="full", limit=1000, optional=None):
    optional_results = (
        {SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY:result}
        if optional is None else optional
    )
    report_path, _word_appendix_path = write_excel_report(
        tmp_path / f"{level}-{limit}",
        writer_config(limit),
        {}, [], [], [],
        known_modification_candidates=[{"Candidate_ID":"formal-1"}],
        known_modification_summary=[{"Summary_Type":"formal-summary"}],
        optional_results=optional_results,
        audit_policy=AuditPolicy.from_level(level),
    )
    return report_path


def sheet_names(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def read_sheet(path, name):
    return pd.read_excel(path, sheet_name=name, header=2)


def test_optional_key_sheet_names_and_policy_registry_are_stable():
    assert SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY == "sciex_intact_mass_comparison"
    assert SCIEX_MASS_COMPARISON_SUMMARY_SHEET == "SCIEX_Intact_Mass_Comp_Summary"
    assert SCIEX_MASS_COMPARISON_DETAIL_SHEET == "SCIEX_Intact_Mass_Comparison"
    names = [SCIEX_MASS_COMPARISON_SUMMARY_SHEET, SCIEX_MASS_COMPARISON_DETAIL_SHEET]
    assert included_sheet_names(names, AuditPolicy.from_level("standard"))[0] == []
    assert included_sheet_names(names, AuditPolicy.from_level("audit"))[0] == names
    assert included_sheet_names(names, AuditPolicy.from_level("full"))[0] == names


@pytest.mark.parametrize("level,expected", [
    ("standard", set()),
    ("audit", {SCIEX_MASS_COMPARISON_SUMMARY_SHEET, SCIEX_MASS_COMPARISON_DETAIL_SHEET}),
    ("full", {SCIEX_MASS_COMPARISON_SUMMARY_SHEET, SCIEX_MASS_COMPARISON_DETAIL_SHEET}),
])
def test_workbook_policy_for_terminal_comparison(tmp_path, level, expected):
    names = set(sheet_names(write_comparison(tmp_path, comparison(), level=level)))
    actual = names & {SCIEX_MASS_COMPARISON_SUMMARY_SHEET, SCIEX_MASS_COMPARISON_DETAIL_SHEET}
    assert actual == expected


@pytest.mark.parametrize("level", ["standard", "audit", "full"])
@pytest.mark.parametrize("optional", [{}, {SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY:None}])
def test_missing_or_none_result_never_creates_comparison_sheets(tmp_path, level, optional):
    report = write_comparison(tmp_path, None, level=level, optional=optional)
    names = set(sheet_names(report))
    assert SCIEX_MASS_COMPARISON_SUMMARY_SHEET not in names
    assert SCIEX_MASS_COMPARISON_DETAIL_SHEET not in names


def test_synthetic_excel_summary_and_detail_round_trip(tmp_path):
    result = comparison()
    detail_snapshot = result.details()
    summary_snapshot = result.summaries()
    report = write_comparison(tmp_path, result)
    summary = read_sheet(report, SCIEX_MASS_COMPARISON_SUMMARY_SHEET)
    detail = read_sheet(report, SCIEX_MASS_COMPARISON_DETAIL_SHEET)
    assert len(summary) == 1
    assert len(detail) == len(detail_snapshot)
    assert list(summary.columns) == SUMMARY_COLUMNS
    assert list(detail.columns) == DETAIL_COLUMNS
    row = summary.iloc[0]
    assert row["Sequence_Source"] == "synthetic-excel-test"
    assert row["Sequence_Length"] == len(SEQUENCE)
    assert len(row["Sequence_SHA256"]) == 64
    assert bool(row["Ends_With_CCA"])
    assert row["CCA_Policy"] == "AS_PROVIDED"
    assert row["Identity_Gate_Status"] == "CONFIRMED"
    assert bool(row["Identity_Gate_Passed"])
    assert row["Observed_Mass_Type"] == "UNKNOWN"
    assert row["Theoretical_Mass_Type"] == "MONOISOTOPIC_NEUTRAL"
    assert row["Mass_Definition_Compatibility"] == "UNKNOWN"
    assert not bool(row["Calibration_Applied"])
    assert row["Calibration_Method"] == "NONE"
    assert row["Terminal_Candidate_Count"] == 4
    assert row["Mass_Equivalent_Group_Count"] == 3
    assert row["Strict_Tolerance_Da"] == 1.0
    assert row["Exploratory_Tolerance_Da"] == 5.0
    assert bool(row["Mass_Match_Only"])
    assert all(not bool(row[key]) for key in FALSE_CERTAINTY_FLAGS + FORMAL_FLAGS)
    exact = detail[
        (detail["Peak_ID"] == "exact")
        & (detail["Candidate_ID"] == "TERM_LINEAR_5P_3OH")
    ].iloc[0]
    assert exact["Preferred_Mass_Field"] == "Apex_Mass"
    assert exact["Apex_Delta_Da"] == pytest.approx(0.0)
    assert exact["Centroid_Delta_Da"] == pytest.approx(0.2)
    assert exact["Theoretical_Formula"]
    assert exact["Five_Prime_State"] == "MONOPHOSPHATE"
    assert exact["Three_Prime_State"] == "OH"
    assert exact["Match_Tolerance_Class"] == "STRICT"
    assert exact["Peak_Eligibility"] == "PRIMARY"
    exploratory = detail[
        (detail["Peak_ID"] == "exploratory")
        & (detail["Candidate_ID"] == "TERM_LINEAR_5P_3OH")
    ].iloc[0]
    assert exploratory["Match_Tolerance_Class"] == "EXPLORATORY"
    broad = detail[detail["Peak_ID"] == "broad"].iloc[0]
    assert broad["Peak_Eligibility"] == "MATCHABLE_WITH_SHAPE_WARNING"
    edge = detail[detail["Peak_ID"] == "edge"].iloc[0]
    assert edge["Peak_Eligibility"] == "NOT_PRIMARY_ELIGIBLE"
    assert "edge_peak" in edge["Peak_Eligibility_Reason"]
    assert result.details() == detail_snapshot
    assert result.summaries() == summary_snapshot


def test_isobaric_terminal_candidates_remain_visibly_ambiguous_in_excel(tmp_path):
    detail = read_sheet(
        write_comparison(tmp_path, comparison(peaks=[synthetic_peak("P", reference_mass())])),
        SCIEX_MASS_COMPARISON_DETAIL_SHEET,
    )
    rows = detail[detail["Candidate_ID"].isin(["TERM_LINEAR_5P_3OH", "TERM_LINEAR_5OH_3P"])]
    assert len(rows) == 2
    assert rows["Mass_Equivalent_Candidate_Group_ID"].nunique() == 1
    assert set(rows["Candidate_Ambiguity_Count"]) == {2}
    assert rows["Terminal_State_Ambiguous"].map(bool).all()
    assert set(rows["Mass_Equivalent_Candidate_IDs"]) == {
        "TERM_LINEAR_5OH_3P;TERM_LINEAR_5P_3OH"
    }


def test_ambiguous_identity_warning_and_false_certainty_survive_excel(tmp_path):
    report = write_comparison(tmp_path, comparison(identity_status="AMBIGUOUS"))
    summary = read_sheet(report, SCIEX_MASS_COMPARISON_SUMMARY_SHEET).iloc[0]
    detail = read_sheet(report, SCIEX_MASS_COMPARISON_DETAIL_SHEET)
    assert summary["Identity_Gate_Status"] == "AMBIGUOUS"
    assert not bool(summary["Identity_Gate_Passed"])
    assert "IDENTITY_ASSUMPTION_DEPENDENT" in summary["Identity_Warnings"]
    assert detail["Interpretation_Warnings"].str.contains("IDENTITY_ASSUMPTION_DEPENDENT").all()
    assert detail["Mass_Match_Only"].map(bool).all()
    for key in FALSE_CERTAINTY_FLAGS + FORMAL_FLAGS:
        assert not detail[key].map(bool).any()


def test_identity_conflict_writes_one_summary_and_no_candidate_detail(tmp_path):
    result = comparison(identity_status="CONFLICT")
    report = write_comparison(tmp_path, result)
    names = sheet_names(report)
    assert SCIEX_MASS_COMPARISON_SUMMARY_SHEET in names
    assert SCIEX_MASS_COMPARISON_DETAIL_SHEET not in names
    summary = read_sheet(report, SCIEX_MASS_COMPARISON_SUMMARY_SHEET)
    assert len(summary) == 1
    row = summary.iloc[0]
    assert row["Comparison_Status"] == "IDENTITY_CONFLICT"
    assert row["Identity_Gate_Status"] == "CONFLICT"
    assert not bool(row["Identity_Gate_Passed"])
    assert row["Terminal_Candidate_Count"] == 0
    assert row["Strict_Match_Count"] == 0
    assert row["Exploratory_Match_Count"] == 0
    assert result.details() == []


class CellEnum(Enum):
    VALUE = "enum-value"


def test_excel_safe_serialization_is_deterministic_and_nonmutating(tmp_path):
    detail = {
        "Candidate_ID":CellEnum.VALUE,
        "Source_File":Path("/tmp/synthetic.txt"),
        "Apex_Mass":np.float64(123.5),
        "Centroid_Mass":float("nan"),
        "Interpretation_Warnings":({"z", "a"}, ["b", "a"], {"b":2, "a":1}),
        "Mass_Equivalent_Candidate_IDs":set(["TERM_B", "TERM_A"]),
    }
    summary = {
        "Source_File":Path("/tmp/synthetic.txt"),
        "Identity_Warnings":tuple(["second", "first"]),
        "Terminal_Candidate_Count":np.int64(4),
    }
    original_warning = detail["Interpretation_Warnings"]
    value = {"detail_rows":[detail], "summary_rows":[summary]}
    first = _sciex_mass_comparison_excel_sheets(value)
    second = _sciex_mass_comparison_excel_sheets(value)
    detail_frame = first[SCIEX_MASS_COMPARISON_DETAIL_SHEET]
    summary_frame = first[SCIEX_MASS_COMPARISON_SUMMARY_SHEET]
    assert detail_frame.at[0, "Candidate_ID"] == "enum-value"
    assert detail_frame.at[0, "Source_File"] == "/tmp/synthetic.txt"
    assert detail_frame.at[0, "Apex_Mass"] == 123.5
    assert detail_frame.at[0, "Centroid_Mass"] == ""
    assert detail_frame.at[0, "Interpretation_Warnings"] == '[["a","z"],["b","a"],{"a":1,"b":2}]'
    assert detail_frame.at[0, "Mass_Equivalent_Candidate_IDs"] == '["TERM_A","TERM_B"]'
    assert summary_frame.at[0, "Identity_Warnings"] == '["second","first"]'
    assert summary_frame.at[0, "Terminal_Candidate_Count"] == 4
    pd.testing.assert_frame_equal(detail_frame, second[SCIEX_MASS_COMPARISON_DETAIL_SHEET])
    pd.testing.assert_frame_equal(summary_frame, second[SCIEX_MASS_COMPARISON_SUMMARY_SHEET])
    report = write_comparison(tmp_path, value)
    written_detail = read_sheet(report, SCIEX_MASS_COMPARISON_DETAIL_SHEET)
    assert written_detail.at[0, "Candidate_ID"] == "enum-value"
    assert written_detail.at[0, "Source_File"] == "/tmp/synthetic.txt"
    assert written_detail.at[0, "Interpretation_Warnings"] == '[["a","z"],["b","a"],{"a":1,"b":2}]'
    assert pd.isna(written_detail.at[0, "Centroid_Mass"])
    assert detail["Interpretation_Warnings"] is original_warning


def test_comparison_sheets_use_existing_truncation_and_run_summary(tmp_path):
    result = comparison()
    report = write_comparison(tmp_path, result, limit=3)
    detail = read_sheet(report, SCIEX_MASS_COMPARISON_DETAIL_SHEET)
    assert len(detail) == 3
    run_summary = read_sheet(report, "Run_summary")
    value = run_summary.loc[run_summary["Item"] == "Truncated sheets", "Value"].iloc[0]
    assert f"{SCIEX_MASS_COMPARISON_DETAIL_SHEET}: {len(result.details())} -> 3" in value


def test_noncomparison_policy_registrations_are_unchanged():
    sciex_names = [
        "SCIEX_Profile_Input",
        SCIEX_INTACT_DIAGNOSTIC_SHEET,
        SCIEX_INTACT_PEAK_SHEET,
    ]
    assert included_sheet_names(sciex_names, AuditPolicy.from_level("standard"))[0] == []
    assert included_sheet_names(sciex_names, AuditPolicy.from_level("audit"))[0] == [
        SCIEX_INTACT_DIAGNOSTIC_SHEET,
        SCIEX_INTACT_PEAK_SHEET,
    ]
    assert included_sheet_names(sciex_names, AuditPolicy.from_level("full"))[0] == sciex_names


def test_standard_formal_workbook_is_unchanged_by_optional_comparison(tmp_path):
    without = write_comparison(tmp_path / "without", None, level="standard", optional={})
    with_result = write_comparison(tmp_path / "with", comparison(), level="standard")
    assert sheet_names(without) == sheet_names(with_result)
    for name in ("Known_Modification_Candidates", "Known_Modification_Summary"):
        pd.testing.assert_frame_equal(read_sheet(without, name), read_sheet(with_result, name))


def test_comparison_sheet_content_is_deterministic_across_workbooks(tmp_path):
    result = comparison()
    first = write_comparison(tmp_path / "first", result)
    second = write_comparison(tmp_path / "second", result)
    for name in (SCIEX_MASS_COMPARISON_SUMMARY_SHEET, SCIEX_MASS_COMPARISON_DETAIL_SHEET):
        pd.testing.assert_frame_equal(read_sheet(first, name), read_sheet(second, name))