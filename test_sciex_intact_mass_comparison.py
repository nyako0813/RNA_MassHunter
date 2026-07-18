from dataclasses import dataclass
from types import SimpleNamespace

import pandas as pd
import pytest

import main as main_module
from rna_masshunter.audit_policy import AUDIT_DETAIL, AUDIT_SUMMARY, AuditPolicy, included_sheet_names, sheet_category
from rna_masshunter.config import validate_config
from rna_masshunter.excel_report import (
    SCIEX_INTACT_OPTIONAL_RESULT_KEY, SCIEX_MASS_COMPARISON_DETAIL_SHEET,
    SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY, SCIEX_MASS_COMPARISON_SUMMARY_SHEET,
    _sciex_mass_comparison_excel_sheets, write_excel_report,
)
from rna_masshunter.models import IntactMassCandidate, RunConfig
from rna_masshunter.sciex_intact_mass_comparison import (
    DETAIL_COLUMNS, SUMMARY_COLUMNS, compare_sciex_intact_masses,
)

class Detection:
    def __init__(self, peaks, status="DETECTION_COMPLETED", input_status="SUPPORTED_INPUT", **diagnostics):
        self._peaks=[dict(x) for x in peaks]
        self._diagnostics={"Detection_Status":status,"Input_Status":input_status,**diagnostics}
    def peak_rows(self): return [dict(x) for x in self._peaks]
    def diagnostics_row(self): return dict(self._diagnostics)

def peak(mass, intensity=10, peak_id="P", **extra):
    return {"Peak_ID":peak_id,"Apex_Mass":mass,"Apex_Intensity_Raw":intensity,"Detection_Tier":"STRICT","Prominence":5.0,"Half_Prominence_Width_Da":2.0,**extra}

def compare(masses, theory=100.0, existing=(), **kwargs):
    return compare_sciex_intact_masses(Detection([peak(m, peak_id=f"P{i}") for i,m in enumerate(masses)]),theory,existing,**kwargs)

@pytest.mark.parametrize("mass,status",[(99.0,"STRICT_MATCH"),(101.0,"STRICT_MATCH"),(95.0,"BROAD_MATCH"),(105.0,"BROAD_MATCH"),(105.0001,"NO_MATCH")])
def test_status_boundaries_are_inclusive(mass,status):
    assert compare([mass]).details()[0]["Comparison_Status"]==status

@pytest.mark.parametrize("mass,sign",[(99.0,-1),(100.0,0),(101.0,1)])
def test_delta_and_ppm_sign_follow_observed_minus_theoretical(mass,sign):
    row=compare([mass]).details()[0]
    assert row["Delta_Mass"]==mass-100
    assert row["Absolute_Delta_Mass"]==abs(mass-100)
    assert (row["Delta_ppm"]>0)-(row["Delta_ppm"]<0)==sign

def test_nearest_existing_mass_and_signed_deltas_are_independent_of_theory_status():
    row=compare([104.0],existing=[102.0,110.0]).details()[0]
    assert row["Comparison_Status"]=="BROAD_MATCH"
    assert row["Nearest_Existing_Intact_Mass"]==102
    assert row["Nearest_Existing_Intact_Delta"]==2

def test_nearest_existing_tie_breaks_to_lower_mass():
    assert compare([105.0],existing=[110.0,100.0]).details()[0]["Nearest_Existing_Intact_Mass"]==100

def test_no_existing_result_retains_theory_comparison():
    row=compare([100.2]).details()[0]
    assert row["Comparison_Status"]=="STRICT_MATCH"
    assert row["Existing_Intact_Comparison_Status"]=="NO_EXISTING_INTACT_RESULT"
    assert row["Nearest_Existing_Intact_Mass"] is None

def test_dataclass_and_mapping_existing_results_are_supported():
    candidate=IntactMassCandidate(99.0,1,[1],1,1.0)
    result=compare_sciex_intact_masses(Detection([peak(101.0)]),100,[candidate,{"Reconstructed_Mass":101.25}])
    assert result.details()[0]["Nearest_Existing_Intact_Mass"]==101.25
    assert result.summaries()[0]["Existing_Intact_Mass_Count"]==2

def test_peak_fields_are_preserved_and_formal_flags_are_false():
    result=compare_sciex_intact_masses(Detection([peak(100.0,Possible_Shoulder=True,Broad_Peak_Flag=True,Peak_Area_Raw=12.0)]),100)
    row=result.details()[0]
    assert row["Possible_Shoulder"] is True and row["Broad_Peak_Flag"] is True and row["Peak_Area_Raw"]==12
    assert row["Shadow_Only"] is True
    assert not row["Applied_To_Formal_Score"] and not row["Applied_To_Ranking"] and not row["Applied_To_Candidate_Filtering"]
    assert not row["Molecular_Identity_Assigned"] and not row["Modification_Lookup_Performed"]

def test_output_is_mass_sorted_and_input_is_not_mutated():
    source=[peak(102,peak_id="late"),peak(98,peak_id="early")]
    snapshot=[dict(x) for x in source]
    rows=compare_sciex_intact_masses(Detection(source),100).details()
    assert [x["Peak_ID"] for x in rows]==["early","late"] and source==snapshot

def test_closest_peak_tie_uses_intensity_then_mass():
    detection=Detection([peak(99,2,"low"),peak(101,5,"high")])
    assert compare_sciex_intact_masses(detection,100).summaries()[0]["Closest_Peak_ID"]=="high"

def test_strongest_peak_tie_uses_lower_mass():
    detection=Detection([peak(101,5,"highmass"),peak(99,5,"lowmass")])
    assert compare_sciex_intact_masses(detection,100).summaries()[0]["Strongest_Peak_ID"]=="lowmass"

def test_noncompleted_detector_is_not_eligible_when_called_directly():
    result=compare_sciex_intact_masses(Detection([peak(100)],status="INVALID_AXIS"),100)
    assert result.details()[0]["Comparison_Status"]=="NOT_ELIGIBLE"
    assert result.summaries()[0]["Comparison_Status"]=="NOT_ELIGIBLE"

def test_no_theoretical_mass_returns_explicit_noneligible_rows_and_summary():
    result=compare([100],theory=None)
    assert result.details()[0]["Comparison_Status"]=="NO_THEORETICAL_MASS"
    assert result.details()[0]["Comparison_Eligible"] is False
    assert result.summaries()[0]["Comparison_Status"]=="NO_THEORETICAL_MASS"

@pytest.mark.parametrize("strict,broad,message",[(0,5,"strict"),(-1,5,"strict"),(1,.5,"broad"),(1,float("nan"),"broad")])
def test_invalid_tolerances_fail(strict,broad,message):
    with pytest.raises(ValueError,match=message): compare([100],strict_tolerance_da=strict,broad_tolerance_da=broad)

def config(enabled=True,comparison=True,strict=1,broad=5):
    return RunConfig(sciex_profile={"enabled":enabled,"path":"x","intact_peak_detection":{"enabled":True},"intact_mass_comparison":{"enabled":comparison,"strict_tolerance_da":strict,"broad_tolerance_da":broad}})

def wrapper(status="DETECTION_COMPLETED",peaks=None):
    return {SCIEX_INTACT_OPTIONAL_RESULT_KEY:{"result":Detection([peak(100)] if peaks is None else peaks,status),"source_file":"x"}}

@pytest.mark.parametrize("comparison,status,peaks",[(False,"DETECTION_COMPLETED",None),(True,"DETECTION_COMPLETED_WITH_WARNINGS",None),(True,"INVALID_AXIS",None),(True,"DETECTION_COMPLETED",[])])
def test_routing_skips_disabled_incomplete_or_zero_peak_inputs(comparison,status,peaks):
    assert main_module.build_sciex_intact_mass_comparison_optional_results(config(comparison=comparison),wrapper(status,peaks),100,[],[])=={}

def test_routing_builds_internal_result_and_accepts_missing_existing_results():
    result=main_module.build_sciex_intact_mass_comparison_optional_results(config(),wrapper(),100,[],[])
    assert SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY in result
    assert result[SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY].details()[0]["Existing_Intact_Comparison_Status"]=="NO_EXISTING_INTACT_RESULT"

def test_comparison_exception_isolated_and_detector_wrapper_retained(monkeypatch):
    source=wrapper(); warnings=[]
    monkeypatch.setattr(main_module,"compare_sciex_intact_masses",lambda *_a,**_k: (_ for _ in ()).throw(RuntimeError("boom")))
    assert main_module.build_sciex_intact_mass_comparison_optional_results(config(),source,100,[],warnings)=={}
    assert SCIEX_INTACT_OPTIONAL_RESULT_KEY in source
    assert warnings[-1]["Source"]=="sciex_intact_mass_comparison"

@pytest.mark.parametrize("strict,broad",[(0,5),(1,.5)])
def test_config_rejects_invalid_comparison_tolerances(strict,broad):
    with pytest.raises(ValueError): validate_config(config(strict=strict,broad=broad))

def test_disabled_sciex_config_does_not_validate_new_comparison_or_warn():
    cfg=RunConfig(sciex_profile={"enabled":False,"path":None,"intact_peak_detection":{"enabled":True},"intact_mass_comparison":"future"})
    warnings=[];validate_config(cfg,warnings)
    assert not any("sciex" in x["Message"] for x in warnings)

@pytest.mark.parametrize("level,present",[("standard",False),("audit",True),("full",True)])
def test_sheet_policy_is_standard_hidden_and_audit_full_visible(level,present):
    names=[SCIEX_MASS_COMPARISON_SUMMARY_SHEET,SCIEX_MASS_COMPARISON_DETAIL_SHEET]
    assert (included_sheet_names(names,AuditPolicy.from_level(level))[0]==names) is present


def writer_config():
    return SimpleNamespace(analysis={"mode":"full"},project={"name":"comparison-excel"},input={},organism={},sequence={},experiment={},instrument={},sciex_profile={},reconstruction={"enabled":False},digestion={"enabled":False},alkaline_phosphatase={},fragment_mapping={},modification_search={},peak_filtering={},p1_annotation={},ms2_annotation={},modification_evidence_ranking={},biological_context={},performance={},reporting={"max_excel_rows_per_sheet":1000,"truncate_large_sheets":True})

@pytest.mark.parametrize("level,expected",[("standard",False),("audit",True),("full",True)])
def test_real_excel_writer_routes_comparison_sheets_by_policy(tmp_path,level,expected):
    result=compare([100])
    report=write_excel_report(tmp_path/level,writer_config(),{},[],[],[],optional_results={SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY:result},audit_policy=AuditPolicy.from_level(level))
    from openpyxl import load_workbook
    workbook=load_workbook(report,read_only=True,data_only=True)
    try: names=workbook.sheetnames
    finally: workbook.close()
    assert (SCIEX_MASS_COMPARISON_DETAIL_SHEET in names) is expected
    assert (SCIEX_MASS_COMPARISON_SUMMARY_SHEET in names) is expected
    if expected:
        detail=pd.read_excel(report,sheet_name=SCIEX_MASS_COMPARISON_DETAIL_SHEET,header=2)
        assert len(detail)==1 and list(detail.columns)==DETAIL_COLUMNS

def test_sheet_categories_and_excel_columns_are_fixed():
    assert sheet_category(SCIEX_MASS_COMPARISON_SUMMARY_SHEET)==AUDIT_SUMMARY
    assert sheet_category(SCIEX_MASS_COMPARISON_DETAIL_SHEET)==AUDIT_DETAIL
    sheets=_sciex_mass_comparison_excel_sheets(compare([100]))
    assert list(sheets[SCIEX_MASS_COMPARISON_DETAIL_SHEET].columns)==DETAIL_COLUMNS
    assert list(sheets[SCIEX_MASS_COMPARISON_SUMMARY_SHEET].columns)==SUMMARY_COLUMNS
    assert len(SCIEX_MASS_COMPARISON_SUMMARY_SHEET)<=31


TERMINAL_SEQUENCE = "ACGUCCA"
TERMINAL_IDS = [
    "TERM_LINEAR_5P_3OH",
    "TERM_LINEAR_5OH_3OH",
    "TERM_LINEAR_5OH_3P",
    "TERM_LINEAR_5OH_3CYCLIC",
]


def terminal_compare(peaks, *, identity_status="CONFIRMED", sequence=TERMINAL_SEQUENCE, **kwargs):
    return compare_sciex_intact_masses(
        Detection(peaks),
        1.0,
        sequence=sequence,
        sequence_source="synthetic-unit-test",
        identity_status=identity_status,
        **kwargs,
    )


def candidate_rows(result, peak_id="P"):
    return [row for row in result.details() if row["Peak_ID"] == peak_id]


def reference_mass():
    seed = terminal_compare([peak(1.0, Centroid_Mass=1.0)])
    row = next(row for row in seed.details() if row["Candidate_ID"] == "TERM_LINEAR_5P_3OH")
    return row["Theoretical_Mass"]


def test_explicit_sequence_uses_terminal_model_not_legacy_theoretical_mass():
    result = terminal_compare([peak(reference_mass(), Centroid_Mass=reference_mass())])
    rows = result.details()
    assert len(rows) == 4
    assert all(row["Candidate_Category"] == "UNMODIFIED_TERMINAL_STATE" for row in rows)
    assert all(row["Theoretical_Mass"] != 1.0 for row in rows)
    summary = result.summaries()[0]
    assert summary["Sequence_Source"] == "synthetic-unit-test"
    assert summary["Sequence_Length"] == len(TERMINAL_SEQUENCE)
    assert len(summary["Sequence_SHA256"]) == 64
    assert summary["Ends_With_CCA"] is True
    assert summary["CCA_Policy"] == "AS_PROVIDED"
    assert set(DETAIL_COLUMNS).issubset(rows[0])
    assert set(SUMMARY_COLUMNS).issubset(result.summaries()[0])


def test_default_terminal_candidates_are_complete_and_deterministic():
    first = terminal_compare([peak(1.0)]).details()
    second = terminal_compare([peak(1.0)]).details()
    assert [row["Candidate_ID"] for row in first] == TERMINAL_IDS
    assert first == second
    states = [(row["Five_Prime_State"], row["Three_Prime_State"], row["Topology"]) for row in first]
    assert states == [
        ("MONOPHOSPHATE", "OH", "LINEAR"),
        ("OH", "OH", "LINEAR"),
        ("OH", "MONOPHOSPHATE", "LINEAR"),
        ("OH", "CYCLIC_PHOSPHATE", "LINEAR"),
    ]
    assert all(row["Theoretical_Formula"] for row in first)


def test_five_prime_and_three_prime_phosphate_are_one_mass_equivalent_group():
    rows = terminal_compare([peak(1.0)]).details()
    phosphate_rows = [rows[0], rows[2]]
    assert phosphate_rows[0]["Theoretical_Mass"] == phosphate_rows[1]["Theoretical_Mass"]
    assert phosphate_rows[0]["Mass_Equivalent_Candidate_Group_ID"] == phosphate_rows[1]["Mass_Equivalent_Candidate_Group_ID"]
    assert {row["Candidate_Ambiguity_Count"] for row in phosphate_rows} == {2}
    assert all(row["Terminal_State_Ambiguous"] for row in phosphate_rows)
    assert all(row["Mass_Equivalent_Candidate_IDs"] == "TERM_LINEAR_5OH_3P;TERM_LINEAR_5P_3OH" for row in phosphate_rows)
    assert terminal_compare([peak(1.0)]).summaries()[0]["Mass_Equivalent_Group_Count"] == 3


@pytest.mark.parametrize("offset,tolerance,status", [
    (0.0, "STRICT", "MATCH_STRICT_MASS_ONLY"),
    (1.0, "STRICT", "MATCH_STRICT_MASS_ONLY"),
    (-1.0, "STRICT", "MATCH_STRICT_MASS_ONLY"),
    (5.0, "EXPLORATORY", "MATCH_EXPLORATORY_MASS_ONLY"),
    (-5.0, "EXPLORATORY", "MATCH_EXPLORATORY_MASS_ONLY"),
    (5.0001, "NO_MATCH", "NO_MATCH"),
])
def test_terminal_match_boundaries_use_apex_mass(offset, tolerance, status):
    mass = reference_mass()
    result = terminal_compare([peak(mass + offset, Centroid_Mass=mass + 100.0)])
    row = next(row for row in result.details() if row["Candidate_ID"] == TERMINAL_IDS[0])
    assert row["Match_Tolerance_Class"] == tolerance
    assert row["Match_Status"] == status
    assert row["Apex_Delta_Da"] == pytest.approx(offset)
    assert row["Apex_Delta_ppm"] == pytest.approx(offset / mass * 1_000_000)
    assert row["Preferred_Mass_Field"] == "Apex_Mass"


def test_centroid_is_secondary_and_nan_does_not_block_apex_comparison():
    mass = reference_mass()
    diagnostic = terminal_compare([peak(mass, Centroid_Mass=mass + 2.5)]).details()[0]
    assert diagnostic["Apex_Delta_Da"] == pytest.approx(0.0)
    assert diagnostic["Centroid_Delta_Da"] == pytest.approx(2.5)
    nan_row = terminal_compare([peak(mass, Centroid_Mass=float("nan"))]).details()[0]
    assert nan_row["Match_Tolerance_Class"] == "STRICT"
    assert nan_row["Centroid_Delta_Da"] is None
    assert nan_row["Centroid_Complete"] is False


def test_identity_gate_confirmed_ambiguous_unknown_and_conflict():
    confirmed = terminal_compare([peak(1.0)], identity_status="CONFIRMED")
    assert confirmed.summaries()[0]["Identity_Gate_Passed"] is True
    assert confirmed.summaries()[0]["Terminal_Candidate_Count"] == 4
    for status, warning in [("AMBIGUOUS", "IDENTITY_ASSUMPTION_DEPENDENT"), ("UNKNOWN", "IDENTITY_UNKNOWN")]:
        result = terminal_compare([peak(1.0)], identity_status=status)
        assert len(result.details()) == 4
        assert result.summaries()[0]["Identity_Gate_Passed"] is False
        assert warning in result.summaries()[0]["Identity_Warnings"]
    conflict = terminal_compare([peak(reference_mass())], identity_status="CONFLICT")
    summary = conflict.summaries()[0]
    assert conflict.details() == []
    assert summary["Comparison_Status"] == "IDENTITY_CONFLICT"
    assert summary["Terminal_Candidate_Count"] == 0
    assert summary["Closest_Observed_Mass"] is None


def test_filename_match_is_not_treated_as_confirmed_sequence_identity():
    audit = {"Audit_Status":"MATCH", "Configured_Sequence":TERMINAL_SEQUENCE}
    result = compare_sciex_intact_masses(Detection([peak(1.0)]), 1.0, input_identity_audit=audit)
    summary = result.summaries()[0]
    assert summary["Identity_Gate_Status"] == "AMBIGUOUS"
    assert summary["Identity_Gate_Passed"] is False
    assert "FILENAME_IDENTITY_DOES_NOT_CONFIRM_SEQUENCE_IDENTITY" in summary["Identity_Warnings"]


@pytest.mark.parametrize("extra,expected", [
    ({"Detection_Tier":"STRICT", "Centroid_Mass":1.0}, "PRIMARY"),
    ({"Detection_Tier":"SENSITIVE", "Centroid_Mass":1.0}, "EXPLORATORY"),
    ({"Centroid_Mass":1.0, "Broad_Peak_Flag":True}, "MATCHABLE_WITH_SHAPE_WARNING"),
    ({"Centroid_Mass":1.0, "Possible_Shoulder":True}, "MATCHABLE_WITH_SHAPE_WARNING"),
    ({"Centroid_Mass":1.0, "Edge_Peak_Flag":True}, "NOT_PRIMARY_ELIGIBLE"),
    ({"Centroid_Mass":1.0, "Peak_Area_Complete":False}, "NOT_PRIMARY_ELIGIBLE"),
])
def test_peak_eligibility_preserves_every_peak_candidate_row(extra, expected):
    values = peak(1.0, **extra)
    rows = terminal_compare([values]).details()
    assert len(rows) == 4
    assert {row["Peak_Eligibility"] for row in rows} == {expected}


def test_false_certainty_mass_type_calibration_and_formal_flags_are_fixed():
    row = terminal_compare([peak(reference_mass(), Centroid_Mass=reference_mass())]).details()[0]
    assert row["Mass_Match_Only"] is True
    assert row["Observed_Mass_Type"] == "UNKNOWN"
    assert row["Theoretical_Mass_Type"] == "MONOISOTOPIC_NEUTRAL"
    assert row["Mass_Definition_Compatibility"] == "UNKNOWN"
    assert row["SCIEX_Reconstruction_Settings_Available"] is False
    assert row["Observed_Mass_Raw"] == row["Observed_Mass_Calibrated"]
    assert row["Calibration_Applied"] is False
    assert row["Calibration_Method"] == "NONE"
    assert row["Calibration_Offset_Da"] == 0.0
    for key in (
        "Structure_Identity_Assigned", "Position_Assigned", "Molecular_Identity_Assigned",
        "Modification_Assigned", "SCIEX_Intact_Mass_Matching_Applied_To_Formal_Score",
        "SCIEX_Intact_Mass_Matching_Applied_To_Ranking",
        "SCIEX_Intact_Mass_Matching_Applied_To_Candidate_Filtering",
        "SCIEX_Intact_Mass_Matching_Applied_To_Final_Consensus",
    ):
        assert row[key] is False


def test_existing_formal_inputs_and_detection_data_are_not_mutated():
    source_peak = peak(1.0, Centroid_Mass=1.0)
    formal = {"Reconstructed_Mass":2.0, "Applied_To_Formal_Score":True}
    peak_snapshot = dict(source_peak)
    formal_snapshot = dict(formal)
    first = terminal_compare([source_peak], existing_intact_results=[formal]).details()
    second = terminal_compare([source_peak], existing_intact_results=[formal]).details()
    assert first == second
    assert source_peak == peak_snapshot
    assert formal == formal_snapshot


def test_empty_invalid_mz_and_ineligible_inputs_are_explicit():
    empty = terminal_compare([])
    assert empty.details() == []
    assert empty.summaries()[0]["Comparison_Status"] == "NOT_ELIGIBLE"
    with pytest.raises(ValueError, match="invalid canonical RNA base"):
        terminal_compare([peak(1.0)], sequence="ACGX")
    mz = compare_sciex_intact_masses(
        Detection([peak(1.0)], Profile_Type="MZ_PROFILE"), 1.0,
        sequence=TERMINAL_SEQUENCE, sequence_source="synthetic", identity_status="CONFIRMED",
    )
    assert mz.details() == []
    assert mz.summaries()[0]["Comparison_Status"] == "NOT_ELIGIBLE"
    invalid_detector = compare_sciex_intact_masses(
        Detection([peak(1.0)], status="INVALID_AXIS"), 1.0,
        sequence=TERMINAL_SEQUENCE, sequence_source="synthetic", identity_status="CONFIRMED",
    )
    assert len(invalid_detector.details()) == 4
    assert {row["Comparison_Status"] for row in invalid_detector.details()} == {"NOT_ELIGIBLE"}
