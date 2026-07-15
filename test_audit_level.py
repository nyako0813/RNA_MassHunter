from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from main import parse_args, resolve_config_path
from rna_masshunter.audit_policy import (
    AUDIT_DETAIL, AUDIT_GROUP, AUDIT_LEVEL_DEFAULT, AUDIT_LEVELS, AUDIT_STATUS_COLUMNS,
    AUDIT_SUMMARY, FORMAL_CORE, AuditPolicy, append_audit_level_diagnostics,
    audit_status_row, included_sheet_names, sheet_category, unclassified_sheets,
)
from rna_masshunter.excel_report import AUDIT_TOP_SHADOW_COLUMNS, SHEET_DESCRIPTIONS, write_excel_report
from rna_masshunter.models import RunConfig


def policy(level): return AuditPolicy.from_level(level)


def config(output):
    return RunConfig(
        project={"name":"audit-test","output_dir":str(output)}, input={}, organism={}, sequence={},
        experiment={}, instrument={}, reconstruction={"enabled":False}, digestion={"enabled":False},
        alkaline_phosphatase={}, fragment_mapping={}, modification_search={}, peak_filtering={},
        p1_annotation={}, ms2_annotation={}, modification_evidence_ranking={}, biological_context={},
        performance={}, reporting={"max_excel_rows_per_sheet":1000,"truncate_large_sheets":True},
    )


def status_rows(p):
    return [audit_status_row("synthetic","MS1",p,p.run_shadow_audits,True,True,1.25,2.5)]


def make_report(tmp_path, level):
    p=policy(level); out=tmp_path/level; out.mkdir()
    diagnostics=append_audit_level_diagnostics([{"Formal_Diagnostic":"same"}],p,status_rows(p),3)
    top=pd.DataFrame([{
        "Review_Rank":1,"Review_Priority":"A_strong_review","Modification_ID":"m1",
        "Best_Final_Score":7.0,"Best_Final_Confidence":"High",
        "MS1_Truncation_Affected":True,"Tier_Top50_MS1_Support":4,
    }])
    optional={
        "Top_Modification_Candidates":top,
        "MS2_Unmatched_Ion_Diagnostics":diagnostics,
        "MS1_Truncation_Summary":[{"Audit_Mode":"shadow","Applied_To_Final_Score":False}],
        "MS1_Truncation_Audit":[{"Fragment_ID":"F1","Applied_To_Final_Score":False}],
        "MS1_Truncation_Detail":[{"Fragment_ID":"F1","Applied_To_Formal_Result":False}],
        "MS1_CrossFrag_Summary":[{"Applied_To_Formal_Result":False}],
        "MS1_CrossFrag_Ambiguity":[{"Applied_To_Formal_Result":False}],
        "MS1_CrossFrag_Detail":[{"Applied_To_Formal_Result":False}],
        "Audit_Status":status_rows(p),
        "PT_Cross_Run_Runs":[{"Run_ID":"r1","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "PT_Cross_Run_Summary":[{"Cross_Run_Candidate_Key":"k","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "PT_Cross_Run_Pairs":[{"Pair_Key":"p","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "PT_Cross_Run_Detail":[{"Run_ID":"r1","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "Mod_Hypothesis_Summary":[{"Hypothesis_ID":"h","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "Mod_Hypothesis_Cross_Run":[{"Hypothesis_ID":"h","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "Mod_Hypothesis_Invalid":[],
        "Mod_Hypothesis_Structure_Map":[{"Position_Hypothesis_ID":"h","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "Mod_Hypothesis_ID_Audit":[{"Hypothesis_ID":"h","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "Mod_Oxidation_Family":[{"Modification_Family_ID":"f","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "Mod_Hypothesis_Detail":[{"Hypothesis_ID":"h","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
        "Mod_Hypothesis_Alternatives":[{"Hypothesis_ID":"h","Applied_To_Formal_Result":False,"Formal_Change_Ready":False,"Formal_Result_Changed":False}],
    }
    return write_excel_report(
        output_dir=out,config=config(out),diagnostics={},intact_results=[],charge_state_peaks=[],
        warnings=[],modifications=[],rule_set={},pathways=[],theoretical_fragments=[],
        fragment_ms1_matches=[],known_modification_candidates=[],known_modification_summary=[],
        optional_results=optional,audit_policy=p,
    )


def read(path,sheet): return pd.read_excel(path,sheet_name=sheet,header=2,dtype=object)


def test_default_is_full():
    assert parse_args([]).audit_level=="full" and AUDIT_LEVEL_DEFAULT=="full"


@pytest.mark.parametrize("level",AUDIT_LEVELS)
def test_valid_levels(level): assert policy(level).level==level


def test_invalid_level_cli():
    with pytest.raises(SystemExit): parse_args(["--audit-level","invalid"])


def test_invalid_level_object():
    with pytest.raises(ValueError): policy("invalid")


def test_config_combination_and_argument_order(tmp_path):
    path=tmp_path/"alternate.yaml"
    a=parse_args(["--config",str(path),"--audit-level","audit"])
    b=parse_args(["--audit-level","audit","--config",str(path)])
    assert a==b and resolve_config_path(Path('/repo'),a.config)==path


def test_policy_flags():
    assert not policy("standard").run_shadow_audits
    assert policy("audit").include_summary and not policy("audit").include_detail
    assert policy("full").include_group_tables and policy("full").include_detail


def test_sheet_registry_categories_and_no_documented_gaps():
    assert sheet_category("Run_summary")==FORMAL_CORE
    assert sheet_category("MS1_Truncation_Summary")==AUDIT_SUMMARY
    assert sheet_category("MS1_Truncation_Audit")==AUDIT_GROUP
    assert sheet_category("MS1_Truncation_Detail")==AUDIT_DETAIL
    assert sheet_category("MS2_Modified_Precursor_Candidat") is not None
    assert sheet_category("MS2_Modification_Localization_E") is not None
    assert unclassified_sheets(SHEET_DESCRIPTIONS)==[]


def test_unknown_sheet_full_only():
    assert included_sheet_names(["Unknown"],policy("full"))==(["Unknown"],["Unknown"])
    assert included_sheet_names(["Unknown"],policy("standard"))==([],["Unknown"])


def test_sheet_names_fit_excel_limit(): assert all(len(name)<=31 for name in SHEET_DESCRIPTIONS)


def test_standard_formal_and_no_audit_sheets(tmp_path):
    path=make_report(tmp_path,"standard"); names=load_workbook(path,read_only=True).sheetnames
    assert "Run_summary" in names and "Fragment_MS1_matches" in names and "Top_Modification_Candidates" in names
    assert "MS1_Truncation_Summary" not in names and "MS1_Truncation_Detail" not in names and "Audit_Status" not in names


def test_audit_summary_without_group_or_detail(tmp_path):
    path=make_report(tmp_path,"audit"); names=load_workbook(path,read_only=True).sheetnames
    assert "MS1_Truncation_Summary" in names and "MS1_CrossFrag_Summary" in names and "Audit_Status" in names
    assert "MS1_Truncation_Audit" not in names and "MS1_Truncation_Detail" not in names


def test_full_has_summary_group_detail_and_status(tmp_path):
    path=make_report(tmp_path,"full"); names=load_workbook(path,read_only=True).sheetnames
    assert {"MS1_Truncation_Summary","MS1_Truncation_Audit","MS1_Truncation_Detail","Audit_Status"}<=set(names)


def test_cross_run_excel_sheets_follow_audit_level(tmp_path):
    names={level:set(load_workbook(make_report(tmp_path,level),read_only=True).sheetnames) for level in AUDIT_LEVELS}
    assert not any(name.startswith("PT_Cross_Run") for name in names["standard"])
    assert {"PT_Cross_Run_Runs","PT_Cross_Run_Summary","PT_Cross_Run_Pairs"} <= names["audit"]
    assert "PT_Cross_Run_Detail" not in names["audit"] and "PT_Cross_Run_Detail" in names["full"]

def test_modification_hypothesis_excel_sheets_follow_audit_level(tmp_path):
    names={level:set(load_workbook(make_report(tmp_path,level),read_only=True).sheetnames) for level in AUDIT_LEVELS}
    assert not any(name.startswith("Mod_Hypothesis") or name=="Mod_Oxidation_Family" for name in names["standard"])
    assert {"Mod_Hypothesis_Summary","Mod_Hypothesis_Cross_Run","Mod_Hypothesis_Invalid","Mod_Hypothesis_Structure_Map","Mod_Hypothesis_ID_Audit","Mod_Oxidation_Family"} <= names["audit"]
    assert "Mod_Hypothesis_Detail" not in names["audit"] and "Mod_Hypothesis_Alternatives" not in names["audit"]
    assert {"Mod_Hypothesis_Detail","Mod_Hypothesis_Alternatives"} <= names["full"]

def test_top_formal_columns_stable_and_standard_shadow_removed(tmp_path):
    paths={level:make_report(tmp_path,level) for level in AUDIT_LEVELS}
    frames={level:read(path,"Top_Modification_Candidates") for level,path in paths.items()}
    formal=["Review_Rank","Review_Priority","Modification_ID","Best_Final_Score","Best_Final_Confidence"]
    assert frames["standard"][formal].equals(frames["audit"][formal]) and frames["audit"][formal].equals(frames["full"][formal])
    assert not set(AUDIT_TOP_SHADOW_COLUMNS)&set(frames["standard"].columns)
    assert "MS1_Truncation_Affected" in frames["audit"].columns and "MS1_Truncation_Affected" in frames["full"].columns


def test_formal_fragment_and_known_sheets_identical(tmp_path):
    paths={level:make_report(tmp_path,level) for level in AUDIT_LEVELS}
    for sheet in ["Fragment_MS1_matches","Fragment_MS1_filtered","Fragment_MS1_summary","Known_Modification_Candidates","Known_Modification_Summary"]:
        values=[read(paths[level],sheet) for level in AUDIT_LEVELS]
        assert values[0].equals(values[1]) and values[1].equals(values[2])


def test_run_summary_metadata(tmp_path):
    for level in AUDIT_LEVELS:
        frame=read(make_report(tmp_path,level),"Run_summary"); values=dict(zip(frame.Item,frame.Value))
        assert values["Audit_Level"]==level and values["Audit_Level_Default"]=="full"
        assert values["Formal_Result_Changed_By_Audit_Level"] in {False,0}


def test_diagnostics_not_run_is_not_zero(tmp_path):
    frame=read(make_report(tmp_path,"standard"),"MS2_Unmatched_Ion_Diagnostics")
    assert frame.loc[0,"Audit_Level"]=="standard" and frame.loc[0,"Shadow_Audit_Runtime_Seconds"]=="not_run"
    assert frame.loc[0,"Shadow_Audit_Peak_Memory_MB"]=="not_run"
    assert frame.loc[0,"Formal_Result_Changed_By_Audit_Level"] in {False,0}


def test_audit_status_columns_and_nonapplication(tmp_path):
    frame=read(make_report(tmp_path,"audit"),"Audit_Status")
    assert list(frame.columns)==AUDIT_STATUS_COLUMNS and not bool(frame.loc[0,"Applied_To_Formal_Result"])


def test_report_does_not_change_selected_config_file(tmp_path):
    selected=tmp_path/"alternate.yaml"; selected.write_text("sequence: test\n",encoding="utf-8"); before=selected.read_bytes()
    make_report(tmp_path,"standard")
    assert selected.read_bytes()==before
