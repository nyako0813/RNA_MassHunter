from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from rna_masshunter.config import validate_config
from rna_masshunter.excel_report import write_excel_report
from rna_masshunter.intact_reconstruction import (
    _apply_split_envelope_merge,
    apply_assignment_dry_run,
    run_assignment_sensitivity,
    build_assignment_candidate_audit_rows,
    build_assignment_sensitivity_rows,
    build_assignment_stability_rows,
    build_intact_competition_group_rows,
    build_intact_competition_score_rows,
    build_intact_engine_comparison_rows,
    build_intact_reconstruction_qc,
    build_reconstructed_mass_spectrum_rows,
    reconstruct_intact_masses,
)
from rna_masshunter.masses import mz_from_neutral_mass
from rna_masshunter.models import IntactMassCandidate, Peak, PeakTierResult, RunConfig


UNMODIFIED_MASS = 25082.3
MODIFIED_MASS = 25325.5
ALT_MODIFIED_MASS = 25342.0
LOW_FULL_LENGTH_MASS = 25082.0
FRAGMENT_LIKE_MASS = 4818.0

BASE_QC_CONFIG = {
    "intact_reconstruction": {
        "min_charge_states_for_reliable": 3,
        "min_charge_states_for_review": 2,
        "require_contiguous_charge_states": True,
        "max_neutral_mass_sd_da": 0.5,
        "max_neutral_mass_range_da": 1.5,
        "max_mass_error_ppm": 20,
        "max_envelope_internal_error_ppm": 20,
        "min_relative_envelope_intensity_percent_for_reliable": 1.0,
        "min_relative_envelope_intensity_percent_for_review": 0.1,
        "max_competing_envelopes": 3,
        "comparison_ready_statuses": ["Reliable", "Review"],
        "max_rt_range_min_for_reliable": 0.15,
        "max_rt_range_min_for_review": 0.30,
        "allow_trace_only_reliable": False,
        "search_mode": "untargeted",
        "reference_mass_tolerance_ppm": 20,
    }
}


def _candidate(cluster_id, charges, mass=1000.0, intensity=10000.0, theoretical_mass=1000.0):
    delta = mass - theoretical_mass if theoretical_mass is not None else None
    ppm = delta / theoretical_mass * 1_000_000 if theoretical_mass else None
    return IntactMassCandidate(
        observed_mass=mass,
        charge_state_count=len(charges),
        charge_states=charges,
        supporting_peak_count=len(charges),
        total_intensity=intensity,
        theoretical_mass=theoretical_mass,
        mass_error_da=delta,
        mass_error_ppm=ppm,
        cluster_id=cluster_id,
    )


def _peaks(cluster_id, charges, masses=None, rts=None, intensity=1000.0, tier="Major"):
    masses = masses or [1000.0 for _ in charges]
    rts = rts if rts is not None else [5.0 + index * 0.02 for index, _ in enumerate(charges)]
    return [
        {
            "Cluster_ID": cluster_id,
            "Charge": charge,
            "Neutral_Mass": mass,
            "Intensity": intensity,
            "RT": rt,
            "Peak_Tier": tier,
        }
        for charge, mass, rt in zip(charges, masses, rts)
    ]


def _qc(candidates, peaks, config=None, enabled=True):
    rows, diagnostics = build_intact_reconstruction_qc(candidates, peaks, config or BASE_QC_CONFIG, reconstruction_enabled=enabled)
    return rows, diagnostics[0]


def test_modified_envelope_far_from_unmodified_theory_can_be_reliable():
    candidate = _candidate("C1", [10, 11, 12], mass=MODIFIED_MASS, intensity=50000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("C1", [10, 11, 12], masses=[25325.45, 25325.50, 25325.55], intensity=10000)
    rows, _ = _qc([candidate], peaks)
    row = rows[0]
    assert row["Reconstruction_Status"] == "Reliable"
    assert row["Comparison_Ready_Strict"] is True
    assert row["Unmodified_Theory_Delta_Da"] == candidate.mass_error_da
    assert round(row["Unmodified_Theory_Delta_Da"], 1) == 243.2
    assert abs(row["Unmodified_Theory_Delta_ppm"]) > 9000


def test_large_unmodified_delta_alone_does_not_lower_reliable_status():
    candidate = _candidate("C2", [20, 21, 22], mass=MODIFIED_MASS, intensity=40000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("C2", [20, 21, 22], masses=[25325.49, 25325.50, 25325.51], intensity=9000)
    rows, _ = _qc([candidate], peaks)
    assert rows[0]["Reconstruction_Status"] == "Reliable"
    assert "mass_error_too_large" not in rows[0]["Limiting_Factors"]


def test_envelope_internal_mass_error_prevents_reliable():
    candidate = _candidate("C3", [10, 11, 12], mass=1000.0, intensity=30000)
    peaks = _peaks("C3", [10, 11, 12], masses=[1000.0, 1001.0, 1002.0], intensity=8000)
    rows, _ = _qc([candidate], peaks)
    assert rows[0]["Reconstruction_Status"] == "Review"
    assert "internal_mass_error_too_large" in rows[0]["Limiting_Factors"]
    assert rows[0]["Comparison_Ready"] is False


def test_same_rt_three_contiguous_charge_states_are_reliable():
    candidate = _candidate("C4", [7, 8, 9], mass=1500.0, intensity=25000)
    peaks = _peaks("C4", [7, 8, 9], masses=[1499.99, 1500.0, 1500.01], rts=[3.00, 3.04, 3.08], intensity=7000)
    rows, _ = _qc([candidate], peaks)
    assert rows[0]["Charge_State_Continuity"] == "contiguous"
    assert rows[0]["RT_Consistency"] == "consistent"
    assert rows[0]["Reconstruction_Status"] == "Reliable"


def test_rt_range_too_large_adds_rt_inconsistent():
    candidate = _candidate("C5", [10, 11, 12], mass=1000.0, intensity=30000)
    peaks = _peaks("C5", [10, 11, 12], masses=[999.99, 1000.0, 1000.01], rts=[1.0, 1.3, 1.7], intensity=9000)
    rows, _ = _qc([candidate], peaks)
    assert rows[0]["RT_Consistency"] == "inconsistent"
    assert "rt_inconsistent" in rows[0]["Limiting_Factors"]
    assert rows[0]["Reconstruction_Status"] == "Review"


def test_trace_only_envelope_is_not_reliable_by_default():
    candidate = _candidate("C6", [10, 11, 12], mass=1000.0, intensity=30000)
    peaks = _peaks("C6", [10, 11, 12], masses=[999.99, 1000.0, 1000.01], intensity=9000, tier="Trace")
    rows, _ = _qc([candidate], peaks)
    assert rows[0]["Trace_Only_Envelope"] is True
    assert rows[0]["Reconstruction_Status"] == "Review"
    assert "trace_only_envelope" in rows[0]["Limiting_Factors"]


def test_highest_intensity_envelope_is_dominant():
    low = _candidate("LOW", [10, 11, 12], mass=1000.0, intensity=10000)
    high = _candidate("HIGH", [13, 14, 15], mass=2000.0, intensity=90000)
    peaks = _peaks("LOW", [10, 11, 12], masses=[999.99, 1000.0, 1000.01], intensity=3000) + _peaks(
        "HIGH", [13, 14, 15], masses=[1999.99, 2000.0, 2000.01], intensity=30000
    )
    _, diag = _qc([low, high], peaks)
    assert diag["Dominant_Envelope_Mass"] == 2000.0
    assert diag["Dominant_Envelope_Intensity"] == 90000


def test_comparison_ready_review_allows_minor_noncontiguous_case():
    candidate = _candidate("C7", [10, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("C7", [10, 12], masses=[25325.49, 25325.51], rts=[2.0, 2.1], intensity=10000)
    rows, _ = _qc([candidate], peaks)
    assert rows[0]["Reconstruction_Status"] == "Review"
    assert rows[0]["Comparison_Ready_Strict"] is False
    assert rows[0]["Comparison_Ready_Review"] is True
    assert rows[0]["Comparison_Ready"] is True
    assert rows[0]["In_Neutral_Mass_Search_Range"] is True


def test_multiple_limiting_factors_are_preserved():
    candidate = _candidate("C8", [10, 12, 14], mass=1000.0, intensity=30000)
    peaks = _peaks("C8", [10, 12, 14], masses=[1000.0, 1002.0, 1004.0], rts=[1.0, 1.4, 1.8], intensity=8000, tier="Trace")
    rows, _ = _qc([candidate], peaks)
    factors = rows[0]["Limiting_Factors"]
    assert "non_contiguous_charge_states" in factors
    assert "mass_spread_too_large" in factors
    assert "rt_inconsistent" in factors
    assert "trace_only_envelope" in factors
    assert rows[0]["Num_Limiting_Factors"] >= 4


def test_reference_mass_match_is_calculated():
    config = {"intact_reconstruction": {**BASE_QC_CONFIG["intact_reconstruction"], "reference_masses": [{"label": "SCiex_major_25325", "mass_da": 25325.5}]}}
    candidate = _candidate("C9", [10, 11, 12], mass=25325.52, intensity=40000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("C9", [10, 11, 12], masses=[25325.51, 25325.52, 25325.53], intensity=10000)
    rows, diag = _qc([candidate], peaks, config=config)
    assert rows[0]["Best_Reference_Label"] == "SCiex_major_25325"
    assert abs(rows[0]["Reference_Mass_Error_Da"] - 0.02) < 1e-6
    assert rows[0]["Reference_Mass_Matched"] is True
    assert diag["Reference_Match_Count"] == 1


def test_reference_not_configured_is_safe():
    candidate = _candidate("C10", [10, 11, 12], mass=1000.0, intensity=30000)
    peaks = _peaks("C10", [10, 11, 12], masses=[999.99, 1000.0, 1000.01], intensity=9000)
    rows, diag = _qc([candidate], peaks, config={"intact_reconstruction": {}})
    assert rows[0]["Best_Reference_Label"] == "not_configured"
    assert rows[0]["Reference_Mass_Matched"] is False
    assert diag["Reference_Masses_Used"] == "not_configured"


def test_python_defaults_work_without_new_qc_settings():
    candidate = _candidate("C11", [10, 11, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("C11", [10, 11, 12], masses=[25325.49, 25325.5, 25325.51], intensity=9000)
    rows, diag = _qc([candidate], peaks, config={"intact_reconstruction": {}})
    assert rows[0]["Reconstruction_Status"] == "Reliable"
    assert rows[0]["In_Neutral_Mass_Search_Range"] is True
    assert rows[0]["Neutral_Mass_Search_Min_Da"] == 20000.0
    assert rows[0]["Neutral_Mass_Search_Max_Da"] == 30000.0
    assert diag["Search_Mode"] == "untargeted"
    assert diag["Neutral_Mass_Search_Min_Da"] == 20000.0
    assert diag["Neutral_Mass_Search_Max_Da"] == 30000.0
    assert diag["Min_Relative_Envelope_Intensity_Percent_For_Reliable"] == 1.0


def test_fragment_like_4818_da_is_outside_default_neutral_mass_range():
    candidate = _candidate("FRAG", [10, 11, 12], mass=FRAGMENT_LIKE_MASS, intensity=50000)
    peaks = _peaks("FRAG", [10, 11, 12], masses=[4817.99, 4818.0, 4818.01], intensity=10000)
    rows, diag = _qc([candidate], peaks)
    assert rows[0]["In_Neutral_Mass_Search_Range"] is False
    assert rows[0]["Neutral_Mass_Range_Status"] == "outside_range"
    assert rows[0]["Comparison_Ready"] is False
    assert "outside_neutral_mass_search_range" in rows[0]["Limiting_Factors"]
    assert diag["Total_Candidates_Outside_Mass_Range"] == 1
    assert diag["Total_Candidates_In_Mass_Range"] == 0


def test_full_length_masses_are_inside_default_neutral_mass_range():
    candidates = [
        _candidate("M25082", [10, 11, 12], mass=LOW_FULL_LENGTH_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS),
        _candidate("M25325", [10, 11, 12], mass=MODIFIED_MASS, intensity=40000, theoretical_mass=UNMODIFIED_MASS),
        _candidate("M25342", [10, 11, 12], mass=ALT_MODIFIED_MASS, intensity=35000, theoretical_mass=UNMODIFIED_MASS),
    ]
    peaks = (
        _peaks("M25082", [10, 11, 12], masses=[25081.99, 25082.0, 25082.01], intensity=7000)
        + _peaks("M25325", [10, 11, 12], masses=[25325.49, 25325.5, 25325.51], intensity=8000)
        + _peaks("M25342", [10, 11, 12], masses=[25341.99, 25342.0, 25342.01], intensity=7500)
    )
    rows, diag = _qc(candidates, peaks)
    assert {row["Reconstructed_Mass"]: row["In_Neutral_Mass_Search_Range"] for row in rows} == {
        LOW_FULL_LENGTH_MASS: True,
        MODIFIED_MASS: True,
        ALT_MODIFIED_MASS: True,
    }
    assert diag["Total_Candidates_In_Mass_Range"] == 3
    assert diag["Total_Candidates_Outside_Mass_Range"] == 0


def test_neutral_mass_range_can_be_changed_and_changes_readiness():
    config = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "neutral_mass_range": {"enabled": True, "min_da": 24000, "max_da": 26000},
        }
    }
    in_candidate = _candidate("IN", [10, 11, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    out_candidate = _candidate("OUT", [10, 11, 12], mass=23000.0, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("IN", [10, 11, 12], masses=[25325.49, 25325.5, 25325.51], intensity=9000) + _peaks(
        "OUT", [10, 11, 12], masses=[22999.99, 23000.0, 23000.01], intensity=9000
    )
    rows, diag = _qc([in_candidate, out_candidate], peaks, config=config)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["IN"]["In_Neutral_Mass_Search_Range"] is True
    assert by_cluster["IN"]["Comparison_Ready"] is True
    assert by_cluster["OUT"]["In_Neutral_Mass_Search_Range"] is False
    assert by_cluster["OUT"]["Comparison_Ready"] is False
    assert diag["Neutral_Mass_Search_Min_Da"] == 24000.0
    assert diag["Neutral_Mass_Search_Max_Da"] == 26000.0
    assert diag["Total_Candidates_In_Mass_Range"] == 1
    assert diag["Total_Candidates_Outside_Mass_Range"] == 1


def test_neutral_mass_range_change_can_exclude_previous_default_candidate():
    default_candidate = _candidate("DEFAULT", [10, 11, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("DEFAULT", [10, 11, 12], masses=[25325.49, 25325.5, 25325.51], intensity=9000)
    default_rows, _ = _qc([default_candidate], peaks)
    assert default_rows[0]["Comparison_Ready"] is True

    narrowed = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "neutral_mass_range": {"enabled": True, "min_da": 26000, "max_da": 30000},
        }
    }
    narrowed_rows, narrowed_diag = _qc([default_candidate], peaks, config=narrowed)
    assert narrowed_rows[0]["In_Neutral_Mass_Search_Range"] is False
    assert narrowed_rows[0]["Comparison_Ready"] is False
    assert narrowed_diag["Total_Candidates_In_Mass_Range"] == 0


def test_dominant_envelope_in_search_range_ignores_out_of_range_overall_dominant():
    fragment = _candidate("FRAG_DOM", [10, 11, 12], mass=FRAGMENT_LIKE_MASS, intensity=90000)
    in_range_low = _candidate("IN_LOW", [10, 11, 12], mass=LOW_FULL_LENGTH_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    in_range_high = _candidate("IN_HIGH", [10, 11, 12], mass=MODIFIED_MASS, intensity=50000, theoretical_mass=UNMODIFIED_MASS)
    peaks = (
        _peaks("FRAG_DOM", [10, 11, 12], masses=[4817.99, 4818.0, 4818.01], intensity=30000)
        + _peaks("IN_LOW", [10, 11, 12], masses=[25081.99, 25082.0, 25082.01], intensity=9000)
        + _peaks("IN_HIGH", [10, 11, 12], masses=[25325.49, 25325.5, 25325.51], intensity=16000)
    )
    _, diag = _qc([fragment, in_range_low, in_range_high], peaks)
    assert diag["Dominant_Envelope_Overall_Mass"] == FRAGMENT_LIKE_MASS
    assert diag["Dominant_Envelope_Overall_Intensity"] == 90000
    assert diag["Dominant_Envelope_In_Mass_Range_Mass"] == MODIFIED_MASS
    assert diag["Dominant_Envelope_In_Mass_Range_Intensity"] == 50000


def test_bad_high_intensity_in_range_candidate_is_raw_dominant_not_intact_dominant():
    bad = _candidate("BAD_RAW", [10, 11, 12], mass=22286.0, intensity=90000, theoretical_mass=UNMODIFIED_MASS)
    good = _candidate("GOOD", [10, 11, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("BAD_RAW", [10, 11, 12], masses=[22284.0, 22286.0, 22288.0], rts=[1.0, 1.4, 1.8], intensity=30000) + _peaks(
        "GOOD", [10, 11, 12], masses=[25325.49, 25325.50, 25325.51], intensity=10000
    )
    rows, diag = _qc([bad, good], peaks)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert diag["Dominant_Envelope_In_Search_Range_Raw_Mass"] == 22286.0
    assert by_cluster["BAD_RAW"]["Envelope_QC_Eligible"] is False
    assert by_cluster["BAD_RAW"]["Comparison_Ready"] is False
    assert diag["Dominant_Intact_Eligible_Envelope_Mass"] == MODIFIED_MASS
    assert by_cluster["GOOD"]["Dominant_Intact_Envelope_Flag"] is True


def test_good_qc_candidate_can_be_intact_eligible_despite_lower_intensity():
    bad = _candidate("BAD", [10, 11, 12], mass=22286.0, intensity=90000, theoretical_mass=UNMODIFIED_MASS)
    good = _candidate("LOW_GOOD", [10, 11, 12], mass=LOW_FULL_LENGTH_MASS, intensity=15000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("BAD", [10, 11, 12], masses=[22284.0, 22286.0, 22288.0], rts=[1.0, 1.4, 1.8], intensity=30000) + _peaks(
        "LOW_GOOD", [10, 11, 12], masses=[25081.99, 25082.0, 25082.01], intensity=5000
    )
    rows, _ = _qc([bad, good], peaks)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["BAD"]["Intact_Strict_Eligible"] is False
    assert by_cluster["LOW_GOOD"]["Intact_Strict_Eligible"] is True
    assert by_cluster["LOW_GOOD"]["Comparison_Ready_Strict"] is True


def test_strict_candidate_is_preferred_over_review_candidate():
    strict = _candidate("STRICT", [10, 11, 12], mass=MODIFIED_MASS, intensity=20000, theoretical_mass=UNMODIFIED_MASS)
    review = _candidate("REVIEW", [10, 12], mass=LOW_FULL_LENGTH_MASS, intensity=80000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("STRICT", [10, 11, 12], masses=[25325.49, 25325.50, 25325.51], intensity=7000) + _peaks(
        "REVIEW", [10, 12], masses=[25081.99, 25082.01], rts=[3.0, 3.1], intensity=40000
    )
    rows, diag = _qc([strict, review], peaks)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["STRICT"]["Intact_Strict_Eligible"] is True
    assert by_cluster["REVIEW"]["Intact_Review_Eligible"] is True
    assert diag["Dominant_Intact_Eligible_Envelope_Mass"] == MODIFIED_MASS
    assert by_cluster["STRICT"]["Dominant_Intact_Envelope_Flag"] is True


def test_review_top_is_dominant_when_no_strict_candidate_exists():
    better_review = _candidate("REVIEW_A", [10, 12, 13], mass=MODIFIED_MASS, intensity=25000, theoretical_mass=UNMODIFIED_MASS)
    lower_review = _candidate("REVIEW_B", [10, 12], mass=LOW_FULL_LENGTH_MASS, intensity=50000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("REVIEW_A", [10, 12, 13], masses=[25325.49, 25325.50, 25325.51], rts=[2.0, 2.1, 2.2], intensity=8000) + _peaks(
        "REVIEW_B", [10, 12], masses=[25081.99, 25082.01], rts=[3.0, 3.1], intensity=25000
    )
    rows, diag = _qc([better_review, lower_review], peaks)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert diag["Intact_Strict_Eligible_Count"] == 0
    assert by_cluster["REVIEW_A"]["Intact_Review_Eligible"] is True
    assert by_cluster["REVIEW_B"]["Intact_Review_Eligible"] is True
    assert diag["Dominant_Intact_Review_Envelope_Mass"] == MODIFIED_MASS
    assert diag["Dominant_Intact_Eligible_Envelope_Mass"] == MODIFIED_MASS


def test_trace_only_alone_does_not_remove_review_eligibility():
    candidate = _candidate("TRACE_REVIEW", [10, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("TRACE_REVIEW", [10, 12], masses=[25325.49, 25325.51], rts=[2.0, 2.1], intensity=10000, tier="Trace")
    rows, _ = _qc([candidate], peaks)
    row = rows[0]
    assert row["Trace_Only_Envelope"] is True
    assert row["Intact_Review_Eligible"] is True
    assert row["Comparison_Ready_Review"] is True
    assert "trace_only_envelope" in row["Comparison_Readiness_Reason"] or row["Comparison_Readiness_Reason"] == "review"


def test_relative_intact_eligible_intensity_is_normalized_within_eligible_candidates():
    bad = _candidate("BAD", [10, 11, 12], mass=22286.0, intensity=90000, theoretical_mass=UNMODIFIED_MASS)
    eligible_max = _candidate("ELIGIBLE_MAX", [10, 11, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    eligible_half = _candidate("ELIGIBLE_HALF", [10, 12], mass=LOW_FULL_LENGTH_MASS, intensity=15000, theoretical_mass=UNMODIFIED_MASS)
    peaks = (
        _peaks("BAD", [10, 11, 12], masses=[22284.0, 22286.0, 22288.0], rts=[1.0, 1.4, 1.8], intensity=30000)
        + _peaks("ELIGIBLE_MAX", [10, 11, 12], masses=[25325.49, 25325.5, 25325.51], intensity=10000)
        + _peaks("ELIGIBLE_HALF", [10, 12], masses=[25081.99, 25082.01], intensity=7500)
    )
    rows, _ = _qc([bad, eligible_max, eligible_half], peaks)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["BAD"]["Relative_Intact_Eligible_Intensity_Percent"] == 0.0
    assert by_cluster["ELIGIBLE_MAX"]["Relative_Intact_Eligible_Intensity_Percent"] == 100.0
    assert by_cluster["ELIGIBLE_HALF"]["Relative_Intact_Eligible_Intensity_Percent"] == 50.0


def test_target_review_range_enabled_disabled_and_counts():
    candidates = [
        _candidate("M22286", [10, 11, 12], mass=22286.0, intensity=10000, theoretical_mass=UNMODIFIED_MASS),
        _candidate("M25082", [10, 11, 12], mass=LOW_FULL_LENGTH_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS),
        _candidate("M25325", [10, 11, 12], mass=MODIFIED_MASS, intensity=40000, theoretical_mass=UNMODIFIED_MASS),
        _candidate("M25343", [10, 11, 12], mass=25343.0, intensity=35000, theoretical_mass=UNMODIFIED_MASS),
    ]
    peaks = (
        _peaks("M22286", [10, 11, 12], masses=[22285.99, 22286.0, 22286.01], intensity=3000)
        + _peaks("M25082", [10, 11, 12], masses=[25081.99, 25082.0, 25082.01], intensity=7000)
        + _peaks("M25325", [10, 11, 12], masses=[25325.49, 25325.5, 25325.51], intensity=8000)
        + _peaks("M25343", [10, 11, 12], masses=[25342.99, 25343.0, 25343.01], intensity=7500)
    )
    default_rows, default_diag = _qc(candidates, peaks)
    assert {row["Target_Review_Mass_Range_Status"] for row in default_rows} == {"not_configured"}
    assert default_diag["Target_Review_Mass_Range_Settings"] == "disabled"

    config = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "target_review_mass_range": {"enabled": True, "min_da": 24000, "max_da": 26000},
        }
    }
    rows, diag = _qc(candidates, peaks, config=config)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["M22286"]["In_Target_Review_Mass_Range"] is False
    assert by_cluster["M22286"]["Target_Review_Mass_Range_Status"] == "outside_range"
    assert by_cluster["M25082"]["In_Target_Review_Mass_Range"] is True
    assert by_cluster["M25325"]["In_Target_Review_Mass_Range"] is True
    assert by_cluster["M25343"]["In_Target_Review_Mass_Range"] is True
    assert diag["Target_Review_Candidate_Count"] == 3


def test_sciex_reference_mass_error_for_25325_581():
    config = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "reference_masses": [
                {"label": "SCiex_25325_5", "mass_da": 25325.5},
                {"label": "SCiex_25342", "mass_da": 25342.0},
                {"label": "SCiex_25343", "mass_da": 25343.0},
            ],
            "reference_mass_tolerance_ppm": 20,
        }
    }
    candidate = _candidate("REF", [10, 11, 12], mass=25325.581, intensity=40000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("REF", [10, 11, 12], masses=[25325.571, 25325.581, 25325.591], intensity=10000)
    rows, _ = _qc([candidate], peaks, config=config)
    row = rows[0]
    assert row["Reference_Mass_Matched"] is True
    assert row["Best_Reference_Label"] == "SCiex_25325_5"
    assert round(row["Reference_Mass_Error_Da"], 3) == 0.081
    assert round(row["Reference_Mass_Error_ppm"], 1) == 3.2
    assert row["Reconstruction_Status"] == "Reliable"


def test_generic_logic_works_without_reference_masses_or_trna_specific_values():
    config = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "neutral_mass_range": {"enabled": True, "min_da": 15000, "max_da": 16000},
        }
    }
    good = _candidate("GENERIC_GOOD", [6, 7, 8], mass=15555.5, intensity=42000, theoretical_mass=12345.6)
    out = _candidate("GENERIC_OUT", [6, 7, 8], mass=17000.0, intensity=84000, theoretical_mass=12345.6)
    peaks = _peaks("GENERIC_GOOD", [6, 7, 8], masses=[15555.49, 15555.50, 15555.51], intensity=14000) + _peaks(
        "GENERIC_OUT", [6, 7, 8], masses=[16999.99, 17000.00, 17000.01], intensity=28000
    )
    rows, diag = _qc([good, out], peaks, config=config)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["GENERIC_GOOD"]["Best_Reference_Label"] == "not_configured"
    assert by_cluster["GENERIC_GOOD"]["Intact_Strict_Eligible"] is True
    assert by_cluster["GENERIC_GOOD"]["Comparison_Ready"] is True
    assert by_cluster["GENERIC_OUT"]["In_Neutral_Mass_Search_Range"] is False
    assert by_cluster["GENERIC_OUT"]["Comparison_Ready"] is False
    assert diag["Dominant_Intact_Eligible_Envelope_Mass"] == 15555.5


def test_reference_annotations_do_not_change_qc_score_rank_or_dominant_selection():
    candidates_without_ref = [
        _candidate("A", [6, 7, 8], mass=15555.5, intensity=40000, theoretical_mass=12345.6),
        _candidate("B", [6, 8], mass=15620.0, intensity=50000, theoretical_mass=12345.6),
    ]
    peaks = _peaks("A", [6, 7, 8], masses=[15555.49, 15555.50, 15555.51], intensity=13000) + _peaks(
        "B", [6, 8], masses=[15619.99, 15620.01], rts=[3.0, 3.1], intensity=25000
    )
    base_config = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "neutral_mass_range": {"enabled": True, "min_da": 15000, "max_da": 16000},
        }
    }
    ref_config = {
        "intact_reconstruction": {
            **base_config["intact_reconstruction"],
            "reference_masses": [{"label": "external_check_A", "mass_da": 15555.5}],
            "reference_mass_tolerance_ppm": 20,
        }
    }
    rows_no_ref, diag_no_ref = _qc(candidates_without_ref, peaks, config=base_config)
    candidates_with_ref = [
        _candidate("A", [6, 7, 8], mass=15555.5, intensity=40000, theoretical_mass=12345.6),
        _candidate("B", [6, 8], mass=15620.0, intensity=50000, theoretical_mass=12345.6),
    ]
    rows_with_ref, diag_with_ref = _qc(candidates_with_ref, peaks, config=ref_config)
    no_ref = {row["Cluster_ID"]: row for row in rows_no_ref}
    with_ref = {row["Cluster_ID"]: row for row in rows_with_ref}
    for cluster_id in ["A", "B"]:
        for field in [
            "Reconstruction_Status",
            "Envelope_QC_Eligible",
            "Intact_Strict_Eligible",
            "Intact_Review_Eligible",
            "Comparison_Ready",
            "Intact_Envelope_QC_Score",
            "Intact_Envelope_QC_Rank",
            "Dominant_Intact_Envelope_Flag",
        ]:
            assert with_ref[cluster_id][field] == no_ref[cluster_id][field]
    assert diag_with_ref["Dominant_Intact_Eligible_Envelope_Mass"] == diag_no_ref["Dominant_Intact_Eligible_Envelope_Mass"]
    assert with_ref["A"]["Reference_Mass_Matched"] is True
    assert no_ref["A"]["Reference_Mass_Matched"] is False


def test_target_review_range_enabled_without_bounds_is_not_configured_and_non_filtering():
    config = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "neutral_mass_range": {"enabled": True, "min_da": 15000, "max_da": 16000},
            "target_review_mass_range": {"enabled": True},
        }
    }
    candidate = _candidate("NO_BOUNDS", [6, 7, 8], mass=15555.5, intensity=40000, theoretical_mass=12345.6)
    peaks = _peaks("NO_BOUNDS", [6, 7, 8], masses=[15555.49, 15555.50, 15555.51], intensity=13000)
    rows, diag = _qc([candidate], peaks, config=config)
    assert rows[0]["Target_Review_Mass_Range_Status"] == "not_configured"
    assert rows[0]["Comparison_Ready"] is True
    assert diag["Target_Review_Mass_Range_Settings"] == "not_configured"
    assert diag["Target_Review_Candidate_Count"] == 0


def _generic_config(**intact_overrides):
    intact = {
        **BASE_QC_CONFIG["intact_reconstruction"],
        "neutral_mass_range": {"enabled": True, "min_da": 10000, "max_da": 20000},
    }
    intact.update(intact_overrides)
    return {"intact_reconstruction": intact}


def test_exact_duplicate_peak_set_forms_group_and_one_representative():
    candidates = [
        _candidate("DUP", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0),
        _candidate("DUP", [6, 7, 8], mass=15000.1, intensity=35000, theoretical_mass=12000.0),
    ]
    peaks = _peaks("DUP", [6, 7, 8], masses=[14999.99, 15000.0, 15000.01], intensity=10000)
    rows, diag = _qc(candidates, peaks, config=_generic_config())
    assert {row["Exact_Duplicate_Group_ID"] for row in rows} == {"ED00001"}
    assert all(row["Exact_Duplicate_Count"] == 2 for row in rows)
    assert sum(1 for row in rows if row["Is_Exact_Duplicate_Representative"]) == 1
    assert len({row["Intact_Envelope_Group_ID"] for row in rows}) == 1
    assert sum(1 for row in rows if row["Comparison_Representative"]) == 1
    assert diag["Exact_Duplicate_Group_Count"] == 1
    assert diag["Exact_Duplicate_Candidate_Count"] == 2
    assert diag["Candidates_Removed_As_Exact_Duplicates"] == 1


def test_qc_eligible_candidate_beats_high_intensity_bad_group_member():
    good = _candidate("GGOOD", [6, 7, 8], mass=15000.0, intensity=20000, theoretical_mass=12000.0)
    bad = _candidate("GBAD", [6, 7, 8], mass=15000.2, intensity=90000, theoretical_mass=12000.0)
    shared = [
        {"Cluster_ID": "GGOOD", "Charge": 6, "Neutral_Mass": 14999.99, "Intensity": 5000, "RT": 5.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "GGOOD", "Charge": 7, "Neutral_Mass": 15000.0, "Intensity": 5000, "RT": 5.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "GGOOD", "Charge": 8, "Neutral_Mass": 15000.01, "Intensity": 5000, "RT": 5.04, "Peak_Tier": "Major", "mz": 1875.0, "Scan_ID": "s3"},
        {"Cluster_ID": "GBAD", "Charge": 6, "Neutral_Mass": 14998.5, "Intensity": 30000, "RT": 5.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "GBAD", "Charge": 7, "Neutral_Mass": 15000.5, "Intensity": 30000, "RT": 5.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "GBAD", "Charge": 8, "Neutral_Mass": 15002.0, "Intensity": 30000, "RT": 5.04, "Peak_Tier": "Major", "mz": 1875.0, "Scan_ID": "s3"},
    ]
    rows, _ = _qc([good, bad], shared, config=_generic_config())
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["GGOOD"]["Group_Representative"] is True
    assert by_cluster["GBAD"]["Group_Representative"] is False
    assert by_cluster["GGOOD"]["Comparison_Representative"] is True


def test_nearby_shared_peak_candidates_group_but_no_shared_peak_stays_separate():
    a = _candidate("A", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0)
    b = _candidate("B", [6, 7, 8], mass=15000.4, intensity=25000, theoretical_mass=12000.0)
    c = _candidate("C", [6, 7, 8], mass=15000.5, intensity=25000, theoretical_mass=12000.0)
    peaks = [
        {"Cluster_ID": "A", "Charge": 6, "Neutral_Mass": 14999.99, "Intensity": 10000, "RT": 5.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "A", "Charge": 7, "Neutral_Mass": 15000.0, "Intensity": 10000, "RT": 5.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "A", "Charge": 8, "Neutral_Mass": 15000.01, "Intensity": 10000, "RT": 5.04, "Peak_Tier": "Major", "mz": 1875.0, "Scan_ID": "s3"},
        {"Cluster_ID": "B", "Charge": 6, "Neutral_Mass": 15000.39, "Intensity": 9000, "RT": 5.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "B", "Charge": 7, "Neutral_Mass": 15000.4, "Intensity": 9000, "RT": 5.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "B", "Charge": 8, "Neutral_Mass": 15000.41, "Intensity": 9000, "RT": 5.05, "Peak_Tier": "Major", "mz": 1875.5, "Scan_ID": "s4"},
        {"Cluster_ID": "C", "Charge": 6, "Neutral_Mass": 15000.49, "Intensity": 9000, "RT": 5.01, "Peak_Tier": "Major", "mz": 2501.0, "Scan_ID": "u1"},
        {"Cluster_ID": "C", "Charge": 7, "Neutral_Mass": 15000.5, "Intensity": 9000, "RT": 5.03, "Peak_Tier": "Major", "mz": 2143.0, "Scan_ID": "u2"},
        {"Cluster_ID": "C", "Charge": 8, "Neutral_Mass": 15000.51, "Intensity": 9000, "RT": 5.05, "Peak_Tier": "Major", "mz": 1876.0, "Scan_ID": "u3"},
    ]
    rows, _ = _qc([a, b, c], peaks, config=_generic_config())
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["A"]["Intact_Envelope_Group_ID"] == by_cluster["B"]["Intact_Envelope_Group_ID"]
    assert by_cluster["A"]["Intact_Envelope_Group_ID"] != by_cluster["C"]["Intact_Envelope_Group_ID"]


def test_rt_shared_peak_and_shared_charge_thresholds_control_grouping():
    base = _candidate("BASE", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0)
    far_rt = _candidate("FAR", [6, 7, 8], mass=15000.2, intensity=30000, theoretical_mass=12000.0)
    low_charge = _candidate("LOWCHG", [6, 9, 10], mass=15000.3, intensity=30000, theoretical_mass=12000.0)
    peaks = [
        {"Cluster_ID": "BASE", "Charge": 6, "Neutral_Mass": 14999.99, "Intensity": 10000, "RT": 5.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "BASE", "Charge": 7, "Neutral_Mass": 15000.0, "Intensity": 10000, "RT": 5.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "BASE", "Charge": 8, "Neutral_Mass": 15000.01, "Intensity": 10000, "RT": 5.04, "Peak_Tier": "Major", "mz": 1875.0, "Scan_ID": "s3"},
        {"Cluster_ID": "FAR", "Charge": 6, "Neutral_Mass": 15000.19, "Intensity": 10000, "RT": 6.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "FAR", "Charge": 7, "Neutral_Mass": 15000.2, "Intensity": 10000, "RT": 6.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "FAR", "Charge": 8, "Neutral_Mass": 15000.21, "Intensity": 10000, "RT": 6.04, "Peak_Tier": "Major", "mz": 1875.0, "Scan_ID": "s3"},
        {"Cluster_ID": "LOWCHG", "Charge": 6, "Neutral_Mass": 15000.29, "Intensity": 10000, "RT": 5.01, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "LOWCHG", "Charge": 9, "Neutral_Mass": 15000.3, "Intensity": 10000, "RT": 5.03, "Peak_Tier": "Major", "mz": 1666.0, "Scan_ID": "s9"},
        {"Cluster_ID": "LOWCHG", "Charge": 10, "Neutral_Mass": 15000.31, "Intensity": 10000, "RT": 5.05, "Peak_Tier": "Major", "mz": 1500.0, "Scan_ID": "s10"},
    ]
    rows, _ = _qc([base, far_rt, low_charge], peaks, config=_generic_config(envelope_grouping={"enabled": True, "mass_tolerance_da": 1.0, "rt_tolerance_min": 0.15, "min_shared_peak_fraction": 0.5, "min_shared_charge_fraction": 0.5, "require_peak_overlap": True}))
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["BASE"]["Intact_Envelope_Group_ID"] != by_cluster["FAR"]["Intact_Envelope_Group_ID"]
    assert by_cluster["BASE"]["Intact_Envelope_Group_ID"] != by_cluster["LOWCHG"]["Intact_Envelope_Group_ID"]


def test_reference_and_target_do_not_change_global_group_representative():
    candidates = [
        _candidate("A", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0),
        _candidate("B", [6, 8], mass=15000.2, intensity=50000, theoretical_mass=12000.0),
    ]
    peaks = _peaks("A", [6, 7, 8], masses=[14999.99, 15000.0, 15000.01], intensity=10000) + _peaks("B", [6, 8], masses=[15000.19, 15000.21], intensity=25000)
    base_rows, _ = _qc(candidates, peaks, config=_generic_config())
    ref_target_rows, _ = _qc(
        [_candidate("A", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0), _candidate("B", [6, 8], mass=15000.2, intensity=50000, theoretical_mass=12000.0)],
        peaks,
        config=_generic_config(reference_masses=[{"label": "external", "mass_da": 15000.2}], target_review_mass_range={"enabled": True, "min_da": 14999, "max_da": 15001}),
    )
    base_rep = sorted(row["Cluster_ID"] for row in base_rows if row["Group_Representative"])
    ref_target_rep = sorted(row["Cluster_ID"] for row in ref_target_rows if row["Group_Representative"])
    assert ref_target_rep == base_rep


def test_grouping_disabled_keeps_existing_qc_and_singleton_groups():
    candidates = [_candidate(f"C{i}", [6, 7, 8], mass=15000.0 + i, intensity=10000 + i, theoretical_mass=12000.0) for i in range(3)]
    peaks = []
    for i in range(3):
        peaks.extend(_peaks(f"C{i}", [6, 7, 8], masses=[15000.0 + i - 0.01, 15000.0 + i, 15000.0 + i + 0.01], intensity=3000))
    rows, diag = _qc(candidates, peaks, config=_generic_config(envelope_grouping={"enabled": False}))
    assert len({row["Intact_Envelope_Group_ID"] for row in rows}) == 3
    assert all(row["Envelope_Group_Size"] == 1 for row in rows)
    assert diag["Intact_Envelope_Group_Count"] == 3


def test_grouping_performance_for_ten_thousand_candidates():
    candidates = [_candidate(f"P{i}", [6, 7, 8], mass=10000.0 + i * 2.0, intensity=1000.0, theoretical_mass=9000.0) for i in range(10000)]
    rows, diag = _qc(candidates, [], config=_generic_config(envelope_grouping={"enabled": True, "mass_tolerance_da": 0.5, "rt_tolerance_min": 0.1}))
    assert len(rows) == 10000
    assert diag["Intact_Envelope_Group_Count"] == 10000



def _excel_config(reconstruction, analysis=None):
    return SimpleNamespace(
        analysis=analysis or {"mode": "full"},
        project={"name": "test"},
        input={},
        organism={},
        sequence={},
        experiment={},
        instrument={},
        reconstruction=reconstruction,
        digestion={"enabled": True},
        alkaline_phosphatase={},
        fragment_mapping={},
        modification_search={},
        peak_filtering={},
        p1_annotation={},
        ms2_annotation={},
        modification_evidence_ranking={},
        biological_context={},
        performance={},
        reporting={"max_excel_rows_per_sheet": 1000, "truncate_large_sheets": True},
    )


def _write_empty_excel(tmp_path, reconstruction, analysis=None):
    return write_excel_report(
        output_dir=tmp_path,
        config=_excel_config(reconstruction, analysis=analysis),
        diagnostics={},
        intact_results=[],
        charge_state_peaks=[],
        warnings=[],
        modifications=[],
        rule_set={},
        pathways=[],
        theoretical_fragments=[],
        fragment_ms1_matches=[],
        known_modification_candidates=[],
        known_modification_summary=[],
        optional_results={},
    )


def _diagnostic_row(workbook):
    diagnostics = workbook["Intact_Reconstruction_Diag"]
    headers = [cell.value for cell in next(diagnostics.iter_rows(min_row=3, max_row=3))]
    values = [cell.value for cell in next(diagnostics.iter_rows(min_row=4, max_row=4))]
    return dict(zip(headers, values))


def _assert_intact_qc_sheet_names(workbook):
    assert "Intact_Reconstruction_QC" in workbook.sheetnames
    assert "Intact_Reconstruction_Diag" in workbook.sheetnames
    old_sheet_name = "Intact_Reconstruction_" + "Diagnostics"
    assert old_sheet_name not in workbook.sheetnames
    assert all(len(name) <= 31 for name in workbook.sheetnames)


def test_zero_candidates_excel_output_succeeds_with_qc_sheets(tmp_path, recwarn):
    report = _write_empty_excel(tmp_path, {"enabled": True, **BASE_QC_CONFIG})
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        assert "Intact_mass_reconstruction" in workbook.sheetnames
        _assert_intact_qc_sheet_names(workbook)
        row = _diagnostic_row(workbook)
        assert row["Total_Reconstruction_Candidates"] == 0
        assert "no_charge_state_candidates" in row["Failure_Reason_Counts"]
        assert not [warning for warning in recwarn if "Title is more than 31 characters" in str(warning.message)]
    finally:
        workbook.close()


def test_reconstruction_disabled_still_writes_diagnostic_sheet(tmp_path):
    report = _write_empty_excel(tmp_path, {"enabled": False, **BASE_QC_CONFIG})
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        assert "Intact_mass_reconstruction" not in workbook.sheetnames
        _assert_intact_qc_sheet_names(workbook)
        row = _diagnostic_row(workbook)
        assert row["Reconstruction_Enabled"] is False
        assert "reconstruction_disabled" in row["Failure_Reason_Counts"]
    finally:
        workbook.close()


def test_existing_excel_sheets_and_new_columns_are_present(tmp_path):
    report = _write_empty_excel(tmp_path, {"enabled": True, **BASE_QC_CONFIG})
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        for sheet in ["Run_summary", "Input_parameters", "mzML_diagnostics", "Intact_mass_reconstruction", "Charge_state_peaks"]:
            assert sheet in workbook.sheetnames
        qc_headers = [cell.value for cell in next(workbook["Intact_Reconstruction_QC"].iter_rows(min_row=3, max_row=3))]
        diag_headers = [cell.value for cell in next(workbook["Intact_Reconstruction_Diag"].iter_rows(min_row=3, max_row=3))]
        assert "Comparison_Ready_Strict" in qc_headers
        assert "Comparison_Ready_Review" in qc_headers
        assert "In_Neutral_Mass_Search_Range" in qc_headers
        assert "Neutral_Mass_Search_Min_Da" in qc_headers
        assert "Neutral_Mass_Search_Max_Da" in qc_headers
        assert "Neutral_Mass_Range_Status" in qc_headers
        assert "Limiting_Factors" in qc_headers
        assert "Envelope_QC_Eligible" in qc_headers
        assert "Intact_Review_Eligible" in qc_headers
        assert "Intact_Strict_Eligible" in qc_headers
        assert "Intact_Envelope_QC_Score" in qc_headers
        assert "Relative_Intact_Eligible_Intensity_Percent" in qc_headers
        assert "In_Target_Review_Mass_Range" in qc_headers
        assert "Dominant_Envelope_Mass" in diag_headers
        assert "Dominant_Envelope_Overall_Mass" in diag_headers
        assert "Dominant_Envelope_In_Mass_Range_Mass" in diag_headers
        assert "Dominant_Envelope_In_Search_Range_Raw_Mass" in diag_headers
        assert "Dominant_Intact_Eligible_Envelope_Mass" in diag_headers
        assert "Envelope_QC_Eligible_Count" in diag_headers
        assert "Intact_Strict_Eligible_Count" in diag_headers
        assert "Intact_Review_Eligible_Count" in diag_headers
        assert "Target_Review_Candidate_Count" in diag_headers
        assert "Exact_Duplicate_Group_Count" in diag_headers
        assert "Intact_Envelope_Group_Count" in diag_headers
        assert "Comparison_Representative_Count" in diag_headers
        assert "Dominant_Comparison_Representative_Mass" in diag_headers
        assert "Total_Candidates_In_Mass_Range" in diag_headers
        assert "Total_Candidates_Outside_Mass_Range" in diag_headers
        assert "Intact_Envelope_Groups" in workbook.sheetnames
        assert "Intact_Comparison_Candidates" in workbook.sheetnames
        assert "Target_Review_Candidates" in workbook.sheetnames
    finally:
        workbook.close()



def test_reconstructed_mass_spectrum_defaults_to_total_supporting_intensity_and_mass_sort():
    low_mass = _candidate("LOW_MASS", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0)
    high_mass = _candidate("HIGH_MASS", [6, 7, 8], mass=15500.0, intensity=60000, theoretical_mass=12000.0)
    peaks = _peaks("LOW_MASS", [6, 7, 8], masses=[14999.99, 15000.0, 15000.01], intensity=10000) + _peaks(
        "HIGH_MASS", [6, 7, 8], masses=[15499.99, 15500.0, 15500.01], intensity=20000
    )
    rows, _ = _qc([high_mass, low_mass], peaks, config=_generic_config())
    spectrum = build_reconstructed_mass_spectrum_rows(rows, _generic_config())
    assert [row["Reconstructed_Mass_Da"] for row in spectrum] == [15000.0, 15500.0]
    assert [row["Reconstructed_Envelope_Intensity"] for row in spectrum] == [30000.0, 60000.0]
    assert [row["Intensity_Method"] for row in spectrum] == ["total_supporting_intensity", "total_supporting_intensity"]
    assert {row["Cluster_ID"]: row["Spectrum_Point_Rank"] for row in spectrum} == {"HIGH_MASS": 1, "LOW_MASS": 2}
    assert max(row["Relative_Intensity_Percent"] for row in spectrum) == 100.0
    assert {row["Cluster_ID"]: row["Relative_Intensity_Percent"] for row in spectrum}["LOW_MASS"] == 50.0


def test_reconstructed_mass_spectrum_intensity_methods_are_configurable():
    candidate = _candidate("METHOD", [6, 7, 8], mass=15000.0, intensity=60.0, theoretical_mass=12000.0)
    peaks = _peaks("METHOD", [6, 7, 8], masses=[14999.99, 15000.0, 15000.01], intensity=10.0)
    peaks[1]["Intensity"] = 20.0
    peaks[2]["Intensity"] = 30.0
    for method, expected in [("total_supporting_intensity", 60.0), ("mean_supporting_intensity", 20.0), ("max_supporting_intensity", 30.0)]:
        config = _generic_config(mass_spectrum_output={"intensity_method": method})
        rows, _ = _qc([_candidate("METHOD", [6, 7, 8], mass=15000.0, intensity=60.0, theoretical_mass=12000.0)], peaks, config=config)
        spectrum = build_reconstructed_mass_spectrum_rows(rows, config)
        assert spectrum[0]["Reconstructed_Envelope_Intensity"] == expected
        assert spectrum[0]["Intensity_Method"] == method


def test_reconstructed_mass_spectrum_filters_range_representatives_and_readiness():
    good = _candidate("GOOD", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0)
    bad = _candidate("BAD", [6, 7, 8], mass=15000.2, intensity=90000, theoretical_mass=12000.0)
    out = _candidate("OUT", [6, 7, 8], mass=21000.0, intensity=120000, theoretical_mass=12000.0)
    peaks = [
        {"Cluster_ID": "GOOD", "Charge": 6, "Neutral_Mass": 14999.99, "Intensity": 10000, "RT": 5.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "GOOD", "Charge": 7, "Neutral_Mass": 15000.0, "Intensity": 10000, "RT": 5.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "GOOD", "Charge": 8, "Neutral_Mass": 15000.01, "Intensity": 10000, "RT": 5.04, "Peak_Tier": "Major", "mz": 1875.0, "Scan_ID": "s3"},
        {"Cluster_ID": "BAD", "Charge": 6, "Neutral_Mass": 14998.5, "Intensity": 30000, "RT": 5.0, "Peak_Tier": "Major", "mz": 2500.0, "Scan_ID": "s1"},
        {"Cluster_ID": "BAD", "Charge": 7, "Neutral_Mass": 15000.5, "Intensity": 30000, "RT": 5.02, "Peak_Tier": "Major", "mz": 2142.0, "Scan_ID": "s2"},
        {"Cluster_ID": "BAD", "Charge": 8, "Neutral_Mass": 15002.0, "Intensity": 30000, "RT": 5.04, "Peak_Tier": "Major", "mz": 1875.0, "Scan_ID": "s3"},
        {"Cluster_ID": "OUT", "Charge": 6, "Neutral_Mass": 20999.99, "Intensity": 40000, "RT": 8.0, "Peak_Tier": "Major", "mz": 3500.0, "Scan_ID": "o1"},
        {"Cluster_ID": "OUT", "Charge": 7, "Neutral_Mass": 21000.0, "Intensity": 40000, "RT": 8.02, "Peak_Tier": "Major", "mz": 3000.0, "Scan_ID": "o2"},
        {"Cluster_ID": "OUT", "Charge": 8, "Neutral_Mass": 21000.01, "Intensity": 40000, "RT": 8.04, "Peak_Tier": "Major", "mz": 2625.0, "Scan_ID": "o3"},
    ]
    config = _generic_config(mass_spectrum_output={"representatives_only": True, "comparison_ready_only": True, "include_qc_ineligible": False})
    rows, _ = _qc([good, bad, out], peaks, config=config)
    spectrum = build_reconstructed_mass_spectrum_rows(rows, config)
    assert [row["Cluster_ID"] for row in spectrum] == ["GOOD"]
    assert all(row["Comparison_Ready"] for row in spectrum)
    assert all(row["Group_Representative"] for row in spectrum)


def test_reconstructed_mass_spectrum_empty_when_no_candidates():
    assert build_reconstructed_mass_spectrum_rows([], _generic_config()) == []


def test_excel_includes_reconstructed_mass_spectrum_and_workflow_summary(tmp_path):
    report = _write_empty_excel(
        tmp_path,
        {"enabled": True, **BASE_QC_CONFIG},
        analysis={"mode": "intact_only"},
    )
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        assert "Workflow_Summary" in workbook.sheetnames
        assert "Reconstructed_Mass_Spectrum" in workbook.sheetnames
        assert "Theoretical_fragments" not in workbook.sheetnames
        assert "Known_Modification_Candidates" not in workbook.sheetnames
        spectrum_headers = [cell.value for cell in next(workbook["Reconstructed_Mass_Spectrum"].iter_rows(min_row=3, max_row=3))]
        assert "Reconstructed_Envelope_Intensity" in spectrum_headers
        assert "Relative_Intensity_Percent" in spectrum_headers
        assert "Intensity_Method" in spectrum_headers
        assert all(len(name) <= 31 for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_intact_only_config_requires_reconstruction_enabled():
    config = RunConfig(
        analysis={"mode": "intact_only"},
        instrument={"polarity": "negative"},
        input={},
        sequence={"sequence": "ACG"},
        reconstruction={"enabled": False, "intact_reconstruction": {}},
        digestion={"enabled": True},
        fragment_mapping={},
        alkaline_phosphatase={},
    )
    with pytest.raises(ValueError, match="analysis.mode=intact_only requires intact_reconstruction.enabled=true"):
        validate_config(config)


def test_invalid_reconstructed_spectrum_intensity_method_is_rejected():
    config = RunConfig(
        analysis={"mode": "full"},
        instrument={"polarity": "negative"},
        input={},
        sequence={"sequence": "ACG"},
        reconstruction={"enabled": True, "intact_reconstruction": {"mass_spectrum_output": {"intensity_method": "median"}}},
        digestion={"enabled": True},
        fragment_mapping={},
        alkaline_phosphatase={},
    )
    with pytest.raises(ValueError, match="intensity_method must be one of"):
        validate_config(config)



def _rt_config(**overrides):
    intact = {
        "engine": "rt_localized",
        "neutral_mass_range": {"enabled": True, "min_da": 10000, "max_da": 20000},
        "rt_localized": {
            "rt_window_min": 0.10,
            "rt_step_min": 0.05,
            "min_scans_per_window": 1,
            "peak_aggregation": "max",
            "mz_merge_tolerance_ppm": 10,
            "adjacent_charge_mz_tolerance_ppm": 20,
            "max_charge_gap": 1,
            "min_charge_states": 2,
            "min_consecutive_charge_states": 2,
            "require_consecutive_for_candidate": True,
            "neutral_mass_estimator": "intensity_weighted_mean",
            "merge_across_windows": {"enabled": True, "mass_tolerance_ppm": 10, "rt_overlap_required": True, "min_shared_charge_fraction": 0.5},
        },
    }
    for key, value in overrides.items():
        if key == "rt_localized":
            intact["rt_localized"].update(value)
        else:
            intact[key] = value
    return {"enabled": True, "min_charge": 6, "max_charge": 9, "min_charge_states": 2, "intact_reconstruction": intact}


def _rt_peak(mass, charge, intensity=1000.0, rt=5.0, tier="Major", scan=None):
    return Peak(mz=mz_from_neutral_mass(mass, charge, "negative"), intensity=intensity, rt=rt, scan_id=scan or f"s{charge}_{rt}", tier=tier)


def _run_rt(peaks, below=None, config=None):
    tier = PeakTierResult(major=peaks, below_threshold=below or [])
    return reconstruct_intact_masses(tier, config or _rt_config(), {"polarity": "negative"}, theoretical_mass=None)


def test_rt_localized_reconstructs_same_window_consecutive_charges():
    candidates, charge_peaks, meta = _run_rt([_rt_peak(15000.0, z, intensity=1000 * z, rt=5.0 + z * 0.005) for z in [6, 7, 8]])
    assert len(candidates) == 1
    assert abs(candidates[0].observed_mass - 15000.0) < 0.01
    assert candidates[0].charge_states == [6, 7, 8]
    assert candidates[0].reconstruction_engine == "rt_localized"
    assert candidates[0].longest_consecutive_charge_run == 3
    assert candidates[0].comparison_ready is True
    assert meta["stats"]["Num_RT_Windows"] == 1
    assert meta["stats"]["Num_Candidates_After_RT_Window_Merge"] == 1
    assert len(charge_peaks) == 3


def test_rt_localized_does_not_mix_distant_rt_peaks():
    peaks = [_rt_peak(15000.0, 6, rt=5.0), _rt_peak(15000.0, 7, rt=8.0)]
    candidates, _, _ = _run_rt(peaks)
    assert candidates == []


def test_rt_localized_adjacent_charge_prediction_finds_neighbor():
    candidates, _, meta = _run_rt([_rt_peak(15000.0, 6, rt=5.0), _rt_peak(15000.0, 7, rt=5.02)])
    assert len(candidates) == 1
    assert candidates[0].charge_states == [6, 7]
    assert candidates[0].charge_coverage_fraction >= 2 / 3
    assert meta["stats"]["Num_Anchor_Peaks_Evaluated"] >= 2


def test_rt_localized_charge_gap_and_missing_charge_no_peak():
    config = _rt_config(rt_localized={"max_charge_gap": 2, "require_consecutive_for_candidate": False})
    candidates, _, meta = _run_rt([_rt_peak(15000.0, 6, rt=5.0), _rt_peak(15000.0, 8, rt=5.02)], config=config)
    assert len(candidates) == 1
    candidate = candidates[0]
    assert candidate.charge_gap_count == 1
    assert candidate.missing_charge_states == "7"
    assert candidate.missing_charge_predicted_mz
    missing = meta["missing_charge_diagnostics"]
    assert missing[0]["Missing_Charge"] == 7
    assert missing[0]["Detection_Status"] == "no_peak_in_tolerance"


def test_rt_localized_missing_charge_weak_peak_is_distinguished():
    config = _rt_config(rt_localized={"max_charge_gap": 2, "require_consecutive_for_candidate": False})
    weak = [_rt_peak(15000.0, 7, intensity=10.0, rt=5.01, tier="BelowThreshold")]
    candidates, _, meta = _run_rt([_rt_peak(15000.0, 6, rt=5.0), _rt_peak(15000.0, 8, rt=5.02)], below=weak, config=config)
    assert len(candidates) == 1
    missing = meta["missing_charge_diagnostics"]
    assert missing[0]["Detection_Status"] == "below_intensity_threshold"
    assert missing[0]["Nearest_Intensity"] == 10.0
    assert meta["stats"]["Num_Missing_Charges_With_Weak_Peaks"] == 1


def test_rt_localized_single_charge_candidates_are_filtered():
    candidates, _, meta = _run_rt([_rt_peak(15000.0, 6, rt=5.0), _rt_peak(16000.0, 7, rt=5.0)])
    assert candidates == []
    assert meta["stats"]["Num_Raw_Envelope_Candidates"] > 0
    assert meta["stats"]["Num_Candidates_After_Charge_Filter"] == 0


def test_rt_localized_local_relative_intensity_and_weighted_mass():
    mass = 15000.0
    peaks = [_rt_peak(mass, 6, intensity=100.0, rt=5.0), _rt_peak(mass + 0.6, 7, intensity=900.0, rt=5.01)]
    candidates, _, _ = _run_rt(peaks, config=_rt_config(rt_localized={"neutral_mass_estimator": "intensity_weighted_mean", "adjacent_charge_mz_tolerance_ppm": 100}))
    assert len(candidates) == 1
    expected = (mass * 100.0 + (mass + 0.6) * 900.0) / 1000.0
    assert abs(candidates[0].observed_mass - expected) < 0.01
    assert candidates[0].local_window_max_intensity == 900.0
    assert round(candidates[0].local_relative_peak_intensity_percent, 1) in {11.1, 100.0}


def test_rt_localized_median_estimator():
    peaks = [_rt_peak(15000.0, 6, rt=5.0), _rt_peak(15001.0, 7, rt=5.01), _rt_peak(15010.0, 8, rt=5.02)]
    candidates, _, _ = _run_rt(peaks, config=_rt_config(rt_localized={"neutral_mass_estimator": "median", "adjacent_charge_mz_tolerance_ppm": 1000}))
    assert candidates
    best = min(candidates, key=lambda candidate: abs(candidate.observed_mass - 15001.0))
    assert abs(best.observed_mass - 15001.0) < 0.01
    assert best.neutral_mass_estimator == "median"


def test_rt_localized_merges_adjacent_windows_without_duplicate_intensity():
    peaks = [_rt_peak(15000.0, 6, intensity=1000, rt=5.00, scan="s6"), _rt_peak(15000.0, 7, intensity=2000, rt=5.04, scan="s7")]
    candidates, _, _ = _run_rt(peaks, config=_rt_config(rt_localized={"rt_window_min": 0.08, "rt_step_min": 0.04}))
    assert len(candidates) == 1
    assert candidates[0].merged_across_rt_windows is True
    assert candidates[0].total_intensity == 3000.0


def test_rt_localized_reference_and_target_do_not_change_generation():
    peaks = [_rt_peak(15000.0, z, rt=5.0 + z * 0.005) for z in [6, 7, 8]]
    base_candidates, _, _ = _run_rt(peaks)
    ref_config = _rt_config(reference_masses=[{"label": "external", "mass_da": 15000.0}], target_review_mass_range={"enabled": True, "min_da": 14900, "max_da": 15100})
    ref_candidates, _, _ = _run_rt(peaks, config=ref_config)
    assert len(base_candidates) == len(ref_candidates) == 1
    assert base_candidates[0].observed_mass == ref_candidates[0].observed_mass
    assert base_candidates[0].charge_states == ref_candidates[0].charge_states


def test_legacy_engine_is_maintained():
    config = {"enabled": True, "min_charge": 6, "max_charge": 8, "min_charge_states": 2, "mass_cluster_tolerance_da": 1.0, "intact_reconstruction": {"engine": "legacy_cluster", "neutral_mass_range": {"enabled": True, "min_da": 10000, "max_da": 20000}}}
    candidates, _, meta = reconstruct_intact_masses(PeakTierResult(major=[_rt_peak(15000.0, z, rt=5.0) for z in [6, 7]]), config, {"polarity": "negative"}, None)
    assert candidates
    assert all(candidate.reconstruction_engine == "legacy_cluster" for candidate in candidates)
    assert meta["engine"] == "legacy_cluster"


def test_rt_localized_outputs_to_reconstructed_mass_spectrum():
    candidates, charge_peaks, _ = _run_rt([_rt_peak(15000.0, z, intensity=1000*z, rt=5.0) for z in [6, 7, 8]])
    rows, _ = build_intact_reconstruction_qc(candidates, charge_peaks, _rt_config(), reconstruction_enabled=True)
    spectrum = build_reconstructed_mass_spectrum_rows(rows, _rt_config())
    assert len(spectrum) == 1
    assert spectrum[0]["Reconstruction_Engine"] == "rt_localized"
    assert abs(spectrum[0]["Reconstructed_Mass_Da"] - 15000.0) < 0.01


def test_rt_localized_excel_sheets_are_present(tmp_path):
    optional = {
        "RT_Envelope_Diagnostics": [{"Cluster_ID": "RTL00001", "Reconstruction_Engine": "rt_localized"}],
        "Missing_Charge_Diagnostics": [{"Cluster_ID": "RTL00001", "Missing_Charge": 7, "Detection_Status": "no_peak_in_tolerance"}],
        "Intact_Engine_Comparison": [],
    }
    report = write_excel_report(
        output_dir=tmp_path,
        config=_excel_config({"enabled": True, **_rt_config()}),
        diagnostics={},
        intact_results=[],
        charge_state_peaks=[],
        warnings=[],
        modifications=[],
        rule_set={},
        pathways=[],
        theoretical_fragments=[],
        fragment_ms1_matches=[],
        known_modification_candidates=[],
        known_modification_summary=[],
        optional_results=optional,
    )
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        assert "RT_Envelope_Diagnostics" in workbook.sheetnames
        assert "Missing_Charge_Diagnostics" in workbook.sheetnames
        assert "Intact_Engine_Comparison" in workbook.sheetnames
        assert all(len(name) <= 31 for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_rt_localized_config_defaults_and_validation():
    config = RunConfig(
        analysis={"mode": "full"},
        instrument={"polarity": "negative"},
        input={},
        sequence={"sequence": "ACG"},
        reconstruction={"enabled": True, "intact_reconstruction": {"engine": "rt_localized", "rt_localized": {}}},
        digestion={"enabled": True},
        fragment_mapping={},
        alkaline_phosphatase={},
    )
    validate_config(config)
    bad = RunConfig(
        analysis={"mode": "full"},
        instrument={"polarity": "negative"},
        input={},
        sequence={"sequence": "ACG"},
        reconstruction={"enabled": True, "intact_reconstruction": {"engine": "bad"}},
        digestion={"enabled": True},
        fragment_mapping={},
        alkaline_phosphatase={},
    )
    with pytest.raises(ValueError, match="intact_reconstruction.engine"):
        validate_config(bad)



def test_mvp597_quality_tiers_and_failure_matrix_are_reported():
    tier1 = _candidate("TIER1", [10, 11, 12], mass=MODIFIED_MASS, intensity=50000, theoretical_mass=UNMODIFIED_MASS)
    tier2 = _candidate("TIER2", [10, 11], mass=LOW_FULL_LENGTH_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    tier3 = _candidate("TIER3", [10, 11], mass=25343.0, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    tier4 = _candidate("TIER4", [10], mass=25344.0, intensity=10000, theoretical_mass=UNMODIFIED_MASS)
    peaks = (
        _peaks("TIER1", [10, 11, 12], masses=[25325.49, 25325.50, 25325.51], intensity=10000)
        + _peaks("TIER2", [10, 11], masses=[25081.99, 25082.01], intensity=10000)
        + _peaks("TIER3", [10, 11], masses=[25342.0, 25344.0], intensity=10000)
        + _peaks("TIER4", [10], masses=[25344.0], intensity=10000)
    )
    rows, diag = _qc([tier1, tier2, tier3, tier4], peaks)
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["TIER1"]["Intact_Quality_Tier"] == "Tier_1_high_quality"
    assert by_cluster["TIER2"]["Intact_Quality_Tier"] == "Tier_2_supported"
    assert by_cluster["TIER3"]["Intact_Quality_Tier"] == "Tier_3_weak"
    assert by_cluster["TIER4"]["Intact_Quality_Tier"] == "Tier_4_rejected"
    assert by_cluster["TIER1"]["Comparison_Ready_Strict"] is True
    assert by_cluster["TIER2"]["Comparison_Ready_Review"] is True
    assert by_cluster["TIER3"]["Comparison_Ready"] is False
    assert by_cluster["TIER3"]["Pass_Neutral_Mass_Range"] is False
    assert "neutral_mass_range" in by_cluster["TIER3"]["Review_Failure_Reasons"]
    assert "Tier_1_high_quality:1" in diag["Candidate_Count_By_Quality_Tier"]
    assert "Tier_4_rejected:1" in diag["Candidate_Count_By_Quality_Tier"]


def test_mvp597_comparison_ready_tiers_are_configurable():
    config = {
        "intact_reconstruction": {
            **BASE_QC_CONFIG["intact_reconstruction"],
            "comparison_ready_tiers": {"strict": ["Tier_1_high_quality"], "review": ["Tier_1_high_quality", "Tier_2_supported", "Tier_3_weak"]},
        }
    }
    weak = _candidate("WEAK", [10, 11], mass=25343.0, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    rows, _ = _qc([weak], _peaks("WEAK", [10, 11], masses=[25342.0, 25344.0], intensity=10000), config=config)
    assert rows[0]["Intact_Quality_Tier"] == "Tier_3_weak"
    assert rows[0]["Comparison_Ready_Strict"] is False
    assert rows[0]["Comparison_Ready_Review"] is True
    assert rows[0]["Comparison_Ready"] is True


def test_mvp597_reference_and_target_review_do_not_change_quality_tier_or_comparison_ready():
    base = _generic_config()
    annotated = _generic_config(
        reference_masses=[{"label": "external", "mass_da": 15000.0}],
        target_review_mass_range={"enabled": True, "min_da": 14900, "max_da": 15100},
    )
    candidate_a = _candidate("GEN", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0)
    candidate_b = _candidate("GEN", [6, 7, 8], mass=15000.0, intensity=30000, theoretical_mass=12000.0)
    peaks = _peaks("GEN", [6, 7, 8], masses=[14999.99, 15000.0, 15000.01], intensity=10000)
    base_rows, _ = _qc([candidate_a], peaks, config=base)
    annotated_rows, _ = _qc([candidate_b], peaks, config=annotated)
    for field in ["Intact_Quality_Tier", "Quality_Tier_Rank", "Comparison_Ready", "Intact_Strict_Eligible", "Intact_Envelope_QC_Score"]:
        assert annotated_rows[0][field] == base_rows[0][field]
    assert annotated_rows[0]["Reference_Mass_Matched"] is True
    assert annotated_rows[0]["In_Target_Review_Mass_Range"] is True


def test_mvp597_charge_extension_reports_weak_and_missing_neighbor_charges():
    peaks = [_rt_peak(15000.0, 6, intensity=1000, rt=5.0), _rt_peak(15000.0, 7, intensity=900, rt=5.01)]
    below = [_rt_peak(15000.0, 8, intensity=5, rt=5.01, tier="Below")]
    candidates, _, meta = _run_rt(peaks, below=below, config=_rt_config(rt_localized={"charge_extension": {"enabled": True, "max_extension_charges": 2, "weak_peak_tolerance_ppm": 30, "weak_peak_min_local_relative_percent": 0.01, "add_weak_peaks_to_envelope": False}}))
    assert candidates
    candidate = candidates[0]
    assert "8" in candidate.extended_upper_charges_evaluated
    assert "8" in candidate.extended_weak_charges_detected
    assert candidate.charge_extension_improved_envelope is True
    assert meta["stats"]["Missing_Charge_Status_Count"] in {"", None} or isinstance(meta["stats"]["Missing_Charge_Status_Count"], str)


def test_mvp597_split_envelope_merge_recalculates_mass_and_gap_metrics():
    rt_config = _rt_config()["intact_reconstruction"]["rt_localized"]
    rt_config["split_envelope_merge"] = {"enabled": True, "mass_tolerance_ppm": 20, "rt_tolerance_min": 0.2, "max_charge_gap": 0}
    window = {"RT_Window_ID": "RT00001", "center": 5.0}
    record_a = {
        "mass": 15000.0,
        "charges": [6, 7],
        "observed": {6: {"mz": mz_from_neutral_mass(15000.0, 6, "negative"), "Intensity": 1000, "Source_Peak_IDs": ["a6"], "Local_Peak_ID": "a6"}, 7: {"mz": mz_from_neutral_mass(15000.0, 7, "negative"), "Intensity": 900, "Source_Peak_IDs": ["a7"], "Local_Peak_ID": "a7"}},
        "rt_window_ids": ["RT00001"],
        "window": window,
        "local_max": 1000,
    }
    record_b = {
        "mass": 15000.0,
        "charges": [8, 9],
        "observed": {8: {"mz": mz_from_neutral_mass(15000.0, 8, "negative"), "Intensity": 800, "Source_Peak_IDs": ["b8"], "Local_Peak_ID": "b8"}, 9: {"mz": mz_from_neutral_mass(15000.0, 9, "negative"), "Intensity": 700, "Source_Peak_IDs": ["b9"], "Local_Peak_ID": "b9"}},
        "rt_window_ids": ["RT00001"],
        "window": window,
        "local_max": 1000,
    }
    merged = _apply_split_envelope_merge([record_a, record_b], rt_config, "negative")
    assert len(merged) == 1
    assert merged[0]["charges"] == [6, 7, 8, 9]
    assert merged[0]["split_envelope_merged"] is True
    assert merged[0]["charge_gaps_after_merge"] == 0
    assert abs(merged[0]["mass"] - 15000.0) < 0.01


def test_mvp597_peak_sharing_can_degrade_quality_tier():
    candidate = _candidate("SHARED", [10, 11, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    candidate.max_peak_usage_count = 10
    candidate.mean_peak_usage_count = 10
    candidate.num_highly_shared_peaks = 3
    candidate.highly_shared_peak_fraction = 1.0
    peaks = _peaks("SHARED", [10, 11, 12], masses=[25325.49, 25325.50, 25325.51], intensity=10000)
    rows, _ = _qc([candidate], peaks)
    assert rows[0]["Peak_Sharing_Status"] == "highly_shared"
    assert rows[0]["Pass_Peak_Sharing"] is False
    assert rows[0]["Intact_Quality_Tier"] == "Tier_3_weak"
    assert rows[0]["Comparison_Ready"] is False


def test_mvp597_engine_comparison_reports_match_and_mismatch_statuses():
    legacy = [_candidate("L1", [6, 7, 8], mass=15000.0, intensity=1000, theoretical_mass=None), _candidate("L2", [6, 7, 8], mass=16000.0, intensity=1000, theoretical_mass=None)]
    rt = [_candidate("R1", [6, 7, 8], mass=15000.01, intensity=1000, theoretical_mass=None), _candidate("R2", [6, 8], mass=17000.0, intensity=1000, theoretical_mass=None)]
    rows = build_intact_engine_comparison_rows(legacy, [], rt, [], _generic_config(engine_comparison={"mass_tolerance_ppm": 20, "rt_tolerance_min": 0.15, "min_shared_charge_fraction": 0.5, "require_mass_match": True}))
    statuses = {row["Legacy_Cluster_ID"] or row["RT_Localized_Cluster_ID"]: row["Engine_Match_Status"] for row in rows}
    assert statuses["L1"] == "matched"
    assert statuses["L2"] == "mass_mismatch"
    assert statuses["R2"] == "rt_localized_only"
    assert all(row.get("Mass_Delta_Da") is None for row in rows if row["Engine_Match_Status"] != "matched")


def test_mvp597_spectrum_minimum_quality_tier_filters_weak_candidates():
    tier1 = _candidate("SPEC1", [10, 11, 12], mass=MODIFIED_MASS, intensity=50000, theoretical_mass=UNMODIFIED_MASS)
    tier3 = _candidate("SPEC3", [10, 11], mass=25343.0, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("SPEC1", [10, 11, 12], masses=[25325.49, 25325.50, 25325.51], intensity=10000) + _peaks("SPEC3", [10, 11], masses=[25342.0, 25344.0], intensity=10000)
    rows, _ = _qc([tier1, tier3], peaks)
    spectrum = build_reconstructed_mass_spectrum_rows(rows, {"intact_reconstruction": {**BASE_QC_CONFIG["intact_reconstruction"], "mass_spectrum_output": {"enabled": True, "representatives_only": False, "minimum_quality_tier": "Tier_2_supported"}}})
    assert [round(row["Reconstructed_Mass_Da"], 1) for row in spectrum] == [MODIFIED_MASS]
    assert "Intact_Quality_Tier" in spectrum[0]
    assert "Peak_Sharing_Status" in spectrum[0]


def test_mvp597_rt_engine_qc_summary_sheet_is_written(tmp_path):
    candidate = _candidate("SUMMARY", [10, 11, 12], mass=MODIFIED_MASS, intensity=30000, theoretical_mass=UNMODIFIED_MASS)
    peaks = _peaks("SUMMARY", [10, 11, 12], masses=[25325.49, 25325.50, 25325.51], intensity=10000)
    rows, _ = _qc([candidate], peaks)
    report = write_excel_report(
        output_dir=tmp_path,
        config=_excel_config({"enabled": True, **BASE_QC_CONFIG}),
        diagnostics={},
        intact_results=[candidate],
        charge_state_peaks=peaks,
        warnings=[],
        modifications=[],
        rule_set={},
        pathways=[],
        theoretical_fragments=[],
        fragment_ms1_matches=[],
        known_modification_candidates=[],
        known_modification_summary=[],
        optional_results={},
    )
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        assert "RT_Engine_QC_Summary" in workbook.sheetnames
        headers = [cell.value for cell in next(workbook["RT_Engine_QC_Summary"].iter_rows(min_row=3, max_row=3))]
        assert headers == ["Metric", "Value", "Notes"]
        values = [row[0] for row in workbook["RT_Engine_QC_Summary"].iter_rows(min_row=4, values_only=True)]
        assert "Candidate_Count_By_Quality_Tier" in values
        assert all(len(name) <= 31 for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_mvp597_invalid_spectrum_minimum_quality_tier_is_rejected():
    config = RunConfig(
        analysis={"mode": "full"},
        instrument={"polarity": "negative"},
        input={},
        sequence={"sequence": "ACG"},
        reconstruction={"enabled": True, "intact_reconstruction": {"mass_spectrum_output": {"minimum_quality_tier": "Tier_X"}}},
        digestion={"enabled": True},
        fragment_mapping={},
        alkaline_phosphatase={},
    )
    with pytest.raises(ValueError, match="minimum_quality_tier"):
        validate_config(config)



def _competition_peak(cluster_id, charge, local_peak_id, mass=15000.0, intensity=1000.0, rt=5.0, neutral_offset=0.0):
    return {
        "Cluster_ID": cluster_id,
        "Charge": charge,
        "Neutral_Mass": mass + neutral_offset,
        "Intensity": intensity,
        "RT": rt,
        "Peak_Tier": "Major",
        "mz": mz_from_neutral_mass(mass + neutral_offset, charge, "negative"),
        "Scan_ID": f"scan_{local_peak_id}",
        "Local_Peak_ID": local_peak_id,
    }


def _competition_candidate(cluster_id, charges, mass=15000.0, intensity=30000.0):
    return _candidate(cluster_id, charges, mass=mass, intensity=intensity, theoretical_mass=12000.0)


def _parse_component_total(text):
    total = 0.0
    for item in str(text or "").split(";"):
        item = item.strip()
        if item and "=" in item:
            total += float(item.split("=", 1)[1])
    return total


def test_mvp598a_nonshared_candidates_get_separate_competition_groups():
    a = _competition_candidate("A", [6, 7])
    b = _competition_candidate("B", [6, 7], mass=15100.0)
    peaks = [_competition_peak("A", 6, "p1"), _competition_peak("A", 7, "p2"), _competition_peak("B", 6, "p3", mass=15100.0), _competition_peak("B", 7, "p4", mass=15100.0)]
    rows, diag = _qc([a, b], peaks, config=_generic_config())
    assert rows[0]["Competing_Envelope_Group_ID"] != rows[1]["Competing_Envelope_Group_ID"]
    assert all(row["Is_Noncompeting_Candidate"] is True for row in rows)
    assert diag["Competing_Envelope_Group_Count"] == 2
    assert diag["MultiCandidate_Competition_Group_Count"] == 0


def test_mvp598a_one_shared_peak_groups_candidates_and_shared_fraction():
    a = _competition_candidate("A", [6, 7])
    b = _competition_candidate("B", [6, 7], mass=15000.2)
    peaks = [_competition_peak("A", 6, "shared"), _competition_peak("A", 7, "a7"), _competition_peak("B", 6, "shared", neutral_offset=0.2), _competition_peak("B", 7, "b7", neutral_offset=0.2)]
    rows, _ = _qc([a, b], peaks, config=_generic_config())
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["A"]["Competing_Envelope_Group_ID"] == by_cluster["B"]["Competing_Envelope_Group_ID"]
    assert by_cluster["A"]["Competing_Envelope_Group_Size"] == 2
    assert by_cluster["A"]["Maximum_Shared_Peak_Fraction"] == 0.5
    assert by_cluster["A"]["Mean_Shared_Peak_Fraction"] == 0.5
    assert by_cluster["A"]["Shared_Peak_Competitor_Count"] == 1
    assert by_cluster["A"]["Is_Noncompeting_Candidate"] is False


def test_mvp598a_connected_component_groups_transitive_shared_peaks():
    a = _competition_candidate("A", [6, 7])
    b = _competition_candidate("B", [6, 7], mass=15000.1)
    c = _competition_candidate("C", [6, 7], mass=15000.2)
    peaks = [
        _competition_peak("A", 6, "ab"), _competition_peak("A", 7, "a7"),
        _competition_peak("B", 6, "ab", neutral_offset=0.1), _competition_peak("B", 7, "bc", neutral_offset=0.1),
        _competition_peak("C", 6, "bc", neutral_offset=0.2), _competition_peak("C", 7, "c7", neutral_offset=0.2),
    ]
    rows, diag = _qc([a, b, c], peaks, config=_generic_config())
    assert len({row["Competing_Envelope_Group_ID"] for row in rows}) == 1
    assert all(row["Competing_Envelope_Group_Size"] == 3 for row in rows)
    assert diag["Largest_Competition_Group_Size"] == 3


def test_mvp598a_single_candidate_group_has_unique_group_id_and_rank_one():
    candidate = _competition_candidate("SOLO", [6, 7, 8])
    peaks = [_competition_peak("SOLO", z, f"s{z}") for z in [6, 7, 8]]
    rows, _ = _qc([candidate], peaks, config=_generic_config())
    row = rows[0]
    assert row["Competing_Envelope_Group_ID"].startswith("CG")
    assert row["Competing_Envelope_Group_Size"] == 1
    assert row["Evidence_Score_Rank_In_Competition"] == 1
    assert row["Is_Noncompeting_Candidate"] is True


def test_mvp598a_evidence_score_components_penalties_and_total_match():
    candidate = _competition_candidate("SCORE", [6, 7, 8], intensity=30000)
    peaks = [_competition_peak("SCORE", 6, "p6", neutral_offset=-0.5), _competition_peak("SCORE", 7, "p7"), _competition_peak("SCORE", 8, "p8", neutral_offset=0.5)]
    rows, _ = _qc([candidate], peaks, config=_generic_config())
    row = rows[0]
    assert "charge_count=" in row["Evidence_Score_Components"]
    assert "consecutive_run=" in row["Evidence_Score_Components"]
    assert "internal_error=" in row["Evidence_Score_Penalties"]
    total = _parse_component_total(row["Evidence_Score_Components"]) + _parse_component_total(row["Evidence_Score_Penalties"])
    assert round(total, 6) == row["Envelope_Evidence_Score"]


def test_mvp598a_group_rank_descending_and_tie_breaker_deterministic():
    better = _competition_candidate("A_BETTER", [6, 7, 8], intensity=40000)
    weaker = _competition_candidate("B_WEAKER", [6, 7], mass=15000.1, intensity=20000)
    tie_a = _competition_candidate("TIE_A", [6, 7], mass=15100.0, intensity=20000)
    tie_b = _competition_candidate("TIE_B", [6, 7], mass=15100.0, intensity=20000)
    peaks = [
        _competition_peak("A_BETTER", 6, "shared1"), _competition_peak("A_BETTER", 7, "a7"), _competition_peak("A_BETTER", 8, "a8"),
        _competition_peak("B_WEAKER", 6, "shared1", neutral_offset=0.1), _competition_peak("B_WEAKER", 7, "b7", neutral_offset=0.1),
        _competition_peak("TIE_A", 6, "shared2", mass=15100.0), _competition_peak("TIE_A", 7, "ta7", mass=15100.0),
        _competition_peak("TIE_B", 6, "shared2", mass=15100.0), _competition_peak("TIE_B", 7, "tb7", mass=15100.0),
    ]
    rows, _ = _qc([weaker, better, tie_b, tie_a], peaks, config=_generic_config())
    by_cluster = {row["Cluster_ID"]: row for row in rows}
    assert by_cluster["A_BETTER"]["Evidence_Score_Rank_In_Competition"] == 1
    assert by_cluster["B_WEAKER"]["Evidence_Score_Rank_In_Competition"] == 2
    assert by_cluster["TIE_A"]["Evidence_Score_Rank_In_Competition"] == 1
    assert by_cluster["TIE_B"]["Evidence_Score_Rank_In_Competition"] == 2


def test_mvp598a_reference_and_target_review_do_not_change_competition_score_or_rank():
    candidates = [_competition_candidate("A", [6, 7]), _competition_candidate("B", [6, 7], mass=15000.1)]
    peaks = [_competition_peak("A", 6, "shared"), _competition_peak("A", 7, "a7"), _competition_peak("B", 6, "shared", neutral_offset=0.1), _competition_peak("B", 7, "b7", neutral_offset=0.1)]
    base_rows, _ = _qc(candidates, peaks, config=_generic_config())
    annotated_rows, _ = _qc(
        [_competition_candidate("A", [6, 7]), _competition_candidate("B", [6, 7], mass=15000.1)],
        peaks,
        config=_generic_config(reference_masses=[{"label": "ref", "mass_da": 15000.0}], target_review_mass_range={"enabled": True, "min_da": 14900, "max_da": 15100}),
    )
    base = {row["Cluster_ID"]: row for row in base_rows}
    annotated = {row["Cluster_ID"]: row for row in annotated_rows}
    for cluster_id in ["A", "B"]:
        for field in ["Competing_Envelope_Group_Size", "Envelope_Evidence_Score", "Evidence_Score_Rank_In_Competition", "Maximum_Shared_Peak_Fraction"]:
            assert annotated[cluster_id][field] == base[cluster_id][field]


def test_mvp598a_zero_candidates_and_disabled_competition_are_safe():
    rows, diag = _qc([], [], config=_generic_config())
    assert rows == []
    assert diag["Competing_Envelope_Group_Count"] == 0
    config = _generic_config(competitive_assignment={"enabled": False})
    candidates = [_competition_candidate("A", [6, 7]), _competition_candidate("B", [6, 7])]
    peaks = [_competition_peak("A", 6, "shared"), _competition_peak("A", 7, "a7"), _competition_peak("B", 6, "shared"), _competition_peak("B", 7, "b7")]
    rows, diag = _qc(candidates, peaks, config=config)
    assert len({row["Competing_Envelope_Group_ID"] for row in rows}) == 2
    assert diag["MultiCandidate_Competition_Group_Count"] == 0


def test_mvp598a_legacy_engine_runs_with_competition_columns():
    config = {"enabled": True, "min_charge": 6, "max_charge": 7, "min_charge_states": 2, "mass_cluster_tolerance_da": 1.0, "intact_reconstruction": {"engine": "legacy_cluster", "neutral_mass_range": {"enabled": True, "min_da": 10000, "max_da": 20000}}}
    candidates, peaks, _ = reconstruct_intact_masses(PeakTierResult(major=[_rt_peak(15000.0, z, rt=5.0) for z in [6, 7]]), config, {"polarity": "negative"}, None)
    rows, _ = build_intact_reconstruction_qc(candidates, peaks, config)
    assert rows
    assert "Competing_Envelope_Group_ID" in rows[0]
    assert "Envelope_Evidence_Score" in rows[0]


def test_mvp598a_competition_excel_sheets_are_present(tmp_path):
    candidate = _competition_candidate("EXCEL", [6, 7, 8])
    peaks = [_competition_peak("EXCEL", z, f"p{z}") for z in [6, 7, 8]]
    build_intact_reconstruction_qc([candidate], peaks, _generic_config())
    report = write_excel_report(
        output_dir=tmp_path,
        config=_excel_config({"enabled": True, **_generic_config()}),
        diagnostics={},
        intact_results=[candidate],
        charge_state_peaks=peaks,
        warnings=[],
        modifications=[],
        rule_set={},
        pathways=[],
        theoretical_fragments=[],
        fragment_ms1_matches=[],
        known_modification_candidates=[],
        known_modification_summary=[],
        optional_results={},
    )
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        assert "Intact_Competition_Groups" in workbook.sheetnames
        assert "Intact_Competition_Scores" in workbook.sheetnames
        assert "Intact_Assignment_Dry_Run" in workbook.sheetnames
        assert "Competition_Dry_Run_Summary" in workbook.sheetnames
        assert all(len(name) <= 31 for name in workbook.sheetnames)
        qc_headers = [cell.value for cell in next(workbook["Intact_Reconstruction_QC"].iter_rows(min_row=3, max_row=3))]
        spectrum_headers = [cell.value for cell in next(workbook["Reconstructed_Mass_Spectrum"].iter_rows(min_row=3, max_row=3))]
        assert "Competing_Envelope_Group_ID" in qc_headers
        assert "Envelope_Evidence_Score" in spectrum_headers
    finally:
        workbook.close()


def test_mvp598a_competition_group_rows_and_score_rows_are_diagnostic_only():
    a = _competition_candidate("A", [6, 7])
    b = _competition_candidate("B", [6, 7], mass=15000.1)
    peaks = [_competition_peak("A", 6, "shared"), _competition_peak("A", 7, "a7"), _competition_peak("B", 6, "shared", neutral_offset=0.1), _competition_peak("B", 7, "b7", neutral_offset=0.1)]
    rows, _ = _qc([a, b], peaks, config=_generic_config())
    group_rows = build_intact_competition_group_rows(rows)
    score_rows = build_intact_competition_score_rows(rows)
    assert len(group_rows) == 1
    assert group_rows[0]["MultiCandidate_Group"] is True
    assert len(score_rows) == 2
    assert all(row["Comparison_Ready"] == rows[index]["Comparison_Ready"] for index, row in enumerate(rows))


def test_mvp598a_competition_performance_smoke():
    candidates = [_competition_candidate(f"P{i}", [6, 7], mass=15000.0 + i * 0.01, intensity=1000.0 + i) for i in range(500)]
    peaks = []
    for i in range(500):
        shared = f"shared_{i // 5}"
        peaks.append(_competition_peak(f"P{i}", 6, shared, neutral_offset=i * 0.01))
        peaks.append(_competition_peak(f"P{i}", 7, f"unique_{i}", neutral_offset=i * 0.01))
    rows, diag = _qc(candidates, peaks, config=_generic_config())
    assert len(rows) == 500
    assert diag["Competing_Envelope_Group_Count"] <= 500
    assert diag["Competitive_Scoring_Time_Seconds"] < 5.0


def _dry_row(cluster_id, peaks, charges, score, group="CG00001", peak_charge_map=None):
    return {
        "Cluster_ID": cluster_id,
        "Competing_Envelope_Group_ID": group,
        "Competing_Envelope_Group_Size": 1,
        "Envelope_Evidence_Score": score,
        "Quality_Tier_Rank": 1,
        "Num_Supporting_Charge_States": len(charges),
        "Envelope_Internal_Error_ppm": 0.0,
        "Reconstructed_Mass": 15000.0,
        "_supporting_local_peak_id_set": set(peaks),
        "_supporting_charge_set": set(charges),
        "_supporting_peak_charge_map": peak_charge_map or {},
        "Intact_Quality_Tier": "Tier_1_high_quality",
        "Comparison_Ready": True,
        "Comparison_Representative": True,
    }


def _dry_config(**overrides):
    return {"competitive_assignment": {
        "enabled": True,
        "dry_run": True,
        "min_independent_peak_fraction": 0.5,
        "min_independent_charge_states": 2,
        "allow_shared_peaks_between_selected": False,
        "minimum_score_margin_for_exclusive_selection": 1.0,
        **overrides,
    }}


def test_mvp598b_noncompeting_and_group_local_selection_order():
    rows = [
        _dry_row("A", {"a1", "a2"}, {6, 7}, 10, group="G1"),
        _dry_row("B", {"b1", "b2"}, {6, 7}, 9, group="G2"),
    ]
    apply_assignment_dry_run(rows, _dry_config())
    assert [row["Dry_Run_Assignment_Status"] for row in rows] == ["noncompeting", "noncompeting"]
    assert [row["Dry_Run_Selection_Order"] for row in rows] == [1, 1]


def test_mvp598b_direct_competitors_only_and_transitive_nonsharing_survives():
    rows = [
        _dry_row("A", {"ab", "a2"}, {6, 7}, 30),
        _dry_row("B", {"ab", "bc"}, {6, 7}, 20),
        _dry_row("C", {"bc", "c2"}, {6, 7}, 10),
    ]
    apply_assignment_dry_run(rows, _dry_config(min_independent_peak_fraction=0.6))
    by_id = {row["Cluster_ID"]: row for row in rows}
    assert by_id["A"]["Dry_Run_Assignment_Status"] == "selected_primary"
    assert by_id["A"]["Direct_Competitor_Cluster_IDs"] == "B"
    assert by_id["C"]["Direct_Competitor_Cluster_IDs"] == "B"
    assert by_id["C"]["Dry_Run_Assignment_Status"] == "selected_primary"
    assert by_id["C"]["Dry_Run_Selected"] is True


def test_mvp598b_primary_independent_peak_shortage_and_charge_shortage():
    independent = [
        _dry_row("A", {"shared", "a2", "a3"}, {6, 7, 8}, 30,
                 peak_charge_map={"shared": {6}, "a2": {7}, "a3": {8}}),
        _dry_row("B", {"shared", "b2", "b3"}, {6, 9, 10}, 20,
                 peak_charge_map={"shared": {6}, "b2": {9}, "b3": {10}}),
    ]
    apply_assignment_dry_run(independent, _dry_config())
    assert independent[0]["Dry_Run_Assignment_Status"] == "selected_primary"
    assert independent[1]["Dry_Run_Assignment_Status"] == "selected_independent"

    peak_short = [_dry_row("A", {"p1", "p2"}, {6, 7}, 30), _dry_row("B", {"p1", "p2"}, {8, 9}, 20)]
    apply_assignment_dry_run(peak_short, _dry_config())
    assert peak_short[1]["Dry_Run_Assignment_Status"] == "would_exclude_peak_reuse"

    charge_short = [
        _dry_row("A", {"p1", "a2"}, {6, 7}, 30, peak_charge_map={"p1": {6}, "a2": {7}}),
        _dry_row("B", {"p1", "b2"}, {6}, 20, peak_charge_map={"p1": {6}, "b2": {6}}),
    ]
    apply_assignment_dry_run(charge_short, _dry_config())
    assert charge_short[1]["Dry_Run_Assignment_Status"] == "would_exclude_independent_charge_shortage"


def test_mvp598b_close_score_ambiguity_and_confident_margin():
    close = [_dry_row("A", {"p"}, {6, 7}, 10.5), _dry_row("B", {"p"}, {6, 7}, 10.0)]
    apply_assignment_dry_run(close, _dry_config())
    assert close[1]["Close_Score_Ambiguity"] is True
    assert close[1]["Assignment_Confidence"] == "ambiguous"
    assert close[1]["Dry_Run_Exclusion_Reason"] == "ambiguous_close_score"
    clear = [_dry_row("A", {"p"}, {6, 7}, 20), _dry_row("B", {"p"}, {6, 7}, 10)]
    apply_assignment_dry_run(clear, _dry_config())
    assert clear[1]["Close_Score_Ambiguity"] is False
    assert clear[1]["Assignment_Confidence"] == "high"


def test_mvp598b_observed_peak_and_actual_peak_charge_are_distinct():
    rows = [
        _dry_row("A", {"same"}, {6}, 20, peak_charge_map={"same": {6}}),
        _dry_row("B", {"same"}, {7}, 10, peak_charge_map={"same": {7}}),
    ]
    apply_assignment_dry_run(rows, _dry_config(min_independent_charge_states=1))
    assert rows[0]["Shared_Observed_Peak_Count"] == 1
    assert rows[0]["Shared_Peak_Charge_Assignment_Count"] == 0
    assert rows[1]["Independent_Observed_Peak_Count"] == 0


def test_mvp598b_allow_shared_peaks_allows_selected_secondary():
    rows = [_dry_row("A", {"p"}, {6, 7}, 20), _dry_row("B", {"p"}, {6, 7}, 10)]
    apply_assignment_dry_run(rows, _dry_config(allow_shared_peaks_between_selected=True))
    assert rows[1]["Dry_Run_Assignment_Status"] == "selected_independent"
    assert rows[1]["Independent_Supporting_Peak_Fraction"] == 1.0


def test_mvp598b_empty_disabled_and_legacy_shape_are_safe():
    stats = apply_assignment_dry_run([], _dry_config())
    assert stats["Dry_Run_Selected_Candidate_Count"] == 0
    row = _dry_row("LEGACY", {"p1", "p2"}, {6, 7}, 1)
    apply_assignment_dry_run([row], _dry_config(enabled=False))
    assert row["Dry_Run_Assignment_Status"] == "not_evaluated"
    assert row["Dry_Run_Selected"] is False


def test_mvp598b_dry_run_does_not_change_qc_representative_or_spectrum():
    candidates = [_competition_candidate("A", [6, 7, 8]), _competition_candidate("B", [6, 7], mass=15000.1)]
    peaks = [
        _competition_peak("A", 6, "shared"), _competition_peak("A", 7, "a7"), _competition_peak("A", 8, "a8"),
        _competition_peak("B", 6, "shared", neutral_offset=0.1), _competition_peak("B", 7, "b7", neutral_offset=0.1),
    ]
    rows, _ = _qc(candidates, peaks, config=_generic_config())
    before = [(r["Intact_Quality_Tier"], r["Comparison_Ready"], r["Comparison_Representative"]) for r in rows]
    spectrum_before = build_reconstructed_mass_spectrum_rows(rows, _generic_config()["intact_reconstruction"])
    apply_assignment_dry_run(rows, _dry_config())
    after = [(r["Intact_Quality_Tier"], r["Comparison_Ready"], r["Comparison_Representative"]) for r in rows]
    spectrum_after = build_reconstructed_mass_spectrum_rows(rows, _generic_config()["intact_reconstruction"])
    assert after == before
    assert spectrum_after == spectrum_before


def test_mvp598b_reference_and_target_annotations_do_not_change_assignment():
    peaks = [_competition_peak("A", 6, "shared"), _competition_peak("A", 7, "a7"), _competition_peak("B", 6, "shared", neutral_offset=0.1), _competition_peak("B", 7, "b7", neutral_offset=0.1)]
    def result(config):
        rows, _ = _qc([_competition_candidate("A", [6, 7]), _competition_candidate("B", [6, 7], mass=15000.1)], peaks, config=config)
        return [(r["Cluster_ID"], r["Dry_Run_Assignment_Status"], r["Dry_Run_Selected"], r["Assignment_Confidence"]) for r in rows]
    base = result(_generic_config())
    annotated = result(_generic_config(reference_masses=[{"label": "ref", "mass_da": 15000.0}], target_review_mass_range={"enabled": True, "min_da": 14900, "max_da": 15100}))
    assert annotated == base


def test_mvp598b_large_component_keeps_distant_candidates_and_is_fast():
    rows = []
    for index in range(300):
        peaks = {f"edge_{index - 1}", f"edge_{index}"}
        rows.append(_dry_row(f"C{index:04d}", peaks, {6, 7}, 1000 - index))
    stats = apply_assignment_dry_run(rows, _dry_config())
    assert rows[0]["Dry_Run_Selected"] is True
    assert rows[2]["Dry_Run_Selected"] is True
    assert stats["Dry_Run_Selected_Candidate_Count"] >= 150
    assert stats["Assignment_Dry_Run_Time_Seconds"] < 5.0



def _sensitivity_config(audit=None, enabled=True):
    competitive = _dry_config()["competitive_assignment"]
    competitive["sensitivity_analysis"] = {"enabled": enabled, "scenarios": ["strict", "balanced", "sensitive", "permissive"]}
    competitive["audit_masses"] = audit or {"enabled": False, "tolerance_da": 2.0, "masses": []}
    return {"competitive_assignment": competitive}


def _prepare_dry_graph(rows):
    apply_assignment_dry_run(rows, _dry_config())
    return rows


def test_mvp598c_scenario_selection_monotonicity_and_defaults():
    rows = _prepare_dry_graph([
        _dry_row("A", {"p1", "a2"}, {6, 7}, 20, peak_charge_map={"p1": {6}, "a2": {7}}),
        _dry_row("B", {"p1", "b2"}, {6, 8}, 10, peak_charge_map={"p1": {6}, "b2": {8}}),
    ])
    stats = run_assignment_sensitivity(rows, _sensitivity_config())
    summary = {row["Scenario"]: row for row in stats["_sensitivity_summary_rows"]}
    assert summary["strict"]["Selected_Total"] <= summary["balanced"]["Selected_Total"]
    assert summary["permissive"]["Selected_Total"] >= summary["balanced"]["Selected_Total"]
    assert stats["Sensitivity_Scenario_Count"] == 4


def test_mvp598c_stability_statuses_and_noncompeting():
    stable_selected = _prepare_dry_graph([_dry_row("S", {"s1", "s2"}, {6, 7}, 20)])
    run_assignment_sensitivity(stable_selected, _sensitivity_config())
    assert stable_selected[0]["Selection_Stability_Status"] == "noncompeting"
    assert stable_selected[0]["Selection_Stability_Count"] == 4

    excluded = _prepare_dry_graph([
        _dry_row("A", {"p"}, {6}, 20, peak_charge_map={"p": {6}}),
        _dry_row("E", {"p"}, set(), 10, peak_charge_map={}),
    ])
    run_assignment_sensitivity(excluded, _sensitivity_config())
    assert excluded[1]["Selection_Stability_Status"] == "stable_excluded"

    sensitive = _prepare_dry_graph([
        _dry_row("A", {"p", "a"}, {6, 7}, 20, peak_charge_map={"p": {6}, "a": {7}}),
        _dry_row("T", {"p", "t"}, {6, 8}, 10, peak_charge_map={"p": {6}, "t": {8}}),
    ])
    run_assignment_sensitivity(sensitive, _sensitivity_config())
    assert sensitive[1]["Selection_Stability_Status"] == "threshold_sensitive"

    ambiguous = _prepare_dry_graph([
        _dry_row("A", {"p"}, {6, 7}, 10.75),
        _dry_row("Q", {"p"}, {6, 7}, 10.0),
    ])
    run_assignment_sensitivity(ambiguous, _sensitivity_config())
    assert ambiguous[1]["Selection_Stability_Status"] == "ambiguous_across_scenarios"


def test_mvp598c_scenarios_preserve_original_qc_and_reuse_graph():
    rows = _prepare_dry_graph([
        _dry_row("A", {"p", "a"}, {6, 7}, 20),
        _dry_row("B", {"p", "b"}, {6, 8}, 10),
    ])
    protected = [{key: row.get(key) for key in ["Competing_Envelope_Group_ID", "Envelope_Evidence_Score", "Direct_Competitor_Cluster_IDs", "Dry_Run_Assignment_Status", "Dry_Run_Selected", "Intact_Quality_Tier", "Comparison_Ready", "Comparison_Representative"]} for row in rows]
    run_assignment_sensitivity(rows, _sensitivity_config())
    after = [{key: row.get(key) for key in item} for row, item in zip(rows, protected)]
    assert after == protected


def test_mvp598c_audit_disabled_empty_tolerance_and_no_assignment_effect():
    rows = _prepare_dry_graph([_dry_row("A", {"a1", "a2"}, {6, 7}, 20)])
    baseline = (rows[0]["Dry_Run_Assignment_Status"], rows[0]["Dry_Run_Selected"])
    disabled = run_assignment_sensitivity(rows, _sensitivity_config())
    assert disabled["_audit_rows"] == []
    empty = run_assignment_sensitivity(rows, _sensitivity_config({"enabled": True, "tolerance_da": 1.0, "masses": []}))
    assert empty["_audit_rows"] == []
    enabled = run_assignment_sensitivity(rows, _sensitivity_config({"enabled": True, "tolerance_da": 0.5, "masses": [{"label": "review", "mass_da": 15000.2}]}))
    assert len(enabled["_audit_rows"]) == 1
    assert enabled["_audit_rows"][0]["Audit_Mass_Label"] == "review"
    assert (rows[0]["Dry_Run_Assignment_Status"], rows[0]["Dry_Run_Selected"]) == baseline


def test_mvp598c_reference_target_tier_ready_and_spectrum_invariants():
    peaks = [_competition_peak("A", 6, "shared"), _competition_peak("A", 7, "a7"), _competition_peak("B", 6, "shared", neutral_offset=0.1), _competition_peak("B", 7, "b7", neutral_offset=0.1)]
    def result(config):
        rows, _ = _qc([_competition_candidate("A", [6, 7]), _competition_candidate("B", [6, 7], mass=15000.1)], peaks, config=config)
        spectrum = build_reconstructed_mass_spectrum_rows(rows, config["intact_reconstruction"])
        return rows, spectrum
    base_rows, base_spectrum = result(_generic_config())
    annotated_rows, annotated_spectrum = result(_generic_config(reference_masses=[{"label": "ref", "mass_da": 15000}], target_review_mass_range={"enabled": True, "min_da": 14900, "max_da": 15100}))
    fields = ["Selected_Strict", "Selected_Balanced", "Selected_Sensitive", "Selected_Permissive", "Selection_Stability_Status", "Intact_Quality_Tier", "Comparison_Ready"]
    assert [[r[f] for f in fields] for r in annotated_rows] == [[r[f] for f in fields] for r in base_rows]
    assert len(annotated_spectrum) == len(base_spectrum)


def test_mvp598c_empty_disabled_and_legacy_engine_safe():
    assert run_assignment_sensitivity([], _sensitivity_config())["Sensitivity_Scenario_Count"] == 4
    row = _prepare_dry_graph([_dry_row("A", {"a"}, {6}, 1)])[0]
    disabled = _sensitivity_config(); disabled["competitive_assignment"]["enabled"] = False
    assert run_assignment_sensitivity([row], disabled)["Sensitivity_Scenario_Count"] == 0
    config = {"enabled": True, "min_charge": 6, "max_charge": 7, "min_charge_states": 2, "mass_cluster_tolerance_da": 1.0, "intact_reconstruction": {"engine": "legacy_cluster", "neutral_mass_range": {"enabled": True, "min_da": 10000, "max_da": 20000}}}
    candidates, peaks, _ = reconstruct_intact_masses(PeakTierResult(major=[_rt_peak(15000.0, z, rt=5.0) for z in [6, 7]]), config, {"polarity": "negative"}, None)
    rows, diagnostics = build_intact_reconstruction_qc(candidates, peaks, config)
    assert rows and "Selection_Stability_Status" in rows[0]
    assert diagnostics[0]["Sensitivity_Scenario_Count"] == 4


def test_mvp598c_excel_sheets_empty_audit_and_names(tmp_path):
    candidate = _competition_candidate("EXCELC", [6, 7, 8])
    peaks = [_competition_peak("EXCELC", z, f"p{z}") for z in [6, 7, 8]]
    report = write_excel_report(output_dir=tmp_path, config=_excel_config({"enabled": True, **_generic_config()}), diagnostics={}, intact_results=[candidate], charge_state_peaks=peaks, warnings=[], modifications=[], rule_set={}, pathways=[], theoretical_fragments=[], fragment_ms1_matches=[], known_modification_candidates=[], known_modification_summary=[], optional_results={})
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        for name in ["Assignment_Sensitivity", "Assignment_Stability", "Assignment_Candidate_Audit"]:
            assert name in workbook.sheetnames
        assert workbook["Assignment_Candidate_Audit"].max_row == 3
        assert all(len(name) <= 31 for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_mvp598c_performance_smoke():
    rows = []
    for index in range(500):
        rows.append(_dry_row(f"C{index:04d}", {f"edge_{index - 1}", f"edge_{index}"}, {6, 7}, 1000-index))
    _prepare_dry_graph(rows)
    stats = run_assignment_sensitivity(rows, _sensitivity_config())
    assert stats["Sensitivity_Scenario_Count"] == 4
    assert stats["Sensitivity_Analysis_Time_Seconds"] < 5.0


def test_mvp598c_competing_candidate_can_be_stable_selected():
    rows = _prepare_dry_graph([
        _dry_row("A", {"shared", "a2", "a3", "a4"}, {6, 7, 8, 9}, 20, peak_charge_map={"shared": {6}, "a2": {7}, "a3": {8}, "a4": {9}}),
        _dry_row("B", {"shared", "b2", "b3", "b4"}, {6, 10, 11, 12}, 10, peak_charge_map={"shared": {6}, "b2": {10}, "b3": {11}, "b4": {12}}),
    ])
    run_assignment_sensitivity(rows, _sensitivity_config())
    assert rows[1]["Selection_Stability_Status"] == "stable_selected"
    assert rows[1]["Selection_Stability_Count"] == 4


def test_mvp598c_sensitivity_does_not_call_grouping_or_evidence_scoring(monkeypatch):
    import rna_masshunter.intact_reconstruction as intact_module
    rows = _prepare_dry_graph([_dry_row("A", {"a1"}, {6}, 10)])
    monkeypatch.setattr(intact_module, "apply_competitive_assignment", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("grouping/scoring recalculated")))
    monkeypatch.setattr(intact_module, "_evidence_score", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("evidence score recalculated")))
    stats = run_assignment_sensitivity(rows, _sensitivity_config())
    assert stats["Sensitivity_Scenario_Count"] == 4
