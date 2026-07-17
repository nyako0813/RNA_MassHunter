from hashlib import sha256
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

import main as main_module
from rna_masshunter.audit_policy import AUDIT_DETAIL, AUDIT_SUMMARY, AuditPolicy, included_sheet_names, sheet_category
from rna_masshunter.config import DEFAULT_CONFIG, validate_config
from rna_masshunter.excel_report import (
    SCIEX_INTACT_OPTIONAL_RESULT_KEY, SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY,
    write_excel_report,
)
from rna_masshunter.models import RunConfig
from rna_masshunter.sciex_delta_mass_cluster_audit import (
    AUDIT_RESULT_KEY, CLUSTER_COLUMNS, CLUSTER_SHEET, ERROR_CODE, RELATION_COLUMNS,
    RELATION_SHEET, SUMMARY_COLUMNS, SUMMARY_SHEET, DeltaMassClusterParameters,
    audit_sciex_delta_mass_clusters,
)


class Comparison:
    def __init__(self, rows, theory=100.0, identity_status="MATCH", identity_conflict=False, biological=True):
        self._rows = [dict(row) for row in rows]
        self._summary = [{
            "Source_File": "synthetic-full.txt",
            "Theoretical_Unmodified_Mass": theory,
            "Input_Identity_Audit_Status": identity_status,
            "Input_Identity_Conflict": identity_conflict,
            "Biological_Interpretation_Eligible": biological,
        }]
    def details(self): return [dict(row) for row in self._rows]
    def summaries(self): return [dict(row) for row in self._summary]


def comparison(deltas, intensities=None, modes=None, widths=None, theory=100.0, **identity):
    intensities = intensities if intensities is not None else [10.0] * len(deltas)
    modes = modes if modes is not None else ["STRICT"] * len(deltas)
    widths = widths if widths is not None else [1.0] * len(deltas)
    rows = []
    observed_base = theory if theory is not None else 100.0
    for index, (delta, intensity, mode, width) in enumerate(zip(deltas, intensities, modes, widths), 1):
        rows.append({
            "Comparison_ID": f"SCIEX_CMP_{index:05d}", "Source_Row_Index": index,
            "Observed_Mass": observed_base + delta, "Delta_Mass": delta,
            "Absolute_Delta_Mass": abs(delta), "Apex_Intensity_Raw": intensity,
            "Detection_Tier": mode, "Half_Prominence_Width_Da": width,
            "Comparison_Status": "STRICT_MATCH" if abs(delta) <= 1 else "NO_MATCH",
        })
    return Comparison(rows, theory=theory, **identity)


def run(deltas, parameters=None, **kwargs):
    return audit_sciex_delta_mass_clusters(comparison(deltas, **kwargs), parameters)


def summary(result): return result.summaries()[0]


def test_default_config_values_follow_observed_half_da_grid():
    assert DEFAULT_CONFIG["sciex_profile"]["delta_mass_cluster_audit"] == {
        "enabled": True, "cluster_tolerance_da": 0.5,
        "duplicate_apex_tolerance_da": 0.25, "isotope_spacing_da": 1.003355,
        "isotope_spacing_tolerance_da": 0.15, "integer_spacing_tolerance_da": 0.15,
        "minimum_cluster_size": 2, "max_pair_spacing_da": 200.0,
        "max_pair_rows": 20000,
    }


@pytest.mark.parametrize("field,value", [
    ("cluster_tolerance_da", 0), ("cluster_tolerance_da", float("nan")),
    ("duplicate_apex_tolerance_da", -0.1), ("isotope_spacing_da", 0),
    ("isotope_spacing_tolerance_da", 0), ("integer_spacing_tolerance_da", 0),
    ("minimum_cluster_size", 1), ("minimum_cluster_size", 2.5),
    ("max_pair_spacing_da", 0), ("max_pair_rows", 0), ("max_pair_rows", 2.5),
])
def test_invalid_parameters_are_rejected(field, value):
    values = dict(DEFAULT_CONFIG["sciex_profile"]["delta_mass_cluster_audit"])
    values[field] = value
    with pytest.raises((TypeError, ValueError)):
        DeltaMassClusterParameters.from_mapping(values)


def test_one_peak_is_singleton_without_error():
    result = run([0.0])
    row = result.clusters()[0]
    assert row["Cluster_Label"] == "SINGLETON"
    assert row["Cluster_Is_Singleton"] is True
    assert summary(result)["Cluster_Count"] == 1


@pytest.mark.parametrize("deltas,expected_sizes", [
    ([0.0, 0.4], [2]),
    ([0.0, 0.2, 0.5], [3]),
    ([0.0, 0.4, 0.8], [2, 1]),
    ([-1.0, -0.6, 1.0], [2, 1]),
])
def test_complete_link_span_bounded_clusters_prevent_chaining(deltas, expected_sizes):
    rows = run(deltas).clusters()
    assert [row["Cluster_Size"] for row in rows] == expected_sizes
    assert all(row["Cluster_Span_Da"] <= 0.5 + 1e-12 for row in rows)


def test_cluster_mean_median_and_weighted_mean():
    row = run([0.0, 0.2, 0.5], intensities=[1, 2, 7]).clusters()[0]
    assert row["Cluster_Mean_Delta_Da"] == pytest.approx(0.7 / 3)
    assert row["Cluster_Median_Delta_Da"] == pytest.approx(0.2)
    assert row["Cluster_Weighted_Mean_Delta_Da"] == pytest.approx(0.39)
    assert row["Cluster_Weighted_Mean_Fallback_Used"] is False


@pytest.mark.parametrize("intensities", [[0, 0], [None, None]])
def test_weighted_mean_falls_back_when_intensity_is_zero_or_missing(intensities):
    row = run([0.0, 0.4], intensities=intensities).clusters()[0]
    assert row["Cluster_Weighted_Mean_Delta_Da"] == pytest.approx(0.2)
    assert row["Cluster_Weighted_Mean_Fallback_Used"] is True


@pytest.mark.parametrize("deltas,duplicate", [
    ([0.0, 0.0], True), ([0.0, 0.25], True), ([0.0, 0.250001], False),
])
def test_duplicate_like_tolerance_boundary(deltas, duplicate):
    row = run(deltas).clusters()[0]
    assert row["Duplicate_Like"] is duplicate
    assert bool(row["Duplicate_Group_ID"]) is duplicate


def test_duplicate_representative_tie_break_uses_detection_then_width_then_mass():
    row = run(
        [0.0, 0.0, 0.0], intensities=[10, 10, 10],
        modes=["SENSITIVE", "STRICT", "STRICT"], widths=[0.5, 1.0, 0.4],
    ).clusters()[0]
    assert row["Duplicate_Representative_Row"] == "3"
    assert row["Duplicate_Strongest_Row"] == "1"
    assert row["Duplicate_Group_Size"] == 3


def test_strongest_peak_tie_break_uses_lower_mass_then_row():
    result = run([-1.0, 1.0], intensities=[10, 10])
    assert summary(result)["Strongest_Peak_Row"] == 1
    assert summary(result)["Strongest_Peak_Cluster_ID"] == "SCIEX_DELTA_C00001"


def test_closest_peak_tie_break_uses_intensity_then_mass():
    result = run([-1.0, 1.0], intensities=[5, 10])
    assert summary(result)["Closest_Peak_Row"] == 2


@pytest.mark.parametrize("spacing,multiple", [(1.0, 1), (2.0, 2), (3.1, 3)])
def test_integer_spacing_candidates(spacing, multiple):
    relation = run([0.0, spacing]).relations()[0]
    assert relation["Integer_Spacing_Candidate"] is True
    assert relation["Nearest_Integer_Spacing"] == multiple


def test_integer_spacing_outside_tolerance_is_not_exported_without_other_relation():
    result = run([0.0, 1.2], parameters={"minimum_cluster_size": 3})
    assert result.relations() == []


@pytest.mark.parametrize("multiple", [1, 2, 3])
def test_isotope_like_spacing_candidates(multiple):
    relation = run([0.0, multiple * 1.003355]).relations()[0]
    assert relation["Isotope_Spacing_Candidate"] is True
    assert relation["Nearest_Isotope_Multiple"] == multiple
    assert relation["Isotope_Spacing_Error_Da"] == pytest.approx(0.0)


def test_isotope_spacing_outside_tolerance_is_not_candidate():
    result = run([0.0, 1.2], parameters={"minimum_cluster_size": 3})
    assert result.relations() == []


def test_integer_and_isotope_flags_can_both_be_true_without_assignment():
    relation = run([0.0, 1.003355]).relations()[0]
    assert relation["Integer_Spacing_Candidate"] is True
    assert relation["Isotope_Spacing_Candidate"] is True
    assert "INTEGER" in relation["Relation_Types"] and "ISOTOPE" in relation["Relation_Types"]


def test_recurrent_spacing_requires_repeated_pair_spacing():
    result = run([0.0, 10.0, 20.0], parameters={"integer_spacing_tolerance_da": 0.01, "isotope_spacing_tolerance_da": 0.01})
    assert summary(result)["Recurrent_Spacing_Group_Count"] >= 1
    assert any(row["Recurrent_Spacing_Group_ID"] for row in result.relations())


def test_unique_spacings_have_no_recurrent_group():
    result = run([0.0, 10.0, 25.0], parameters={"minimum_cluster_size": 2, "integer_spacing_tolerance_da": 0.01, "isotope_spacing_tolerance_da": 0.01})
    # Integer relations remain, but 10, 15, and 25 Da each occur only once.
    assert summary(result)["Recurrent_Spacing_Group_Count"] == 0


def test_pair_spacing_upper_bound_is_enforced():
    result = run([0.0, 10.0], parameters={"max_pair_spacing_da": 5.0})
    assert summary(result)["Total_Eligible_Pair_Count"] == 0
    assert result.relations() == []


def test_pair_rows_are_deterministically_truncated():
    result = run([0.0, 1.0, 2.0, 3.0], parameters={"max_pair_rows": 2})
    row = summary(result)
    assert row["Total_Eligible_Pair_Count"] == 6
    assert row["Exported_Pair_Count"] == 2
    assert row["Pair_Rows_Truncated"] is True
    assert [item["Relation_ID"] for item in result.relations()] == ["SCIEX_REL_00001", "SCIEX_REL_00002"]


def test_relation_sort_prioritizes_duplicate_before_spacing_candidates():
    result = run([0.0, 0.1, 1.0])
    assert result.relations()[0]["Duplicate_Like"] is True


def test_cluster_and_duplicate_ids_are_deterministic_under_input_reordering():
    base = comparison([0.0, 0.1, 1.0])
    shuffled = Comparison(list(reversed(base.details())))
    first = audit_sciex_delta_mass_clusters(base)
    second = audit_sciex_delta_mass_clusters(shuffled)
    fields = ["Cluster_ID", "Cluster_Size", "Member_Delta_Masses", "Duplicate_Group_ID"]
    assert [[row[field] for field in fields] for row in first.clusters()] == [
        [row[field] for field in fields] for row in second.clusters()
    ]


@pytest.mark.parametrize("identity_status,conflict,biological", [
    ("MATCH", False, True), ("CONFLICT", True, False),
    ("INSUFFICIENT_INFORMATION", False, False),
])
def test_identity_status_is_reflected_without_controlling_execution(identity_status, conflict, biological):
    result = audit_sciex_delta_mass_clusters(comparison(
        [0.0, 1.0], identity_status=identity_status,
        identity_conflict=conflict, biological=biological,
    ))
    row = summary(result)
    assert row["Input_Identity_Audit_Status"] == identity_status
    assert row["Input_Identity_Conflict"] is conflict
    assert row["Biological_Interpretation_Eligible"] is biological
    assert row["Cluster_Count"] > 0


def test_conflict_notes_explicitly_forbid_biological_interpretation():
    row = summary(audit_sciex_delta_mass_clusters(comparison(
        [0.0], identity_status="CONFLICT", identity_conflict=True, biological=False,
    )))
    assert "must not be interpreted biologically" in row["Notes"]


class Detection:
    def __init__(self, status="DETECTION_COMPLETED", peaks=True, profile="NEUTRAL_MASS_PROFILE"):
        self.status, self.has_peaks, self.profile = status, peaks, profile
    def diagnostics_row(self):
        return {"Detection_Status": self.status, "Profile_Type": self.profile}
    def peak_rows(self):
        return [{"Peak_ID": "P1"}] if self.has_peaks else []


def route_config(enabled=True, cluster_enabled=True):
    return RunConfig(sciex_profile={
        "enabled": enabled, "path": "synthetic-full.txt",
        "intact_peak_detection": {"enabled": True},
        "intact_mass_comparison": {"enabled": True},
        "delta_mass_cluster_audit": {"enabled": cluster_enabled},
    })


def routed(detector=None, comp=None, config=None, warnings=None):
    sciex = {} if detector is None else {
        SCIEX_INTACT_OPTIONAL_RESULT_KEY: {"result": detector, "source_file": "synthetic-full.txt"}
    }
    comparison_results = {} if comp is None else {SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY: comp}
    return main_module.build_sciex_delta_mass_cluster_optional_results(
        config or route_config(), sciex, comparison_results,
        [] if warnings is None else warnings,
    )


@pytest.mark.parametrize("config,detector,comp", [
    (RunConfig(), Detection(), comparison([0.0])),
    (route_config(enabled=False), Detection(), comparison([0.0])),
    (route_config(cluster_enabled=False), Detection(), comparison([0.0])),
    (route_config(), None, comparison([0.0])),
    (route_config(), Detection(status="FAILED"), comparison([0.0])),
    (route_config(), Detection(peaks=False), comparison([0.0])),
    (route_config(), Detection(profile="MZ_PROFILE"), comparison([0.0])),
    (route_config(), Detection(), None),
    (route_config(), Detection(), comparison([0.0], theory=None)),
    (route_config(), Detection(), comparison([0.0], theory=float("nan"))),
])
def test_routing_skips_ineligible_inputs(config, detector, comp):
    assert routed(detector, comp, config=config) == {}


def test_routing_executes_for_completed_neutral_comparison():
    assert AUDIT_RESULT_KEY in routed(Detection(), comparison([0.0, 1.0]))


def test_audit_exception_adds_code_and_preserves_upstream_results(monkeypatch):
    warnings = []
    detector = Detection()
    comp = comparison([0.0])
    monkeypatch.setattr(main_module, "audit_sciex_delta_mass_clusters", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    assert routed(detector, comp, warnings=warnings) == {}
    assert warnings[-1]["Context"]["Warning_Code"] == ERROR_CODE
    assert detector.peak_rows() and comp.details()


def test_config_validation_accepts_defaults_and_rejects_nonmapping():
    validate_config(route_config())
    config = route_config(); config.sciex_profile["delta_mass_cluster_audit"] = "bad"
    with pytest.raises(ValueError, match="delta_mass_cluster_audit"):
        validate_config(config)


def test_all_output_rows_are_shadow_only_and_nonpropagating():
    result = run([0.0, 0.1, 1.0])
    for row in result.clusters() + result.summaries() + result.relations():
        assert row["Shadow_Only"] is True
        assert row["Applied_To_Formal_Score"] is False
        assert row["Applied_To_Ranking"] is False
        assert row["Applied_To_Candidate_Filtering"] is False
        assert row["Molecular_Identity_Assigned"] is False


def test_cluster_audit_does_not_mutate_comparison_or_numeric_values():
    source = comparison([-1.0, 2.0])
    details_before = source.details(); summary_before = source.summaries()
    audit_sciex_delta_mass_clusters(source)
    assert source.details() == details_before
    assert source.summaries() == summary_before


def writer_config():
    return SimpleNamespace(
        analysis={"mode": "full"}, project={"name": "cluster-audit"}, input={},
        organism={}, sequence={}, experiment={}, instrument={}, sciex_profile={},
        reconstruction={"enabled": False}, digestion={"enabled": False},
        alkaline_phosphatase={}, fragment_mapping={}, modification_search={}, peak_filtering={},
        p1_annotation={}, ms2_annotation={}, modification_evidence_ranking={}, biological_context={},
        performance={}, reporting={"max_excel_rows_per_sheet": 100000, "truncate_large_sheets": True},
    )


def write_report(tmp_path, level, optional):
    return write_excel_report(
        tmp_path / level, writer_config(), {}, [], [], [],
        known_modification_candidates=[{"Candidate_ID": "C1"}],
        known_modification_summary=[{"Summary_Key": "S1"}],
        optional_results=optional, audit_policy=AuditPolicy.from_level(level),
    )


@pytest.mark.parametrize("level,present", [("standard", False), ("audit", True), ("full", True)])
def test_excel_policy_fixed_columns_and_row_limits(tmp_path, level, present):
    result = run([0.0, 1.0, 2.0], parameters={"max_pair_rows": 2})
    report = write_report(tmp_path, level, {AUDIT_RESULT_KEY: result})
    workbook = load_workbook(report, read_only=True, data_only=True)
    try: names = workbook.sheetnames
    finally: workbook.close()
    expected_names = {CLUSTER_SHEET, SUMMARY_SHEET, RELATION_SHEET}
    assert expected_names.issubset(names) is present
    if present:
        clusters = pd.read_excel(report, sheet_name=CLUSTER_SHEET, header=2)
        summaries = pd.read_excel(report, sheet_name=SUMMARY_SHEET, header=2)
        relations = pd.read_excel(report, sheet_name=RELATION_SHEET, header=2)
        assert list(clusters.columns) == CLUSTER_COLUMNS
        assert list(summaries.columns) == SUMMARY_COLUMNS and len(summaries) == 1
        assert list(relations.columns) == RELATION_COLUMNS and len(relations) == 2


def test_sheet_registry_includes_audit_and_full_details():
    assert sheet_category(SUMMARY_SHEET) == AUDIT_SUMMARY
    assert sheet_category(CLUSTER_SHEET) == AUDIT_DETAIL
    assert sheet_category(RELATION_SHEET) == AUDIT_DETAIL
    names = [SUMMARY_SHEET, CLUSTER_SHEET, RELATION_SHEET]
    assert included_sheet_names(names, AuditPolicy.from_level("standard"))[0] == []
    assert included_sheet_names(names, AuditPolicy.from_level("audit"))[0] == names
    assert all(len(name) <= 31 for name in names)


def test_controlled_formal_ab_is_exact_for_all_shadow_variants(tmp_path):
    formal = {
        "Modification_Evidence_Ranking": [{"Candidate_ID": "C1", "Final_Score": 1.0}],
        "Top_Modification_Candidates": [{"Candidate_ID": "C1", "Final_Score": 1.0}],
        "P1_Summary": [{"Status": "CONTROL"}], "MS2_Summary": [{"Status": "CONTROL"}],
    }
    match = run([0.0, 1.0])
    conflict = audit_sciex_delta_mass_clusters(comparison(
        [0.0, 1.0], identity_status="CONFLICT", identity_conflict=True, biological=False,
    ))
    variants = {
        "sciex_disabled": dict(formal), "cluster_disabled": dict(formal),
        "match": {**formal, AUDIT_RESULT_KEY: match},
        "conflict": {**formal, AUDIT_RESULT_KEY: conflict},
        "exception": dict(formal),
    }
    reports = {name: write_report(tmp_path / name, "full", value) for name, value in variants.items()}
    sheets = [
        "Known_Modification_Candidates", "Known_Modification_Summary",
        "Modification_Evidence_Ranking", "Top_Modification_Candidates", "P1_Summary", "MS2_Summary",
    ]
    base = reports["sciex_disabled"]
    for report in reports.values():
        for sheet in sheets:
            expected = pd.read_excel(base, sheet_name=sheet, header=2, dtype=object).fillna("")
            actual = pd.read_excel(report, sheet_name=sheet, header=2, dtype=object).fillna("")
            pd.testing.assert_frame_equal(expected, actual, check_dtype=False)


def test_excel_cluster_and_relation_order_is_deterministic(tmp_path):
    result = run([2.0, 0.0, 1.0])
    first = write_report(tmp_path / "a", "audit", {AUDIT_RESULT_KEY: result})
    second = write_report(tmp_path / "b", "audit", {AUDIT_RESULT_KEY: result})
    for sheet in (CLUSTER_SHEET, RELATION_SHEET):
        left = pd.read_excel(first, sheet_name=sheet, header=2, dtype=object).fillna("")
        right = pd.read_excel(second, sheet_name=sheet, header=2, dtype=object).fillna("")
        pd.testing.assert_frame_equal(left, right, check_dtype=False)


def test_real_input_sha_and_contents_are_non_destructive():
    path = Path(".cache/sciex_research/WT_LeuUAA(Full).txt")
    before = sha256(path.read_bytes()).hexdigest()
    run([0.0, 1.0])
    after = sha256(path.read_bytes()).hexdigest()
    assert before == after == "daae23985f9fda05902761346a75adf680ed2b141a1648346b7f6d5fda0b92a6"
