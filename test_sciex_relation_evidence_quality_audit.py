from hashlib import sha256
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

import main as main_module
from rna_masshunter.audit_policy import AUDIT_DETAIL, AUDIT_SUMMARY, AuditPolicy, included_sheet_names, sheet_category
from rna_masshunter.config import DEFAULT_CONFIG, validate_config
from rna_masshunter.excel_report import write_excel_report
from rna_masshunter.models import RunConfig
from rna_masshunter.sciex_delta_mass_cluster_audit import SciexDeltaMassClusterAuditResult
from rna_masshunter.sciex_relation_evidence_quality_audit import (
    ALGORITHM_VERSION, AUDIT_RESULT_KEY, DEFAULT_PARAMETERS, DETAIL_COLUMNS,
    DETAIL_SHEET, ERROR_CODE, FORMAL_FALSE, SUMMARY_COLUMNS, SUMMARY_SHEET,
    RelationEvidenceQualityParameters, annotate_cluster_summary,
    annotate_resolution_summary, audit_sciex_relation_evidence_quality,
)
from rna_masshunter.sciex_spacing_resolution_audit import SciexSpacingResolutionAuditResult

REAL_INPUT = Path(".cache/sciex_research/WT_LeuUAA(Full).txt")
REAL_SHA256 = "daae23985f9fda05902761346a75adf680ed2b141a1648346b7f6d5fda0b92a6"
CLUSTER_KEY = "sciex_delta_mass_cluster_audit"
RESOLUTION_KEY = "sciex_spacing_resolution_audit"
CLUSTER_PARAMETERS = {
    "integer_spacing_tolerance_da": 0.15,
    "isotope_spacing_tolerance_da": 0.15,
}


def relation(
    relation_id="R1", integer=True, isotope=False, integer_error=0.0,
    isotope_error=0.0, group="", spacing=1.0,
):
    return {
        "Relation_ID": relation_id,
        "Peak_Row_A": 1,
        "Peak_Row_B": 2,
        "Observed_Mass_A": 100.0,
        "Observed_Mass_B": 100.0 + spacing,
        "Pair_Mass_Spacing_Da": spacing,
        "Pair_Delta_Spacing_Da": spacing,
        "Integer_Spacing_Candidate": integer,
        "Nearest_Integer_Spacing": 1 if integer else None,
        "Integer_Spacing_Error_Da": integer_error if integer else None,
        "Isotope_Spacing_Candidate": isotope,
        "Nearest_Isotope_Multiple": 1 if isotope else None,
        "Isotope_Spacing_Error_Da": isotope_error if isotope else None,
        "Recurrent_Spacing_Group_ID": group,
    }


def cluster_result(rows=None, count=2):
    rows = [relation()] if rows is None else rows
    return SciexDeltaMassClusterAuditResult(
        tuple({"Cluster_ID": f"C{i + 1}"} for i in range(count)),
        ({
            "SCIEX_Source_File": "synthetic.txt", "Cluster_Count": count,
            "Relation_Evidence_Audit_Status": "NOT_RUN",
            "Highest_Available_Evidence_Tier": "",
            "Resolution_Ambiguous_Relation_Count": 0,
            "Interpretation_Eligible_Relation_Count": 0,
        },),
        tuple(rows),
    )


def resolution_result(
    status="NOT_DISTINGUISHABLE_BOTH", distinguishable=False, margin=0.00671,
    identity="MATCH", conflict=False, biological=True, chemical=False,
    include_details=True,
):
    summary = {
        "SCIEX_Source_File": "synthetic.txt",
        "Estimated_Effective_Grid_Da": 0.5,
        "Grid_Confidence": "HIGH",
        "Resolution_Status": status,
        "Theoretically_Distinguishable": distinguishable,
        "Single_Step_Grid_Steps_Per_Separation": margin,
        "Isotope_Interpretation_Eligible": False,
        "Numerical_Spacing_Interpretation_Eligible": True,
        "Chemical_Interpretation_Eligible": chemical,
        "Input_Identity_Audit_Status": identity,
        "Input_Identity_Conflict": conflict,
        "Biological_Interpretation_Eligible": biological,
        "Relation_Evidence_Audit_Status": "NOT_RUN",
        "Tier_2_Resolution_Limited_Count": 0,
        "Tier_3_Resolution_Supported_Count": 0,
        "Tier_4_Interpretation_Eligible_Count": 0,
    }
    details = ({
        "Spacing_Multiple": 1,
        "Resolution_Status": status,
        "Theoretically_Distinguishable": distinguishable,
        "Grid_Steps_Per_Target_Separation": margin,
    },) if include_details else ()
    return SciexSpacingResolutionAuditResult((summary,), details)


def run(rows=None, resolution=None, cluster_parameters=None, parameters=None):
    return audit_sciex_relation_evidence_quality(
        cluster_result(rows), resolution or resolution_result(),
        CLUSTER_PARAMETERS if cluster_parameters is None else cluster_parameters,
        parameters,
    )


def detail(result, index=0): return result.details()[index]
def summary(result): return result.summaries()[0]


def test_default_config_schema():
    assert DEFAULT_CONFIG["sciex_profile"]["relation_evidence_quality_audit"] == DEFAULT_PARAMETERS


@pytest.mark.parametrize("field,value", [
    ("high_error_fraction_threshold", 0),
    ("high_error_fraction_threshold", float("nan")),
    ("low_error_fraction_threshold", 0),
    ("low_error_fraction_threshold", 1.1),
    ("low_error_fraction_threshold", float("inf")),
    ("minimum_recurrent_support_pairs", 1),
    ("minimum_recurrent_support_pairs", 2.5),
    ("minimum_recurrent_support_pairs", True),
    ("minimum_interpretable_resolution_margin", 0),
    ("minimum_interpretable_resolution_margin", -1),
    ("minimum_interpretable_resolution_margin", float("nan")),
])
def test_invalid_parameters(field, value):
    values = dict(DEFAULT_PARAMETERS); values[field] = value
    with pytest.raises((TypeError, ValueError)):
        RelationEvidenceQualityParameters.from_mapping(values)


def test_threshold_order_is_validated():
    with pytest.raises(ValueError, match="less than"):
        RelationEvidenceQualityParameters.from_mapping({
            "high_error_fraction_threshold": .8,
            "low_error_fraction_threshold": .7,
        })


def route_config(enabled=True, audit_enabled=True):
    return RunConfig(sciex_profile={
        "enabled": enabled, "path": "synthetic.txt",
        "intact_peak_detection": {"enabled": True},
        "intact_mass_comparison": {"enabled": True},
        "delta_mass_cluster_audit": {"enabled": True, **CLUSTER_PARAMETERS},
        "spacing_resolution_audit": {"enabled": True},
        "relation_evidence_quality_audit": {**DEFAULT_PARAMETERS, "enabled": audit_enabled},
    })


def test_config_validation_accepts_defaults(): validate_config(route_config())


def test_config_validation_rejects_nonmapping():
    config=route_config(); config.sciex_profile["relation_evidence_quality_audit"]="bad"
    with pytest.raises(ValueError, match="relation_evidence_quality_audit"):
        validate_config(config)


def test_valid_integer_candidate():
    row=detail(run([relation(integer=True,isotope=False)]))
    assert row["Integer_Numerical_Evidence"] is True
    assert row["Integer_Numerical_Interpretation_Eligible"] is True
    assert row["Integer_Chemical_Interpretation_Eligible"] is False


def test_valid_isotope_candidate():
    row=detail(run([relation(integer=False,isotope=True)]))
    assert row["Isotope_Numerical_Proximity"] is True
    assert row["Isotope_Assignment_Eligible"] is False


def test_dual_candidate():
    row=detail(run([relation(integer=True,isotope=True)]))
    assert row["Integer_Spacing_Candidate"] and row["Isotope_Spacing_Candidate"]


def test_neither_candidate_is_numerical_only():
    row=detail(run([relation(integer=False,isotope=False)]))
    assert row["Evidence_Tier"] == "TIER_1_NUMERICAL_ONLY"
    assert row["Best_Numerical_Relation"] == "NONE"


@pytest.mark.parametrize("kind", ["integer", "isotope"])
def test_exact_zero_error_fraction(kind):
    row=detail(run([relation(integer=kind=="integer", isotope=kind=="isotope")]))
    assert row[f"{kind.title()}_Error_Fraction"] == 0
    assert row["Numerical_Fit_Quality"] == "EXCELLENT"


def test_integer_error_fraction():
    assert detail(run([relation(integer_error=.075)]))["Integer_Error_Fraction"] == pytest.approx(.5)


def test_isotope_error_fraction():
    row=detail(run([relation(integer=False,isotope=True,isotope_error=.03)]))
    assert row["Isotope_Error_Fraction"] == pytest.approx(.2)


@pytest.mark.parametrize("fraction,expected", [
    (.25,"EXCELLENT"),(.5,"GOOD"),(.9,"WEAK"),(1.1,"OUTSIDE_TOLERANCE"),
])
def test_fit_quality_boundaries(fraction,expected):
    row=detail(run([relation(integer_error=.15*fraction)]))
    assert row["Numerical_Fit_Quality"] == expected


@pytest.mark.parametrize("tolerance", [0,-.1])
def test_nonpositive_tolerance_is_not_applicable(tolerance):
    row=detail(run([relation()],cluster_parameters={
        "integer_spacing_tolerance_da":tolerance,"isotope_spacing_tolerance_da":.15,
    }))
    assert row["Integer_Error_Fraction"] is None
    assert row["Numerical_Fit_Quality"] == "NOT_APPLICABLE"


@pytest.mark.parametrize("error", [float("nan"),float("inf"),"bad"])
def test_nonfinite_or_invalid_error_is_not_applicable(error):
    row=detail(run([relation(integer_error=error)]))
    assert row["Integer_Error_Fraction"] is None


def test_tier_zero_for_missing_required_mass():
    row=relation(); row["Observed_Mass_A"]=None
    assert detail(run([row]))["Evidence_Tier"] == "TIER_0_UNUSABLE"


def test_tier_one_for_non_candidate_numerical_relation():
    assert detail(run([relation(integer=False,isotope=False)]))["Evidence_Tier"] == "TIER_1_NUMERICAL_ONLY"


def test_tier_two_for_ambiguous_candidate():
    assert detail(run())["Evidence_Tier"] == "TIER_2_RESOLUTION_LIMITED"


def test_tier_three_for_supported_candidate():
    supported=resolution_result("DISTINGUISHABLE",True,3.0)
    assert detail(run(resolution=supported))["Evidence_Tier"] == "TIER_3_RESOLUTION_SUPPORTED"


def test_tier_four_is_unreachable_by_default():
    supported=resolution_result("DISTINGUISHABLE",True,3.0,chemical=True)
    assert summary(run(resolution=supported))["Tier_4_Count"] == 0


def test_resolution_ambiguous_flag():
    row=detail(run()); assert row["Resolution_Ambiguous"] is True and row["Resolution_Supported"] is False


def test_resolution_supported_flag():
    row=detail(run(resolution=resolution_result("DISTINGUISHABLE",True,3)))
    assert row["Resolution_Ambiguous"] is False and row["Resolution_Supported"] is True


def test_minimum_resolution_margin_blocks_marginal_support():
    row=detail(run(resolution=resolution_result("MARGINALLY_DISTINGUISHABLE",True,1.5)))
    assert row["Resolution_Supported"] is False


def test_identity_conflict_flag_and_audit_continuation():
    row=detail(run(resolution=resolution_result(identity="CONFLICT",conflict=True,biological=False)))
    assert row["Identity_Blocked"] is True and row["Input_Identity_Conflict"] is True


def test_identity_missing_blocks_identity_but_audit_runs():
    row=detail(run(resolution=resolution_result(identity="NOT_RUN",biological=False)))
    assert row["Identity_Blocked"] is True
    assert "INPUT_IDENTITY_UNAVAILABLE" in row["Interpretation_Block_Reasons"]


def test_biological_interpretation_blocked():
    assert detail(run(resolution=resolution_result(biological=False)))["Biological_Interpretation_Blocked"] is True


def test_chemical_interpretation_always_blocked():
    row=detail(run()); assert row["Chemical_Interpretation_Blocked"] is True and row["Chemical_Interpretation_Eligible"] is False


def test_multiple_block_reasons_have_deterministic_order():
    row=detail(run(resolution=resolution_result(identity="CONFLICT",conflict=True,biological=False)))
    assert row["Interpretation_Block_Reasons"] == (
        "RESOLUTION_NOT_DISTINGUISHABLE; INPUT_IDENTITY_CONFLICT; "
        "BIOLOGICAL_INTERPRETATION_BLOCKED; CHEMICAL_ASSIGNMENT_DISABLED"
    )


def test_fit_quality_cannot_override_resolution_ambiguity():
    row=detail(run([relation(integer_error=0)]))
    assert row["Numerical_Fit_Quality"]=="EXCELLENT"
    assert row["Evidence_Tier"]=="TIER_2_RESOLUTION_LIMITED"


def test_best_relation_integer():
    row=detail(run([relation(integer=True,isotope=True,integer_error=.01,isotope_error=.1)]))
    assert row["Best_Numerical_Relation"]=="INTEGER"


def test_best_relation_isotope():
    row=detail(run([relation(integer=True,isotope=True,integer_error=.1,isotope_error=.01)]))
    assert row["Best_Numerical_Relation"]=="ISOTOPE_LIKE"


def test_best_relation_tie_breaks_to_integer():
    row=detail(run([relation(integer=True,isotope=True,integer_error=.01,isotope_error=.01)]))
    assert row["Best_Numerical_Relation"]=="INTEGER"


@pytest.mark.parametrize("count,expected_support,expected_level", [
    (1,False,"LOW"),(2,True,"MODERATE"),(4,True,"HIGH"),
])
def test_recurrent_support_levels(count,expected_support,expected_level):
    rows=[relation(f"R{i}",group="G1") for i in range(count)]
    result=run(rows)
    assert all(row["Recurrent_Support"] is expected_support for row in result.details())
    assert all(row["Recurrent_Support_Level"]==expected_level for row in result.details())
    assert all(row["Recurrent_Group_Pair_Count"]==count for row in result.details())


def test_no_recurrent_group_is_none():
    row=detail(run()); assert row["Recurrent_Support"] is False and row["Recurrent_Support_Level"]=="NONE"


def test_recurrent_never_enables_chemical_interpretation():
    row=detail(run([relation(group="G1"),relation("R2",group="G1")]))
    assert row["Recurrent_Numerical_Support"] is True
    assert row["Recurrent_Chemical_Interpretation_Eligible"] is False


def test_deterministic_evidence_ids_follow_source_order():
    rows=[relation("R9"),relation("R3")]
    result=run(rows)
    assert [r["Relation_Evidence_ID"] for r in result.details()] == ["SCIEX_REL_EVID_00001","SCIEX_REL_EVID_00002"]
    assert [r["Source_Relation_ID"] for r in result.details()] == ["R9","R3"]


def test_reordering_preserves_aggregate_result():
    rows=[relation("R1",integer=True),relation("R2",integer=False,isotope=True)]
    first=summary(run(rows)); second=summary(run(list(reversed(rows))))
    keys=["Total_Relation_Count","Tier_2_Count","Integer_Candidate_Count","Isotope_Candidate_Count"]
    assert [first[k] for k in keys]==[second[k] for k in keys]


def test_relation_row_count_and_order_are_preserved():
    rows=[relation("R4"),relation("R2"),relation("R8")]
    result=run(rows)
    assert len(result.details())==len(rows)
    assert [r["Source_Relation_ID"] for r in result.details()]==[r["Relation_ID"] for r in rows]


def test_existing_truncated_export_subset_is_reused_exactly():
    exported=[relation("R1"),relation("R2")]
    assert [r["Source_Relation_ID"] for r in run(exported).details()]==["R1","R2"]


def test_summary_tier_counts_sum_to_total():
    rows=[relation("R0"),relation("R1",integer=False,isotope=False)]
    broken=relation("R2"); broken["Observed_Mass_A"]=None; rows.append(broken)
    row=summary(run(rows))
    assert sum(row[f"Tier_{i}_Count"] for i in range(5))==row["Total_Relation_Count"]


def test_summary_candidate_counts():
    rows=[relation("D",True,True),relation("I",True,False),relation("S",False,True)]
    row=summary(run(rows))
    assert (row["Integer_Candidate_Count"],row["Isotope_Candidate_Count"],row["Dual_Candidate_Count"])==(2,2,1)


def test_summary_fit_counts_sum_to_total():
    rows=[relation("E",integer_error=0),relation("G",integer_error=.075),relation("W",integer_error=.14),relation("O",integer_error=.2),relation("N",False,False)]
    row=summary(run(rows))
    assert sum(row[k] for k in ["Excellent_Fit_Count","Good_Fit_Count","Weak_Fit_Count","Outside_Tolerance_Count","Not_Applicable_Fit_Count"])==5


def test_summary_assignment_and_tier_four_are_zero():
    row=summary(run())
    assert row["Isotope_Assignment_Eligible_Count"]==0 and row["Tier_4_Count"]==0


def test_summary_highest_tier():
    row=summary(run(resolution=resolution_result("DISTINGUISHABLE",True,3)))
    assert row["Highest_Available_Evidence_Tier"]=="TIER_3_RESOLUTION_SUPPORTED"


def test_formal_flags_all_false():
    result=run()
    for row in result.details()+result.summaries():
        assert row["Shadow_Only"] is True
        assert row["Applied_To_Formal_Score"] is False
        assert row["Applied_To_Ranking"] is False
        assert row["Applied_To_Candidate_Filtering"] is False
        assert row["Molecular_Identity_Assigned"] is False


def test_source_relations_are_not_mutated():
    source=cluster_result([relation("R1"),relation("R2")]); before=source.relations()
    audit_sciex_relation_evidence_quality(source,resolution_result(),CLUSTER_PARAMETERS)
    assert source.relations()==before


def test_cluster_annotation_preserves_clusters_and_relations():
    source=cluster_result([relation("R1")]); evidence=run([relation("R1")])
    updated=annotate_cluster_summary(source,evidence)
    assert updated.clusters()==source.clusters() and updated.relations()==source.relations()
    assert updated.summaries()[0]["Relation_Evidence_Audit_Status"]=="AUDIT_COMPLETED"


def test_resolution_annotation_preserves_detail_and_core_status():
    source=resolution_result(); before=source.details(); evidence=run()
    updated=annotate_resolution_summary(source,evidence)
    assert updated.details()==before
    assert updated.summaries()[0]["Resolution_Status"]=="NOT_DISTINGUISHABLE_BOTH"
    assert updated.summaries()[0]["Tier_2_Resolution_Limited_Count"]==1


def test_result_rows_are_defensive_copies():
    result=run(); rows=result.details(); rows[0]["Evidence_Tier"]="CHANGED"
    assert result.details()[0]["Evidence_Tier"]!="CHANGED"


def routed(config=None,cluster=None,resolution=None,warnings=None):
    cluster_values={} if cluster is False else {CLUSTER_KEY: cluster or cluster_result()}
    resolution_values={} if resolution is False else {RESOLUTION_KEY: resolution or resolution_result()}
    return main_module.build_sciex_relation_evidence_optional_results(
        config or route_config(),cluster_values,resolution_values,
        [] if warnings is None else warnings,
    )


@pytest.mark.parametrize("config,cluster,resolution", [
    (RunConfig(),cluster_result(),resolution_result()),
    (route_config(enabled=False),cluster_result(),resolution_result()),
    (route_config(audit_enabled=False),cluster_result(),resolution_result()),
    (route_config(),False,resolution_result()),
    (route_config(),cluster_result([]),resolution_result()),
    (route_config(),cluster_result(),False),
])
def test_routing_skips_ineligible_inputs(config,cluster,resolution):
    assert routed(config,cluster,resolution)=={}


def test_routing_executes_with_relations_and_resolution():
    assert AUDIT_RESULT_KEY in routed()


def test_normal_ambiguity_adds_no_duplicate_warning():
    warnings=[]; routed(warnings=warnings); assert warnings==[]


def test_schema_mismatch_adds_error_warning():
    warnings=[]
    broken=SciexSpacingResolutionAuditResult(({"Resolution_Status":"X"},),())
    assert routed(resolution=broken,warnings=warnings)=={}
    assert warnings[-1]["Context"]["Warning_Code"]==ERROR_CODE


def test_audit_exception_preserves_upstream(monkeypatch):
    warnings=[]; cluster=cluster_result(); resolution=resolution_result()
    monkeypatch.setattr(main_module,"audit_sciex_relation_evidence_quality",lambda *_a,**_k:(_ for _ in ()).throw(RuntimeError("boom")))
    assert routed(cluster=cluster,resolution=resolution,warnings=warnings)=={}
    assert cluster.relations() and resolution.summaries()
    assert warnings[-1]["Context"]["Warning_Code"]==ERROR_CODE


def test_sheet_registry_and_policy():
    assert sheet_category(DETAIL_SHEET)==AUDIT_DETAIL
    assert sheet_category(SUMMARY_SHEET)==AUDIT_SUMMARY
    names=[DETAIL_SHEET,SUMMARY_SHEET]
    assert included_sheet_names(names,AuditPolicy.from_level("standard"))[0]==[]
    assert included_sheet_names(names,AuditPolicy.from_level("audit"))[0]==names
    assert included_sheet_names(names,AuditPolicy.from_level("full"))[0]==names
    assert all(len(name)<=31 for name in names)


def writer_config():
    return SimpleNamespace(
        analysis={"mode":"full"},project={"name":"relation-evidence"},input={},organism={},sequence={},experiment={},instrument={},sciex_profile={},
        reconstruction={"enabled":False},digestion={"enabled":False},alkaline_phosphatase={},fragment_mapping={},modification_search={},peak_filtering={},
        p1_annotation={},ms2_annotation={},modification_evidence_ranking={},biological_context={},performance={},reporting={"max_excel_rows_per_sheet":100000,"truncate_large_sheets":True},
    )


def write_report(tmp_path,level,optional):
    return write_excel_report(
        tmp_path/level,writer_config(),{},[],[],[],
        known_modification_candidates=[{"Candidate_ID":"C1"}],
        known_modification_summary=[{"Summary_Key":"S1"}],optional_results=optional,
        audit_policy=AuditPolicy.from_level(level),
    )


@pytest.mark.parametrize("level,present",[("standard",False),("audit",True),("full",True)])
def test_excel_policy_columns_rows_and_order(tmp_path,level,present):
    result=run([relation("R2"),relation("R1")])
    report=write_report(tmp_path,level,{AUDIT_RESULT_KEY:result})
    wb=load_workbook(report,read_only=True,data_only=True)
    try:names=wb.sheetnames
    finally:wb.close()
    assert (DETAIL_SHEET in names) is present and (SUMMARY_SHEET in names) is present
    if present:
        details=pd.read_excel(report,sheet_name=DETAIL_SHEET,header=2)
        summaries=pd.read_excel(report,sheet_name=SUMMARY_SHEET,header=2)
        assert list(details.columns)==DETAIL_COLUMNS and list(summaries.columns)==SUMMARY_COLUMNS
        assert details["Source_Relation_ID"].tolist()==["R2","R1"]
        assert len(details)==2 and len(summaries)==1


def formal_fixture():
    return {
        "Known_Modification_Candidates":[{"ID":"K","Value":1}],
        "Known_Modification_Summary":[{"ID":"KS","Value":2}],
        "Modification_Evidence_Ranking":[{"ID":"R","Value":3}],
        "Top_Modification_Candidates":[{"ID":"T","Value":4}],
        "P1_Summary":[{"ID":"P","Value":5}],
        "MS2_Summary":[{"ID":"M","Value":6}],
    }


@pytest.mark.parametrize("variant",[
    "sciex_disabled_enabled","audit_disabled_enabled","ambiguous_supported",
    "identity_match_conflict","audit_success_exception",
])
def test_controlled_ab_formal_excel_is_identical(tmp_path,variant):
    first_shadow=None; second_shadow=run()
    if variant=="ambiguous_supported":
        first_shadow=run(resolution=resolution_result("DISTINGUISHABLE",True,3))
    elif variant=="identity_match_conflict":
        first_shadow=run(resolution=resolution_result(identity="MATCH",conflict=False,biological=True))
        second_shadow=run(resolution=resolution_result(identity="CONFLICT",conflict=True,biological=False))
    first=formal_fixture(); second=formal_fixture()
    if first_shadow is not None:first[AUDIT_RESULT_KEY]=first_shadow
    if second_shadow is not None:second[AUDIT_RESULT_KEY]=second_shadow
    report_a=write_report(tmp_path/variant/"a","full",first)
    report_b=write_report(tmp_path/variant/"b","full",second)
    for name in formal_fixture():
        a=pd.read_excel(report_a,sheet_name=name,header=2)
        b=pd.read_excel(report_b,sheet_name=name,header=2)
        pd.testing.assert_frame_equal(a,b,check_dtype=False,check_exact=True)


def test_real_input_sha_and_non_destructive_read():
    import pytest
    if not REAL_INPUT.is_file():
        pytest.skip("Local SCIEX real-data fixture is not available: .cache/sciex_research/WT_LeuUAA(Full).txt")
    assert REAL_INPUT.exists(); before=sha256(REAL_INPUT.read_bytes()).hexdigest()
    assert before==REAL_SHA256; run(); after=sha256(REAL_INPUT.read_bytes()).hexdigest()
    assert after==before


def test_algorithm_and_column_contracts():
    result=run()
    assert all(row["Algorithm_Version"]==ALGORITHM_VERSION for row in result.details()+result.summaries())
    for column in FORMAL_FALSE:
        assert column in DETAIL_COLUMNS and column in SUMMARY_COLUMNS
