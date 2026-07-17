from enum import Enum
import json
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from rna_masshunter.audit_policy import (
    AUDIT_DETAIL,
    AUDIT_SUMMARY,
    AuditPolicy,
    included_sheet_names,
    sheet_category,
)
from rna_masshunter.excel_report import (
    SCIEX_INTACT_DIAGNOSTIC_COLUMNS,
    SCIEX_INTACT_DIAGNOSTIC_SHEET,
    SCIEX_INTACT_OPTIONAL_RESULT_KEY,
    SCIEX_INTACT_PEAK_COLUMNS,
    SCIEX_INTACT_PEAK_SHEET,
    write_excel_report,
)
from rna_masshunter.sciex_intact_peak_detection import detect_sciex_intact_peaks


OMITTED = object()
FORMAL_COLUMNS = (
    "SCIEX_Intact_Peak_Detection_Applied_To_Formal_Score",
    "SCIEX_Intact_Peak_Detection_Applied_To_Ranking",
    "SCIEX_Intact_Peak_Detection_Applied_To_Candidate_Filtering",
)


class ExampleEnum(Enum):
    VALUE = "enum-value"


def config(report_limit=1000):
    return SimpleNamespace(
        analysis={"mode": "full"},
        project={"name": "sciex-intact-excel-test"},
        input={},
        organism={},
        sequence={},
        experiment={},
        instrument={},
        reconstruction={"enabled": False},
        digestion={"enabled": False},
        alkaline_phosphatase={},
        fragment_mapping={},
        modification_search={},
        peak_filtering={},
        p1_annotation={},
        ms2_annotation={},
        modification_evidence_ranking={},
        biological_context={},
        performance={},
        reporting={
            "max_excel_rows_per_sheet": report_limit,
            "truncate_large_sheets": True,
        },
    )


@pytest.fixture(scope="module")
def single_peak_result():
    masses = np.arange(0.0, 100.5, 0.5)
    intensities = 1.0 + 20.0 * np.exp(-0.5 * ((masses - 50.0) / 2.0) ** 2)
    result = detect_sciex_intact_peaks(
        masses,
        intensities,
        profile_type="NEUTRAL_MASS_PROFILE",
        input_status="SUPPORTED_INPUT",
        eligible_for_neutral_mass_analysis=True,
    )
    assert len(result.peaks) == 1
    return result


def write_report(tmp_path, level, payload=OMITTED, report_limit=1000, label="report"):
    output_dir = tmp_path / f"{label}-{level}"
    optional_results = {}
    if payload is not OMITTED:
        optional_results[SCIEX_INTACT_OPTIONAL_RESULT_KEY] = payload
    return write_excel_report(
        output_dir=output_dir,
        config=config(report_limit),
        diagnostics={},
        intact_results=[],
        charge_state_peaks=[],
        warnings=[],
        modifications=[],
        rule_set={},
        pathways=[],
        theoretical_fragments=[],
        fragment_ms1_matches=[],
        known_modification_candidates=[],
        known_modification_summary=[],
        optional_results=optional_results,
        audit_policy=AuditPolicy.from_level(level),
    )


def sheet_names(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def read_sheet(path, name):
    return pd.read_excel(path, sheet_name=name, header=2, dtype=object)


def formatted_payload(single_peak_result):
    diagnostics = single_peak_result.diagnostics_row()
    diagnostics.update({
        "Detected_Sensitive_Peak_Count": 2,
        "Detected_Strict_Peak_Count": 1,
        "Broad_Peak_Count": 1,
        "Edge_Peak_Count": 1,
    })
    first = single_peak_result.peak_rows()[0]
    first.update({
        "Peak_ID": "SCIEX_INTACT_000001",
        "Strict_Threshold_Passed": True,
        "Broad_Peak_Flag": False,
        "Edge_Peak_Flag": False,
    })
    second = dict(first)
    second.update({
        "Peak_ID": "SCIEX_INTACT_000002",
        "Apex_Mass": 60.0,
        "Strict_Threshold_Passed": False,
        "Broad_Peak_Flag": True,
        "Severe_Broad_Peak_Flag": False,
        "Edge_Peak_Flag": True,
        "Peak_Area_Raw": 123.0,
        "Peak_Area_Baseline_Corrected": 45.0,
        "Prominence_Base_Left_Mass": 51.0,
        "Left_Boundary_Mass": 55.0,
        "Shared_Valley_Mass": 55.0,
        "Possible_Shoulder": True,
        "Centroid_Complete": False,
        "Peak_Area_Complete": False,
    })
    return {
        "diagnostics": diagnostics,
        "peaks": [first, second],
        "parameter_provenance": single_peak_result.provenance_rows(),
        "warnings": (),
        "source_file": Path("/synthetic/single-full.txt"),
    }


def test_policy_registry_and_existing_sciex_profile_policy_are_stable():
    names = [SCIEX_INTACT_DIAGNOSTIC_SHEET, SCIEX_INTACT_PEAK_SHEET]
    assert sheet_category(SCIEX_INTACT_DIAGNOSTIC_SHEET) == AUDIT_SUMMARY
    assert sheet_category(SCIEX_INTACT_PEAK_SHEET) == AUDIT_DETAIL
    assert included_sheet_names(names, AuditPolicy.from_level("standard"))[0] == []
    assert included_sheet_names(names, AuditPolicy.from_level("audit"))[0] == names
    assert included_sheet_names(names, AuditPolicy.from_level("full"))[0] == names

    profile_names = ["SCIEX_Profile_Diagnostics", "SCIEX_Profile_Input"]
    assert included_sheet_names(profile_names, AuditPolicy.from_level("standard"))[0] == []
    assert included_sheet_names(profile_names, AuditPolicy.from_level("audit"))[0] == [
        "SCIEX_Profile_Diagnostics"
    ]
    assert included_sheet_names(profile_names, AuditPolicy.from_level("full"))[0] == profile_names


def test_missing_none_and_standard_result_do_not_change_sheet_list(tmp_path, single_peak_result):
    for level in ("standard", "audit", "full"):
        missing = sheet_names(write_report(tmp_path, level, label=f"missing-{level}"))
        none = sheet_names(write_report(tmp_path, level, None, label=f"none-{level}"))
        assert SCIEX_INTACT_DIAGNOSTIC_SHEET not in missing
        assert SCIEX_INTACT_PEAK_SHEET not in missing
        assert none == missing

    standard_with_result = sheet_names(
        write_report(tmp_path, "standard", single_peak_result, label="standard-with-result")
    )
    standard_without_result = sheet_names(
        write_report(tmp_path, "standard", label="standard-without-result")
    )
    assert standard_with_result == standard_without_result


@pytest.mark.parametrize("level", ["audit", "full"])
def test_result_object_writes_fixed_columns_and_all_sensitive_peaks(
    tmp_path, single_peak_result, level
):
    payload = {
        "result": single_peak_result,
        "source_file": Path("/synthetic/single-full.txt"),
    }
    report = write_report(tmp_path, level, payload, label=f"object-{level}")
    assert {SCIEX_INTACT_DIAGNOSTIC_SHEET, SCIEX_INTACT_PEAK_SHEET} <= set(sheet_names(report))
    diagnostics = read_sheet(report, SCIEX_INTACT_DIAGNOSTIC_SHEET)
    peaks = read_sheet(report, SCIEX_INTACT_PEAK_SHEET)
    assert list(diagnostics.columns) == SCIEX_INTACT_DIAGNOSTIC_COLUMNS
    assert list(peaks.columns) == SCIEX_INTACT_PEAK_COLUMNS
    assert len(diagnostics) == 1
    assert len(peaks) == single_peak_result.diagnostics["Detected_Sensitive_Peak_Count"]
    assert diagnostics.loc[0, "Source_File_Name"] == "single-full.txt"
    assert bool(peaks.loc[0, "Sensitive_Threshold_Passed"])


def test_diagnostics_only_and_zero_peak_write_headers(tmp_path, single_peak_result):
    diagnostics_only = {
        "diagnostics": single_peak_result.diagnostics_row(),
        "peaks": [],
    }
    report = write_report(tmp_path, "audit", diagnostics_only, label="diagnostics-only")
    assert len(read_sheet(report, SCIEX_INTACT_DIAGNOSTIC_SHEET)) == 1
    peaks = read_sheet(report, SCIEX_INTACT_PEAK_SHEET)
    assert peaks.empty and list(peaks.columns) == SCIEX_INTACT_PEAK_COLUMNS

    masses = np.arange(0.0, 20.5, 0.5)
    zero = detect_sciex_intact_peaks(
        masses,
        np.ones_like(masses),
        profile_type="NEUTRAL_MASS_PROFILE",
        input_status="SUPPORTED_INPUT",
        eligible_for_neutral_mass_analysis=True,
    )
    zero_report = write_report(tmp_path, "audit", zero, label="zero")
    assert read_sheet(zero_report, SCIEX_INTACT_PEAK_SHEET).empty
    assert read_sheet(zero_report, SCIEX_INTACT_DIAGNOSTIC_SHEET).loc[
        0, "Detected_Sensitive_Peak_Count"
    ] == 0


@pytest.mark.parametrize(
    "result,expected_status",
    [
        (
            detect_sciex_intact_peaks(
                [1.0, 2.0, 3.0], [1.0, 2.0, 1.0],
                profile_type="MZ_PROFILE", input_status="SUPPORTED_INPUT",
                eligible_for_neutral_mass_analysis=False,
            ),
            "SKIPPED_INELIGIBLE_PROFILE",
        ),
        (
            detect_sciex_intact_peaks(
                [1.0, 3.0, 2.0, 4.0, 5.0], [1.0, 2.0, 1.0, 2.0, 1.0],
                profile_type="NEUTRAL_MASS_PROFILE", input_status="SUPPORTED_INPUT",
                eligible_for_neutral_mass_analysis=True,
            ),
            "INVALID_AXIS",
        ),
        (
            detect_sciex_intact_peaks(
                [1.0, 2.0], [1.0, 2.0],
                profile_type="NEUTRAL_MASS_PROFILE", input_status="SUPPORTED_INPUT",
                eligible_for_neutral_mass_analysis=True,
            ),
            "INSUFFICIENT_POINTS",
        ),
    ],
)
def test_skip_invalid_and_insufficient_results_write_diagnostics(
    tmp_path, result, expected_status
):
    report = write_report(tmp_path, "audit", result, label=expected_status)
    diagnostics = read_sheet(report, SCIEX_INTACT_DIAGNOSTIC_SHEET)
    assert diagnostics.loc[0, "Detection_Status"] == expected_status
    assert read_sheet(report, SCIEX_INTACT_PEAK_SHEET).empty


def test_peak_semantics_and_shadow_flags_are_preserved(tmp_path, single_peak_result):
    report = write_report(
        tmp_path, "audit", formatted_payload(single_peak_result), label="semantics"
    )
    diagnostics = read_sheet(report, SCIEX_INTACT_DIAGNOSTIC_SHEET)
    peaks = read_sheet(report, SCIEX_INTACT_PEAK_SHEET)
    assert len(peaks) == 2
    for column in FORMAL_COLUMNS:
        assert diagnostics[column].map(bool).sum() == 0
    assert peaks["Sensitive_Threshold_Passed"].map(bool).all()
    assert peaks["Strict_Threshold_Passed"].map(bool).sum() == 1
    assert peaks["Molecular_Identity_Assigned"].map(bool).sum() == 0
    for column in FORMAL_COLUMNS:
        assert peaks[column].map(bool).sum() == 0
    broad = peaks.loc[peaks["Broad_Peak_Flag"].map(bool)].iloc[0]
    assert broad["Peak_Area_Raw"] == 123.0
    assert broad["Peak_Area_Baseline_Corrected"] == 45.0
    assert broad["Prominence_Base_Left_Mass"] == 51.0
    assert broad["Left_Boundary_Mass"] == 55.0
    assert broad["Shared_Valley_Mass"] == 55.0
    assert bool(broad["Edge_Peak_Flag"])
    assert bool(broad["Possible_Shoulder"])
    assert not bool(broad["Peak_Area_Complete"])
    assert "Smoothed_Area" not in peaks.columns


def test_parameter_provenance_and_excel_values_are_safe_and_deterministic(
    tmp_path, single_peak_result
):
    payload = formatted_payload(single_peak_result)
    payload["diagnostics"]["Automatic_Parameter_Fallbacks"] = {
        "items": [Path("fallback/path"), np.int64(3), np.float64(np.nan)]
    }
    payload["parameter_provenance"] = [{
        "Parameter_Name": "unsafe_types",
        "Parameter_Value": {
            "path": Path("parameter/path"),
            "enum": ExampleEnum.VALUE,
            "integer": np.int64(7),
            "float": np.float64(1.25),
            "boolean": np.bool_(True),
            "set": {"b", "a"},
            "nan": np.float64(np.nan),
        },
        "Parameter_Unit": "mixed",
        "Parameter_Source": "test",
        "Parameter_Reason": ("deterministic", "serialization"),
    }]
    first = write_report(tmp_path, "audit", payload, label="safe-first")
    second = write_report(tmp_path, "audit", payload, label="safe-second")
    first_row = read_sheet(first, SCIEX_INTACT_DIAGNOSTIC_SHEET).iloc[0]
    second_row = read_sheet(second, SCIEX_INTACT_DIAGNOSTIC_SHEET).iloc[0]
    pd.testing.assert_frame_equal(
        read_sheet(first, SCIEX_INTACT_DIAGNOSTIC_SHEET),
        read_sheet(second, SCIEX_INTACT_DIAGNOSTIC_SHEET),
    )
    pd.testing.assert_frame_equal(
        read_sheet(first, SCIEX_INTACT_PEAK_SHEET),
        read_sheet(second, SCIEX_INTACT_PEAK_SHEET),
    )
    assert first_row["Parameter_Provenance_JSON"] == second_row["Parameter_Provenance_JSON"]
    provenance = json.loads(first_row["Parameter_Provenance_JSON"])
    value = provenance[0]["Parameter_Value"]
    assert value == {
        "boolean": True,
        "enum": "enum-value",
        "float": 1.25,
        "integer": 7,
        "nan": None,
        "path": "parameter/path",
        "set": ["a", "b"],
    }
    assert first_row["Automatic_Parameter_Fallbacks"] == (
        '{"items":["fallback/path",3,null]}'
    )


def test_max_excel_rows_truncates_detected_peaks_with_warning(tmp_path, single_peak_result):
    payload = formatted_payload(single_peak_result)
    template = payload["peaks"][0]
    payload["peaks"] = [dict(template, Peak_ID=f"P{index:03d}") for index in range(5)]
    report = write_report(tmp_path, "audit", payload, report_limit=2, label="truncated")
    assert len(read_sheet(report, SCIEX_INTACT_PEAK_SHEET)) == 2
    run_summary = read_sheet(report, "Run_summary")
    truncations = run_summary.loc[
        run_summary["Item"] == "Truncated sheets", "Value"
    ].iloc[0]
    assert SCIEX_INTACT_PEAK_SHEET in str(truncations)


def test_writer_does_not_mutate_result_object(tmp_path, single_peak_result):
    before = {
        "diagnostics": single_peak_result.diagnostics_row(),
        "peaks": single_peak_result.peak_rows(),
        "provenance": single_peak_result.provenance_rows(),
        "warnings": single_peak_result.warnings,
        "raw": single_peak_result.raw_intensity,
    }
    write_report(tmp_path, "full", single_peak_result, label="immutable")
    after = {
        "diagnostics": single_peak_result.diagnostics_row(),
        "peaks": single_peak_result.peak_rows(),
        "provenance": single_peak_result.provenance_rows(),
        "warnings": single_peak_result.warnings,
        "raw": single_peak_result.raw_intensity,
    }
    assert after == before
