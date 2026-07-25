from copy import deepcopy
from dataclasses import replace
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from rna_masshunter.sciex_cross_layer_bundle_runner import (
    FULL_PROVENANCE_WARNING,
    REQUIRED_LAYERS,
    XL_SHEET_NAMES,
    CrossLayerBundleRunnerError,
    run_cross_layer_from_bundles,
    validate_cross_layer_bundle_set,
)
from rna_masshunter.sciex_layer_evidence_bundle import (
    LayerEvidenceBundleError,
    export_layer_evidence_bundle,
)
from rna_masshunter.sciex_rna_cross_layer_evidence_reconciliation import OPTIONAL_RESULT_KEY
from tests.test_sciex_layer_evidence_bundle import (
    COMMIT, CREATED, RNA, _full_result, _p1_ms1_result, _p1_ms2_result,
    _rehash, _t1_result,
)


@pytest.fixture
def production_bundles(tmp_path):
    source_dir = tmp_path / "sources"
    source_dir.mkdir()
    sources = {
        "FULL": source_dir / "full.mzML",
        "T1": source_dir / "05 old T1.mzML",
        "P1AP_MS1": source_dir / "p1.mzML",
        "P1AP_MS2": source_dir / "p1.mzML",
    }
    sources["FULL"].write_bytes(b"full synthetic source")
    sources["T1"].write_bytes(b"t1 synthetic source")
    sources["P1AP_MS1"].write_bytes(b"p1 synthetic source")
    results = {
        "FULL": _full_result(),
        "T1": _t1_result(tmp_path),
        "P1AP_MS1": _p1_ms1_result(tmp_path),
        "P1AP_MS2": _p1_ms2_result(tmp_path),
    }
    independence = {
        "FULL": "INDEPENDENCE_FULL",
        "T1": "INDEPENDENCE_T1",
        "P1AP_MS1": "INDEPENDENCE_P1",
        "P1AP_MS2": "INDEPENDENCE_P1",
    }
    shared = {
        "FULL": "SOURCE_FULL",
        "T1": "SOURCE_T1",
        "P1AP_MS1": "SOURCE_P1",
        "P1AP_MS2": "SOURCE_P1",
    }
    bundle_dir = tmp_path / "bundles"
    bundle_dir.mkdir()
    bundles = {}
    paths = {}
    for layer in REQUIRED_LAYERS:
        experiment = {
            "condition_name": "wild_type",
            "digest_type": layer,
            "layer": layer,
            "independence_group": independence[layer],
            "shared_source_group": shared[layer],
        }
        path = bundle_dir / f"{layer}.json"
        bundle = export_layer_evidence_bundle(
            results[layer], layer=layer, source_path=sources[layer], rna=RNA,
            experiment=experiment, producer_commit=COMMIT,
            created_at_utc=CREATED, output_path=path,
            run_id="P1_RUN" if layer.startswith("P1AP") else f"{layer}_RUN",
            sample_id="SAMPLE", biological_sample_id="BIO_SAMPLE",
        )
        bundles[layer] = bundle
        paths[layer] = path
    return {
        "bundles": bundles, "paths": paths, "results": results,
        "sources": sources, "tmp_path": tmp_path,
    }


def _changed(bundle, change):
    output = deepcopy(bundle)
    change(output)
    _rehash(output)
    return output


def test_valid_four_bundle_run(production_bundles):
    run = run_cross_layer_from_bundles(production_bundles["paths"])
    assert run.compatibility_report.compatible
    assert set(run.restored_result_types) == set(REQUIRED_LAYERS)
    assert OPTIONAL_RESULT_KEY in run.cross_layer_optional_results
    assert run.safeguard_summary["verified"] is True


def test_mapping_bundle_inputs_are_supported(production_bundles):
    run = run_cross_layer_from_bundles(production_bundles["bundles"])
    assert run.compatibility_report.compatible
    assert all(path is None for path in run.input_bundle_paths.values())


def test_input_order_independence(production_bundles):
    ordered = list(production_bundles["paths"].values())
    left = run_cross_layer_from_bundles(ordered)
    right = run_cross_layer_from_bundles(reversed(ordered))
    assert left.aggregate_counts == right.aggregate_counts
    assert left.consensus_status == right.consensus_status
    assert left.confidence == right.confidence


@pytest.mark.parametrize("missing", REQUIRED_LAYERS)
def test_missing_layer_rejected(production_bundles, missing):
    inputs = {key: value for key, value in production_bundles["paths"].items() if key != missing}
    with pytest.raises(CrossLayerBundleRunnerError, match="missing required layer"):
        run_cross_layer_from_bundles(inputs)


def test_duplicate_layer_rejected(production_bundles):
    paths = list(production_bundles["paths"].values())
    paths.append(production_bundles["paths"]["FULL"])
    with pytest.raises(CrossLayerBundleRunnerError, match="duplicate layer"):
        run_cross_layer_from_bundles(paths)


def test_old_serializer_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    del bundles["FULL"]["serializer_format_version"]
    _rehash(bundles["FULL"])
    with pytest.raises(LayerEvidenceBundleError, match="serializer_format_version"):
        run_cross_layer_from_bundles(bundles)


def test_old_antigravity_bundle_rejected(production_bundles):
    values = list(production_bundles["bundles"].values())
    values[0] = {"schema_version": "rna-masshunter-cross-layer-bundle-v1", "layer": "FULL"}
    with pytest.raises(LayerEvidenceBundleError, match="unknown schema_version"):
        run_cross_layer_from_bundles(values)


@pytest.mark.parametrize(
    "field,value,error",
    [
        ("name", "OTHER_RNA", "RNA_NAME_MISMATCH"),
        ("sequence", "ACGU", "RNA_SEQUENCE_MISMATCH"),
        ("anticodon", "AAA", "RNA_ANTICODON_MISMATCH"),
    ],
)
def test_rna_identity_mismatch_rejected(production_bundles, field, value, error):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["T1"] = _changed(bundles["T1"], lambda item: item["rna"].__setitem__(field, value))
    with pytest.raises(CrossLayerBundleRunnerError, match=error):
        run_cross_layer_from_bundles(bundles)


def test_condition_mismatch_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["T1"] = _changed(
        bundles["T1"], lambda item: item["experiment"].__setitem__("condition_name", "treated")
    )
    with pytest.raises(CrossLayerBundleRunnerError, match="UNSUPPORTED_CONDITION_RELATIONSHIP"):
        run_cross_layer_from_bundles(bundles)


def test_p1ap_source_mismatch_rejected(production_bundles, tmp_path):
    other = tmp_path / "other-p1.mzML"
    other.write_bytes(b"different p1 source")
    original = production_bundles["bundles"]["P1AP_MS2"]
    experiment = dict(original["experiment"])
    replacement = export_layer_evidence_bundle(
        production_bundles["results"]["P1AP_MS2"], layer="P1AP_MS2",
        source_path=other, rna=RNA, experiment=experiment,
        producer_commit=COMMIT, created_at_utc=CREATED,
        run_id="P1_RUN", sample_id="SAMPLE", biological_sample_id="BIO_SAMPLE",
    )
    bundles = dict(production_bundles["bundles"], P1AP_MS2=replacement)
    with pytest.raises(CrossLayerBundleRunnerError, match="P1AP_SOURCE_(SHA256|BASENAME)_MISMATCH"):
        run_cross_layer_from_bundles(bundles)


def test_p1ap_run_mismatch_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["P1AP_MS2"] = _changed(
        bundles["P1AP_MS2"], lambda item: item["source"].__setitem__("run_id", "OTHER_RUN")
    )
    with pytest.raises(CrossLayerBundleRunnerError, match="P1AP_SOURCE_RUN_ID_MISMATCH"):
        run_cross_layer_from_bundles(bundles)


def test_p1ap_shared_group_mismatch_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["P1AP_MS2"] = _changed(
        bundles["P1AP_MS2"],
        lambda item: item["experiment"].__setitem__("shared_source_group", "OTHER_SOURCE"),
    )
    with pytest.raises(CrossLayerBundleRunnerError, match="SHARED_SOURCE_GROUP_MISMATCH"):
        run_cross_layer_from_bundles(bundles)


def test_forbidden_04_new_t1_substitution_rejected(production_bundles, tmp_path):
    source = tmp_path / "04 new T1.mzML"
    source.write_bytes(b"wrong t1 sample")
    original = production_bundles["bundles"]["T1"]
    replacement = export_layer_evidence_bundle(
        production_bundles["results"]["T1"], layer="T1", source_path=source,
        rna=RNA, experiment=dict(original["experiment"]), producer_commit=COMMIT,
        created_at_utc=CREATED, run_id="T1_RUN", sample_id="SAMPLE",
        biological_sample_id="BIO_SAMPLE",
    )
    bundles = dict(production_bundles["bundles"], T1=replacement)
    with pytest.raises(CrossLayerBundleRunnerError, match="T1_FORBIDDEN_SOURCE_SUBSTITUTION"):
        run_cross_layer_from_bundles(bundles)


def test_safeguard_true_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["FULL"] = _changed(
        bundles["FULL"], lambda item: item["safeguards"].__setitem__("formal_propagation", True)
    )
    with pytest.raises(LayerEvidenceBundleError, match="unsafe safeguard"):
        run_cross_layer_from_bundles(bundles)


def test_empty_bundle_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    root = bundles["P1AP_MS1"]["result"]
    root["fields"]["matches"] = {"__bundle_type__": "tuple", "items": []}
    _rehash(bundles["P1AP_MS1"])
    with pytest.raises(LayerEvidenceBundleError, match="empty production result"):
        run_cross_layer_from_bundles(bundles)


def test_wrong_root_class_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["P1AP_MS1"]["result"] = deepcopy(bundles["FULL"]["result"])
    _rehash(bundles["P1AP_MS1"])
    with pytest.raises(LayerEvidenceBundleError, match="restored result class mismatch"):
        run_cross_layer_from_bundles(bundles)


def test_wrong_optional_result_key_rejected(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["FULL"] = _changed(
        bundles["FULL"], lambda item: item.__setitem__("optional_result_key", "wrong_key")
    )
    with pytest.raises(LayerEvidenceBundleError, match="optional_result_key"):
        run_cross_layer_from_bundles(bundles)


def test_raw_parser_and_producers_are_not_invoked(production_bundles, monkeypatch):
    def forbidden(*args, **kwargs):
        raise AssertionError("raw parser or producer invoked")

    import rna_masshunter.sciex_profile_parser as parser
    import rna_masshunter.sciex_intact_oxygen_water_state_audit as full_module
    import rna_masshunter.sciex_t1_fragment_state_series_audit as t1_module
    import rna_masshunter.sciex_p1ap_nucleoside_state_audit as p1_module
    import rna_masshunter.sciex_p1ap_nucleoside_ms2_identity_audit as ms2_module
    monkeypatch.setattr(parser, "parse_sciex_profile", forbidden)
    monkeypatch.setattr(full_module, "audit_oxygen_water_state_series", forbidden)
    monkeypatch.setattr(t1_module, "audit_t1_fragment_state_series", forbidden)
    monkeypatch.setattr(p1_module, "audit_p1ap_nucleoside_state_series", forbidden)
    monkeypatch.setattr(ms2_module, "audit_p1ap_nucleoside_ms2_identity", forbidden)
    run = run_cross_layer_from_bundles(production_bundles["paths"])
    assert run.compatibility_report.compatible


def test_aggregate_counts_consensus_and_independence_preserved(production_bundles):
    run = run_cross_layer_from_bundles(production_bundles["paths"])
    result = run.cross_layer_optional_results[OPTIONAL_RESULT_KEY]
    assert run.aggregate_counts == {
        "nodes": len(result.nodes), "edges": len(result.edges),
        "hypotheses": len(result.hypotheses),
        "layer_summaries": len(result.layer_summaries), "consensus": 1,
        "next_evidence": len(result.next_evidence),
        "independence_groups": result.consensus.independence_group_count,
    }
    assert run.consensus_status == result.consensus.cross_layer_evidence_status
    assert run.confidence == result.consensus.cross_layer_confidence
    groups = run.compatibility_report.independence_groups
    assert groups["P1AP_MS1"] == groups["P1AP_MS2"]
    assert len(set(groups.values())) == 3


def test_full_provenance_warning_and_bundle_provenance(production_bundles):
    run = run_cross_layer_from_bundles(production_bundles["paths"])
    report = run.compatibility_report
    assert FULL_PROVENANCE_WARNING in report.warnings
    assert report.node_level_provenance["FULL"] == "UNAVAILABLE_IN_RESULT"
    assert report.bundle_level_provenance_verified is True
    assert report.bundle_provenance["FULL"]["source_path"].endswith("full.mzML")
    result = run.cross_layer_optional_results[OPTIONAL_RESULT_KEY]
    full_nodes = [node for node in result.nodes if node.layer == "FULL_LENGTH"]
    assert full_nodes and all(node.source_file_id == "UNKNOWN" for node in full_nodes)


def test_cross_sample_relationship_does_not_claim_independent_support(production_bundles):
    bundles = deepcopy(production_bundles["bundles"])
    bundles["T1"] = _changed(
        bundles["T1"], lambda item: item["source"].__setitem__("sample_id", "OTHER_SAMPLE")
    )
    _, report = validate_cross_layer_bundle_set(bundles)
    relation = next(
        row for row in report.sample_relationships
        if {row["left_layer"], row["right_layer"]} == {"FULL", "T1"}
    )
    assert relation["different_sample"] is True
    assert relation["independent_support"] is False


def test_aggregate_json_atomic_and_deterministic(production_bundles, tmp_path):
    left = tmp_path / "left" / "aggregate.json"
    right = tmp_path / "right" / "aggregate.json"
    run_cross_layer_from_bundles(production_bundles["paths"], output_json_path=left)
    run_cross_layer_from_bundles(production_bundles["paths"], output_json_path=right)
    assert left.read_bytes() == right.read_bytes()
    payload = json.loads(left.read_text(encoding="utf-8"))
    assert payload["compatibility_report"]["compatible"] is True
    assert payload["safeguards"]["formal_propagation"] is False
    assert payload["safeguards"]["shadow_analysis_only"] is True
    assert not list(left.parent.glob("*.tmp"))


def test_existing_output_rejected_without_partial_output(production_bundles, tmp_path):
    existing = tmp_path / "aggregate.json"
    existing.write_text("keep", encoding="utf-8")
    excel = tmp_path / "new.xlsx"
    with pytest.raises(CrossLayerBundleRunnerError, match="already exists"):
        run_cross_layer_from_bundles(
            production_bundles["paths"], output_json_path=existing,
            output_excel_path=excel,
        )
    assert existing.read_text(encoding="utf-8") == "keep"
    assert not excel.exists()


def test_combined_output_failure_leaves_no_partial_output(production_bundles, tmp_path):
    base = tmp_path / "base.xlsx"
    workbook = Workbook()
    workbook.active.title = "XL_Nodes"
    workbook.save(base)
    workbook.close()
    output_json = tmp_path / "aggregate.json"
    output_excel = tmp_path / "aggregate.xlsx"
    with pytest.raises(CrossLayerBundleRunnerError, match="already exists in workbook"):
        run_cross_layer_from_bundles(
            production_bundles["paths"], output_json_path=output_json,
            output_excel_path=output_excel, base_workbook_path=base,
        )
    assert not output_json.exists()
    assert not output_excel.exists()
    assert not list(tmp_path.glob("*.runner-stage-*"))


def test_malformed_output_directory_rejected(production_bundles, tmp_path):
    parent = tmp_path / "not-a-directory"
    parent.write_text("file", encoding="utf-8")
    with pytest.raises(CrossLayerBundleRunnerError, match="malformed output directory"):
        run_cross_layer_from_bundles(
            production_bundles["paths"], output_json_path=parent / "aggregate.json"
        )


def test_xl_only_six_sheet_output_and_readback(production_bundles, tmp_path):
    output = tmp_path / "cross-layer.xlsx"
    run = run_cross_layer_from_bundles(
        production_bundles["paths"], output_excel_path=output
    )
    assert run.output_excel_path == str(output)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert tuple(workbook.sheetnames) == XL_SHEET_NAMES
        assert all(len(name) <= 31 for name in workbook.sheetnames)
        expected_rows = {
            "XL_Nodes": run.aggregate_counts["nodes"],
            "XL_Edges": run.aggregate_counts["edges"],
            "XL_Hypotheses": run.aggregate_counts["hypotheses"],
            "XL_Layer_Summary": run.aggregate_counts["layer_summaries"],
            "XL_Consensus": 1,
            "XL_Next_Evidence": run.aggregate_counts["next_evidence"],
        }
        for name, count in expected_rows.items():
            assert workbook[name].max_row - 1 == count
    finally:
        workbook.close()


def test_excel_ab_formal_nonpropagation_and_input_immutability(production_bundles, tmp_path):
    baseline = tmp_path / "baseline.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formal_Result"
    sheet["A1"] = "Final Score"
    sheet["B1"] = 123.456
    sheet["A2"] = "Final Confidence"
    sheet["B2"] = "HIGH"
    workbook.save(baseline)
    workbook.close()
    before = baseline.read_bytes()
    bundle_before = deepcopy(production_bundles["bundles"])
    output = tmp_path / "enabled.xlsx"
    run_cross_layer_from_bundles(
        production_bundles["bundles"], output_excel_path=output,
        base_workbook_path=baseline,
    )
    assert baseline.read_bytes() == before
    assert production_bundles["bundles"] == bundle_before
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Formal_Result"]["A1"].value == "Final Score"
        assert workbook["Formal_Result"]["B1"].value == 123.456
        assert workbook["Formal_Result"]["B2"].value == "HIGH"
        assert set(workbook.sheetnames) == {"Formal_Result", *XL_SHEET_NAMES}
    finally:
        workbook.close()
