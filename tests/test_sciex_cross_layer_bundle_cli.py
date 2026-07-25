from copy import deepcopy
import json
from pathlib import Path
import shutil

from openpyxl import load_workbook
import pytest
import yaml

import rna_masshunter.sciex_cross_layer_bundle_cli as cli
from rna_masshunter.sciex_cross_layer_bundle_cli import (
    EXIT_AGGREGATION,
    EXIT_ARGUMENT_CONFIG,
    EXIT_BUNDLE_VALIDATION,
    EXIT_COMPATIBILITY,
    EXIT_OUTPUT,
    EXIT_SUCCESS,
    load_cross_layer_bundle_cli_config,
    main,
)
from rna_masshunter.sciex_cross_layer_bundle_runner import (
    FULL_PROVENANCE_WARNING,
    CrossLayerBundleRunnerError,
    XL_SHEET_NAMES,
)
from rna_masshunter.sciex_layer_evidence_bundle import (
    LayerEvidenceBundleError,
    canonical_json_bytes,
    export_layer_evidence_bundle,
)
from tests.test_sciex_cross_layer_bundle_runner import production_bundles
from tests.test_sciex_layer_evidence_bundle import COMMIT, CREATED, RNA, _rehash


def _arguments(paths):
    return [
        "--full-bundle", str(paths["FULL"]),
        "--t1-bundle", str(paths["T1"]),
        "--p1ap-ms1-bundle", str(paths["P1AP_MS1"]),
        "--p1ap-ms2-bundle", str(paths["P1AP_MS2"]),
    ]


def _invoke(arguments, capsys):
    code = main(arguments)
    captured = capsys.readouterr()
    stdout = json.loads(captured.out) if captured.out.strip() else None
    stderr = json.loads(captured.err) if captured.err.strip() else None
    return code, stdout, stderr


def _write_bundle(path, bundle):
    path.write_bytes(canonical_json_bytes(bundle))
    return path


def test_help_success(capsys):
    with pytest.raises(SystemExit) as exc:
        main(["--help"])
    assert exc.value.code == 0
    text = capsys.readouterr().out
    assert "--full-bundle" in text
    assert "--dry-run" in text
    assert "without reading raw mzML" in text


def test_valid_four_bundle_cli_run(production_bundles, tmp_path, capsys):
    aggregate = tmp_path / "out" / "aggregate.json"
    excel = tmp_path / "out" / "cross-layer.xlsx"
    code, summary, error = _invoke(
        _arguments(production_bundles["paths"]) +
        ["--aggregate-json", str(aggregate), "--excel-output", str(excel)],
        capsys,
    )
    assert code == EXIT_SUCCESS and error is None
    assert summary["status"] == "PASSED"
    assert summary["compatibility"]["compatible"] is True
    assert aggregate.is_file() and excel.is_file()


def test_dry_run_success_and_no_writes(production_bundles, tmp_path, capsys):
    aggregate = tmp_path / "aggregate.json"
    excel = tmp_path / "cross-layer.xlsx"
    code, summary, error = _invoke(
        _arguments(production_bundles["paths"]) + [
            "--aggregate-json", str(aggregate), "--excel-output", str(excel), "--dry-run",
        ],
        capsys,
    )
    assert code == EXIT_SUCCESS and error is None
    assert summary["status"] == "DRY_RUN_PASSED"
    assert summary["dry_run"] is True
    assert summary["outputs"] == {"aggregate_json": None, "excel": None, "summary_json": None}
    assert not aggregate.exists() and not excel.exists()


def test_missing_bundle_argument(production_bundles, tmp_path, capsys):
    args = _arguments(production_bundles["paths"])
    del args[0:2]
    code, _, error = _invoke(args + ["--aggregate-json", str(tmp_path / "out.json")], capsys)
    assert code == EXIT_ARGUMENT_CONFIG
    assert error["error_category"] == "ARGUMENT_ERROR"
    assert "FULL" in error["blocking_reason"]


def test_no_arguments_is_argument_error(capsys):
    code, _, error = _invoke([], capsys)
    assert code == EXIT_ARGUMENT_CONFIG
    assert error["error_category"] == "ARGUMENT_ERROR"


def test_missing_output_argument(production_bundles, capsys):
    code, _, error = _invoke(_arguments(production_bundles["paths"]), capsys)
    assert code == EXIT_ARGUMENT_CONFIG
    assert "--aggregate-json or --excel-output" in error["blocking_reason"]


def test_invalid_bundle_path_has_layer_and_exit_3(production_bundles, tmp_path, capsys):
    paths = dict(production_bundles["paths"], FULL=tmp_path / "missing.json")
    code, _, error = _invoke(_arguments(paths) + ["--dry-run"], capsys)
    assert code == EXIT_BUNDLE_VALIDATION
    assert error["error_category"] == "BUNDLE_VALIDATION_ERROR"
    assert "FULL" in error["blocking_reason"]


def test_raw_mzml_path_is_argument_error(production_bundles, capsys):
    paths = dict(production_bundles["paths"], FULL=production_bundles["sources"]["FULL"])
    code, _, error = _invoke(_arguments(paths) + ["--dry-run"], capsys)
    assert code == EXIT_ARGUMENT_CONFIG
    assert "raw mzML" in error["blocking_reason"]


def test_old_serializer_bundle_rejection(production_bundles, tmp_path, capsys):
    bundle = deepcopy(production_bundles["bundles"]["FULL"])
    del bundle["serializer_format_version"]
    _rehash(bundle)
    paths = dict(production_bundles["paths"], FULL=_write_bundle(tmp_path / "legacy.json", bundle))
    code, _, error = _invoke(_arguments(paths) + ["--dry-run"], capsys)
    assert code == EXIT_BUNDLE_VALIDATION
    assert "serializer_format_version" in error["blocking_reason"]


def test_rna_mismatch_is_compatibility_exit_4(production_bundles, tmp_path, capsys):
    bundle = deepcopy(production_bundles["bundles"]["T1"])
    bundle["rna"]["sequence"] = "ACGU"
    _rehash(bundle)
    paths = dict(production_bundles["paths"], T1=_write_bundle(tmp_path / "wrong-rna.json", bundle))
    code, _, error = _invoke(_arguments(paths) + ["--dry-run"], capsys)
    assert code == EXIT_COMPATIBILITY
    assert error["error_category"] == "COMPATIBILITY_ERROR"
    assert "RNA_SEQUENCE_MISMATCH:T1" in error["blocking_reason"]


def test_p1ap_source_mismatch_is_compatibility_error(production_bundles, tmp_path, capsys):
    source = tmp_path / "other-p1.mzML"
    source.write_bytes(b"different source")
    original = production_bundles["bundles"]["P1AP_MS2"]
    path = tmp_path / "other-ms2.json"
    export_layer_evidence_bundle(
        production_bundles["results"]["P1AP_MS2"], layer="P1AP_MS2",
        source_path=source, rna=RNA, experiment=dict(original["experiment"]),
        producer_commit=COMMIT, created_at_utc=CREATED, output_path=path,
        run_id="P1_RUN", sample_id="SAMPLE", biological_sample_id="BIO_SAMPLE",
    )
    paths = dict(production_bundles["paths"], P1AP_MS2=path)
    code, _, error = _invoke(_arguments(paths) + ["--dry-run"], capsys)
    assert code == EXIT_COMPATIBILITY
    assert "P1AP_SOURCE_" in error["blocking_reason"]


def test_existing_output_rejection(production_bundles, tmp_path, capsys):
    output = tmp_path / "exists.json"
    output.write_text("keep", encoding="utf-8")
    code, _, error = _invoke(
        _arguments(production_bundles["paths"]) + ["--aggregate-json", str(output)], capsys
    )
    assert code == EXIT_OUTPUT
    assert error["error_category"] == "OUTPUT_ERROR"
    assert output.read_text(encoding="utf-8") == "keep"


def test_partial_output_cleanup(production_bundles, tmp_path, capsys, monkeypatch):
    aggregate = tmp_path / "aggregate.json"
    excel = tmp_path / "cross-layer.xlsx"

    def fail_after_stage(*args, **kwargs):
        path = Path(kwargs["output_json_path"])
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("partial", encoding="utf-8")
        raise RuntimeError("simulated aggregation failure")

    monkeypatch.setattr(cli, "run_cross_layer_from_bundles", fail_after_stage)
    code, _, error = _invoke(
        _arguments(production_bundles["paths"]) + [
            "--aggregate-json", str(aggregate), "--excel-output", str(excel),
        ], capsys,
    )
    assert code == EXIT_AGGREGATION
    assert error["error_category"] == "AGGREGATION_ERROR"
    assert not aggregate.exists() and not excel.exists()
    assert not list(tmp_path.glob(".*cli-stage*"))


def test_final_rename_failure_rolls_back_committed_outputs(
    production_bundles, tmp_path, capsys, monkeypatch,
):
    aggregate = tmp_path / "aggregate.json"
    excel = tmp_path / "cross-layer.xlsx"
    summary = tmp_path / "summary.json"
    original_runner = cli.run_cross_layer_from_bundles

    def staged_runner(inputs, **kwargs):
        result = original_runner(inputs)
        Path(kwargs["output_json_path"]).write_text("staged json", encoding="utf-8")
        Path(kwargs["output_excel_path"]).write_bytes(b"staged excel")
        return result

    real_replace = cli.os.replace
    commit_count = 0

    def fail_second_commit(source, destination):
        nonlocal commit_count
        if ".cli-stage-" in Path(source).name:
            commit_count += 1
            if commit_count == 2:
                raise OSError("simulated final rename failure")
        return real_replace(source, destination)

    monkeypatch.setattr(cli, "run_cross_layer_from_bundles", staged_runner)
    monkeypatch.setattr(cli.os, "replace", fail_second_commit)
    code, _, error = _invoke(
        _arguments(production_bundles["paths"]) + [
            "--aggregate-json", str(aggregate), "--excel-output", str(excel),
            "--summary-json", str(summary),
        ], capsys,
    )
    assert code == EXIT_AGGREGATION
    assert error["error_category"] == "AGGREGATION_ERROR"
    assert not aggregate.exists() and not excel.exists() and not summary.exists()
    assert not list(tmp_path.glob(".*cli-stage*"))


def test_input_output_path_collision(production_bundles, capsys):
    collision = production_bundles["paths"]["FULL"]
    code, _, error = _invoke(
        _arguments(production_bundles["paths"]) + ["--aggregate-json", str(collision)], capsys
    )
    assert code == EXIT_OUTPUT
    assert "input/output path collision" in error["blocking_reason"]


def test_symlink_output_collision(production_bundles, tmp_path, capsys):
    link = tmp_path / "output-link.json"
    link.symlink_to(production_bundles["paths"]["FULL"])
    code, _, error = _invoke(
        _arguments(production_bundles["paths"]) + ["--aggregate-json", str(link)], capsys
    )
    assert code == EXIT_OUTPUT
    assert "collision" in error["blocking_reason"]


@pytest.mark.parametrize(
    "exception,expected_code,category",
    [
        (LayerEvidenceBundleError("invalid enum bundle"), EXIT_BUNDLE_VALIDATION, "BUNDLE_VALIDATION_ERROR"),
        (CrossLayerBundleRunnerError("incompatible bundle set: RNA mismatch"), EXIT_COMPATIBILITY, "COMPATIBILITY_ERROR"),
        (CrossLayerBundleRunnerError("output file already exists"), EXIT_OUTPUT, "OUTPUT_ERROR"),
        (CrossLayerBundleRunnerError("unsafe cross-layer safeguards"), EXIT_AGGREGATION, "AGGREGATION_ERROR"),
    ],
)
def test_exit_code_mapping(production_bundles, tmp_path, capsys, monkeypatch, exception, expected_code, category):
    def fail(*args, **kwargs):
        raise exception

    monkeypatch.setattr(cli, "run_cross_layer_from_bundles", fail)
    code, _, error = _invoke(
        _arguments(production_bundles["paths"]) + ["--aggregate-json", str(tmp_path / "out.json")], capsys
    )
    assert code == expected_code
    assert error["error_category"] == category


def test_summary_stdout_contains_contract(production_bundles, capsys):
    code, summary, _ = _invoke(_arguments(production_bundles["paths"]) + ["--dry-run"], capsys)
    assert code == EXIT_SUCCESS
    assert set(summary["bundle_sha256"]) == {"FULL", "T1", "P1AP_MS1", "P1AP_MS2"}
    assert summary["aggregate_counts"]["consensus"] == 1
    assert summary["consensus"]
    assert summary["confidence"]


def test_summary_json_atomic_readback(production_bundles, tmp_path, capsys):
    output = tmp_path / "reports" / "summary.json"
    aggregate = tmp_path / "reports" / "aggregate.json"
    code, summary, error = _invoke(
        _arguments(production_bundles["paths"]) + [
            "--aggregate-json", str(aggregate), "--summary-json", str(output),
        ], capsys
    )
    assert code == EXIT_SUCCESS and error is None
    assert aggregate.is_file()
    assert json.loads(output.read_text(encoding="utf-8")) == summary
    assert not list(output.parent.glob(".*tmp*"))


def test_summary_json_alone_is_not_a_primary_output(production_bundles, tmp_path, capsys):
    output = tmp_path / "summary.json"
    code, _, error = _invoke(
        _arguments(production_bundles["paths"]) + ["--summary-json", str(output)], capsys
    )
    assert code == EXIT_ARGUMENT_CONFIG
    assert "--aggregate-json or --excel-output" in error["blocking_reason"]
    assert not output.exists()


def test_full_provenance_warning_and_safeguards(production_bundles, capsys):
    code, summary, _ = _invoke(_arguments(production_bundles["paths"]) + ["--dry-run"], capsys)
    assert code == EXIT_SUCCESS
    assert FULL_PROVENANCE_WARNING in summary["warnings"]
    assert summary["compatibility"]["bundle_level_provenance_verified"] is True
    assert summary["compatibility"]["node_level_provenance"]["FULL"] == "UNAVAILABLE_IN_RESULT"
    assert summary["safeguards"]["verified"] is True
    assert summary["safeguards"]["formal_propagation"] is False
    assert summary["safeguards"]["shadow_analysis_only"] is True


def test_raw_parser_and_producers_not_invoked(production_bundles, capsys, monkeypatch):
    def forbidden(*args, **kwargs):
        raise AssertionError("raw parser or producer invoked")

    import rna_masshunter.sciex_profile_parser as parser
    import rna_masshunter.sciex_intact_oxygen_water_state_audit as full_module
    import rna_masshunter.sciex_t1_fragment_state_series_audit as t1_module
    import rna_masshunter.sciex_p1ap_nucleoside_state_audit as p1_module
    import rna_masshunter.sciex_p1ap_nucleoside_ms2_identity_audit as ms2_module
    monkeypatch.setattr(parser, "parse_sciex_profile", forbidden)
    monkeypatch.setattr(full_module, "audit_oxygen_water_state_series", forbidden)
    monkeypatch.setattr(t1_module, "audit_t1_fragment_state_series", forbidden)
    monkeypatch.setattr(p1_module, "audit_p1ap_nucleoside_state_series", forbidden)
    monkeypatch.setattr(ms2_module, "audit_p1ap_nucleoside_ms2_identity", forbidden)
    code, summary, _ = _invoke(_arguments(production_bundles["paths"]) + ["--dry-run"], capsys)
    assert code == EXIT_SUCCESS
    assert summary["status"] == "DRY_RUN_PASSED"


def test_config_section_absent_is_backward_compatible_noop(tmp_path, capsys):
    config = tmp_path / "config.yaml"
    config.write_text("analysis:\n  mode: full\n", encoding="utf-8")
    loaded = load_cross_layer_bundle_cli_config(config)
    assert loaded.enabled is False
    code, summary, error = _invoke(["--config", str(config)], capsys)
    assert code == EXIT_SUCCESS and error is None
    assert summary["status"] == "DISABLED"


def test_config_enabled_false_is_noop(tmp_path, capsys):
    config = tmp_path / "config.yaml"
    config.write_text(
        "cross_layer_bundle_runner:\n  enabled: false\n  bundles:\n    full: missing.json\n",
        encoding="utf-8",
    )
    code, summary, _ = _invoke(["--config", str(config)], capsys)
    assert code == EXIT_SUCCESS
    assert summary["status"] == "DISABLED"


def test_config_overwrite_true_is_rejected(tmp_path, capsys):
    config = tmp_path / "config.yaml"
    config.write_text(
        "cross_layer_bundle_runner:\n  enabled: true\n  overwrite: true\n",
        encoding="utf-8",
    )
    code, _, error = _invoke(["--config", str(config)], capsys)
    assert code == EXIT_ARGUMENT_CONFIG
    assert "overwrite=true" in error["blocking_reason"]


def test_config_relative_path_resolution(production_bundles, tmp_path, capsys):
    config = tmp_path / "runner.yaml"
    relative = {
        layer: path.relative_to(tmp_path)
        for layer, path in production_bundles["paths"].items()
    }
    payload = {
        "cross_layer_bundle_runner": {
            "enabled": True,
            "bundles": {
                "full": str(relative["FULL"]), "t1": str(relative["T1"]),
                "p1ap_ms1": str(relative["P1AP_MS1"]),
                "p1ap_ms2": str(relative["P1AP_MS2"]),
            },
            "output": {"aggregate_json": "relative/output.json"},
            "overwrite": False,
        }
    }
    config.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
    loaded = load_cross_layer_bundle_cli_config(config)
    assert loaded.bundles["FULL"] == tmp_path / relative["FULL"]
    assert loaded.aggregate_json == tmp_path / "relative/output.json"
    code, summary, error = _invoke(["--config", str(config), "--dry-run"], capsys)
    assert code == EXIT_SUCCESS and error is None
    assert summary["status"] == "DRY_RUN_PASSED"
    assert not loaded.aggregate_json.exists()


def test_paths_with_spaces_and_wsl_style_are_supported(production_bundles, tmp_path, capsys):
    folder = tmp_path / "Windows Style Folder"
    folder.mkdir()
    paths = {}
    for layer, source in production_bundles["paths"].items():
        destination = folder / f"{layer} bundle.json"
        shutil.copy2(source, destination)
        paths[layer] = destination
    code, summary, error = _invoke(_arguments(paths) + ["--dry-run"], capsys)
    assert code == EXIT_SUCCESS and error is None
    assert summary["compatibility"]["compatible"] is True


def test_cli_excel_output_readback(production_bundles, tmp_path, capsys):
    output = tmp_path / "cross-layer.xlsx"
    code, summary, error = _invoke(
        _arguments(production_bundles["paths"]) + ["--excel-output", str(output)], capsys
    )
    assert code == EXIT_SUCCESS and error is None
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert tuple(workbook.sheetnames) == XL_SHEET_NAMES
        assert workbook["XL_Nodes"].max_row - 1 == summary["aggregate_counts"]["nodes"]
    finally:
        workbook.close()


def test_cli_aggregate_json_readback(production_bundles, tmp_path, capsys):
    output = tmp_path / "aggregate.json"
    code, summary, error = _invoke(
        _arguments(production_bundles["paths"]) + ["--aggregate-json", str(output)], capsys
    )
    assert code == EXIT_SUCCESS and error is None
    payload = json.loads(output.read_text(encoding="utf-8"))
    assert payload["aggregate_counts"] == summary["aggregate_counts"]
    assert payload["consensus"]["cross_layer_evidence_status"] == summary["consensus"]
