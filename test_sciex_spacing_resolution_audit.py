from hashlib import sha256
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

import main as main_module
from rna_masshunter.audit_policy import AUDIT_DETAIL, AUDIT_SUMMARY, AuditPolicy, included_sheet_names, sheet_category
from rna_masshunter.config import DEFAULT_CONFIG, validate_config
from rna_masshunter.excel_report import SCIEX_INTACT_OPTIONAL_RESULT_KEY, write_excel_report
from rna_masshunter.models import RunConfig
from rna_masshunter.sciex_delta_mass_cluster_audit import SciexDeltaMassClusterAuditResult
from rna_masshunter.sciex_spacing_resolution_audit import (
    ALGORITHM_VERSION, AUDIT_RESULT_KEY, DEFAULT_PARAMETERS, DETAIL_COLUMNS,
    DETAIL_SHEET, ERROR_CODE, FORMAL_FALSE, SUMMARY_COLUMNS, SUMMARY_SHEET,
    WARNING_CODE, SciexSpacingResolutionAuditResult, SpacingResolutionParameters,
    annotate_cluster_summary, audit_sciex_spacing_resolution,
)

REAL_INPUT = Path(".cache/sciex_research/WT_LeuUAA(Full).txt")
REAL_SHA256 = "daae23985f9fda05902761346a75adf680ed2b141a1648346b7f6d5fda0b92a6"
CLUSTER_PARAMETERS = {
    "isotope_spacing_da": 1.003355,
    "isotope_spacing_tolerance_da": 0.15,
    "integer_spacing_tolerance_da": 0.15,
}


def relation(integer=False, isotope=False):
    return {
        "Integer_Spacing_Candidate": integer,
        "Isotope_Spacing_Candidate": isotope,
    }


def cluster_result(relations=None, status="MATCH", conflict=False, biological=True):
    summary = {
        "Cluster_Count": 2,
        "Input_Identity_Audit_Status": status,
        "Input_Identity_Conflict": conflict,
        "Biological_Interpretation_Eligible": biological,
        "Spacing_Resolution_Audit_Status": "NOT_RUN",
        "Estimated_Effective_Grid_Da": None,
        "Integer_Isotope_Distinguishable": False,
        "Isotope_Interpretation_Eligible": False,
        "Spacing_Resolution_Warning_Code": "",
    }
    clusters = ({"Cluster_ID": "C1"}, {"Cluster_ID": "C2"})
    return SciexDeltaMassClusterAuditResult(clusters, (summary,), tuple(relations or ()))


def audit(grid=None, apex=None, relations=None, cluster_parameters=None, parameters=None, **identity):
    grid = grid if grid is not None else [100.0 + 0.5 * i for i in range(25)]
    apex = apex if apex is not None else [100.0 + 1.5 * i for i in range(20)]
    return audit_sciex_spacing_resolution(
        grid, apex, cluster_result(relations, **identity),
        cluster_parameters or CLUSTER_PARAMETERS,
        parameters,
        source_file="synthetic.txt",
    )


def summary(result):
    return result.summaries()[0]


def detail(result, n=1):
    return result.details()[n - 1]


def test_default_config_schema():
    assert DEFAULT_CONFIG["sciex_profile"]["spacing_resolution_audit"] == DEFAULT_PARAMETERS


@pytest.mark.parametrize("field,value", [
    ("minimum_spacing_sample_count", 1),
    ("minimum_spacing_sample_count", 2.5),
    ("minimum_spacing_sample_count", True),
    ("quantization_tolerance_da", 0),
    ("quantization_tolerance_da", -0.1),
    ("quantization_tolerance_da", float("nan")),
    ("distinguishability_margin_factor", 0),
    ("distinguishability_margin_factor", float("inf")),
    ("maximum_spacing_multiple", 0),
    ("maximum_spacing_multiple", -1),
    ("maximum_spacing_multiple", 1.5),
])
def test_invalid_parameters(field, value):
    values = dict(DEFAULT_PARAMETERS)
    values[field] = value
    with pytest.raises((TypeError, ValueError)):
        SpacingResolutionParameters.from_mapping(values)


def test_config_validation_accepts_defaults():
    validate_config(route_config())


def test_config_validation_rejects_nonmapping():
    config = route_config()
    config.sciex_profile["spacing_resolution_audit"] = "bad"
    with pytest.raises(ValueError, match="spacing_resolution_audit"):
        validate_config(config)


def test_uniform_half_da_grid():
    row = summary(audit())
    assert row["Input_Grid_Median_Da"] == pytest.approx(0.5)
    assert row["Input_Grid_Mode_Da"] == pytest.approx(0.5)
    assert row["Input_Grid_Is_Uniform"] is True


def test_uniform_tenth_da_grid():
    row = summary(audit(grid=[100 + 0.1 * i for i in range(25)]))
    assert row["Estimated_Effective_Grid_Da"] == pytest.approx(0.1)


def test_nonuniform_grid():
    row = summary(audit(grid=[0, .5, 1, 1.75, 2.25, 3.1], apex=[.13, .47]))
    assert row["Input_Grid_Is_Uniform"] is False
    assert row["Input_Grid_Unique_Spacing_Count"] > 1


def test_duplicate_mass_points_are_excluded_from_spacings():
    row = summary(audit(grid=[0, .5, .5, 1.0], apex=[0, .5]))
    assert row["Input_Mass_Point_Count"] == 4
    assert row["Input_Adjacent_Spacing_Count"] == 2


def test_nonfinite_mass_points_are_excluded():
    row = summary(audit(grid=[0, .5, float("nan"), float("inf"), 1], apex=[0, .5]))
    assert row["Input_Mass_Point_Count"] == 3


def test_grid_min_mean_median_mode_max():
    row = summary(audit(grid=[0, .5, 1, 1.75], apex=[.13, .47]))
    assert row["Input_Grid_Min_Da"] == .5
    assert row["Input_Grid_Max_Da"] == .75
    assert row["Input_Grid_Mean_Da"] == pytest.approx(7 / 12)
    assert row["Input_Grid_Median_Da"] == .5
    assert row["Input_Grid_Mode_Da"] == .5


def test_grid_uniformity_fraction():
    row = summary(audit(grid=[0, .5, 1, 1.53], apex=[.13, .47]))
    assert row["Input_Grid_Uniformity_Fraction"] == pytest.approx(2 / 3)


def test_apex_quantization_half_da():
    row = summary(audit(apex=[100 + .5 * i for i in range(20)]))
    assert row["Apex_Quantization_Step_Da"] == .5


def test_apex_quantization_tenth_da():
    row = summary(audit(grid=[], apex=[100 + .1 * i for i in range(20)]))
    assert row["Apex_Quantization_Step_Da"] == .1


def test_apex_quantization_unknown():
    row = summary(audit(grid=[], apex=[0.013, .147, .386, .729]))
    assert row["Apex_Quantization_Step_Da"] is None
    assert row["Apex_Quantization_Confidence"] == "NONE"


def test_grid_confidence_high():
    assert summary(audit())["Grid_Confidence"] == "HIGH"


def test_grid_confidence_medium():
    values = [0.0]
    for i in range(21):
        values.append(values[-1] + (.53 if i in {9, 19} else .5))
    assert summary(audit(grid=values))["Grid_Confidence"] == "MEDIUM"


def test_grid_confidence_low():
    assert summary(audit(grid=[0, .4, 1.1], apex=[.013, .147]))["Grid_Confidence"] == "LOW"


def test_tolerance_windows_overlap():
    row = detail(audit())
    assert row["Tolerance_Windows_Overlap"] is True
    assert row["Tolerance_Window_Overlap_Da"] > 0


def test_tolerance_windows_do_not_overlap():
    params = dict(CLUSTER_PARAMETERS, isotope_spacing_tolerance_da=.001, integer_spacing_tolerance_da=.001)
    assert detail(audit(cluster_parameters=params))["Tolerance_Windows_Overlap"] is False


def test_grid_limited_status():
    params = dict(CLUSTER_PARAMETERS, isotope_spacing_tolerance_da=.001, integer_spacing_tolerance_da=.001)
    assert detail(audit(cluster_parameters=params))["Resolution_Status"] == "NOT_DISTINGUISHABLE_GRID_LIMITED"


def test_tolerance_limited_status():
    grid = [i * .001 for i in range(30)]
    assert detail(audit(grid=grid))["Resolution_Status"] == "NOT_DISTINGUISHABLE_TOLERANCE_OVERLAP"


def test_both_limited_status():
    assert detail(audit())["Resolution_Status"] == "NOT_DISTINGUISHABLE_BOTH"


def test_distinguishable_status():
    params = {"isotope_spacing_da": 1.5, "isotope_spacing_tolerance_da": .01, "integer_spacing_tolerance_da": .01}
    assert detail(audit(grid=[i * .1 for i in range(30)], cluster_parameters=params))["Resolution_Status"] == "DISTINGUISHABLE"


def test_marginally_distinguishable_status():
    params = {"isotope_spacing_da": 1.15, "isotope_spacing_tolerance_da": .01, "integer_spacing_tolerance_da": .01}
    assert detail(audit(grid=[i * .1 for i in range(30)], cluster_parameters=params))["Resolution_Status"] == "MARGINALLY_DISTINGUISHABLE"


def test_insufficient_information_status():
    assert detail(audit(grid=[], apex=[]))["Resolution_Status"] == "INSUFFICIENT_INFORMATION"


def test_target_separation_n1():
    assert detail(audit(), 1)["Target_Separation_Da"] == pytest.approx(.003355)


def test_target_separation_n2():
    assert detail(audit(), 2)["Target_Separation_Da"] == pytest.approx(.00671)


def test_maximum_spacing_multiple_controls_detail_count():
    result = audit(parameters={"maximum_spacing_multiple": 4})
    assert len(result.details()) == 4
    assert result.details()[-1]["Spacing_Multiple"] == 4


def test_relation_flag_counts():
    rows = [relation(True, True), relation(True, False), relation(False, True), relation(False, False)]
    row = summary(audit(relations=rows))
    assert row["Total_Relation_Count"] == 4
    assert row["Integer_Candidate_Count"] == 2
    assert row["Isotope_Candidate_Count"] == 2
    assert row["Dual_Integer_And_Isotope_Count"] == 1
    assert row["Integer_Only_Count"] == 1
    assert row["Isotope_Only_Count"] == 1
    assert row["Neither_Count"] == 1
    assert row["Dual_Flag_Fraction"] == .25


def test_isotope_interpretation_false_when_unresolved():
    assert summary(audit(relations=[relation(True, True)]))["Isotope_Interpretation_Eligible"] is False


def test_numerical_interpretation_true_with_relations():
    assert summary(audit(relations=[relation(True, False)]))["Numerical_Spacing_Interpretation_Eligible"] is True


def test_numerical_interpretation_false_without_relations():
    assert summary(audit(relations=[]))["Numerical_Spacing_Interpretation_Eligible"] is False


def test_chemical_interpretation_always_false():
    assert summary(audit(relations=[relation(True, True)]))["Chemical_Interpretation_Eligible"] is False


@pytest.mark.parametrize("status,conflict,biological", [
    ("MATCH", False, True), ("CONFLICT", True, False),
])
def test_identity_is_reflected_without_controlling_audit(status, conflict, biological):
    row = summary(audit(status=status, conflict=conflict, biological=biological))
    assert row["Audit_Status"] == "AUDIT_COMPLETED"
    assert row["Input_Identity_Audit_Status"] == status
    assert row["Input_Identity_Conflict"] is conflict
    assert row["Biological_Interpretation_Eligible"] is biological


def test_warning_fields_require_unresolved_dual_candidate():
    row = summary(audit(relations=[relation(True, True)]))
    assert row["Warning_Code"] == WARNING_CODE
    assert "cannot distinguish" in row["Warning_Message"]


def test_candidate_without_dual_flag_has_no_warning():
    assert summary(audit(relations=[relation(False, True)]))["Warning_Code"] == ""


def test_no_candidate_has_no_warning():
    assert summary(audit(relations=[]))["Warning_Code"] == ""


def test_warning_message_is_deterministic():
    first = summary(audit(relations=[relation(True, True)]))["Warning_Message"]
    second = summary(audit(relations=[relation(True, True)]))["Warning_Message"]
    assert first == second


def test_effective_grid_is_deterministic():
    assert summary(audit())["Estimated_Effective_Grid_Da"] == summary(audit())["Estimated_Effective_Grid_Da"]


def test_input_reordering_does_not_change_grid_result():
    values = [100 + .5 * i for i in range(25)]
    first = summary(audit(grid=values))
    second = summary(audit(grid=list(reversed(values))))
    keys = ["Input_Grid_Median_Da", "Input_Grid_Mode_Da", "Estimated_Effective_Grid_Da", "Resolution_Status"]
    assert [first[key] for key in keys] == [second[key] for key in keys]


def test_all_rows_are_shadow_only_and_nonpropagating():
    result = audit(relations=[relation(True, True)])
    for row in result.summaries() + result.details():
        assert row["Shadow_Only"] is True
        assert row["Applied_To_Formal_Score"] is False
        assert row["Applied_To_Ranking"] is False
        assert row["Applied_To_Candidate_Filtering"] is False
        assert row["Molecular_Identity_Assigned"] is False


def test_column_contracts_include_formal_flags():
    for column in FORMAL_FALSE:
        assert column in SUMMARY_COLUMNS and column in DETAIL_COLUMNS
    assert len(DETAIL_SHEET) == 31


def test_result_rows_are_defensive_copies():
    result = audit()
    copied = result.summaries()
    copied[0]["Resolution_Status"] = "CHANGED"
    assert result.summaries()[0]["Resolution_Status"] != "CHANGED"


def test_cluster_summary_annotation_preserves_cluster_and_relation_rows():
    original = cluster_result([relation(True, True)])
    updated = annotate_cluster_summary(original, audit(relations=[relation(True, True)]))
    assert updated.clusters() == original.clusters()
    assert updated.relations() == original.relations()
    assert updated.summaries()[0]["Spacing_Resolution_Audit_Status"] == "NOT_DISTINGUISHABLE_BOTH"
    assert updated.summaries()[0]["Estimated_Effective_Grid_Da"] == .5


def test_cluster_summary_annotation_does_not_mutate_source():
    original = cluster_result([relation(True, True)])
    before = original.summaries()
    annotate_cluster_summary(original, audit(relations=[relation(True, True)]))
    assert original.summaries() == before


class Detection:
    def __init__(self, status="DETECTION_COMPLETED", profile="NEUTRAL_MASS_PROFILE", peaks=2, valid=True):
        self.status, self.profile, self.peaks, self.valid = status, profile, peaks, valid
    def diagnostics_row(self):
        return {"Detection_Status": self.status, "Profile_Type": self.profile}
    def peak_rows(self):
        return [{"Apex_Mass": 100 + i * .5 if self.valid else "bad"} for i in range(self.peaks)]


class Parsed:
    def __init__(self, masses=None):
        self.input_rows = [{"Neutral_Mass": value} for value in (masses or [100 + .5 * i for i in range(25)])]


def route_config(enabled=True, resolution_enabled=True, cluster_enabled=True):
    return RunConfig(sciex_profile={
        "enabled": enabled, "path": "synthetic.txt",
        "intact_peak_detection": {"enabled": True},
        "intact_mass_comparison": {"enabled": True},
        "delta_mass_cluster_audit": {**CLUSTER_PARAMETERS, "enabled": cluster_enabled},
        "spacing_resolution_audit": {**DEFAULT_PARAMETERS, "enabled": resolution_enabled},
    })


def routed(config=None, detection=None, parsed=None, cluster=None, warnings=None):
    detection = detection if detection is not None else Detection()
    parsed = parsed if parsed is not None else Parsed()
    sciex = {SCIEX_INTACT_OPTIONAL_RESULT_KEY: {
        "result": detection, "parsed_result": parsed, "source_file": "synthetic.txt",
    }} if detection is not False else {}
    clusters = {"sciex_delta_mass_cluster_audit": cluster if cluster is not None else cluster_result([relation(True, True)])}
    return main_module.build_sciex_spacing_resolution_optional_results(
        config or route_config(), sciex, clusters, [] if warnings is None else warnings,
    )


@pytest.mark.parametrize("config,detection,parsed,cluster", [
    (RunConfig(), Detection(), Parsed(), cluster_result()),
    (route_config(enabled=False), Detection(), Parsed(), cluster_result()),
    (route_config(resolution_enabled=False), Detection(), Parsed(), cluster_result()),
    (route_config(cluster_enabled=False), False, Parsed(), cluster_result()),
    (route_config(), False, Parsed(), cluster_result()),
    (route_config(), Detection(status="FAILED"), Parsed(), cluster_result()),
    (route_config(), Detection(profile="MZ_PROFILE"), Parsed(), cluster_result()),
    (route_config(), Detection(peaks=0), Parsed(), cluster_result()),
    (route_config(), Detection(peaks=1), Parsed(), cluster_result()),
    (route_config(), Detection(valid=False), Parsed(), cluster_result()),
    (route_config(), Detection(), Parsed([100]), cluster_result()),
])
def test_routing_skips_ineligible_inputs(config, detection, parsed, cluster):
    assert routed(config, detection, parsed, cluster) == {}


def test_routing_skips_missing_cluster_result():
    sciex = {SCIEX_INTACT_OPTIONAL_RESULT_KEY: {"result": Detection(), "parsed_result": Parsed(), "source_file": "x"}}
    assert main_module.build_sciex_spacing_resolution_optional_results(route_config(), sciex, {}, []) == {}


def test_routing_executes_for_completed_neutral_inputs():
    assert AUDIT_RESULT_KEY in routed()


def test_routing_warning_is_added_once():
    warnings = []
    routed(warnings=warnings)
    routed(warnings=warnings)
    matching = [row for row in warnings if row.get("Context", {}).get("Warning_Code") == WARNING_CODE]
    assert len(matching) == 1


def test_routing_exception_preserves_upstream_and_adds_error(monkeypatch):
    warnings = []
    detector, parsed, clusters = Detection(), Parsed(), cluster_result([relation(True, True)])
    monkeypatch.setattr(main_module, "audit_sciex_spacing_resolution", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    assert routed(detection=detector, parsed=parsed, cluster=clusters, warnings=warnings) == {}
    assert detector.peak_rows() and parsed.input_rows and clusters.relations()
    assert warnings[-1]["Context"]["Warning_Code"] == ERROR_CODE


def test_sheet_registry_and_policy():
    assert sheet_category(SUMMARY_SHEET) == AUDIT_SUMMARY
    assert sheet_category(DETAIL_SHEET) == AUDIT_DETAIL
    names = [SUMMARY_SHEET, DETAIL_SHEET]
    assert included_sheet_names(names, AuditPolicy.from_level("standard"))[0] == []
    assert included_sheet_names(names, AuditPolicy.from_level("audit"))[0] == names
    assert included_sheet_names(names, AuditPolicy.from_level("full"))[0] == names


def writer_config():
    return SimpleNamespace(
        analysis={"mode": "full"}, project={"name": "resolution-audit"}, input={},
        organism={}, sequence={}, experiment={}, instrument={}, sciex_profile={},
        reconstruction={"enabled": False}, digestion={"enabled": False},
        alkaline_phosphatase={}, fragment_mapping={}, modification_search={}, peak_filtering={},
        p1_annotation={}, ms2_annotation={}, modification_evidence_ranking={}, biological_context={},
        performance={}, reporting={"max_excel_rows_per_sheet": 100000, "truncate_large_sheets": True},
    )


def write_report(tmp_path, level, optional):
    return write_excel_report(
        tmp_path / level, writer_config(), {}, [], [], [],
        known_modification_candidates=[{"Candidate_ID": "C1"}],
        known_modification_summary=[{"Summary_Key": "S1"}],
        optional_results=optional, audit_policy=AuditPolicy.from_level(level),
    )


@pytest.mark.parametrize("level,present", [("standard", False), ("audit", True), ("full", True)])
def test_excel_sheet_presence_columns_and_order(tmp_path, level, present):
    result = audit(parameters={"maximum_spacing_multiple": 4})
    report = write_report(tmp_path, level, {AUDIT_RESULT_KEY: result})
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        names = workbook.sheetnames
    finally:
        workbook.close()
    assert (SUMMARY_SHEET in names) is present
    assert (DETAIL_SHEET in names) is present
    if present:
        summary_frame = pd.read_excel(report, sheet_name=SUMMARY_SHEET, header=2)
        detail_frame = pd.read_excel(report, sheet_name=DETAIL_SHEET, header=2)
        assert list(summary_frame.columns) == SUMMARY_COLUMNS
        assert list(detail_frame.columns) == DETAIL_COLUMNS
        assert len(summary_frame) == 1 and len(detail_frame) == 4
        assert detail_frame["Spacing_Multiple"].tolist() == [1, 2, 3, 4]


def formal_fixture():
    return {
        "Known_Modification_Candidates": [{"ID": "K1", "Value": 1.0}],
        "Known_Modification_Summary": [{"ID": "KS", "Value": 2.0}],
        "Modification_Evidence_Ranking": [{"ID": "R1", "Value": 3.0}],
        "Top_Modification_Candidates": [{"ID": "T1", "Value": 4.0}],
        "P1_Summary": [{"ID": "P1", "Value": 5.0}],
        "MS2_Summary": [{"ID": "M1", "Value": 6.0}],
    }


@pytest.mark.parametrize("variant", [
    "sciex_disabled_vs_enabled", "audit_disabled_vs_enabled",
    "distinguishable_vs_indistinguishable", "identity_match_vs_conflict",
    "audit_success_vs_exception",
])
def test_controlled_ab_formal_results_are_identical(tmp_path, variant):
    first_shadow = None
    second_shadow = audit(relations=[relation(True, True)])
    if variant == "distinguishable_vs_indistinguishable":
        first_shadow = audit(
            cluster_parameters={
                "isotope_spacing_da": 1.5,
                "isotope_spacing_tolerance_da": .01,
                "integer_spacing_tolerance_da": .01,
            }
        )
    elif variant == "identity_match_vs_conflict":
        first_shadow = audit(status="MATCH", conflict=False, biological=True)
        second_shadow = audit(status="CONFLICT", conflict=True, biological=False)

    first_optional = formal_fixture()
    second_optional = formal_fixture()
    if first_shadow is not None:
        first_optional[AUDIT_RESULT_KEY] = first_shadow
    if second_shadow is not None:
        second_optional[AUDIT_RESULT_KEY] = second_shadow
    first_report = write_report(tmp_path / variant / "first", "full", first_optional)
    second_report = write_report(tmp_path / variant / "second", "full", second_optional)

    for sheet_name in formal_fixture():
        first_frame = pd.read_excel(first_report, sheet_name=sheet_name, header=2)
        second_frame = pd.read_excel(second_report, sheet_name=sheet_name, header=2)
        pd.testing.assert_frame_equal(
            first_frame, second_frame, check_dtype=False, check_exact=True,
        )


def test_real_input_sha256_and_non_destructive_read():
    assert REAL_INPUT.exists()
    before = sha256(REAL_INPUT.read_bytes()).hexdigest()
    assert before == REAL_SHA256
    audit()
    after = sha256(REAL_INPUT.read_bytes()).hexdigest()
    assert after == before


def test_algorithm_version_and_notes_are_explicit():
    result = audit()
    assert all(row["Algorithm_Version"] == ALGORITHM_VERSION for row in result.summaries() + result.details())
    assert "unchanged" in result.summaries()[0]["Notes"]
