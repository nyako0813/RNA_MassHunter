from hashlib import sha256
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

import main as main_module
from rna_masshunter.audit_policy import AUDIT_SUMMARY, AuditPolicy, included_sheet_names, sheet_category
from rna_masshunter.excel_report import write_excel_report
from rna_masshunter.models import RunConfig
from rna_masshunter.sciex_input_identity_audit import (
    AUDIT_RESULT_KEY, OUTPUT_COLUMNS, SHEET_NAME, WARNING_CODE,
    audit_sciex_input_identity, extract_sciex_identity_tokens,
)
from rna_masshunter.sciex_intact_mass_comparison import compare_sciex_intact_masses


def audit(filename, name="tRNA-Leu-UAA", anticodon="UAA", sequence="ACGU"):
    return audit_sciex_input_identity(
        filename, sequence_name=name, anticodon=anticodon, sequence=sequence,
        organism_group="archaea", species="example", condition_name="wild_type",
    )


@pytest.mark.parametrize("filename,name,anticodon", [
    ("LeuUAA.txt", "tRNA-Leu-UAA", "UAA"),
    ("Leu-UAA.txt", "tRNA^Leu-UAA", "UAA"),
    ("leu_uaa_profile.txt", "TRNA-LEU-UAA", "UAA"),
    ("ｔＲＮＡ－Ｌｅｕ－ＵＡＡ（Ｆｕｌｌ）.txt", "tRNA-Leu-UAA", "UAA"),
    ("WT.Leu[UAA].mass.txt", "LeuUAA", "UAA"),
])
def test_equivalent_trna_spellings_match(filename, name, anticodon):
    row = audit(filename, name, anticodon).row()
    assert row["Audit_Status"] == "MATCH"
    assert row["Identity_Conflict"] is False


@pytest.mark.parametrize("filename,name,anticodon", [
    ("Leu_sample.txt", "tRNA-Leu-UAA", "UAA"),
    ("UAA_spectrum.txt", "tRNA-Leu-UAA", "UAA"),
    ("Glu_profile.txt", "MA_tRNA^Glu-UUC", "UUC"),
])
def test_one_identity_component_only_is_partial(filename, name, anticodon):
    assert audit(filename, name, anticodon).row()["Audit_Status"] == "PARTIAL_MATCH"


@pytest.mark.parametrize("filename,name,anticodon", [
    ("sample_01.txt", "tRNA-Leu-UAA", "UAA"),
    ("profile.txt", "tRNA-Leu-UAA", "UAA"),
    ("LeuUAA.txt", "target_tRNA", ""),
    ("sample.txt", "target_tRNA", ""),
])
def test_missing_identity_is_insufficient_not_conflict(filename, name, anticodon):
    row = audit(filename, name, anticodon).row()
    assert row["Audit_Status"] == "INSUFFICIENT_INFORMATION"
    assert row["Identity_Conflict"] is False


def test_no_config_name_or_sequence_is_not_eligible():
    row = audit_sciex_input_identity("LeuUAA.txt", sequence_name="", sequence="", anticodon="UAA").row()
    assert row["Audit_Status"] == "NOT_ELIGIBLE"
    assert row["Audit_Eligible"] is False


@pytest.mark.parametrize("filename,name,anticodon,amino,anti", [
    ("Leu.txt", "tRNA-Glu", "", False, None),
    ("UAA.txt", "target_tRNA", "UUC", None, False),
    ("LeuUAA.txt", "tRNA-Leu-UUC", "UUC", True, False),
    ("LeuUAA.txt", "tRNA-Glu-UAA", "UAA", False, True),
    ("WT_LeuUAA(Full).txt", "MA_tRNA^Glu-UUC", "UUC", False, False),
])
def test_explicit_component_conflicts(filename, name, anticodon, amino, anti):
    row = audit(filename, name, anticodon).row()
    assert row["Audit_Status"] == "CONFLICT"
    assert row["Amino_Acid_Match"] is amino
    assert row["Anticodon_Match"] is anti
    assert row["Identity_Conflict"] is True


def test_real_mismatch_has_expected_high_evidence_and_tokens():
    row = audit("WT_LeuUAA(Full).txt", "MA_tRNA^Glu-UUC", "UUC").row()
    assert row["SCIEX_Amino_Acid_Tokens"] == "Leu"
    assert row["SCIEX_Anticodon_Tokens"] == "UAA"
    assert row["SCIEX_Combined_Identity_Tokens"] == "LeuUAA"
    assert row["Configured_Amino_Acid_Tokens"] == "Glu"
    assert row["Configured_Anticodon_Tokens"] == "UUC"
    assert row["Configured_Combined_Identity_Tokens"] == "GluUUC"
    assert row["Identity_Evidence_Level"] == "HIGH"
    assert row["Warning_Code"] == WARNING_CODE
    assert "tRNA-Leu-UAA" in row["Warning_Message"]
    assert "tRNA-Glu-UUC" in row["Warning_Message"]


def test_combined_token_match_is_recorded():
    row = audit("LeuUAA.txt", "LeuUAA", "").row()
    assert row["Combined_Identity_Match"] is True
    assert row["Audit_Status"] == "MATCH"


@pytest.mark.parametrize("filename", [
    "glucose_UAA.txt", "blue_UAA.txt", "glutamate_UAA.txt", "leucine_UAA.txt",
])
def test_substrings_do_not_create_amino_acid_tokens(filename):
    tokens = extract_sciex_identity_tokens(filename)
    assert not tokens.amino_acids
    assert not tokens.combined


def test_noise_run_replicate_and_extension_tokens_are_removed():
    tokens = extract_sciex_identity_tokens("WT_LeuUAA_Full_profile_run_02_rep3_R4.txt")
    assert tokens.filename_tokens == ("leu", "leuuaa", "uaa")


def test_parent_directory_does_not_influence_filename_identity():
    row = audit("LeuUAA_directory/sample.txt").row()
    assert row["SCIEX_Parent_Directory"] == "LeuUAA_directory"
    assert row["SCIEX_Amino_Acid_Tokens"] == ""
    assert row["Audit_Status"] == "INSUFFICIENT_INFORMATION"


def test_token_output_is_sorted_and_deterministic():
    first = extract_sciex_identity_tokens("UAA-Leu-LeuUAA.txt")
    second = extract_sciex_identity_tokens("UAA-Leu-LeuUAA.txt")
    assert first == second
    assert first.filename_tokens == tuple(sorted(first.filename_tokens))


def route_config(enabled=True, name="MA_tRNA^Glu-UUC", anticodon="UUC"):
    return RunConfig(
        sciex_profile={"enabled": enabled, "path": "WT_LeuUAA(Full).txt"},
        sequence={"name": name, "sequence": "ACGU", "anticodon": anticodon},
        organism={"group": "archaea", "species": "example"},
        experiment={"condition_name": "wild_type"},
    )


def test_absent_or_disabled_sciex_produces_no_result_or_warning():
    for config in (RunConfig(), route_config(enabled=False)):
        warnings = []
        assert main_module.build_sciex_input_identity_audit_optional_results(config, warnings) == {}
        assert warnings == []


def test_conflict_adds_one_warning_and_duplicate_call_does_not_repeat_it():
    warnings = []
    config = route_config()
    first = main_module.build_sciex_input_identity_audit_optional_results(config, warnings)
    second = main_module.build_sciex_input_identity_audit_optional_results(config, warnings)
    assert AUDIT_RESULT_KEY in first and AUDIT_RESULT_KEY in second
    assert len(warnings) == 1
    assert warnings[0]["Context"]["Warning_Code"] == WARNING_CODE


@pytest.mark.parametrize("filename,name,anticodon", [
    ("LeuUAA.txt", "tRNA-Leu-UAA", "UAA"),
    ("Leu.txt", "tRNA-Leu-UAA", "UAA"),
    ("sample.txt", "tRNA-Leu-UAA", "UAA"),
])
def test_nonconflict_statuses_do_not_add_warning(filename, name, anticodon):
    config = route_config(name=name, anticodon=anticodon)
    config.sciex_profile["path"] = filename
    warnings = []
    main_module.build_sciex_input_identity_audit_optional_results(config, warnings)
    assert warnings == []


def test_audit_exception_isolated_and_comparison_can_continue(monkeypatch):
    warnings = []
    monkeypatch.setattr(main_module, "audit_sciex_input_identity", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    assert main_module.build_sciex_input_identity_audit_optional_results(route_config(), warnings) == {}
    assert warnings[-1]["Source"] == "sciex_input_identity_audit"
    assert warnings[-1]["Level"] == "ERROR"
    comparison = compare_sciex_intact_masses(Detection(), 100.5).summaries()[0]
    assert comparison["Closest_Delta_Mass"] == -0.5
    assert comparison["Input_Identity_Audit_Status"] == "NOT_RUN"


class Detection:
    def peak_rows(self):
        return [{"Peak_ID": "P1", "Apex_Mass": 100.0, "Apex_Intensity_Raw": 10.0}]
    def diagnostics_row(self):
        return {"Detection_Status": "DETECTION_COMPLETED", "Input_Status": "SUPPORTED_INPUT"}


def test_comparison_conflict_returns_diagnostic_summary_without_numeric_matching():
    identity = audit("LeuUAA.txt", "tRNA-Glu-UUC", "UUC")
    result = compare_sciex_intact_masses(Detection(), 100.5, input_identity_audit=identity)
    linked = result.summaries()[0]
    assert result.details() == []
    assert linked["Comparison_Status"] == "IDENTITY_CONFLICT"
    assert linked["Comparison_Eligible"] is False
    assert linked["Closest_Observed_Mass"] is None
    assert linked["Closest_Delta_Mass"] is None
    assert linked["Closest_Delta_ppm"] is None
    assert linked["Input_Identity_Audit_Status"] == "CONFLICT"
    assert linked["Input_Identity_Conflict"] is True
    assert linked["Input_Identity_Warning_Code"] == WARNING_CODE
    assert linked["Biological_Interpretation_Eligible"] is False


def test_formal_and_identity_flags_are_nonpropagating():
    row = audit("LeuUAA.txt").row()
    assert row["Shadow_Only"] is True
    assert row["Applied_To_Formal_Score"] is False
    assert row["Applied_To_Ranking"] is False
    assert row["Applied_To_Candidate_Filtering"] is False
    assert row["Molecular_Identity_Assigned"] is False


def writer_config():
    return SimpleNamespace(
        analysis={"mode": "full"}, project={"name": "identity-audit"}, input={},
        organism={}, sequence={}, experiment={}, instrument={}, sciex_profile={},
        reconstruction={"enabled": False}, digestion={"enabled": False},
        alkaline_phosphatase={}, fragment_mapping={}, modification_search={}, peak_filtering={},
        p1_annotation={}, ms2_annotation={}, modification_evidence_ranking={}, biological_context={},
        performance={}, reporting={"max_excel_rows_per_sheet": 1000, "truncate_large_sheets": True},
    )


def write_report(tmp_path, level, optional):
    report_path, _word_appendix_path = write_excel_report(
        tmp_path / level, writer_config(), {}, [], [], [],
        known_modification_candidates=[{"Candidate_ID": "C1"}],
        known_modification_summary=[{"Summary_Key": "S1"}],
        optional_results=optional, audit_policy=AuditPolicy.from_level(level),
    )
    return report_path


@pytest.mark.parametrize("level,present", [("standard", False), ("audit", True), ("full", True)])
def test_excel_sheet_policy_and_fixed_columns(tmp_path, level, present):
    report = write_report(tmp_path, level, {AUDIT_RESULT_KEY: audit("LeuUAA.txt")})
    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        names = workbook.sheetnames
    finally:
        workbook.close()
    assert (SHEET_NAME in names) is present
    if present:
        frame = pd.read_excel(report, sheet_name=SHEET_NAME, header=2)
        assert len(frame) == 1
        assert list(frame.columns) == OUTPUT_COLUMNS


def test_sheet_is_registered_as_audit_summary():
    assert len(SHEET_NAME) <= 31
    assert sheet_category(SHEET_NAME) == AUDIT_SUMMARY
    assert included_sheet_names([SHEET_NAME], AuditPolicy.from_level("standard"))[0] == []
    assert included_sheet_names([SHEET_NAME], AuditPolicy.from_level("audit"))[0] == [SHEET_NAME]


def test_controlled_formal_ab_is_exact_for_disabled_match_conflict_and_error(tmp_path):
    formal = {
        "Modification_Evidence_Ranking": [{"Candidate_ID": "C1", "Final_Score": 1.0}],
        "Top_Modification_Candidates": [{"Candidate_ID": "C1", "Final_Score": 1.0}],
        "P1_Summary": [{"Status": "CONTROL"}],
        "MS2_Summary": [{"Status": "CONTROL"}],
    }
    variants = {
        "disabled": dict(formal),
        "match": {**formal, AUDIT_RESULT_KEY: audit("LeuUAA.txt")},
        "conflict": {**formal, AUDIT_RESULT_KEY: audit("LeuUAA.txt", "tRNA-Glu-UUC", "UUC")},
        "error": dict(formal),
    }
    reports = {name: write_report(tmp_path / name, "full", optional) for name, optional in variants.items()}
    sheets = [
        "Known_Modification_Candidates", "Known_Modification_Summary",
        "Modification_Evidence_Ranking", "Top_Modification_Candidates", "P1_Summary", "MS2_Summary",
    ]
    base = reports["disabled"]
    for report in reports.values():
        for sheet in sheets:
            expected = pd.read_excel(base, sheet_name=sheet, header=2, dtype=object).fillna("")
            actual = pd.read_excel(report, sheet_name=sheet, header=2, dtype=object).fillna("")
            pd.testing.assert_frame_equal(expected, actual, check_dtype=False)


def test_real_input_sha_and_contents_are_non_destructive():
    import pytest
    path = Path(".cache/sciex_research/WT_LeuUAA(Full).txt")
    if not path.is_file():
        pytest.skip("Local SCIEX real-data fixture is not available: .cache/sciex_research/WT_LeuUAA(Full).txt")
    before = sha256(path.read_bytes()).hexdigest()
    audit(path, "MA_tRNA^Glu-UUC", "UUC")
    after = sha256(path.read_bytes()).hexdigest()
    assert before == after == "daae23985f9fda05902761346a75adf680ed2b141a1648346b7f6d5fda0b92a6"