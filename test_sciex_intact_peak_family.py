from __future__ import annotations

from dataclasses import FrozenInstanceError, replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

from rna_masshunter.intact_rna_average_mass import PROTON_MASS_DA
from rna_masshunter.intact_rna_candidate_generation import generate_candidates_for_measurement
from rna_masshunter.modifications import load_modifications
from rna_masshunter.sciex_intact_peak_family import (
    DeltaComparisonRole,
    DeltaMassDefinition,
    DeltaMatchClass,
    DeltaReference,
    DeltaReferenceCategory,
    PeakFamilyParameters,
    PeakQualityClass,
    analyze_sciex_intact_peak_families,
    build_delta_reference_registry,
    build_peak_families,
    calculate_peak_metrics,
    connect_candidate_baselines,
    generate_delta_pairs,
    match_delta_pairs,
    select_major_peaks,
)
from rna_masshunter.sciex_sample_manifest import get_rna_identity, load_sciex_sample_manifest

ROOT = Path(__file__).parent
MANIFEST_PATH = ROOT / "data" / "sciex_sample_manifest.yaml"
REFERENCE_PATH = ROOT / "data" / "reference" / "methanosarcina_acetivorans_c2a_trna.xlsx"


class FakeDetection:
    def __init__(self, rows, status="DETECTION_COMPLETED"):
        self._rows = tuple(dict(row) for row in rows)
        self._status = status

    def peak_rows(self):
        return [dict(row) for row in self._rows]

    def diagnostics_row(self):
        return {"Detection_Status": self._status}


def detector_row(
    peak_id,
    mass,
    intensity,
    area,
    prominence,
    fwhm,
    *,
    shoulder=False,
    centroid_offset=0.1,
):
    return {
        "Peak_ID": peak_id,
        "Apex_Mass": float(mass),
        "Centroid_Mass": float(mass + centroid_offset),
        "Apex_Intensity_Raw": float(intensity),
        "Peak_Area_Baseline_Corrected": float(area),
        "Left_Boundary_Mass": float(mass - fwhm),
        "Right_Boundary_Mass": float(mass + fwhm),
        "FWHM_Da": float(fwhm),
        "Prominence": float(prominence),
        "Possible_Shoulder": shoulder,
        "Strict_Threshold_Passed": True,
    }


@pytest.fixture
def quality_detection():
    return FakeDetection([
        detector_row("P1", 100.0, 100.0, 100.0, 100.0, 2.0),
        detector_row("P2", 120.0, 80.0, 80.0, 80.0, 8.0),
        detector_row("P3", 140.0, 2.0, 2.0, 2.0, 2.0),
        detector_row("P4", 160.0, 2.0, 2.0, 2.0, 8.0),
        detector_row("P5", 180.0, 50.0, 50.0, 50.0, 2.0, shoulder=True),
        detector_row("P6", 200.0, 0.5, 0.5, 0.4, 2.0),
    ])


@pytest.fixture
def quality_peaks(quality_detection):
    return calculate_peak_metrics(
        quality_detection,
        source_id="SOURCE",
        measurement_id="MEASUREMENT",
        rna_identity="TRNA_LEU_UAA",
    )


def test_peak_metrics_reuse_detector_fields_and_are_deterministic(quality_detection):
    before = tuple(quality_detection._rows)
    first = calculate_peak_metrics(
        quality_detection, source_id="S", measurement_id="M", rna_identity="R"
    )
    second = calculate_peak_metrics(
        quality_detection, source_id="S", measurement_id="M", rna_identity="R"
    )
    assert first == second
    assert tuple(quality_detection._rows) == before
    peak = first[0]
    assert peak.apex_mass == 100.0
    assert peak.centroid_mass == 100.1
    assert peak.peak_width_da == 4.0
    assert peak.fwhm_da == 2.0
    assert peak.prominence == 100.0
    assert peak.relative_apex_intensity == 1.0
    assert peak.relative_integrated_intensity == 1.0
    assert peak.sharpness_score == 0.5


def test_all_quality_classes_are_reachable(quality_peaks):
    assert [peak.peak_quality_class for peak in quality_peaks] == [
        PeakQualityClass.MAJOR_SHARP,
        PeakQualityClass.MAJOR_BROAD,
        PeakQualityClass.MINOR_SHARP,
        PeakQualityClass.MINOR_BROAD,
        PeakQualityClass.SHOULDER_OR_OVERLAP,
        PeakQualityClass.LOW_SUPPORT,
    ]


def test_threshold_boundary_is_inclusive():
    rows = [
        detector_row("MAX", 100.0, 100.0, 100.0, 100.0, 2.0),
        detector_row("BOUNDARY", 120.0, 1.0, 1.0, 1.0, 6.0),
    ]
    parameters = PeakFamilyParameters(
        minimum_relative_prominence=0.01,
        minor_sharp_minimum_relative_prominence=0.01,
    )
    peaks = calculate_peak_metrics(
        FakeDetection(rows), source_id="S", measurement_id="M", rna_identity="R",
        parameters=parameters,
    )
    assert peaks[1].peak_quality_class is PeakQualityClass.MINOR_SHARP
    assert peaks[1].selected_as_major_peak is True


def test_major_selection_keeps_multiple_quality_supported_peaks(quality_peaks):
    selected = select_major_peaks(quality_peaks)
    assert [peak.peak_id for peak in selected] == ["P1", "P2", "P3"]
    assert len(selected) > 1
    assert "P5" not in {peak.peak_id for peak in selected}


def test_bounded_selection_is_deterministic_and_coverage_aware(quality_peaks):
    parameters = PeakFamilyParameters(maximum_major_peaks_per_profile=2, mass_range_coverage_bins=2)
    first = select_major_peaks(quality_peaks, parameters=parameters)
    second = select_major_peaks(quality_peaks, parameters=parameters)
    assert first == second
    assert len(first) == 2
    assert first[0].apex_mass < first[1].apex_mass


def test_delta_pairs_are_unique_positive_and_complete(quality_peaks):
    selected = tuple(quality_peaks[:3])
    pairs = generate_delta_pairs(selected)
    assert len(pairs) == 3
    assert len({pair.peak_pair_id for pair in pairs}) == 3
    assert all(pair.delta_mass_da > 0 for pair in pairs)
    assert {(pair.lower_peak_id, pair.higher_peak_id) for pair in pairs} == {
        ("P1", "P2"), ("P1", "P3"), ("P2", "P3")
    }
    assert all(pair.observed_centroid_delta_da == pair.delta_mass_da for pair in pairs)


def test_same_peak_id_pair_is_rejected(quality_peaks):
    duplicate = replace(quality_peaks[1], peak_id=quality_peaks[0].peak_id)
    with pytest.raises(ValueError, match="same peak"):
        generate_delta_pairs((quality_peaks[0], duplicate))


def reference_by_name(references, name):
    return next(item for item in references if item.reference_name == name)


def test_delta_reference_registry_derives_cca_terminal_and_separates_mass_definitions():
    refs = build_delta_reference_registry()
    c = reference_by_name(refs, "C_RESIDUE_ADDITION")
    a = reference_by_name(refs, "A_RESIDUE_ADDITION")
    phosphate = reference_by_name(refs, "5_PRIME_OH_TO_MONOPHOSPHATE")
    for item in (c, a, phosphate):
        assert item.delta_mass_definition is DeltaMassDefinition.AVERAGE_DELTA
        assert item.mass_definition_compatible is True
        assert item.eligible_for_family_edge is True
        assert item.comparison_role is DeltaComparisonRole.FAMILY_EDGE_REFERENCE
    assert reference_by_name(refs, "NONE_TO_C").reference_delta_da == c.reference_delta_da
    assert reference_by_name(refs, "C_TO_CC").reference_delta_da == c.reference_delta_da
    assert reference_by_name(refs, "CC_TO_CCA").reference_delta_da == a.reference_delta_da
    proton = reference_by_name(refs, "+PROTON_MASS_DA")
    assert proton.reference_delta_da == PROTON_MASS_DA
    assert proton.delta_mass_definition is DeltaMassDefinition.EXACT_ION_DELTA
    assert proton.mass_definition_compatible is False
    assert proton.eligible_for_family_edge is False

    average_water = reference_by_name(refs, "HYDRATION_H2O_AVERAGE")
    mono_water = reference_by_name(refs, "HYDRATION_H2O_MONOISOTOPIC")
    assert average_water.reference_delta_da == pytest.approx(18.01528)
    assert mono_water.reference_delta_da == pytest.approx(18.01056468403)
    assert average_water.delta_mass_definition is DeltaMassDefinition.AVERAGE_DELTA
    assert mono_water.delta_mass_definition is DeltaMassDefinition.MONOISOTOPIC_DELTA
    assert average_water.reference_delta_da != mono_water.reference_delta_da

    average_o = reference_by_name(refs, "OXIDATION_O_AVERAGE")
    mono_o = reference_by_name(refs, "OXIDATION_O_MONOISOTOPIC")
    assert average_o.reference_delta_da == 15.9994
    assert mono_o.reference_delta_da == pytest.approx(15.99491461957)
    assert average_o.mass_definition_compatible is True
    assert mono_o.mass_definition_compatible is False

    average_na = reference_by_name(refs, "NA_H_EXCHANGE_AVERAGE")
    exact_na = reference_by_name(refs, "NA_H_EXCHANGE_EXACT_ION")
    average_k = reference_by_name(refs, "K_H_EXCHANGE_AVERAGE")
    exact_k = reference_by_name(refs, "K_H_EXCHANGE_EXACT_ION")
    assert average_na.reference_delta_da != exact_na.reference_delta_da
    assert average_k.reference_delta_da != exact_k.reference_delta_da
    assert average_na.delta_mass_definition is DeltaMassDefinition.AVERAGE_DELTA
    assert exact_na.delta_mass_definition is DeltaMassDefinition.EXACT_ION_DELTA
    assert average_k.mass_definition_compatible is True
    assert exact_k.mass_definition_compatible is False


def test_known_modification_references_use_existing_curated_source_only():
    modifications = load_modifications(ROOT / "data" / "modifications.yaml")
    refs = build_delta_reference_registry(modifications[:2])
    known = [
        item for item in refs
        if item.reference_category is DeltaReferenceCategory.KNOWN_RNA_MODIFICATION_DIAGNOSTIC_ONLY
    ]
    assert [item.reference_name for item in known] == [modifications[0].id, modifications[1].id]
    assert [item.reference_delta_da for item in known] == [
        abs(modifications[0].mass_shift_from_unmodified),
        abs(modifications[1].mass_shift_from_unmodified),
    ]
    assert all(item.elemental_difference == "NOT_ASSIGNED" for item in known)
    assert all(item.delta_mass_definition is DeltaMassDefinition.MONOISOTOPIC_DELTA for item in known)
    assert all(item.mass_definition_compatible is False for item in known)
    assert all(item.eligible_for_family_edge is False for item in known)
    assert all(
        item.comparison_role is DeltaComparisonRole.MASS_DEFINITION_MISMATCH_DIAGNOSTIC_ONLY
        for item in known
    )


def custom_reference(delta, *, compatible=True):
    return DeltaReference(
        "REF", "REFERENCE", DeltaReferenceCategory.CCA_OR_TERMINAL_STATE,
        delta, delta, "TEST", "TEST", True,
        DeltaMassDefinition.AVERAGE_DELTA if compatible else DeltaMassDefinition.MONOISOTOPIC_DELTA,
        compatible,
        compatible,
        (
            DeltaComparisonRole.FAMILY_EDGE_REFERENCE
            if compatible else DeltaComparisonRole.MASS_DEFINITION_MISMATCH_DIAGNOSTIC_ONLY
        ),
    )


@pytest.mark.parametrize(
    ("observed", "expected"),
    [(10.4, DeltaMatchClass.STRICT), (10.75, DeltaMatchClass.EXPLORATORY), (12.0, DeltaMatchClass.NO_MATCH)],
)
def test_delta_matching_classes_and_apex_primary(quality_peaks, observed, expected):
    lower = replace(quality_peaks[0], apex_mass=100.0, centroid_mass=100.2, peak_id="LOW")
    higher = replace(quality_peaks[1], apex_mass=100.0 + observed, centroid_mass=110.3, peak_id="HIGH")
    pair = generate_delta_pairs((lower, higher))[0]
    matches = match_delta_pairs((pair,), (custom_reference(10.0),))
    assert matches[0].delta_match_class is expected
    assert matches[0].observed_apex_delta_da == pytest.approx(observed)
    assert matches[0].observed_centroid_delta_da == pytest.approx(10.1)
    if expected is not DeltaMatchClass.NO_MATCH:
        assert matches[0].apex_delta_error_da == pytest.approx(observed - 10.0)
        assert matches[0].centroid_delta_error_da == pytest.approx(0.1)
    assert matches[0].modification_assigned is False
    assert matches[0].position_assigned is False
    assert matches[0].structure_assigned is False


def test_multiple_reference_matches_are_retained(quality_peaks):
    lower = replace(quality_peaks[0], apex_mass=100.0, peak_id="LOW")
    higher = replace(quality_peaks[1], apex_mass=110.0, peak_id="HIGH")
    pair = generate_delta_pairs((lower, higher))[0]
    refs = (custom_reference(10.0), replace(custom_reference(10.0), reference_id="REF2", reference_name="REFERENCE2"))
    matches = match_delta_pairs((pair,), refs)
    assert len(matches) == 2
    assert {item.reference_id for item in matches} == {"REF", "REF2"}


def test_connected_components_isolated_family_and_component_merge(quality_peaks):
    peaks = tuple(replace(peak, peak_id=f"X{i}", apex_mass=100.0 + i * 10) for i, peak in enumerate(quality_peaks[:3]))
    pairs = generate_delta_pairs(peaks)
    refs = (custom_reference(10.0),)
    matches = match_delta_pairs(pairs, refs)
    families = build_peak_families(peaks, pairs, matches, (), rna_identity="TRNA_LEU_UAA")
    assert len(families) == 1
    family = families[0]
    assert family.member_count == 3
    assert family.mass_span_da == 20.0
    assert family.highest_intensity_peak_id == "X0"
    assert family.highest_quality_peak_id == "X0"
    assert family.peak_family_id.startswith("FAMILY__")
    assert family == build_peak_families(peaks, pairs, matches, (), rna_identity="TRNA_LEU_UAA")[0]


def test_unrelated_peak_is_separate_isolated_family(quality_peaks):
    peaks = (
        replace(quality_peaks[0], peak_id="A", apex_mass=100.0),
        replace(quality_peaks[1], peak_id="B", apex_mass=110.0),
        replace(quality_peaks[2], peak_id="C", apex_mass=150.0),
    )
    pairs = generate_delta_pairs(peaks)
    matches = match_delta_pairs(pairs, (custom_reference(10.0),))
    families = build_peak_families(peaks, pairs, matches, (), rna_identity="TRNA_LEU_UAA")
    assert sorted(family.member_count for family in families) == [1, 2]


def test_candidate_baseline_relations_keep_four_modes_separate(quality_peaks):
    manifest = load_sciex_sample_manifest(MANIFEST_PATH)
    candidates = generate_candidates_for_measurement(manifest, "GLU_UUC_WT_FULL")
    relations = connect_candidate_baselines((quality_peaks[0],), candidates)
    assert len(relations) == 4
    assert {item.nearest_reference_mode.value for item in relations} == {
        "AVERAGE_NEUTRAL_M", "AVERAGE_M_PLUS_H", "AVERAGE_M_MINUS_H", "MONOISOTOPIC_NEUTRAL_M"
    }
    assert all(item.observed_output_species == "UNKNOWN" for item in relations)
    assert all(item.observed_output_species_confirmed is False for item in relations)


def test_leu_and_glu_family_hypothesis_safeguards(quality_detection):
    leu = analyze_sciex_intact_peak_families(
        quality_detection, source_id="S", measurement_id="M",
        rna_identity="TRNA_LEU_UAA",
    )
    glu = analyze_sciex_intact_peak_families(
        quality_detection, source_id="S", measurement_id="M",
        rna_identity="TRNA_GLU_UUC",
    )
    assert all(family.co_captured_rna_possible for family in leu.families)
    assert all("CO_CAPTURED_RNA_POSSIBLE" in family.hypotheses for family in leu.families)
    assert all(
        family.primary_biological_hypothesis == "TARGET_TRNA_MODIFICATION_ISOFORM"
        for family in glu.families
    )
    assert all(family.co_captured_rna_excluded is False for family in glu.families)
    assert all(family.hypothesis_confirmed is False for family in glu.families)


@pytest.mark.parametrize("collection_name", ["peaks", "delta_pairs", "delta_matches", "candidate_relations", "families"])
def test_all_analysis_records_keep_false_certainty_and_formal_flags(quality_detection, collection_name):
    manifest = load_sciex_sample_manifest(MANIFEST_PATH)
    candidates = generate_candidates_for_measurement(manifest, "LEU_UAA_WT_FULL")
    result = analyze_sciex_intact_peak_families(
        quality_detection, source_id="S", measurement_id="LEU_UAA_WT_FULL",
        rna_identity="TRNA_LEU_UAA", candidates=candidates,
    )
    records = getattr(result, collection_name)
    assert records
    for item in records:
        assert item.shadow_analysis_only is True
        assert item.mass_evidence_only is True
        assert item.rna_identity_confirmed is False
        assert item.target_rna_identity_confirmed_by_mass is False
        assert item.co_captured_rna_excluded is False
        assert item.modification_assigned is False
        assert item.modification_composition_assigned is False
        assert item.position_assigned is False
        assert item.structure_assigned is False
        assert item.cca_state_confirmed is False
        assert item.terminal_state_confirmed is False
        assert item.biological_cause_assigned is False
        assert item.rnase_t_assigned is False
        assert item.applied_to_formal_score is False
        assert item.applied_to_ranking is False
        assert item.applied_to_candidate_filtering is False
        assert item.applied_to_final_consensus is False


def test_output_species_remains_unknown_unassigned(quality_detection):
    result = analyze_sciex_intact_peak_families(
        quality_detection, source_id="S", measurement_id="M", rna_identity="R"
    )
    assert result.observed_output_species == "UNKNOWN"
    assert result.observed_output_species_confirmed is False
    assert result.output_species_assigned is False


def test_result_records_are_immutable(quality_detection):
    result = analyze_sciex_intact_peak_families(
        quality_detection, source_id="S", measurement_id="M", rna_identity="R"
    )
    with pytest.raises(FrozenInstanceError):
        result.peaks[0].apex_mass = 1.0


def test_reference_spreadsheet_sequences_match_manifest_exactly():
    manifest = load_sciex_sample_manifest(MANIFEST_PATH)
    expected = {
        "tRNA-Glu-TTC-1-1": get_rna_identity(manifest, "TRNA_GLU_UUC").sequence,
        "tRNA-Glu-TTC-1-2": get_rna_identity(manifest, "TRNA_GLU_UUC").sequence,
        "tRNA-Leu-TAA-1-1": get_rna_identity(manifest, "TRNA_LEU_UAA").sequence,
        "tRNA-Leu-TAG-1-1": get_rna_identity(manifest, "TRNA_LEU_UAG").sequence,
    }
    workbook = load_workbook(REFERENCE_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        observed = {}
        for row in worksheet.iter_rows(values_only=True):
            if len(row) > 4 and row[2] in expected:
                observed[str(row[2])] = str(row[4])
    finally:
        workbook.close()
    assert observed == expected


def test_reference_spreadsheet_is_not_mutated_by_validation():
    before = REFERENCE_PATH.stat().st_mtime_ns, REFERENCE_PATH.stat().st_size
    test_reference_spreadsheet_sequences_match_manifest_exactly()
    after = REFERENCE_PATH.stat().st_mtime_ns, REFERENCE_PATH.stat().st_size
    assert after == before


def test_full_profile_observed_delta_definition_is_average(quality_detection):
    result = analyze_sciex_intact_peak_families(
        quality_detection, source_id="S", measurement_id="M", rna_identity="R"
    )
    assert result.observed_delta_mass_definition is DeltaMassDefinition.AVERAGE_DELTA
    assert all(
        pair.observed_delta_mass_definition is DeltaMassDefinition.AVERAGE_DELTA
        for pair in result.delta_pairs
    )
    assert all(
        match.observed_delta_mass_definition is DeltaMassDefinition.AVERAGE_DELTA
        for match in result.delta_matches
    )


def test_incompatible_diagnostic_match_is_retained_but_cannot_form_family_edge(quality_peaks):
    peaks = (
        replace(quality_peaks[0], peak_id="A", apex_mass=100.0),
        replace(quality_peaks[1], peak_id="B", apex_mass=110.0),
    )
    pairs = generate_delta_pairs(peaks)
    matches = match_delta_pairs(pairs, (custom_reference(10.0, compatible=False),))
    assert len(matches) == 1
    assert matches[0].delta_match_class is DeltaMatchClass.STRICT
    assert matches[0].mass_definition_compatible is False
    assert matches[0].eligible_for_family_edge is False
    assert matches[0].comparison_role is DeltaComparisonRole.MASS_DEFINITION_MISMATCH_DIAGNOSTIC_ONLY
    families = build_peak_families(peaks, pairs, matches, (), rna_identity="TRNA_LEU_UAA")
    assert len(families) == 2
    assert all(family.member_count == 1 for family in families)


def test_compatible_average_match_forms_family_edge_and_controls_component_count(quality_peaks):
    peaks = (
        replace(quality_peaks[0], peak_id="A", apex_mass=100.0),
        replace(quality_peaks[1], peak_id="B", apex_mass=110.0),
        replace(quality_peaks[2], peak_id="C", apex_mass=120.0),
    )
    pairs = generate_delta_pairs(peaks)
    compatible = match_delta_pairs(pairs, (custom_reference(10.0),))
    incompatible = match_delta_pairs(pairs, (custom_reference(10.0, compatible=False),))
    compatible_families = build_peak_families(
        peaks, pairs, compatible, (), rna_identity="TRNA_LEU_UAA"
    )
    incompatible_families = build_peak_families(
        peaks, pairs, incompatible, (), rna_identity="TRNA_LEU_UAA"
    )
    assert len(compatible_families) == 1
    assert compatible_families[0].member_count == 3
    assert compatible_families[0].strict_edge_count == 2
    assert len(incompatible_families) == 3


def test_unknown_modification_definition_is_diagnostic_only():
    class UnknownModification:
        id = "UNKNOWN_MOD"
        symbol = "UNKNOWN_MOD"
        mass_shift_from_unmodified = 10.0
        raw = {"mass_basis": "unspecified"}

    ref = reference_by_name(build_delta_reference_registry((UnknownModification(),)), "UNKNOWN_MOD")
    assert ref.delta_mass_definition is DeltaMassDefinition.UNKNOWN
    assert ref.mass_definition_compatible is False
    assert ref.eligible_for_family_edge is False
    assert ref.comparison_role is DeltaComparisonRole.UNKNOWN_MASS_DEFINITION_DIAGNOSTIC_ONLY


def test_curated_composition_creates_independent_average_modification_reference():
    class CompositionModification:
        id = "COMPOSITION_MOD"
        symbol = "COMPOSITION_MOD"
        mass_shift_from_unmodified = 14.01565
        raw = {
            "modified_nucleoside_mass_mono": 1.0,
            "elemental_composition_delta": {"C": 1, "H": 2},
        }

    refs = build_delta_reference_registry((CompositionModification(),))
    mono = reference_by_name(refs, "COMPOSITION_MOD")
    average = reference_by_name(refs, "COMPOSITION_MOD__COMPOSITION_AVERAGE")
    assert mono.delta_mass_definition is DeltaMassDefinition.MONOISOTOPIC_DELTA
    assert mono.mass_definition_compatible is False
    assert average.delta_mass_definition is DeltaMassDefinition.AVERAGE_DELTA
    assert average.reference_delta_da == pytest.approx(14.02658)
    assert average.mass_definition_compatible is True
    assert average.eligible_for_family_edge is True


def test_all_current_curated_modification_shifts_are_mono_and_incompatible():
    modifications = load_modifications(ROOT / "data" / "modifications.yaml")
    refs = build_delta_reference_registry(modifications)
    known = [
        item for item in refs
        if item.reference_category is DeltaReferenceCategory.KNOWN_RNA_MODIFICATION_DIAGNOSTIC_ONLY
    ]
    expected = [
        item for item in modifications
        if item.mass_shift_from_unmodified == item.mass_shift_from_unmodified
        and item.mass_shift_from_unmodified != 0
    ]
    assert len(known) == len(expected)
    assert all(item.delta_mass_definition is DeltaMassDefinition.MONOISOTOPIC_DELTA for item in known)
    assert all(item.mass_definition_compatible is False for item in known)
    assert all(item.eligible_for_family_edge is False for item in known)
    assert not any("COMPOSITION_AVERAGE" in item.reference_name for item in known)
