"""Run cross-layer SCIEX reconciliation from four production bundles only.

The runner requires explicit FULL, T1, P1/AP MS1, and P1/AP MS2 bundles and
never discovers or reparses raw mzML.  RNA identity and sample identity remain
separate: different samples do not create independent support, while P1/AP MS1
and MS2 from one source are explicitly non-independent.  FULL results currently
lack node-level raw-path provenance, so the bundle provenance is retained and a
warning is emitted without rewriting scientific nodes.  All outputs remain
shadow-only, never propagate into formal results, and refuse to overwrite files.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass
from hashlib import sha256
import json
import os
from pathlib import Path
import shutil
from tempfile import NamedTemporaryFile
from typing import Any, Iterable, Mapping

from rna_masshunter.sciex_layer_evidence_bundle import (
    canonical_json_bytes,
    compare_layer_evidence_provenance,
    load_layer_evidence_bundle,
    restore_layer_evidence_result,
    validate_layer_evidence_bundle,
)
from rna_masshunter.sciex_rna_cross_layer_evidence_reconciliation import (
    OPTIONAL_RESULT_KEY,
    CrossLayerEvidenceAuditResult,
    audit_optional_result,
    audit_rna_cross_layer_evidence_reconciliation,
)

REQUIRED_LAYERS = ("FULL", "T1", "P1AP_MS1", "P1AP_MS2")
AGGREGATE_SCHEMA_VERSION = "rna-masshunter-cross-layer-bundle-runner-v1"
FULL_PROVENANCE_WARNING = "FULL_RESULT_LACKS_NODE_LEVEL_SOURCE_ID"
XL_SHEET_NAMES = (
    "XL_Nodes", "XL_Edges", "XL_Hypotheses", "XL_Layer_Summary",
    "XL_Consensus", "XL_Next_Evidence",
)
_FALSE_SAFEGUARDS = (
    "formal_propagation", "chemical_identity_assigned", "modification_assigned",
    "exact_candidate_identity_confirmed", "exact_isomer_identity_confirmed",
    "exact_nucleotide_localization", "exact_atom_localization",
    "reaction_order_assigned", "applied_to_formal_score", "applied_to_ranking",
    "applied_to_candidate_filtering", "applied_to_final_consensus",
)
_TRUE_SAFEGUARDS = ("shadow_analysis_only", "cross_layer_reconciliation_only")


class CrossLayerBundleRunnerError(ValueError):
    """Raised when a bundle set or runner output violates its contract."""


@dataclass(frozen=True)
class CrossLayerBundleCompatibilityReport:
    compatible: bool
    warnings: tuple[str, ...]
    blocking_errors: tuple[str, ...]
    shared_source_relationships: tuple[dict[str, Any], ...]
    independence_groups: dict[str, str]
    sample_relationships: tuple[dict[str, Any], ...]
    bundle_provenance: dict[str, dict[str, Any]]
    node_level_provenance: dict[str, str]
    bundle_level_provenance_verified: bool


@dataclass(frozen=True)
class CrossLayerBundleRunResult:
    compatibility_report: CrossLayerBundleCompatibilityReport
    input_bundle_paths: dict[str, str | None]
    input_bundle_sha256: dict[str, str]
    restored_result_types: dict[str, str]
    cross_layer_optional_results: dict[str, CrossLayerEvidenceAuditResult]
    aggregate_counts: dict[str, int]
    consensus_status: str
    confidence: str
    safeguard_summary: dict[str, Any]
    output_json_path: str | None
    output_excel_path: str | None
    warnings: tuple[str, ...]
    provenance: dict[str, dict[str, Any]]


@dataclass(frozen=True)
class _LoadedBundle:
    layer: str
    path: Path | None
    bundle: dict[str, Any]
    restored: Any
    file_sha256: str


def _file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _qualified_type(value: Any) -> str:
    cls = type(value)
    return f"{cls.__module__}.{cls.__qualname__}"


def _normalize_inputs(
    bundle_inputs: Mapping[str, str | Path | Mapping[str, Any]]
    | Iterable[str | Path | Mapping[str, Any]],
) -> list[tuple[str | None, str | Path | Mapping[str, Any]]]:
    if isinstance(bundle_inputs, Mapping):
        if "schema_version" in bundle_inputs and "layer" in bundle_inputs:
            return [(None, bundle_inputs)]
        return [(str(layer), value) for layer, value in bundle_inputs.items()]
    if isinstance(bundle_inputs, (str, Path)):
        return [(None, bundle_inputs)]
    try:
        return [(None, value) for value in bundle_inputs]
    except TypeError as exc:
        raise CrossLayerBundleRunnerError("bundle inputs must be an explicit mapping or iterable") from exc


def _load_one(
    declared_layer: str | None,
    value: str | Path | Mapping[str, Any],
    source_paths: Mapping[str, str | Path] | None,
) -> _LoadedBundle:
    path: Path | None = None
    source_path: str | Path | None = None
    if isinstance(value, (str, Path)):
        path = Path(value)
        if not path.is_file():
            raise CrossLayerBundleRunnerError(f"bundle file is unavailable: {path}")
        layer_hint = declared_layer
        if layer_hint is None and source_paths:
            try:
                layer_hint = str(json.loads(path.read_text(encoding="utf-8")).get("layer") or "")
            except (OSError, UnicodeError, json.JSONDecodeError, AttributeError):
                layer_hint = None
        source_path = (source_paths or {}).get(layer_hint) if layer_hint else None
        raw = load_layer_evidence_bundle(path, source_path=source_path)
        layer = str(raw.get("layer") or "")
        validate_layer_evidence_bundle(raw, source_path=source_path)
        restored = restore_layer_evidence_result(raw, source_path=source_path)
        file_sha = _file_sha256(path)
    elif isinstance(value, Mapping):
        raw = dict(value)
        layer = str(raw.get("layer") or "")
        source_path = (source_paths or {}).get(layer)
        validate_layer_evidence_bundle(raw, source_path=source_path)
        restored = restore_layer_evidence_result(raw, source_path=source_path)
        file_sha = sha256(canonical_json_bytes(raw)).hexdigest()
    else:
        raise CrossLayerBundleRunnerError(
            f"unsupported bundle input type: {type(value).__name__}"
        )
    if declared_layer is not None and declared_layer != layer:
        raise CrossLayerBundleRunnerError(
            f"declared layer {declared_layer!r} does not match bundle layer {layer!r}"
        )
    return _LoadedBundle(layer, path, raw, restored, file_sha)


def _build_provenance(bundle: Mapping[str, Any]) -> dict[str, Any]:
    source = bundle["source"]
    experiment = bundle["experiment"]
    return {
        "source_path": source["path"],
        "source_basename": source["basename"],
        "source_sha256": source["sha256"],
        "run_id": source["run_id"],
        "sample_id": source["sample_id"],
        "biological_sample_id": source["biological_sample_id"],
        "digest_type": experiment["digest_type"],
        "condition_name": experiment["condition_name"],
        "shared_source_group": experiment["shared_source_group"],
        "independence_group": experiment["independence_group"],
        "canonical_payload_sha256": bundle["canonical_payload_sha256"],
    }


def _sample_relationships(loaded: Mapping[str, _LoadedBundle]) -> tuple[dict[str, Any], ...]:
    rows = []
    for index, left_layer in enumerate(REQUIRED_LAYERS):
        left = loaded[left_layer].bundle
        for right_layer in REQUIRED_LAYERS[index + 1:]:
            right = loaded[right_layer].bundle
            comparison = compare_layer_evidence_provenance(left, right)
            rows.append({
                "left_layer": left_layer,
                "right_layer": right_layer,
                "same_sample": comparison["same_sample"],
                "different_sample": comparison["different_sample"],
                "same_raw_source": comparison["same_raw_source"],
                "same_independence_group": comparison["same_independence_group"],
                "independent_support": comparison["independent_support"],
            })
    return tuple(rows)


def validate_cross_layer_bundle_set(
    bundle_inputs: Mapping[str, str | Path | Mapping[str, Any]]
    | Iterable[str | Path | Mapping[str, Any]],
    *, source_paths: Mapping[str, str | Path] | None = None,
) -> tuple[dict[str, _LoadedBundle], CrossLayerBundleCompatibilityReport]:
    """Load, validate, restore, and compare an explicit four-bundle set."""
    normalized = _normalize_inputs(bundle_inputs)
    loaded: dict[str, _LoadedBundle] = {}
    for declared_layer, value in normalized:
        item = _load_one(declared_layer, value, source_paths)
        if item.layer in loaded:
            raise CrossLayerBundleRunnerError(f"duplicate layer bundle: {item.layer}")
        if item.layer not in REQUIRED_LAYERS:
            raise CrossLayerBundleRunnerError(f"unsupported layer bundle: {item.layer}")
        loaded[item.layer] = item
    missing = [layer for layer in REQUIRED_LAYERS if layer not in loaded]
    if missing:
        raise CrossLayerBundleRunnerError(f"missing required layer bundles: {missing}")

    warnings: list[str] = []
    errors: list[str] = []
    first_rna = loaded["FULL"].bundle["rna"]
    for layer in REQUIRED_LAYERS[1:]:
        rna = loaded[layer].bundle["rna"]
        for field in ("name", "sequence", "anticodon", "wobble_position", "organism_group", "species"):
            if rna[field] != first_rna[field]:
                errors.append(f"RNA_{field.upper()}_MISMATCH:{layer}")
        if loaded[layer].bundle["experiment"]["condition_name"] != loaded["FULL"].bundle["experiment"]["condition_name"]:
            errors.append(f"UNSUPPORTED_CONDITION_RELATIONSHIP:FULL:{layer}")

    t1_basename = loaded["T1"].bundle["source"]["basename"]
    if t1_basename == "04 new T1.mzML":
        errors.append("T1_FORBIDDEN_SOURCE_SUBSTITUTION:04_NEW_FOR_05_OLD")

    ms1 = loaded["P1AP_MS1"].bundle
    ms2 = loaded["P1AP_MS2"].bundle
    p1_fields = (
        ("source", "sha256"), ("source", "basename"), ("source", "run_id"),
        ("source", "sample_id"), ("source", "biological_sample_id"),
        ("experiment", "shared_source_group"), ("experiment", "independence_group"),
    )
    for section, field in p1_fields:
        if ms1[section][field] != ms2[section][field]:
            errors.append(f"P1AP_{section.upper()}_{field.upper()}_MISMATCH")
    p1_comparison = compare_layer_evidence_provenance(ms1, ms2)
    shared = ({
        "left_layer": "P1AP_MS1",
        "right_layer": "P1AP_MS2",
        "same_raw_source": p1_comparison["same_raw_source"],
        "shared_source_relationship": p1_comparison["shared_source_relationship"],
        "same_independence_group": p1_comparison["same_independence_group"],
        "independent_support": p1_comparison["independent_support"],
    },)
    if not p1_comparison["p1ap_ms1_ms2_compatible"]:
        errors.append("P1AP_SHARED_SOURCE_RELATIONSHIP_INCOMPATIBLE")

    full_result = loaded["FULL"].restored
    full_node_source = getattr(full_result, "input_path", None) or getattr(full_result, "source_file_id", None)
    full_node_status = "AVAILABLE" if full_node_source else "UNAVAILABLE_IN_RESULT"
    if not full_node_source:
        warnings.append(FULL_PROVENANCE_WARNING)
    provenance = {layer: _build_provenance(loaded[layer].bundle) for layer in REQUIRED_LAYERS}
    report = CrossLayerBundleCompatibilityReport(
        compatible=not errors,
        warnings=tuple(sorted(set(warnings))),
        blocking_errors=tuple(sorted(set(errors))),
        shared_source_relationships=shared,
        independence_groups={
            layer: loaded[layer].bundle["experiment"]["independence_group"]
            for layer in REQUIRED_LAYERS
        },
        sample_relationships=_sample_relationships(loaded),
        bundle_provenance=provenance,
        node_level_provenance={"FULL": full_node_status, "T1": "AVAILABLE", "P1AP_MS1": "AVAILABLE", "P1AP_MS2": "AVAILABLE"},
        bundle_level_provenance_verified=True,
    )
    if errors:
        raise CrossLayerBundleRunnerError(
            "incompatible bundle set: " + "; ".join(report.blocking_errors)
        )
    return loaded, report


def build_cross_layer_bundle_context(loaded: Mapping[str, _LoadedBundle]) -> dict[str, Any]:
    """Build reconciliation context from restored results without reading raw data."""
    candidates = []
    for layer in ("T1", "P1AP_MS1"):
        result = loaded[layer].restored
        summary = getattr(result, "run_summary", None)
        value = getattr(summary, "rna_identity", None)
        if value:
            candidates.append(str(value))
    runtime_rna = candidates[0] if candidates else str(loaded["FULL"].bundle["rna"]["name"])
    return {
        "RNA_Identity": runtime_rna,
        "Context_Source": "PRODUCTION_LAYER_EVIDENCE_BUNDLES",
        "Context_Confidence": "BUNDLE_PROVENANCE_VERIFIED",
    }


def _validate_cross_layer_safeguards(records: Mapping[str, list[dict[str, Any]]]) -> dict[str, Any]:
    row_count = 0
    for group, rows in records.items():
        if not rows:
            raise CrossLayerBundleRunnerError(f"empty cross-layer record group: {group}")
        for index, row in enumerate(rows):
            row_count += 1
            missing = [name for name in _FALSE_SAFEGUARDS + _TRUE_SAFEGUARDS if name not in row]
            if missing:
                raise CrossLayerBundleRunnerError(
                    f"safeguard fields missing from {group}[{index}]: {missing}"
                )
            bad_false = [name for name in _FALSE_SAFEGUARDS if row[name] is not False]
            bad_true = [name for name in _TRUE_SAFEGUARDS if row[name] is not True]
            if bad_false or bad_true:
                raise CrossLayerBundleRunnerError(
                    f"unsafe cross-layer safeguards in {group}[{index}]: false={bad_false}, true={bad_true}"
                )
    return {
        "verified": True,
        "record_count": row_count,
        **{name: False for name in _FALSE_SAFEGUARDS},
        **{name: True for name in _TRUE_SAFEGUARDS},
    }


def _aggregate_counts(result: CrossLayerEvidenceAuditResult) -> dict[str, int]:
    return {
        "nodes": len(result.nodes),
        "edges": len(result.edges),
        "hypotheses": len(result.hypotheses),
        "layer_summaries": len(result.layer_summaries),
        "consensus": 1,
        "next_evidence": len(result.next_evidence),
        "independence_groups": result.consensus.independence_group_count,
    }


def _reject_unsafe_payload(value: Any, path: str = "aggregate") -> None:
    forbidden = {"raw_peaks", "raw_peak_arrays", "spectrum_arrays", "raw_spectrum_arrays", "mz_array", "intensity_array", "binary", "base64", "binary_payload", "binary_data_array"}
    if isinstance(value, Mapping):
        for key, item in value.items():
            if str(key).lower() in forbidden:
                raise CrossLayerBundleRunnerError(f"forbidden raw/binary aggregate field: {path}.{key}")
            _reject_unsafe_payload(item, f"{path}.{key}")
    elif isinstance(value, (list, tuple)):
        for index, item in enumerate(value):
            _reject_unsafe_payload(item, f"{path}[{index}]")
    elif isinstance(value, float) and not (value == value and abs(value) != float("inf")):
        raise CrossLayerBundleRunnerError(f"non-finite aggregate value: {path}")


def _aggregate_payload(
    loaded: Mapping[str, _LoadedBundle],
    report: CrossLayerBundleCompatibilityReport,
    result: CrossLayerEvidenceAuditResult,
    safeguards: Mapping[str, Any],
) -> dict[str, Any]:
    records = audit_optional_result(result)
    payload = {
        "schema_version": AGGREGATE_SCHEMA_VERSION,
        "input_bundles": {
            layer: {
                "path": str(loaded[layer].path) if loaded[layer].path else None,
                "file_sha256": loaded[layer].file_sha256,
                "canonical_payload_sha256": loaded[layer].bundle["canonical_payload_sha256"],
                "source_provenance": _build_provenance(loaded[layer].bundle),
            }
            for layer in REQUIRED_LAYERS
        },
        "compatibility_report": asdict(report),
        "aggregate_counts": _aggregate_counts(result),
        "consensus": records["consensus_records"][0],
        "safeguards": dict(safeguards),
        "warnings": list(report.warnings),
        "records": records,
    }
    _reject_unsafe_payload(payload)
    return payload


def _output_path(path: str | Path) -> Path:
    destination = Path(path)
    if destination.exists():
        raise CrossLayerBundleRunnerError(f"output file already exists: {destination}")
    parent = destination.parent
    if parent.exists() and not parent.is_dir():
        raise CrossLayerBundleRunnerError(f"malformed output directory: {parent}")
    try:
        parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise CrossLayerBundleRunnerError(f"malformed output directory: {parent}") from exc
    return destination


def write_cross_layer_aggregate_json(
    payload: Mapping[str, Any], output_path: str | Path,
) -> Path:
    """Atomically write deterministic UTF-8 aggregate JSON without overwriting."""
    destination = _output_path(output_path)
    data = canonical_json_bytes(payload) + b"\n"
    temporary: Path | None = None
    try:
        with NamedTemporaryFile("wb", dir=destination.parent, prefix=f".{destination.name}.", suffix=".tmp", delete=False) as handle:
            temporary = Path(handle.name)
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, destination)
    except Exception:
        if temporary is not None and temporary.exists():
            temporary.unlink()
        raise
    return destination


def write_cross_layer_excel(
    result: CrossLayerEvidenceAuditResult,
    output_path: str | Path,
    *, base_workbook_path: str | Path | None = None,
) -> Path:
    """Atomically create an XL-only workbook or append XL sheets to a copied workbook."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from rna_masshunter.excel_report import _sciex_cross_layer_excel_sheets

    destination = _output_path(output_path)
    frames = _sciex_cross_layer_excel_sheets(result)
    if tuple(frames) != XL_SHEET_NAMES:
        raise CrossLayerBundleRunnerError(f"unexpected XL sheet set: {tuple(frames)}")
    temporary: Path | None = None
    try:
        with NamedTemporaryFile("wb", dir=destination.parent, prefix=f".{destination.name}.", suffix=".tmp.xlsx", delete=False) as handle:
            temporary = Path(handle.name)
        if base_workbook_path is None:
            workbook = Workbook()
            workbook.remove(workbook.active)
        else:
            base = Path(base_workbook_path)
            if not base.is_file():
                raise CrossLayerBundleRunnerError(f"base workbook is unavailable: {base}")
            shutil.copy2(base, temporary)
            workbook = load_workbook(temporary)
        try:
            for name, frame in frames.items():
                if name in workbook.sheetnames:
                    raise CrossLayerBundleRunnerError(f"XL sheet already exists in workbook: {name}")
                worksheet = workbook.create_sheet(name)
                for row in dataframe_to_rows(frame, index=False, header=True):
                    worksheet.append(row)
            if any(len(name) > 31 for name in workbook.sheetnames):
                raise CrossLayerBundleRunnerError("Excel sheet name exceeds 31 characters")
            workbook.save(temporary)
        finally:
            workbook.close()
        os.replace(temporary, destination)
    except Exception:
        if temporary is not None and temporary.exists():
            temporary.unlink()
        raise
    return destination


def run_cross_layer_from_bundles(
    bundle_inputs: Mapping[str, str | Path | Mapping[str, Any]]
    | Iterable[str | Path | Mapping[str, Any]],
    *,
    source_paths: Mapping[str, str | Path] | None = None,
    output_json_path: str | Path | None = None,
    output_excel_path: str | Path | None = None,
    base_workbook_path: str | Path | None = None,
) -> CrossLayerBundleRunResult:
    """Validate four bundles, restore results, reconcile, and optionally write outputs."""
    if base_workbook_path is not None and output_excel_path is None:
        raise CrossLayerBundleRunnerError("base_workbook_path requires output_excel_path")
    loaded, report = validate_cross_layer_bundle_set(bundle_inputs, source_paths=source_paths)
    context = build_cross_layer_bundle_context(loaded)
    result = audit_rna_cross_layer_evidence_reconciliation(
        full_length_result=loaded["FULL"].restored,
        t1_result=loaded["T1"].restored,
        p1ap_ms1_result=loaded["P1AP_MS1"].restored,
        p1ap_ms2_result=loaded["P1AP_MS2"].restored,
        runtime_context=context,
    )
    records = audit_optional_result(result)
    safeguards = _validate_cross_layer_safeguards(records)
    payload = _aggregate_payload(loaded, report, result, safeguards)

    json_destination = Path(output_json_path) if output_json_path is not None else None
    excel_destination = Path(output_excel_path) if output_excel_path is not None else None
    for destination in (json_destination, excel_destination):
        if destination is not None and destination.exists():
            raise CrossLayerBundleRunnerError(f"output file already exists: {destination}")
        if destination is not None and destination.parent.exists() and not destination.parent.is_dir():
            raise CrossLayerBundleRunnerError(f"malformed output directory: {destination.parent}")
    if json_destination is not None and excel_destination is not None and json_destination == excel_destination:
        raise CrossLayerBundleRunnerError("JSON and Excel output paths must differ")

    staged: list[tuple[Path, Path]] = []
    try:
        if json_destination is not None and excel_destination is not None:
            json_stage = json_destination.with_name(f".{json_destination.name}.runner-stage-{os.getpid()}")
            excel_stage = excel_destination.with_name(f".{excel_destination.name}.runner-stage-{os.getpid()}.xlsx")
            for stage in (json_stage, excel_stage):
                if stage.exists():
                    raise CrossLayerBundleRunnerError(f"staging output already exists: {stage}")
            write_cross_layer_aggregate_json(payload, json_stage)
            staged.append((json_stage, json_destination))
            write_cross_layer_excel(result, excel_stage, base_workbook_path=base_workbook_path)
            staged.append((excel_stage, excel_destination))
            for stage, destination in staged:
                destination.parent.mkdir(parents=True, exist_ok=True)
                os.replace(stage, destination)
            staged.clear()
        else:
            if json_destination is not None:
                write_cross_layer_aggregate_json(payload, json_destination)
            if excel_destination is not None:
                write_cross_layer_excel(
                    result, excel_destination, base_workbook_path=base_workbook_path,
                )
    except Exception:
        for stage, _ in staged:
            if stage.exists():
                stage.unlink()
        raise
    return CrossLayerBundleRunResult(
        compatibility_report=report,
        input_bundle_paths={
            layer: str(loaded[layer].path) if loaded[layer].path else None
            for layer in REQUIRED_LAYERS
        },
        input_bundle_sha256={layer: loaded[layer].file_sha256 for layer in REQUIRED_LAYERS},
        restored_result_types={layer: _qualified_type(loaded[layer].restored) for layer in REQUIRED_LAYERS},
        cross_layer_optional_results={OPTIONAL_RESULT_KEY: result},
        aggregate_counts=_aggregate_counts(result),
        consensus_status=result.consensus.cross_layer_evidence_status,
        confidence=result.consensus.cross_layer_confidence,
        safeguard_summary=safeguards,
        output_json_path=str(json_destination) if json_destination else None,
        output_excel_path=str(excel_destination) if excel_destination else None,
        warnings=report.warnings,
        provenance=report.bundle_provenance,
    )
