"""Explicit, read-only SCIEX reconstructed-profile registry and shadow routing.

Tracked metadata never contains machine-local absolute paths. Runtime paths are
provided by callers, and only neutral reconstructed full-length profiles can be
routed to intact candidate generation or mass-only shadow comparison.
"""
from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from hashlib import sha256
from math import isfinite
from pathlib import Path
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook

from rna_masshunter.intact_rna_average_mass import (
    ComparisonReferenceRole,
    TheoreticalMassDefinition,
)
from rna_masshunter.intact_rna_candidate_generation import (
    IntactRnaTheoreticalCandidate,
    generate_candidates_for_measurement,
)
from rna_masshunter.sciex_intact_mass_comparison import compare_sciex_intact_masses
from rna_masshunter.sciex_intact_peak_detection import (
    SciexIntactPeakDetectionParameters,
    SciexIntactPeakDetectionResult,
    detect_sciex_intact_peaks,
)
from rna_masshunter.sciex_profile_parser import (
    MZ_PROFILE as PARSER_MZ_PROFILE,
    NEUTRAL_MASS_PROFILE as PARSER_NEUTRAL_MASS_PROFILE,
    parse_sciex_profile,
)
from rna_masshunter.sciex_sample_manifest import (
    AnalyteLevel,
    ExperimentType,
    SCIEXSampleManifest,
    get_measurement,
    get_rna_identity,
    get_sample,
)

PROFILE_REGISTRY_APPLIED_TO_FORMAL_SCORE = False
PROFILE_REGISTRY_APPLIED_TO_RANKING = False
PROFILE_REGISTRY_APPLIED_TO_CANDIDATE_FILTERING = False
PROFILE_REGISTRY_APPLIED_TO_FINAL_CONSENSUS = False


class ProfileRegistryValidationError(ValueError):
    pass


class ProfileSourceType(str, Enum):
    TEXT_NEUTRAL_PROFILE = "TEXT_NEUTRAL_PROFILE"
    EXCEL_SHEET_NEUTRAL_PROFILE = "EXCEL_SHEET_NEUTRAL_PROFILE"
    TEXT_MZ_PROFILE = "TEXT_MZ_PROFILE"
    MZML_DIGEST = "MZML_DIGEST"


class ReconstructedProfileType(str, Enum):
    NEUTRAL_MASS_PROFILE = "NEUTRAL_MASS_PROFILE"
    MZ_PROFILE = "MZ_PROFILE"
    DIGEST_MZML = "DIGEST_MZML"


class UnknownMassMetadata(str, Enum):
    UNKNOWN = "UNKNOWN"


class ObservedMassScale(str, Enum):
    AVERAGE = "AVERAGE"


class ObservedOutputSpecies(str, Enum):
    UNKNOWN = "UNKNOWN"


class ProfileSourceStatus(str, Enum):
    REGISTERED = "REGISTERED"


class ProfileSourceProvenanceStatus(str, Enum):
    USER_PROVIDED = "USER_PROVIDED"


class ProfileRoutingStatus(str, Enum):
    ROUTED = "ROUTED"
    SKIPPED = "SKIPPED"


class ShadowComparisonStatus(str, Enum):
    COMPLETED = "COMPLETED"
    SKIPPED = "SKIPPED"


@dataclass(frozen=True)
class ReconstructedProfileSource:
    profile_source_id: str
    measurement_id: str
    sample_id: str
    rna_identity_id: str
    source_type: ProfileSourceType
    source_file_name: str | None
    source_workbook_name: str | None
    source_sheet_name: str | None
    profile_type: ReconstructedProfileType
    mass_column: str | None
    intensity_column: str | None
    experiment_type: ExperimentType
    expected_analyte_level: AnalyteLevel
    observed_mass_type: UnknownMassMetadata
    mass_definition_compatibility: UnknownMassMetadata
    source_sha256: str | None
    source_size_bytes: int | None
    source_status: ProfileSourceStatus
    source_provenance_status: ProfileSourceProvenanceStatus
    eligible_for_intact_candidate_generation: bool
    eligible_for_intact_mass_comparison: bool
    eligibility_reason: str
    observed_mass_scale: ObservedMassScale | None = None
    observed_output_species: ObservedOutputSpecies | None = None
    observed_output_species_confirmed: bool | None = None


@dataclass(frozen=True)
class ReconstructedProfileRegistry:
    schema_version: int
    sources: tuple[ReconstructedProfileSource, ...]


@dataclass(frozen=True)
class ProfileManifestValidation:
    profile_source_id: str
    measurement_id: str
    sample_id: str
    rna_identity_id: str
    valid: bool
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class LoadedProfileSource:
    profile_source_id: str
    runtime_path: Path
    source_name_matches: bool
    source_sha256: str
    source_size_bytes: int
    header: tuple[str, ...]
    profile_type: ReconstructedProfileType
    input_status: str
    coordinates: tuple[float, ...]
    intensities: tuple[float, ...]
    row_count: int
    coordinate_min: float | None
    coordinate_max: float | None
    median_spacing: float | None
    sheet_name: str | None
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class ProfileCandidateRoutingResult:
    profile_source: ReconstructedProfileSource
    status: ProfileRoutingStatus
    rna_identity_id: str
    candidates: tuple[IntactRnaTheoreticalCandidate, ...]
    reason: str


@dataclass(frozen=True)
class ProfileShadowComparisonRow:
    profile_source_id: str
    measurement_id: str
    rna_identity_id: str
    candidate_id: str
    peak_id: str
    apex_mass: float
    centroid_mass: float | None
    theoretical_mass: float
    apex_delta_da: float
    absolute_apex_delta_da: float
    centroid_delta_da: float | None
    match_tolerance_class: str
    observed_mass_type: str
    theoretical_mass_type: str
    mass_definition_compatibility: str
    calibration_applied: bool
    mass_match_only: bool
    unmodified_candidate: bool
    cca_state_confirmed: bool
    terminal_state_confirmed: bool
    structure_identity_assigned: bool
    position_assigned: bool
    modification_assigned: bool
    biological_cause_assigned: bool
    rnase_t_assigned: bool
    applied_to_formal_score: bool
    applied_to_ranking: bool
    applied_to_candidate_filtering: bool
    applied_to_final_consensus: bool
    theoretical_reference_mode: str
    theoretical_mass_definition: str
    theoretical_output_species: str
    observed_mass_scale: str
    observed_output_species: str
    observed_output_species_confirmed: bool
    comparison_role: str
    candidate_role: str
    native_modifications_expected: bool
    modification_mass_not_yet_applied: bool
    biological_unmodified_state_assigned: bool
    target_rna_identity_confirmed_by_mass: bool
    co_captured_rna_excluded: bool


@dataclass(frozen=True)
class ReferenceModeCounts:
    theoretical_reference_mode: str
    retained_row_count: int
    strict_count: int
    exploratory_count: int
    nearest_no_match_count: int


@dataclass(frozen=True)
class OutputSpeciesInference:
    best_supported_output_species: str
    best_supported_output_species_margin: float
    output_species_inference_ambiguous: bool
    output_species_assigned: bool = False


@dataclass(frozen=True)
class ProfileShadowComparisonResult:
    status: ShadowComparisonStatus
    reason: str
    detection_result: SciexIntactPeakDetectionResult | None
    candidates: tuple[IntactRnaTheoreticalCandidate, ...]
    detail_rows: tuple[ProfileShadowComparisonRow, ...]
    strict_tolerance_da: float
    exploratory_tolerance_da: float
    mode_counts: tuple[ReferenceModeCounts, ...] = ()
    output_species_inference: OutputSpeciesInference | None = None


@dataclass(frozen=True)
class _ComparisonDetectionView:
    result: SciexIntactPeakDetectionResult

    def peak_rows(self) -> list[dict[str, Any]]:
        return self.result.peak_rows()

    def diagnostics_row(self) -> dict[str, Any]:
        values = self.result.diagnostics_row()
        if str(values.get("Detection_Status", "")).startswith("DETECTION_COMPLETED"):
            values["Detection_Status"] = "DETECTION_COMPLETED"
        return values


_DEFAULT_SOURCES = (
    ReconstructedProfileSource(
        "LEU_UAA_WT_FULL_RECONSTRUCTED", "LEU_UAA_WT_FULL", "LEU_UAA_WT", "TRNA_LEU_UAA",
        ProfileSourceType.TEXT_NEUTRAL_PROFILE, "WT_LeuUAA(Full).txt", None, None,
        ReconstructedProfileType.NEUTRAL_MASS_PROFILE, "Mass", "Intensity",
        ExperimentType.FULL_LENGTH, AnalyteLevel.INTACT_RNA,
        UnknownMassMetadata.UNKNOWN, UnknownMassMetadata.UNKNOWN, None, None,
        ProfileSourceStatus.REGISTERED, ProfileSourceProvenanceStatus.USER_PROVIDED,
        True, True, "FULL_LENGTH_NEUTRAL_RECONSTRUCTED_PROFILE",
        ObservedMassScale.AVERAGE, ObservedOutputSpecies.UNKNOWN, False,
    ),
    ReconstructedProfileSource(
        "LEU_UAA_WT_T1_MZ", "LEU_UAA_WT_T1", "LEU_UAA_WT", "TRNA_LEU_UAA",
        ProfileSourceType.TEXT_MZ_PROFILE, "UAA-T1.txt", None, None,
        ReconstructedProfileType.MZ_PROFILE, "Mass/Charge", "Intensity",
        ExperimentType.RNASE_T1_DIGEST, AnalyteLevel.OLIGONUCLEOTIDE,
        UnknownMassMetadata.UNKNOWN, UnknownMassMetadata.UNKNOWN, None, None,
        ProfileSourceStatus.REGISTERED, ProfileSourceProvenanceStatus.USER_PROVIDED,
        False, False, "RNASE_T1_DIGEST_MZ_PROFILE",
    ),
    ReconstructedProfileSource(
        "LEU_UAG_WT_FULL_RECONSTRUCTED", "LEU_UAG_WT_FULL", "LEU_UAG_WT", "TRNA_LEU_UAG",
        ProfileSourceType.TEXT_NEUTRAL_PROFILE, "WT_LeuUAG(Full).txt", None, None,
        ReconstructedProfileType.NEUTRAL_MASS_PROFILE, "Mass", "Intensity",
        ExperimentType.FULL_LENGTH, AnalyteLevel.INTACT_RNA,
        UnknownMassMetadata.UNKNOWN, UnknownMassMetadata.UNKNOWN, None, None,
        ProfileSourceStatus.REGISTERED, ProfileSourceProvenanceStatus.USER_PROVIDED,
        True, True, "FULL_LENGTH_NEUTRAL_RECONSTRUCTED_PROFILE",
        ObservedMassScale.AVERAGE, ObservedOutputSpecies.UNKNOWN, False,
    ),
    ReconstructedProfileSource(
        "LEU_UAG_WT_T1_MZ", "LEU_UAG_WT_T1", "LEU_UAG_WT", "TRNA_LEU_UAG",
        ProfileSourceType.TEXT_MZ_PROFILE, "UAG-T1.txt", None, None,
        ReconstructedProfileType.MZ_PROFILE, "Mass/Charge", "Intensity",
        ExperimentType.RNASE_T1_DIGEST, AnalyteLevel.OLIGONUCLEOTIDE,
        UnknownMassMetadata.UNKNOWN, UnknownMassMetadata.UNKNOWN, None, None,
        ProfileSourceStatus.REGISTERED, ProfileSourceProvenanceStatus.USER_PROVIDED,
        False, False, "RNASE_T1_DIGEST_MZ_PROFILE",
    ),
    ReconstructedProfileSource(
        "GLU_UUC_WT_FULL_RECONSTRUCTED", "GLU_UUC_WT_FULL", "GLU_UUC_WT", "TRNA_GLU_UUC",
        ProfileSourceType.EXCEL_SHEET_NEUTRAL_PROFILE, None, "LC-MS_旧WT.xlsx", "旧WT_kenki_2",
        ReconstructedProfileType.NEUTRAL_MASS_PROFILE, "Mass", "Intensity",
        ExperimentType.FULL_LENGTH, AnalyteLevel.INTACT_RNA,
        UnknownMassMetadata.UNKNOWN, UnknownMassMetadata.UNKNOWN, None, None,
        ProfileSourceStatus.REGISTERED, ProfileSourceProvenanceStatus.USER_PROVIDED,
        True, True, "FULL_LENGTH_NEUTRAL_RECONSTRUCTED_PROFILE",
        ObservedMassScale.AVERAGE, ObservedOutputSpecies.UNKNOWN, False,
    ),
    ReconstructedProfileSource(
        "GLU_UUC_WT_P1_AP_MZML", "GLU_UUC_WT_P1_AP", "GLU_UUC_WT", "TRNA_GLU_UUC",
        ProfileSourceType.MZML_DIGEST, "Nsd01.mzML", None, None,
        ReconstructedProfileType.DIGEST_MZML, None, None,
        ExperimentType.NUCLEASE_P1_AP_DIGEST, AnalyteLevel.NUCLEOSIDE,
        UnknownMassMetadata.UNKNOWN, UnknownMassMetadata.UNKNOWN, None, None,
        ProfileSourceStatus.REGISTERED, ProfileSourceProvenanceStatus.USER_PROVIDED,
        False, False, "P1_AP_DIGEST_NOT_INTACT_RNA",
    ),
)


def build_profile_registry(
    sources: Iterable[ReconstructedProfileSource] = _DEFAULT_SOURCES,
) -> ReconstructedProfileRegistry:
    items = tuple(sources)
    if not items:
        raise ProfileRegistryValidationError("profile registry must not be empty")
    ids: set[str] = set()
    measurements: set[str] = set()
    for source in items:
        if not isinstance(source, ReconstructedProfileSource):
            raise ProfileRegistryValidationError("registry source must be ReconstructedProfileSource")
        if source.profile_source_id in ids:
            raise ProfileRegistryValidationError(f"duplicate profile_source_id: {source.profile_source_id}")
        if source.measurement_id in measurements:
            raise ProfileRegistryValidationError(f"duplicate measurement_id: {source.measurement_id}")
        ids.add(source.profile_source_id)
        measurements.add(source.measurement_id)
        _validate_source_shape(source)
    return ReconstructedProfileRegistry(1, items)


def _validate_source_shape(source: ReconstructedProfileSource) -> None:
    neutral = source.profile_type is ReconstructedProfileType.NEUTRAL_MASS_PROFILE
    mz_text = source.profile_type is ReconstructedProfileType.MZ_PROFILE
    if source.source_type is ProfileSourceType.TEXT_NEUTRAL_PROFILE and not neutral:
        raise ProfileRegistryValidationError("TEXT_NEUTRAL_PROFILE requires NEUTRAL_MASS_PROFILE")
    if source.source_type is ProfileSourceType.EXCEL_SHEET_NEUTRAL_PROFILE:
        if not neutral or not source.source_workbook_name or not source.source_sheet_name:
            raise ProfileRegistryValidationError("Excel neutral source requires workbook, sheet, and neutral type")
    if source.source_type is ProfileSourceType.TEXT_MZ_PROFILE and not mz_text:
        raise ProfileRegistryValidationError("TEXT_MZ_PROFILE requires MZ_PROFILE")
    if source.source_type is ProfileSourceType.MZML_DIGEST:
        if source.profile_type is not ReconstructedProfileType.DIGEST_MZML:
            raise ProfileRegistryValidationError("MZML_DIGEST requires DIGEST_MZML")
    if neutral and (source.mass_column, source.intensity_column) != ("Mass", "Intensity"):
        raise ProfileRegistryValidationError("neutral profile requires Mass and Intensity columns")
    if mz_text and (source.mass_column, source.intensity_column) != ("Mass/Charge", "Intensity"):
        raise ProfileRegistryValidationError("m/z profile requires Mass/Charge and Intensity columns")
    expected_eligible = (
        source.experiment_type is ExperimentType.FULL_LENGTH
        and source.expected_analyte_level is AnalyteLevel.INTACT_RNA
        and neutral
    )
    if source.eligible_for_intact_candidate_generation is not expected_eligible:
        raise ProfileRegistryValidationError("candidate eligibility conflicts with experiment/profile metadata")
    if source.eligible_for_intact_mass_comparison is not expected_eligible:
        raise ProfileRegistryValidationError("comparison eligibility conflicts with experiment/profile metadata")
    if not expected_eligible and not source.eligibility_reason:
        raise ProfileRegistryValidationError("ineligible source requires eligibility_reason")
    if expected_eligible:
        if source.observed_mass_scale is not ObservedMassScale.AVERAGE:
            raise ProfileRegistryValidationError("full neutral profile requires AVERAGE observed scale")
        if source.observed_output_species is not ObservedOutputSpecies.UNKNOWN:
            raise ProfileRegistryValidationError("observed output species must remain UNKNOWN")
        if source.observed_output_species_confirmed is not False:
            raise ProfileRegistryValidationError("observed output species must remain unconfirmed")
    elif any(value is not None for value in (
        source.observed_mass_scale,
        source.observed_output_species,
        source.observed_output_species_confirmed,
    )):
        raise ProfileRegistryValidationError("digest source must not carry intact mass metadata")


def resolve_profile_source(
    registry: ReconstructedProfileRegistry,
    profile_source_id: str,
) -> ReconstructedProfileSource:
    for source in registry.sources:
        if source.profile_source_id == profile_source_id:
            return source
    raise KeyError(f"unknown profile_source_id: {profile_source_id}")


def validate_profile_source_against_manifest(
    registry: ReconstructedProfileRegistry,
    manifest: SCIEXSampleManifest,
    profile_source_id: str,
) -> ProfileManifestValidation:
    source = resolve_profile_source(registry, profile_source_id)
    measurement = get_measurement(manifest, source.measurement_id)
    sample = get_sample(manifest, measurement.sample_id)
    identity = get_rna_identity(manifest, sample.rna_identity_id)
    conflicts = []
    if source.sample_id != sample.sample_id:
        conflicts.append(f"sample_id:{source.sample_id}!={sample.sample_id}")
    if source.rna_identity_id != identity.rna_identity_id:
        conflicts.append(f"rna_identity_id:{source.rna_identity_id}!={identity.rna_identity_id}")
    if source.experiment_type is not measurement.experiment_type:
        conflicts.append("experiment_type")
    if source.expected_analyte_level is not measurement.expected_analyte_level:
        conflicts.append("expected_analyte_level")
    if conflicts:
        raise ProfileRegistryValidationError(
            f"profile source {profile_source_id} conflicts with manifest: {','.join(conflicts)}"
        )
    return ProfileManifestValidation(
        source.profile_source_id,
        measurement.measurement_id,
        sample.sample_id,
        identity.rna_identity_id,
        True,
        (),
    )


def _file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _axis_summary(coordinates: tuple[float, ...]) -> tuple[float | None, float | None, float | None]:
    if not coordinates:
        return None, None, None
    positive_steps = [
        coordinates[index] - coordinates[index - 1]
        for index in range(1, len(coordinates))
        if coordinates[index] > coordinates[index - 1]
    ]
    return (
        min(coordinates),
        max(coordinates),
        median(positive_steps) if positive_steps else None,
    )


def _expected_runtime_name(source: ReconstructedProfileSource) -> str:
    return source.source_workbook_name or source.source_file_name or ""


def load_profile_source(
    source: ReconstructedProfileSource,
    runtime_path: str | Path,
    *,
    sheet_name: str | None = None,
) -> LoadedProfileSource:
    """Read one explicitly routed source without path discovery or data mutation."""
    path = Path(runtime_path)
    if not path.is_file():
        raise FileNotFoundError(f"profile source file not found: {path}")
    warnings = []
    name_matches = path.name == _expected_runtime_name(source)
    if not name_matches:
        warnings.append("SOURCE_NAME_MISMATCH")
    file_hash = _file_sha256(path)
    size = path.stat().st_size

    if source.source_type is ProfileSourceType.MZML_DIGEST:
        if sheet_name is not None:
            raise ProfileRegistryValidationError("sheet_name is invalid for MZML source")
        return LoadedProfileSource(
            source.profile_source_id, path, name_matches, file_hash, size, (),
            ReconstructedProfileType.DIGEST_MZML, "METADATA_ONLY", (), (), 0,
            None, None, None, None, tuple(warnings),
        )

    selected_sheet = None
    if source.source_type is ProfileSourceType.EXCEL_SHEET_NEUTRAL_PROFILE:
        selected_sheet = source.source_sheet_name
        if sheet_name is not None and sheet_name != selected_sheet:
            raise ProfileRegistryValidationError(
                f"wrong sheet: expected {selected_sheet!r}, got {sheet_name!r}"
            )
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if selected_sheet not in workbook.sheetnames:
                raise ProfileRegistryValidationError(f"registered sheet not found: {selected_sheet}")
            worksheet = workbook[selected_sheet]
            rows = worksheet.iter_rows(values_only=True)
            header_raw = next(rows, None)
            header = tuple(str(value).strip() for value in (header_raw or ()) if value is not None)
            if header != ("Mass", "Intensity"):
                raise ProfileRegistryValidationError(f"unexpected Excel header: {header!r}")
            values = []
            for row_number, row in enumerate(rows, 2):
                if not row or all(value is None for value in row):
                    continue
                if len(row) < 2:
                    raise ProfileRegistryValidationError(f"invalid Excel row {row_number}")
                try:
                    values.append((float(row[0]), float(row[1])))
                except (TypeError, ValueError) as exc:
                    raise ProfileRegistryValidationError(f"invalid Excel row {row_number}") from exc
        finally:
            workbook.close()
        profile_type = ReconstructedProfileType.NEUTRAL_MASS_PROFILE
        input_status = "SUPPORTED_INPUT"
    else:
        if sheet_name is not None:
            raise ProfileRegistryValidationError("sheet_name is valid only for Excel sources")
        parsed = parse_sciex_profile(path)
        diagnostic = parsed.diagnostic_rows[0]
        header = tuple(str(diagnostic["Raw_Headers"]).split(";"))
        profile_type = (
            ReconstructedProfileType.NEUTRAL_MASS_PROFILE
            if parsed.profile_type == PARSER_NEUTRAL_MASS_PROFILE
            else ReconstructedProfileType.MZ_PROFILE
            if parsed.profile_type == PARSER_MZ_PROFILE
            else source.profile_type
        )
        input_status = parsed.input_status
        coordinate_field = (
            "Neutral_Mass"
            if profile_type is ReconstructedProfileType.NEUTRAL_MASS_PROFILE
            else "MZ"
        )
        values = [
            (float(row[coordinate_field]), float(row["Intensity"]))
            for row in parsed.input_rows
        ]
    if profile_type is not source.profile_type:
        raise ProfileRegistryValidationError(
            f"profile type mismatch: registered={source.profile_type.value}, observed={profile_type.value}"
        )
    coordinates = tuple(value[0] for value in values)
    intensities = tuple(value[1] for value in values)
    axis_min, axis_max, spacing = _axis_summary(coordinates)
    return LoadedProfileSource(
        source.profile_source_id, path, name_matches, file_hash, size, header,
        profile_type, input_status, coordinates, intensities, len(coordinates),
        axis_min, axis_max, spacing, selected_sheet, tuple(warnings),
    )


def route_profile_source_to_candidates(
    registry: ReconstructedProfileRegistry,
    manifest: SCIEXSampleManifest,
    profile_source_id: str,
    *,
    include_secondary_terminal_state: bool = True,
) -> ProfileCandidateRoutingResult:
    validation = validate_profile_source_against_manifest(
        registry, manifest, profile_source_id
    )
    source = resolve_profile_source(registry, profile_source_id)
    if not source.eligible_for_intact_candidate_generation:
        return ProfileCandidateRoutingResult(
            source, ProfileRoutingStatus.SKIPPED, validation.rna_identity_id, (),
            source.eligibility_reason,
        )
    candidates = generate_candidates_for_measurement(
        manifest,
        source.measurement_id,
        include_secondary_terminal_state=include_secondary_terminal_state,
    )
    if any(item.rna_identity_id != validation.rna_identity_id for item in candidates):
        raise ProfileRegistryValidationError("candidate identity conflicts with profile/manifest routing")
    return ProfileCandidateRoutingResult(
        source, ProfileRoutingStatus.ROUTED, validation.rna_identity_id,
        candidates, "FULL_NEUTRAL_PROFILE_ROUTED",
    )


def _retained_comparison_rows(
    details: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    matched = [
        row for row in details
        if row.get("Match_Tolerance_Class") in {"STRICT", "EXPLORATORY"}
    ]
    if matched:
        return matched
    eligible = [row for row in details if row.get("Absolute_Delta_Mass") is not None]
    return [min(eligible, key=lambda row: row["Absolute_Delta_Mass"])] if eligible else []


def _candidate_mass_references(
    candidate: IntactRnaTheoreticalCandidate,
) -> tuple[tuple[TheoreticalMassDefinition, float, ComparisonReferenceRole, str], ...]:
    return (
        (
            TheoreticalMassDefinition.AVERAGE_NEUTRAL_M,
            candidate.theoretical_average_neutral_molecular_mass_m,
            ComparisonReferenceRole.PRIMARY_CANDIDATE_NEUTRAL_M,
            "M",
        ),
        (
            TheoreticalMassDefinition.AVERAGE_M_PLUS_H,
            candidate.theoretical_average_m_plus_h,
            ComparisonReferenceRole.OUTPUT_SPECIES_DIAGNOSTIC_M_PLUS_H,
            "M_PLUS_H",
        ),
        (
            TheoreticalMassDefinition.AVERAGE_M_MINUS_H,
            candidate.theoretical_average_m_minus_h,
            ComparisonReferenceRole.OUTPUT_SPECIES_DIAGNOSTIC_M_MINUS_H,
            "M_MINUS_H",
        ),
        (
            TheoreticalMassDefinition.MONOISOTOPIC_NEUTRAL_M,
            candidate.theoretical_monoisotopic_neutral_mass,
            ComparisonReferenceRole.MONOISOTOPIC_DIAGNOSTIC_ONLY,
            "M",
        ),
    )


def _mode_counts(rows: list[ProfileShadowComparisonRow]) -> tuple[ReferenceModeCounts, ...]:
    result = []
    modes = (
        TheoreticalMassDefinition.AVERAGE_NEUTRAL_M,
        TheoreticalMassDefinition.AVERAGE_M_PLUS_H,
        TheoreticalMassDefinition.AVERAGE_M_MINUS_H,
        TheoreticalMassDefinition.MONOISOTOPIC_NEUTRAL_M,
    )
    for mode in modes:
        selected = [row for row in rows if row.theoretical_reference_mode == mode.value]
        result.append(ReferenceModeCounts(
            mode.value,
            len(selected),
            sum(row.match_tolerance_class == "STRICT" for row in selected),
            sum(row.match_tolerance_class == "EXPLORATORY" for row in selected),
            sum(row.match_tolerance_class == "NO_MATCH" for row in selected),
        ))
    return tuple(result)


def _infer_output_species(
    rows: list[ProfileShadowComparisonRow],
    counts: tuple[ReferenceModeCounts, ...],
) -> OutputSpeciesInference:
    diagnostic_modes = (
        TheoreticalMassDefinition.AVERAGE_NEUTRAL_M,
        TheoreticalMassDefinition.AVERAGE_M_PLUS_H,
        TheoreticalMassDefinition.AVERAGE_M_MINUS_H,
    )
    by_count = {item.theoretical_reference_mode: item for item in counts}
    ranked = []
    for order, mode in enumerate(diagnostic_modes):
        mode_rows = [row for row in rows if row.theoretical_reference_mode == mode.value]
        nearest_by_candidate = {}
        for row in mode_rows:
            current = nearest_by_candidate.get(row.candidate_id)
            if current is None or row.absolute_apex_delta_da < current:
                nearest_by_candidate[row.candidate_id] = row.absolute_apex_delta_da
        nearest_median = median(nearest_by_candidate.values()) if nearest_by_candidate else float("inf")
        item = by_count[mode.value]
        ranked.append((mode, item, nearest_median, order))
    ranked.sort(key=lambda value: (-value[1].strict_count, -value[1].exploratory_count, value[2], value[3]))
    best, second = ranked[0], ranked[1]
    species = {
        TheoreticalMassDefinition.AVERAGE_NEUTRAL_M: "M",
        TheoreticalMassDefinition.AVERAGE_M_PLUS_H: "M_PLUS_H",
        TheoreticalMassDefinition.AVERAGE_M_MINUS_H: "M_MINUS_H",
    }[best[0]]
    if best[1].strict_count != second[1].strict_count:
        margin = float(best[1].strict_count - second[1].strict_count)
    elif best[1].exploratory_count != second[1].exploratory_count:
        margin = float(best[1].exploratory_count - second[1].exploratory_count)
    else:
        margin = float(second[2] - best[2])
    if not isfinite(margin):
        margin = 0.0
    # Native modifications are expected and are not modeled here.  Even a clear
    # shadow-count leader therefore cannot resolve the SCIEX output convention.
    return OutputSpeciesInference(species, float(margin), True, False)


def compare_loaded_profile_shadow(
    loaded: LoadedProfileSource,
    routing: ProfileCandidateRoutingResult,
    *,
    peak_parameters: SciexIntactPeakDetectionParameters | None = None,
    strict_tolerance_da: float = 1.0,
    exploratory_tolerance_da: float = 5.0,
) -> ProfileShadowComparisonResult:
    """Compare every detected peak in four independently bounded reference modes."""
    if loaded.profile_source_id != routing.profile_source.profile_source_id:
        raise ProfileRegistryValidationError("loaded source and routing source conflict")
    if routing.status is ProfileRoutingStatus.SKIPPED:
        return ProfileShadowComparisonResult(
            ShadowComparisonStatus.SKIPPED, routing.reason, None, (), (),
            strict_tolerance_da, exploratory_tolerance_da,
        )
    if loaded.profile_type is not ReconstructedProfileType.NEUTRAL_MASS_PROFILE:
        return ProfileShadowComparisonResult(
            ShadowComparisonStatus.SKIPPED, "PROFILE_NOT_NEUTRAL_MASS", None, (), (),
            strict_tolerance_da, exploratory_tolerance_da,
        )
    source = routing.profile_source
    if (
        source.observed_mass_scale is not ObservedMassScale.AVERAGE
        or source.observed_output_species is not ObservedOutputSpecies.UNKNOWN
        or source.observed_output_species_confirmed is not False
    ):
        raise ProfileRegistryValidationError("full profile observed mass metadata is invalid")
    detection = detect_sciex_intact_peaks(
        loaded.coordinates,
        loaded.intensities,
        profile_type=PARSER_NEUTRAL_MASS_PROFILE,
        input_status=loaded.input_status,
        eligible_for_neutral_mass_analysis=True,
        parameters=peak_parameters,
    )
    rows: list[ProfileShadowComparisonRow] = []
    for candidate in routing.candidates:
        for mode, theoretical_mass, role, output_species in _candidate_mass_references(candidate):
            comparison = compare_sciex_intact_masses(
                _ComparisonDetectionView(detection),
                theoretical_mass,
                source_file=str(loaded.runtime_path),
                strict_tolerance_da=strict_tolerance_da,
                exploratory_tolerance_da=exploratory_tolerance_da,
                identity_status="CONFIRMED",
                sequence_source="MANIFEST_PROFILE_ROUTING",
                terminal_state_confirmed=False,
            )
            for row in _retained_comparison_rows(comparison.details()):
                centroid = row.get("Centroid_Mass")
                apex_delta = float(row["Apex_Delta_Da"])
                rows.append(ProfileShadowComparisonRow(
                    profile_source_id=source.profile_source_id,
                    measurement_id=source.measurement_id,
                    rna_identity_id=routing.rna_identity_id,
                    candidate_id=candidate.candidate_id,
                    peak_id=str(row.get("Peak_ID") or ""),
                    apex_mass=float(row["Apex_Mass"]),
                    centroid_mass=float(centroid) if centroid is not None else None,
                    theoretical_mass=theoretical_mass,
                    apex_delta_da=apex_delta,
                    absolute_apex_delta_da=abs(apex_delta),
                    centroid_delta_da=(
                        float(row["Centroid_Delta_Da"])
                        if row.get("Centroid_Delta_Da") is not None else None
                    ),
                    match_tolerance_class=str(row["Match_Tolerance_Class"]),
                    observed_mass_type="UNKNOWN",
                    theoretical_mass_type="MONOISOTOPIC_NEUTRAL",
                    mass_definition_compatibility="UNKNOWN",
                    calibration_applied=False,
                    mass_match_only=True,
                    unmodified_candidate=True,
                    cca_state_confirmed=False,
                    terminal_state_confirmed=False,
                    structure_identity_assigned=False,
                    position_assigned=False,
                    modification_assigned=False,
                    biological_cause_assigned=False,
                    rnase_t_assigned=False,
                    applied_to_formal_score=False,
                    applied_to_ranking=False,
                    applied_to_candidate_filtering=False,
                    applied_to_final_consensus=False,
                    theoretical_reference_mode=mode.value,
                    theoretical_mass_definition=mode.value,
                    theoretical_output_species=output_species,
                    observed_mass_scale=source.observed_mass_scale.value,
                    observed_output_species=source.observed_output_species.value,
                    observed_output_species_confirmed=False,
                    comparison_role=role.value,
                    candidate_role=candidate.candidate_role.value,
                    native_modifications_expected=True,
                    modification_mass_not_yet_applied=True,
                    biological_unmodified_state_assigned=False,
                    target_rna_identity_confirmed_by_mass=False,
                    co_captured_rna_excluded=False,
                ))
    counts = _mode_counts(rows)
    inference = _infer_output_species(rows, counts)
    return ProfileShadowComparisonResult(
        ShadowComparisonStatus.COMPLETED,
        "AVERAGE_OUTPUT_SPECIES_SHADOW_DIAGNOSTIC",
        detection,
        routing.candidates,
        tuple(rows),
        strict_tolerance_da,
        exploratory_tolerance_da,
        counts,
        inference,
    )


DEFAULT_PROFILE_REGISTRY = build_profile_registry()
