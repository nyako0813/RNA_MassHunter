from collections.abc import Mapping
from dataclasses import asdict, is_dataclass
from datetime import datetime
from enum import Enum
import json
import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

from rna_masshunter.composite_modification_audit import (
    BACKBONE_COLUMNS, CLEAVAGE_COLUMNS, COMPOSITE_CANDIDATE_COLUMNS,
    COMPOSITE_INVALID_COLUMNS, COMPOSITE_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as COMPOSITE_DIAGNOSTIC_COLUMNS,
)
from rna_masshunter.composite_structure_provenance import (
    POSITION_MAP_COLUMNS as COMPOSITE_STRUCTURE_POSITION_MAP_COLUMNS,
    BOND_MAP_COLUMNS as COMPOSITE_STRUCTURE_BOND_MAP_COLUMNS,
)
from rna_masshunter.composite_observation_audit import (
    BLOCKED_COLUMNS as COMPOSITE_BLOCKED_COLUMNS, COMPARE_COLUMNS as COMPOSITE_COMPARE_COLUMNS,
    FRAGMENT_COLUMNS as COMPOSITE_FRAGMENT_COLUMNS, INVALID_COLUMNS as COMPOSITE_OBS_INVALID_COLUMNS,
    MS1_COLUMNS as COMPOSITE_MS1_COLUMNS, MS1_SUMMARY_COLUMNS as COMPOSITE_MS1_SUMMARY_COLUMNS,
    MS2_ION_COLUMNS as COMPOSITE_MS2_ION_COLUMNS, MS2_MATCH_COLUMNS as COMPOSITE_MS2_MATCH_COLUMNS,
    MS2_ASSIGNMENT_COMPETITION_COLUMNS as COMPOSITE_MS2_ASSIGNMENT_COMPETITION_COLUMNS,
    OBS_SUMMARY_COLUMNS as COMPOSITE_OBS_SUMMARY_COLUMNS, SCORE_COLUMNS as COMPOSITE_SCORE_COLUMNS,
    SUPPORT_COLUMNS as COMPOSITE_SUPPORT_COLUMNS,
)
from rna_masshunter.pt_paired_audit import (
    PT_DISCOVERY_COLUMNS, PT_PAIRED_EVIDENCE_COLUMNS, PT_STATE_SEARCH_COLUMNS, PT_SUMMARY_COLUMNS,
)
from rna_masshunter.audit_policy import (
    AUDIT_STATUS_COLUMNS, DIAGNOSTIC_COLUMNS as AUDIT_LEVEL_DIAGNOSTIC_COLUMNS,
    AuditPolicy, included_sheet_names, sheet_category,
)
from rna_masshunter.warnings_manager import add_warning

from rna_masshunter.intact_reconstruction import (
    ASSIGNMENT_DRY_RUN_COLUMNS,
    ASSIGNMENT_DRY_RUN_SUMMARY_COLUMNS,
    ASSIGNMENT_SENSITIVITY_COLUMNS,
    ASSIGNMENT_STABILITY_COLUMNS,
    ASSIGNMENT_CANDIDATE_AUDIT_COLUMNS,
    ASSIGNMENT_AMBIGUOUS_COLUMNS,
    PREASSIGNMENT_COMPARISON_COLUMNS,
    COMPARISON_CANDIDATE_COLUMNS as INTACT_COMPARISON_CANDIDATE_COLUMNS,
    COMPETITION_GROUP_COLUMNS as INTACT_COMPETITION_GROUP_COLUMNS,
    COMPETITION_SCORE_COLUMNS as INTACT_COMPETITION_SCORE_COLUMNS,
    DIAGNOSTIC_COLUMNS as INTACT_DIAGNOSTIC_COLUMNS,
    ENGINE_COMPARISON_COLUMNS,
    GROUP_COLUMNS as INTACT_GROUP_COLUMNS,
    MISSING_CHARGE_DIAGNOSTIC_COLUMNS,
    QC_COLUMNS as INTACT_QC_COLUMNS,
    RECONSTRUCTED_MASS_SPECTRUM_COLUMNS,
    RT_ENGINE_QC_SUMMARY_COLUMNS,
    RT_ENVELOPE_DIAGNOSTIC_COLUMNS,
    TARGET_REVIEW_CANDIDATE_COLUMNS as INTACT_TARGET_REVIEW_CANDIDATE_COLUMNS,
    build_assignment_dry_run_rows,
    build_assignment_dry_run_summary_rows,
    build_assignment_sensitivity_rows,
    build_assignment_stability_rows,
    build_assignment_candidate_audit_rows,
    build_assignment_ambiguous_rows,
    build_preassignment_comparison_rows,
    build_intact_comparison_candidate_rows,
    build_intact_competition_group_rows,
    build_intact_competition_score_rows,
    build_intact_envelope_group_rows,
    build_intact_reconstruction_qc,
    build_reconstructed_mass_spectrum_rows,
    build_rt_engine_qc_summary_rows,
    build_target_review_candidate_rows,
)
from rna_masshunter.ms2_annotation import (
    MS2_FRAGMENT_EVIDENCE_COLUMNS,
    MS2_ION_MATCH_COLUMNS,
    MS2_MODIFIED_PRECURSOR_COLUMNS,
    MS2_MODIFIED_THEORETICAL_ION_COLUMNS,
    MS2_MODIFIED_ION_MATCH_COLUMNS,
    MS2_LOCALIZATION_EVIDENCE_COLUMNS,
    MS2_PARENT_CANDIDATE_COLUMNS,
    MS2_SPECTRA_COLUMNS,
    MS2_SUMMARY_COLUMNS,
    MS2_THEORETICAL_ION_COLUMNS,
    MS2_UNMATCHED_COLUMNS,
)
from rna_masshunter.evidence_ranking import AMBIGUITY_GROUP_COLUMNS, RANKING_COLUMNS, SUMMARY_COLUMNS
from rna_masshunter.biological_context import CONTEXT_PRIORITY_COLUMNS
from rna_masshunter.biological_position_prior import (
    BIOLOGICAL_PLAUSIBILITY_COLUMNS, DIAGNOSTIC_COLUMNS as BIOLOGICAL_PRIOR_DIAGNOSTIC_COLUMNS,
    POSITION_PRIOR_COLUMNS, SHADOW_RANKING_COLUMNS,
)
from rna_masshunter.ms2_identity_evidence import (IDENTITY_COLUMNS, IDENTITY_SHADOW_COLUMNS, PEAK_ASSIGNMENT_COLUMNS)
from rna_masshunter.ms2_ambiguous_peak_audit import (
    CLUSTER_COLUMNS as MS2_AMBIGUOUS_CLUSTER_COLUMNS,
    DETAIL_COLUMNS as MS2_AMBIGUOUS_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as MS2_AMBIGUITY_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS2_AMBIGUOUS_DIAGNOSTIC_COLUMNS,
    TOP_SHADOW_COLUMNS as MS2_AMBIGUOUS_TOP_COLUMNS,
)
from rna_masshunter.ms2_unmatched_audit import (
    AUDIT_COLUMNS as MS2_UNMATCHED_ION_AUDIT_COLUMNS,
    SUMMARY_COLUMNS as MS2_UNMATCHED_ION_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS2_UNMATCHED_ION_DIAGNOSTIC_COLUMNS,
    TOP_SHADOW_COLUMNS as MS2_UNMATCHED_TOP_COLUMNS,
)
from rna_masshunter.ms2_zero_intensity_audit import (
    SPECTRA_COLUMNS as MS2_ZERO_INTENSITY_SPECTRA_COLUMNS,
    DETAIL_COLUMNS as MS2_ZERO_INTENSITY_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as MS2_ZERO_INTENSITY_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS2_ZERO_INTENSITY_DIAGNOSTIC_COLUMNS,
    TOP_SHADOW_COLUMNS as MS2_ZERO_INTENSITY_TOP_COLUMNS,
)
from rna_masshunter.ms1_match_truncation_audit import (
    AUDIT_COLUMNS as MS1_TRUNCATION_AUDIT_COLUMNS,
    DETAIL_COLUMNS as MS1_TRUNCATION_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as MS1_TRUNCATION_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS1_TRUNCATION_DIAGNOSTIC_COLUMNS,
    TOP_COLUMNS as MS1_TRUNCATION_TOP_COLUMNS,
)
from rna_masshunter.ms1_selection_strategy_audit import (
    STRATEGY_COLUMNS as MS1_SELECTION_STRATEGY_COLUMNS,
    DETAIL_COLUMNS as MS1_SELECTION_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as MS1_SELECTION_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS1_SELECTION_DIAGNOSTIC_COLUMNS,
    TOP_COLUMNS as MS1_SELECTION_TOP_COLUMNS,
)
from rna_masshunter.ms1_top50_dedup_audit import (
    TOP50_COLUMNS as MS1_TOP50_SHADOW_COLUMNS,
    DETAIL_COLUMNS as MS1_PEAK_DEDUP_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as MS1_TOP50_DEDUP_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS1_TOP50_DEDUP_DIAGNOSTIC_COLUMNS,
    TOP_COLUMNS as MS1_TOP50_DEDUP_TOP_COLUMNS,
)
from rna_masshunter.ms1_cross_fragment_ambiguity import (
    AMBIGUITY_COLUMNS as MS1_CROSSFRAG_AMBIGUITY_COLUMNS,
    DETAIL_COLUMNS as MS1_CROSSFRAG_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as MS1_CROSSFRAG_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS1_CROSSFRAG_DIAGNOSTIC_COLUMNS,
    TOP_COLUMNS as MS1_CROSSFRAG_TOP_COLUMNS,
)
from rna_masshunter.ms2_effective_ambiguity import (
    CLUSTER_COLUMNS as MS2_EFFECTIVE_AMBIGUITY_COLUMNS,
    DETAIL_COLUMNS as MS2_EFFECTIVE_AMBIGUITY_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as MS2_EFFECTIVE_AMBIGUITY_SUMMARY_COLUMNS,
    DIAGNOSTIC_COLUMNS as MS2_EFFECTIVE_AMBIGUITY_DIAGNOSTIC_COLUMNS,
    TOP_SHADOW_COLUMNS as MS2_EFFECTIVE_AMBIGUITY_TOP_COLUMNS,
)
from rna_masshunter.rnase_ms2_evidence_synthesis import (
    SUMMARY_COLUMNS as RNASE_MS2_EVIDENCE_SUMMARY_COLUMNS,
    CANDIDATE_EVIDENCE_COLUMNS as RNASE_MS2_CANDIDATE_EVIDENCE_COLUMNS,
    PEAK_EVIDENCE_COLUMNS as RNASE_MS2_PEAK_EVIDENCE_COLUMNS,
)
from rna_masshunter.rnase_ms2_consensus_synthesis import (
    SUMMARY_COLUMNS as RNASE_MS2_CONSENSUS_SUMMARY_COLUMNS,
    EVIDENCE_COLUMNS as RNASE_MS2_CONSENSUS_EVIDENCE_COLUMNS,
)
from rna_masshunter.sciex_profile_parser import (
    DIAGNOSTIC_COLUMNS as SCIEX_PROFILE_DIAGNOSTIC_COLUMNS,
    INPUT_COLUMNS as SCIEX_PROFILE_INPUT_COLUMNS,
)
from rna_masshunter.sciex_intact_mass_comparison import (
    DETAIL_COLUMNS as SCIEX_MASS_COMPARISON_DETAIL_COLUMNS,
    SUMMARY_COLUMNS as SCIEX_MASS_COMPARISON_SUMMARY_COLUMNS,
)
from rna_masshunter.sciex_delta_mass_cluster_audit import (
    AUDIT_RESULT_KEY as SCIEX_DELTA_CLUSTER_RESULT_KEY,
    CLUSTER_COLUMNS as SCIEX_DELTA_CLUSTER_COLUMNS,
    CLUSTER_SHEET as SCIEX_DELTA_CLUSTER_SHEET,
    RELATION_COLUMNS as SCIEX_DELTA_RELATION_COLUMNS,
    RELATION_SHEET as SCIEX_DELTA_RELATION_SHEET,
    SUMMARY_COLUMNS as SCIEX_DELTA_CLUSTER_SUMMARY_COLUMNS,
    SUMMARY_SHEET as SCIEX_DELTA_CLUSTER_SUMMARY_SHEET,
)
from rna_masshunter.sciex_spacing_resolution_audit import (
    AUDIT_RESULT_KEY as SCIEX_SPACING_RESOLUTION_RESULT_KEY,
    DETAIL_COLUMNS as SCIEX_SPACING_RESOLUTION_DETAIL_COLUMNS,
    DETAIL_SHEET as SCIEX_SPACING_RESOLUTION_DETAIL_SHEET,
    SUMMARY_COLUMNS as SCIEX_SPACING_RESOLUTION_SUMMARY_COLUMNS,
    SUMMARY_SHEET as SCIEX_SPACING_RESOLUTION_SUMMARY_SHEET,
)
from rna_masshunter.sciex_relation_evidence_quality_audit import (
    AUDIT_RESULT_KEY as SCIEX_RELATION_EVIDENCE_RESULT_KEY,
    DETAIL_COLUMNS as SCIEX_RELATION_EVIDENCE_DETAIL_COLUMNS,
    DETAIL_SHEET as SCIEX_RELATION_EVIDENCE_DETAIL_SHEET,
    SUMMARY_COLUMNS as SCIEX_RELATION_EVIDENCE_SUMMARY_COLUMNS,
    SUMMARY_SHEET as SCIEX_RELATION_EVIDENCE_SUMMARY_SHEET,
)
from rna_masshunter.sciex_input_identity_audit import (
    AUDIT_RESULT_KEY as SCIEX_IDENTITY_AUDIT_RESULT_KEY,
    OUTPUT_COLUMNS as SCIEX_IDENTITY_AUDIT_COLUMNS,
    SHEET_NAME as SCIEX_IDENTITY_AUDIT_SHEET,
)
from rna_masshunter.sciex_rna_cross_layer_evidence_reconciliation import (
    OPTIONAL_RESULT_KEY as SCIEX_CROSS_LAYER_RESULT_KEY,
    audit_optional_result as audit_cross_layer_optional_result,
)
from rna_masshunter.rnase_ms2_standard_composite_crosswalk import (
    SUMMARY_COLUMNS as RNASE_MS2_STANDARD_COMPOSITE_SUMMARY_COLUMNS,
    CROSSWALK_COLUMNS as RNASE_MS2_STANDARD_COMPOSITE_CROSSWALK_COLUMNS,
)
from rna_masshunter.rnase_ms2_composite_evidence_synthesis import (
    SUMMARY_COLUMNS as RNASE_MS2_COMPOSITE_SUMMARY_COLUMNS,
    EVIDENCE_COLUMNS as RNASE_MS2_COMPOSITE_EVIDENCE_COLUMNS,
    PEAK_EVIDENCE_COLUMNS as RNASE_MS2_COMPOSITE_PEAK_EVIDENCE_COLUMNS,
)
from rna_masshunter.p1_annotation import (
    P1_ANNOTATION_COLUMNS,
    P1_SUMMARY_COLUMNS,
    P1_THEORETICAL_COLUMNS,
    P1_UNMATCHED_COLUMNS,
)
from rna_masshunter.p1_sap_chemical_state_audit import (
    CHEMICAL_STATE_COLUMNS as P1_SAP_CHEMICAL_STATE_COLUMNS,
    COMPETITION_COLUMNS as P1_SAP_COMPETITION_COLUMNS,
    CROSS_ENZYME_COLUMNS as P1_SAP_CROSS_ENZYME_COLUMNS,
    FEATURE_COLUMNS as P1_SAP_FEATURE_COLUMNS,
    MS2_PROVENANCE_COLUMNS as P1_SAP_MS2_PROVENANCE_COLUMNS,
    PT_FAMILY_COLUMNS as P1_SAP_PT_FAMILY_COLUMNS,
    SUMMARY_COLUMNS as P1_SAP_SUMMARY_COLUMNS,
    TERMINAL_COLUMNS as P1_SAP_TERMINAL_COLUMNS,
)
from rna_masshunter.p1_sap_dinucleotide_candidates import (
    SUMMARY_COLUMNS as P1_SAP_DINUC_SUMMARY_COLUMNS,
    GROUP_COLUMNS as P1_SAP_DINUC_GROUP_COLUMNS,
    ASSIGNMENT_COLUMNS as P1_SAP_DINUC_ASSIGNMENT_COLUMNS,
)
from rna_masshunter.p1_sap_dinucleotide_feature_audit import (
    SPECPEAK_COLUMNS as P1_SAP_DINUC_SPECPEAK_COLUMNS,
    FEATURE_COLUMNS as P1_SAP_DINUC_FEATURE_COLUMNS,
    ISOTOPE_COLUMNS as P1_SAP_DINUC_ISOTOPE_COLUMNS,
    COMPETITION_COLUMNS as P1_SAP_DINUC_COMPETITION_COLUMNS,
    MS2_COLUMNS as P1_SAP_DINUC_MS2_COLUMNS,
)
from rna_masshunter.p1_sap_dinucleotide_interpretation import TARGET_COLUMNS as P1_SAP_DINUC_TARGET_COLUMNS
from rna_masshunter.p1_sap_dinucleotide_evidence_synthesis import (
    EVIDENCE_COLUMNS as P1_SAP_DINUC_EVIDENCE_COLUMNS,
    GROUP_EVIDENCE_COLUMNS as P1_SAP_DINUC_GROUP_EVIDENCE_COLUMNS,
    SUMMARY_COLUMNS as P1_SAP_DINUC_EVIDENCE_SUMMARY_COLUMNS,
)
from rna_masshunter.p1_sap_feature_quality import (
    P1_SAP_SPECTRUM_PEAK_COLUMNS,
    P1_SAP_REFINED_FEATURE_COLUMNS,
    P1_SAP_FEATURE_QUALITY_COLUMNS,
    P1_SAP_ISOTOPE_AUDIT_COLUMNS,
    P1_SAP_QUALITY_SUMMARY_COLUMNS,
)
from rna_masshunter.audit_policy import (
    AUDIT_STATUS_COLUMNS, DIAGNOSTIC_COLUMNS as AUDIT_LEVEL_DIAGNOSTIC_COLUMNS,
    AuditPolicy, included_sheet_names, sheet_category,
    FORMAL_CORE, FORMAL_OPTIONAL, AUDIT_SUMMARY, AUDIT_GROUP, AUDIT_DETAIL,
)
from rna_masshunter.report_document_export import (
    WordExportCollector, _is_multi_sentence, write_word_appendix,
)
from rna_masshunter.report_document_export import (
    WordExportCollector, _is_multi_sentence, write_word_appendix,
    SHEETS_EXCLUDED_FROM_WORD_EXPORT,
)


AUDIT_TOP_SHADOW_COLUMNS = list(dict.fromkeys(
    MS2_UNMATCHED_TOP_COLUMNS + MS2_AMBIGUOUS_TOP_COLUMNS
    + MS2_ZERO_INTENSITY_TOP_COLUMNS + MS2_EFFECTIVE_AMBIGUITY_TOP_COLUMNS
    + MS1_TRUNCATION_TOP_COLUMNS + MS1_SELECTION_TOP_COLUMNS
    + MS1_TOP50_DEDUP_TOP_COLUMNS + MS1_CROSSFRAG_TOP_COLUMNS
))


EXCEL_MAX_ROWS = 1_048_576
DATA_START_ROW = 3
EXCEL_DATA_ROW_LIMIT = EXCEL_MAX_ROWS - DATA_START_ROW


SCIEX_INTACT_OPTIONAL_RESULT_KEY = "sciex_intact_peak_detection"
SCIEX_INTACT_DIAGNOSTIC_SHEET = "SCIEX_Intact_Peak_Diagnostics"
SCIEX_INTACT_PEAK_SHEET = "SCIEX_Intact_Detected_Peaks"
SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY = "sciex_intact_mass_comparison"
SCIEX_MASS_COMPARISON_DETAIL_SHEET = "SCIEX_Intact_Mass_Comparison"
SCIEX_MASS_COMPARISON_SUMMARY_SHEET = "SCIEX_Intact_Mass_Comp_Summary"

SCIEX_INTACT_DIAGNOSTIC_COLUMNS = [
    "Source_File", "Source_File_Name", "Profile_Type", "Input_Status",
    "Eligible_For_Neutral_Mass_Analysis", "Input_Validation_Status",
    "Detection_Status", "Detection_Method", "Algorithm_Version",
    "Parsed_Row_Count", "Mass_Min_Da", "Mass_Max_Da", "Mass_Step_Min_Da",
    "Mass_Step_Median_Da", "Mass_Step_Max_Da", "Mass_Axis_Strictly_Increasing",
    "Mass_Axis_Uniform", "Duplicate_Mass_Count", "Missing_Value_Count",
    "Nonfinite_Value_Count", "Negative_Intensity_Count", "Zero_Intensity_Count",
    "Baseline_Method", "Baseline_Quantile", "Baseline_Window_Points",
    "Baseline_Window_Da", "Baseline_Edge_Mode", "Baseline_Min", "Baseline_Median",
    "Baseline_Max", "Baseline_Negative_Residual_Fraction",
    "Noise_Estimation_Method", "Noise_Window_Points", "Noise_Window_Da",
    "Estimated_Noise_Global", "Estimated_Noise_Local_Min",
    "Estimated_Noise_Local_Median", "Estimated_Noise_Local_Max",
    "Height_Threshold_Method", "Prominence_Threshold_Method",
    "Strict_Prominence_Threshold_Method", "Positive_Residual_Quantile_Value",
    "Smoothing_Method", "Smoothing_Window_Points", "Smoothing_Window_Da",
    "Smoothing_Polyorder", "Minimum_Distance_Points", "Minimum_Distance_Da",
    "Minimum_Width_Points", "Minimum_Width_Da", "Boundary_Method",
    "Boundary_Peak_Set_Tier", "Boundary_Recomputed_After_Filtering",
    "Centroid_Method", "Area_Method", "Detected_Sensitive_Peak_Count",
    "Detected_Strict_Peak_Count", "Rejected_Height_Count",
    "Rejected_Prominence_Count", "Rejected_Width_Count",
    "Suppressed_By_Distance_Count", "Shallow_Valley_Neighbor_Count",
    "Possible_Shoulder_Count", "Broad_Peak_Width_Threshold_Da",
    "Broad_Peak_Threshold_Source", "Broad_Peak_Count", "Severe_Broad_Peak_Count",
    "Edge_Peak_Count", "Centroid_Fallback_Count", "Warning_Count",
    "Automatic_Parameter_Fallbacks", "Detection_Warnings",
    "Parameter_Provenance_JSON",
    "SCIEX_Intact_Peak_Detection_Applied_To_Formal_Score",
    "SCIEX_Intact_Peak_Detection_Applied_To_Ranking",
    "SCIEX_Intact_Peak_Detection_Applied_To_Candidate_Filtering",
]

SCIEX_INTACT_PEAK_COLUMNS = [
    "Peak_ID", "Source_File", "Detection_Tier", "Sensitive_Threshold_Passed",
    "Strict_Threshold_Passed", "Molecular_Identity_Assigned", "Apex_Index",
    "Apex_Mass", "Apex_Intensity_Raw", "Local_Baseline_At_Apex",
    "Apex_Intensity_Baseline_Corrected", "Detection_Signal_Apex",
    "Height_Threshold_Local", "Prominence", "Prominence_Threshold_Local",
    "Strict_Prominence_Threshold_Local", "Local_Noise_Sigma",
    "Noise_Estimation_Method_Local", "Prominence_Base_Left_Index",
    "Prominence_Base_Right_Index", "Prominence_Base_Left_Mass",
    "Prominence_Base_Right_Mass", "Half_Prominence_Left_IP",
    "Half_Prominence_Right_IP", "Half_Prominence_Width_Points",
    "Half_Prominence_Width_Da", "FWHM_Points", "FWHM_Da",
    "Left_Boundary_Index", "Right_Boundary_Index", "Left_Boundary_Mass",
    "Right_Boundary_Mass", "Boundary_Width_Da", "Boundary_Method",
    "Boundary_Fallback_Used", "Boundary_Left_Neighbor_Peak_ID",
    "Boundary_Right_Neighbor_Peak_ID", "Boundary_Peak_Set_Tier",
    "Boundary_Recomputed_After_Filtering", "Centroid_Mass",
    "Centroid_Minus_Apex_Da", "Centroid_Fallback_Used", "Centroid_Complete",
    "Peak_Area_Raw", "Peak_Area_Baseline_Corrected", "Peak_Area_Complete",
    "Area_Unit", "Plateau_Start_Index", "Plateau_End_Index",
    "Plateau_Size_Points", "Plateau_Width_Da", "Plateau_Center_Mass",
    "Neighbor_Peak_ID", "Neighbor_Separation_Da", "Shared_Valley_Index",
    "Shared_Valley_Mass", "Valley_To_Smaller_Apex_Ratio",
    "Width_To_Separation_Ratio", "Shallow_Valley_Neighbor_Flag",
    "Shoulder_Diagnostic_Reason", "Possible_Shoulder", "Broad_Peak_Flag",
    "Severe_Broad_Peak_Flag", "Edge_Peak_Flag", "Noise_Fallback_Used",
    "SCIEX_Intact_Peak_Detection_Applied_To_Formal_Score",
    "SCIEX_Intact_Peak_Detection_Applied_To_Ranking",
    "SCIEX_Intact_Peak_Detection_Applied_To_Candidate_Filtering",
]


INTACT_COLUMNS = [
    "Cluster_ID",
    "Reconstructed_Mass",
    "Observed_Mass",
    "In_Neutral_Mass_Search_Range",
    "Neutral_Mass_Search_Min_Da",
    "Neutral_Mass_Search_Max_Da",
    "Neutral_Mass_Range_Status",
    "In_Target_Review_Mass_Range",
    "Target_Review_Mass_Range_Status",
    "Target_Review_Priority",
    "Envelope_QC_Eligible",
    "Intact_Review_Eligible",
    "Intact_Strict_Eligible",
    "Intact_Envelope_QC_Score",
    "Intact_Envelope_QC_Rank",
    "Strict_Eligible_Rank",
    "Review_Eligible_Rank",
    "Dominant_Intact_Envelope_Flag",
    "Supporting_Peak_IDs",
    "Supporting_Peak_Count",
    "Supporting_Scan_IDs",
    "Supporting_RT_Values",
    "Supporting_Charge_States",
    "Exact_Peak_Set_Key",
    "Exact_Duplicate_Group_ID",
    "Exact_Duplicate_Count",
    "Is_Exact_Duplicate_Representative",
    "Intact_Envelope_Group_ID",
    "Envelope_Group_Size",
    "Group_Representative",
    "Group_Ambiguity_Status",
    "Comparison_Representative",
    "Comparison_Representative_Reason",
    "Comparison_Representative_Rank",
    "Excluded_From_Comparison_Reason",
    "Target_Review_Group_Representative",
    "Target_Review_Rank",
    "Dominant_Target_Review_Eligible_Flag",
    "Reconstruction_Status",
    "Reconstruction_Confidence",
    "Reconstruction_Engine",
    "RT_Window_ID",
    "RT_Window_Start_Min",
    "RT_Window_End_Min",
    "RT_Window_Center_Min",
    "Num_MS1_Scans_In_Window",
    "Peak_Aggregation_Method",
    "Anchor_MZ",
    "Anchor_Charge",
    "Predicted_Charge_States",
    "Observed_Charge_States",
    "Missing_Charge_States",
    "Missing_Charge_Predicted_MZ",
    "Num_Predicted_Charges",
    "Num_Observed_Charges",
    "Charge_Coverage_Fraction",
    "Consecutive_Charge_Run_Length",
    "Longest_Consecutive_Charge_Run",
    "Charge_Gap_Count",
    "Charge_Continuity_Fraction",
    "Peak_Usage_Count",
    "Shared_Peak_Count",
    "Shared_Peak_Fraction",
    "Local_Window_Max_Intensity",
    "Local_Relative_Peak_Intensity_Percent",
    "Local_Envelope_Relative_Intensity_Percent",
    "Neutral_Mass_Estimator",
    "Neutral_Mass_Unweighted_Mean",
    "Neutral_Mass_Weighted_Mean",
    "Neutral_Mass_Median",
    "Envelope_Internal_Error_Max_ppm",
    "Envelope_Internal_Error_Mean_ppm",
    "Envelope_Internal_Error_Median_ppm",
    "Source_RT_Window_IDs",
    "Num_Source_RT_Windows",
    "Merged_Across_RT_Windows",
    "Extended_Lower_Charges_Evaluated",
    "Extended_Upper_Charges_Evaluated",
    "Extended_Charges_Detected",
    "Extended_Weak_Charges_Detected",
    "Extended_Charges_Not_Detected",
    "Charge_Extension_Improved_Envelope",
    "Original_Charge_States",
    "Final_Charge_States",
    "Split_Envelope_Group_ID",
    "Split_Envelope_Member_Count",
    "Split_Envelope_Merged",
    "Charge_Gaps_Before_Merge",
    "Charge_Gaps_After_Merge",
    "Max_Peak_Usage_Count",
    "Mean_Peak_Usage_Count",
    "Num_Highly_Shared_Peaks",
    "Highly_Shared_Peak_Fraction",
    "Competing_Candidate_Count",
    "Peak_Sharing_Status",
    "Competing_Envelope_Group_ID",
    "Competing_Envelope_Group_Size",
    "Shared_Peak_Competitor_Count",
    "Maximum_Shared_Peak_Fraction",
    "Mean_Shared_Peak_Fraction",
    "Competitor_Cluster_IDs",
    "Is_Noncompeting_Candidate",
    "Envelope_Evidence_Score",
    "Evidence_Score_Rank_In_Competition",
    "Evidence_Score_Components",
    "Evidence_Score_Penalties",
    "Evidence_Score_Config_Version",
    "Direct_Competitor_Count",
    "Direct_Competitor_Cluster_IDs",
    "Direct_Shared_Peak_Count_Max",
    "Direct_Shared_Peak_Fraction_Max",
    "Competition_Component_Size",
    "Dry_Run_Assignment_Status",
    "Dry_Run_Selected",
    "Dry_Run_Selection_Order",
    "Supporting_Peak_Count_Before_Assignment",
    "Independent_Supporting_Peak_Count",
    "Independent_Supporting_Peak_Fraction",
    "Supporting_Charge_Count_Before_Assignment",
    "Independent_Charge_State_Count",
    "Peaks_Already_Assigned_Count",
    "Charges_Already_Assigned_Count",
    "Excluded_By_Cluster_ID",
    "Dry_Run_Exclusion_Reason",
    "Score_Margin_To_Excluding_Candidate",
    "Close_Score_Ambiguity",
    "Assignment_Confidence",
    "Shared_Observed_Peak_Count",
    "Shared_Peak_Charge_Assignment_Count",
    "Independent_Observed_Peak_Count",
    "Pass_Min_Charge_Count",
    "Pass_Min_Consecutive_Charge_Count",
    "Pass_Charge_Continuity",
    "Pass_Internal_Error",
    "Pass_Neutral_Mass_SD",
    "Pass_Neutral_Mass_Range",
    "Pass_RT_Consistency",
    "Pass_Local_Intensity",
    "Pass_Competing_Envelope",
    "Pass_Peak_Sharing",
    "Num_Strict_Criteria_Passed",
    "Num_Review_Criteria_Passed",
    "Strict_Failure_Reasons",
    "Review_Failure_Reasons",
    "Intact_Quality_Tier",
    "Quality_Tier_Reason",
    "Quality_Tier_Rank",
    "Comparison_Ready_Strict",
    "Comparison_Ready_Review",
    "Comparison_Ready",
    "Comparison_Readiness_Reason",
    "Total_Supporting_Intensity",
    "Mean_Supporting_Intensity",
    "Max_Supporting_Intensity",
    "Reconstructed_Envelope_Intensity",
    "Intensity_Method",
    "Relative_Envelope_Intensity_Percent",
    "Relative_Overall_Envelope_Intensity_Percent",
    "Relative_In_Range_Raw_Intensity_Percent",
    "Relative_Intact_Eligible_Intensity_Percent",
    "Supporting_Peak_Classes",
    "Trace_Only_Envelope",
    "Num_Supporting_Charge_States",
    "Charge_State_Count",
    "Charge_States",
    "Charge_State_Range",
    "Charge_State_Continuity",
    "Supporting_Peak_Count",
    "RT_Min",
    "RT_Max",
    "RT_Mean",
    "RT_Range_Min",
    "Max_RT_Difference_Min",
    "RT_Consistency",
    "Neutral_Mass_SD",
    "Neutral_Mass_Range",
    "Envelope_Internal_Error_ppm",
    "Max_Mass_Error_ppm",
    "Theoretical_Mass",
    "Mass_Error_Da",
    "Mass_Error_ppm",
    "Unmodified_Theory_Delta_Da",
    "Unmodified_Theory_Delta_ppm",
    "Best_Reference_Label",
    "Best_Reference_Mass_Da",
    "Reference_Mass_Error_Da",
    "Reference_Mass_Error_ppm",
    "Reference_Mass_Matched",
    "Competing_Envelope_Count",
    "Limiting_Factors",
    "Severe_Limiting_Factors",
    "Num_Limiting_Factors",
    "Primary_Limiting_Factor",
    "Total_Intensity",
    "Assignment",
    "Confidence",
    "Warnings",
]

CHARGE_COLUMNS = ["Cluster_ID", "mz", "Intensity", "RT", "Scan_ID", "Charge", "Neutral_Mass", "Peak_Tier"]

THEORETICAL_FRAGMENT_COLUMNS = [
    "Fragment_ID",
    "Target_ID",
    "Sequence",
    "Length",
    "Start",
    "End",
    "Enzyme",
    "Missed_Cleavages",
    "Terminal_Form",
    "Unmodified_Mass",
    "Warnings",
]

FRAGMENT_MS1_MATCH_COLUMNS = [
    "Match_ID",
    "Fragment_ID",
    "Target_ID",
    "Sequence",
    "Start",
    "End",
    "Enzyme",
    "Missed_Cleavages",
    "Terminal_Form",
    "Fragment_Mass",
    "Charge",
    "Theoretical_mz",
    "Observed_mz",
    "Mass_Error_Da",
    "Mass_Error_ppm",
    "Intensity",
    "RT",
    "Scan_ID",
    "Peak_Tier",
    "Confidence",
    "Warnings",
]

FRAGMENT_MS1_FILTERED_COLUMNS = [
    "Match_ID",
    "Fragment_ID",
    "Target_ID",
    "Sequence",
    "Length",
    "Start",
    "End",
    "Enzyme",
    "Missed_Cleavages",
    "Terminal_Form",
    "Fragment_Mass",
    "Charge",
    "Theoretical_mz",
    "Observed_mz",
    "Mass_Error_Da",
    "Mass_Error_ppm",
    "Intensity",
    "RT",
    "Scan_ID",
    "Peak_Tier",
    "Confidence",
    "Warnings",
]

FRAGMENT_MS1_SUMMARY_COLUMNS = [
    "Fragment_ID",
    "Target_ID",
    "Sequence",
    "Length",
    "Start",
    "End",
    "Enzyme",
    "Missed_Cleavages",
    "Terminal_Form",
    "Best_Charge",
    "Best_Theoretical_mz",
    "Best_Observed_mz",
    "Best_Mass_Error_ppm",
    "Best_Intensity",
    "Best_RT",
    "Best_Peak_Tier",
    "Best_Confidence",
    "Match_Count",
    "Major_Count",
    "Minor_Count",
    "Trace_Count",
    "High_Count",
    "Medium_Count",
    "Low_Count",
]

KNOWN_MODIFICATION_CANDIDATE_COLUMNS = [
    "candidate_id",
    "source_type",
    "source_id",
    "target_id",
    "sequence",
    "start",
    "end",
    "observed_mz",
    "theoretical_mz",
    "observed_mass",
    "unmodified_mass",
    "mass_error_unmodified_da",
    "mass_error_unmodified_ppm",
    "modification_id",
    "modification_symbol",
    "modification_name",
    "target_base",
    "modification_mass_shift",
    "modified_mass",
    "mass_error_modified_da",
    "mass_error_modified_ppm",
    "charge",
    "intensity",
    "rt",
    "peak_tier",
    "confidence",
    "priority_score",
    "notes",
    "warnings",
]

WORKFLOW_SUMMARY_COLUMNS = [
    "Analysis_Mode",
    "Step_Name",
    "Step_Status",
    "Enabled_By_Config",
    "Executed",
    "Skip_Reason",
    "Output_Sheets",
    "Notes",
]


KNOWN_MODIFICATION_SUMMARY_COLUMNS = [
    "Modification_ID",
    "Modification_Name",
    "Symbol",
    "Target_Base",
    "Candidate_Count",
    "Best_Source_ID",
    "Best_Sequence",
    "Best_Mass_Error_Modified_ppm",
    "Best_Intensity",
    "Best_Peak_Tier",
    "Best_Confidence",
    "Best_Priority_Score",
]

UNKNOWN_MODIFICATION_CANDIDATE_COLUMNS = [
    "candidate_id",
    "source_type",
    "source_id",
    "target_id",
    "sequence",
    "start",
    "end",
    "observed_mz",
    "theoretical_mz",
    "observed_mass",
    "unmodified_mass",
    "mass_error_unmodified_da",
    "mass_error_unmodified_ppm",
    "delta_label",
    "delta_elements",
    "delta_mass_shift",
    "modified_mass",
    "mass_error_modified_da",
    "mass_error_modified_ppm",
    "charge",
    "intensity",
    "rt",
    "peak_tier",
    "confidence",
    "priority_score",
    "notes",
    "warnings",
]

UNKNOWN_MODIFICATION_SUMMARY_COLUMNS = [
    "Delta_Label",
    "Delta_Elements",
    "Delta_Mass_Shift",
    "Candidate_Count",
    "Best_Source_ID",
    "Best_Sequence",
    "Best_Mass_Error_Modified_ppm",
    "Best_Intensity",
    "Best_Peak_Tier",
    "Best_Confidence",
    "Best_Priority_Score",
]

COMPOUND_MODIFICATION_CANDIDATE_COLUMNS = [
    "candidate_id",
    "source_type",
    "source_id",
    "target_id",
    "sequence",
    "start",
    "end",
    "observed_mz",
    "theoretical_mz",
    "observed_mass",
    "unmodified_mass",
    "mass_error_unmodified_da",
    "mass_error_unmodified_ppm",
    "modification_id",
    "modification_symbol",
    "modification_name",
    "target_base",
    "modification_mass_shift",
    "delta_label",
    "delta_elements",
    "delta_mass_shift",
    "combined_mass_shift",
    "modified_mass",
    "mass_error_modified_da",
    "mass_error_modified_ppm",
    "charge",
    "intensity",
    "rt",
    "peak_tier",
    "confidence",
    "priority_score",
    "notes",
    "warnings",
]

COMPOUND_MODIFICATION_SUMMARY_COLUMNS = [
    "Modification_ID",
    "Delta_Label",
    "Combined_Mass_Shift",
    "Candidate_Count",
    "Best_Source_ID",
    "Best_Sequence",
    "Best_Mass_Error_Modified_ppm",
    "Best_Intensity",
    "Best_Peak_Tier",
    "Best_Confidence",
    "Best_Priority_Score",
]

SHEET_DESCRIPTIONS = {
    "Run_summary": "Run-level summary for this RNA_MassHunter MVP-3 report.",
    "Workflow_Summary": "Workflow step execution and skip status for the selected analysis mode.",
    "Input_parameters": "Flattened parameters loaded from config.yaml.",
    "mzML_diagnostics": "mzML scan counts, ranges, precursor metadata, and warnings.",
    "Intact_mass_reconstruction": "Reconstructed intact mass clusters, mass errors, and reconstruction QC fields.",
    "Charge_state_peaks": "Peak and charge-state evidence supporting reconstructed masses.",
    "Intact_Reconstruction_QC": "Per-candidate intact mass reconstruction quality diagnostics.",
    "Intact_Reconstruction_Diag": "Run-level intact reconstruction QC settings, status counts, and limiting reasons.",
    "Intact_Envelope_Groups": "Grouped intact envelope candidates and selected group representatives.",
    "Intact_Comparison_Candidates": "Group representative intact candidates suitable for condition comparison.",
    "Target_Review_Candidates": "Optional target review range candidates when configured.",
    "Reconstructed_Mass_Spectrum": "Neutral-mass reconstructed spectrum points with representative envelope intensities.",
    "RT_Envelope_Diagnostics": "RT-localized reconstruction envelope generation diagnostics.",
    "RT_Engine_QC_Summary": "RT-localized engine quality tier, failure reason, missing charge, split envelope, and engine-match summaries.",
    "Missing_Charge_Diagnostics": "Predicted missing charge-state m/z diagnostics for RT-localized envelopes.",
    "Intact_Engine_Comparison": "Optional comparison between legacy_cluster and rt_localized intact engines.",
    "Intact_Competition_Groups": "Diagnostic groups of intact candidates sharing supporting local peaks; no candidate exclusion is applied.",
    "Intact_Competition_Scores": "Envelope-internal evidence scores and rank details within competition groups.",
    "P1_SAP_Dinuc_Evidence": "Physical-feature P1+SAP dinucleotide evidence synthesis; shadow-only.",
    "P1_SAP_Dinuc_Group_Evidence": "Group-level P1+SAP dinucleotide evidence and targeted MS/MS priority; shadow-only.",
    "P1_SAP_Dinuc_Evidence_Summary": "Run-level P1+SAP dinucleotide evidence synthesis counts; shadow-only.",
    "Intact_Assignment_Dry_Run": "Diagnostic-only dry-run peak assignment for competing intact candidates.",
    "Competition_Dry_Run_Summary": "Component-level summary of diagnostic dry-run assignment outcomes.",
    "Assignment_Sensitivity": "Threshold-only competitive assignment sensitivity scenarios.",
    "Assignment_Stability": "Per-candidate selection stability across assignment scenarios.",
    "Assignment_Candidate_Audit": "Optional audit-mass candidate extraction; does not affect assignment or QC.",
    "Assignment_Ambiguous_Candidates": "Ambiguous and threshold-sensitive candidates retained for assignment review.",
    "Preassignment_Comparison": "Comparison representatives before optional assignment eligibility gating.",
    "Theoretical_fragments": "Theoretical RNase digestion fragments and terminal forms.",
    "Fragment_MS1_matches": "MS1 peak matches for unmodified theoretical fragments.",
    "Fragment_MS1_filtered": "Filtered MS1 fragment matches for practical review.",
    "Fragment_MS1_summary": "Best MS1 match per fragment with match counts.",
    "MS1_Truncation_Audit": "Fragment-level shadow audit of matches retained and discarded by the formal MS1 cap.",
    "MS1_Truncation_Detail": "Pre-truncation Fragment MS1 match detail; formal matching and scoring are unchanged.",
    "MS1_Truncation_Summary": "Run-level unlimited-shadow comparison and truncation risk recommendation.",
    "MS1_Selection_Strategy": "Fragment-by-strategy shadow A/B comparison; formal MS1 selection is unchanged.",
    "MS1_Selection_Detail": "Per-match deterministic membership in current, filter-first, tier-first, and unlimited shadows.",
    "MS1_Selection_Summary": "Run-level selection recovery, candidate/ranking impact, and formal-readiness diagnosis.",
    "MS1_Top50_Shadow": "Full downstream tier-top20/top50/unlimited and exact-ID dedup shadow comparison.",
    "MS1_Peak_Dedup_Detail": "Physical-peak assignment sharing, charge interpretation, near-m/z, and dedup membership detail.",
    "MS1_Top50_Dedup_Summary": "Dataset and fragment-length summary of top50 recovery and physical-peak dedup impact.",
    "MS1_CrossFrag_Ambiguity": "Physical-peak groups with competing fragment assignments and deterministic shadow handling.",
    "MS1_CrossFrag_Detail": "Per-assignment cross-fragment ambiguity, candidate linkage, and shadow weights.",
    "MS1_CrossFrag_Summary": "Dataset and fragment-length cross-fragment ambiguity risk and strategy comparison.",
    "Audit_Status": "Audit execution status, availability, runtime, memory, and formal non-application by audit level.",
    "Composite_Mod_Candidates": "Constraint-valid Phase-1 composite nucleoside states; shadow only.",
    "Composite_Mod_Invalid": "Rejected composite attempts with structured chemical constraint reasons.",
    "Composite_Mod_Summary": "Summary of composite state generation, legacy overlap, and formal non-propagation.",
    "Backbone_Mod_Candidates": "Hypothetical per-bond phosphorothioate candidates; shadow only.",
    "Cleavage_Block_Audit": "Per-enzyme phosphorothioate cleavage constraint and resulting shadow fragment audit.",
    "Composite_Fragment_Masses": "Exact elemental compositions and masses from complete sample structure states.",
    "Composite_MS1_Matches": "Shadow MS1 peak matches for complete-structure fragments.",
    "Composite_MS1_Summary": "Candidate-level MS1 observation and observability counts.",
    "Composite_MS2_Ions": "Position/bond-state propagated complete-structure theoretical MS2 ions.",
    "Composite_MS2_Matches": "Observed MS2 support for propagated complete-structure ions.",
    "Composite_MS2_Assignment_Compe": "Excel alias for all within-tolerance composite MS2 assignment competition provenance.",
    "Composite_Structure_Position_Ma": "Excel alias for position-level component provenance derived directly from complete structure states; shadow-only.",
    "Composite_Structure_Bond_Map": "Bond-level component provenance derived directly from complete structure states; shadow-only.",
    "Composite_Support_Summary": "Candidate-level MS1/MS2/blocked-cleavage shadow support.",
    "Blocked_Cleavage_Matches": "Observed phosphorothioate-blocked cleavage fragments and mechanism comparison.",
    "Legacy_Composite_Compare": "Formal legacy versus valid complete-structure shadow comparison.",
    "Composite_Shadow_Score": "Non-propagating what-if score and rank simulation.",
    "Composite_Obs_Summary": "Phase-2 observation connection status and formal non-propagation summary.",
    "PT_Paired_Summary": "Normal-phosphate/PT paired evidence summary; shadow-only.",
    "PT_Discovery_Candidates": "Enzyme-rule-driven PT cleavage-bond discovery candidates; shadow-only.",
    "PT_Paired_Evidence": "Charge-resolved normal/PT paired exact-mass and physical-peak evidence; shadow-only.",
    "PT_Cross_Run_Runs": "Explicit cross-run manifest metadata, validation, acquisition, and runtime summary; shadow-only.",
    "PT_Cross_Run_Summary": "Charge-specific PT candidate recurrence statistics and qualitative evidence class; shadow-only.",
    "PT_Cross_Run_Neutral": "Neutral-structure recurrence aggregated across charge states; shadow-only.",
    "PT_Cross_Run_Pairs": "Cross-run normal-phosphate/PT counterpart states; shadow-only.",
    "PT_Cross_Run_Decoy": "Reference-only cross-run target/decoy recurrence audit; shadow-only.",
    "PT_Cross_Run_Target": "Targeted bond 10_11 H1-H4 run table; shadow-only.",
    "PT_Cross_Run_Detail": "Run/candidate MS1 peak, continuity, RT, isotope, and competition detail; shadow-only.",
    "PT_Cross_Run_Decoy_Detail": "Run/candidate shifted-m/z decoy matches; reference-only shadow audit.",
    "PT_Cross_Run_MS2_Detail": "Run-resolved precursor-compatible composite/PT MS2 shadow matches.",
    "Mod_Hypothesis_Summary": "Prior-separated, data-driven modification-position hypothesis interpretations; shadow-only.",
    "Mod_Hypothesis_Detail": "Run/fragment/peak evidence provenance for modification-position hypotheses; shadow-only.",
    "Mod_Hypothesis_Alternatives": "Alternative discovery explanations for each explicit position hypothesis; shadow-only.",
    "Mod_Hypothesis_Cross_Run": "Hypothesis-level recurrence projection from the explicit cross-run audit; shadow-only.",
    "Mod_Hypothesis_Invalid": "Schema, identity, component-domain, chemistry, and cross-source ID validation failures.",
    "Mod_Hypothesis_Structure_Map": "Canonical position-hypothesis to sample-structure mapping, including extra-state mismatches; shadow-only.",
    "Mod_Hypothesis_ID_Audit": "Cross-source hypothesis identifier consistency guard; shadow-only.",
    "Mod_Oxidation_Family": "Schema-derived precursor, thioamide, and defined oxidation-state comparison with origin cautions; shadow-only.",
    "PT_State_Search": "Four-state and discovery-state MS1 search detail; shadow-only.",
    "Composite_Obs_Invalid": "Invalid sample structure hypotheses excluded from observation matching.",
    "Known_Modification_Candidates": "Known modification candidates explaining fragment or intact mass shifts.",
    "Known_Modification_Summary": "Grouped summary of known modification candidates.",
    "Modification_Evidence_Summary": "Run-level counts for integrated modification evidence ranking.",
    "Modification_Evidence_Ranking": "Integrated evidence scores for prioritizing modification candidates.",
    "Modification_Ambiguity_Groups": "Position ambiguity groups for shared parent-fragment modification candidates.",
    "Modification_Position_Priors": "Diagnostic input-sequence position priors; no Sprinzl numbering is assumed.",
    "MS2_Biological_Plausibility": "Shadow biological plausibility, parent-base compatibility, and structural ambiguity review.",
    "MS2_Modification_Identity": "Shadow summary separating modified fragment ions, position localization, and structure/isomer resolution.",
    "RNase_MS2_Evidence_Summary": "Run-level RNase MS/MS evidence-synthesis status counts; shadow-only.",
    "RNase_MS2_Candidate_Evidence": "Candidate-level RNase MS/MS identity, localization, structure, and ambiguity synthesis; shadow-only.",
    "RNase_MS2_Peak_Evidence": "Physical-peak RNase MS/MS assignment and ambiguity provenance; shadow-only.",
    "RNase_MS2_Composite_Summary": "Run-level complete-structure RNase MS/MS evidence synthesis; shadow-only.",
    "RNase_MS2_Composite_Evidence": "Complete-structure candidate RNase MS/MS evidence synthesis; shadow-only.",
    "RNase_MS2_Composite_Peak_Eviden": "Excel alias for complete-structure physical-peak RNase MS/MS evidence; shadow-only.",
    "RNase_MS2_Standard_Composite_Su": "Excel alias for standard-to-composite RNase MS/MS crosswalk summary; shadow-only.",
    "RNase_MS2_Standard_Composite_Cr": "Excel alias for standard-to-composite RNase MS/MS crosswalk detail; shadow-only.",
    "RNase_MS2_Consensus_Summary": "Final standard/composite RNase MS/MS consensus summary; shadow-only.",
    "RNase_MS2_Consensus_Evidence": "Candidate-level final standard/composite RNase MS/MS consensus; shadow-only.",
    "SCIEX_Profile_Diagnostics": "SCIEX profile text header, filename expectation, numeric validity, spacing, and routing diagnostics; shadow-only.",
    "SCIEX_Profile_Input": "Parsed SCIEX profile points with neutral-mass and m/z coordinates kept separate; shadow-only.",
    "SCIEX_Intact_Peak_Diagnostics": "SCIEX intact neutral-mass peak detection diagnostics and parameter provenance; shadow-only.",
    "SCIEX_Intact_Detected_Peaks": "Sensitive-tier SCIEX intact neutral-mass peaks with strict flags, boundaries, and areas; shadow-only.",
    "SCIEX_Intact_Mass_Comparison": "Detected SCIEX intact peak proximity to unmodified theory and optional reconstructed intact masses; shadow-only, no identity assignment.",
    "SCIEX_Intact_Mass_Comp_Summary": "Run-level SCIEX intact mass-proximity counts and closest/strongest peaks; shadow-only.",
    "SCIEX_Input_Identity_Audit": "SCIEX filename identity tokens compared with configured RNA metadata; shadow-only and non-propagating.",
    "SCIEX_Delta_Mass_Clusters": "Span-bounded numerical clusters of SCIEX delta masses; shadow-only, no chemical assignment.",
    "SCIEX_Delta_Mass_Clust_Summary": "Run-level SCIEX delta-mass cluster and pair-spacing diagnostics; shadow-only.",
    "SCIEX_Delta_Mass_Relations": "Relevant duplicate-like, integer, isotope-like, and recurrent numerical spacing candidates; shadow-only.",
    "SCIEX_Spacing_Resolution": "Run-level SCIEX mass-grid and integer/isotope spacing distinguishability audit; shadow-only.",
    "SCIEX_Spacing_Resolution_Detail": "Per-multiple target-window overlap and grid-resolution diagnostics; shadow-only.",
    "SCIEX_Relation_Evidence": "Relation-level numerical fit, resolution, recurrence, and interpretation-gate diagnostics; shadow-only.",
    "SCIEX_Relation_Evidence_Summary": "Run-level SCIEX relation evidence-tier and interpretation-block summary; shadow-only.",
    "MS2_Identity_Peak_Assignments": "Candidate-match assignments annotated with candidate-crossing physical observed peak sharing.",
    "MS2_Unmatched_Ion_Audit": "Shadow reason audit for unmatched modified theoretical ions; formal matching is unchanged.",
    "MS2_Unmatched_Ion_Summary": "Candidate-level shadow summary of unmatched modified theoretical ion reasons.",
    "MS2_Unmatched_Ion_Diagnostics": "Run-level availability, threshold, tolerance, and unmatched-reason audit diagnostics.",
    "MS2_Ambiguous_Peak_Clusters": "Shadow theoretical-ion peak clusters for ambiguous nearby raw peaks.",
    "MS2_Ambiguous_Peak_Detail": "Per-physical-peak candidate sharing and theoretical-ion competition within ambiguous clusters.",
    "MS2_Ambiguity_Summary": "Candidate-level shadow summary of ambiguous nearby-peak cluster patterns.",
    "MS2_Zero_Intensity_Spectra": "Per-spectrum shadow trace of decoded, parsed, and annotation-input MS2 intensity states.",
    "MS2_Zero_Intensity_Detail": "Per-raw-peak shadow trace of zero intensity provenance and existing evidence use.",
    "MS2_Zero_Intensity_Summary": "Run-level shadow diagnostic of zero-intensity origin, impact, and nonzero simulation.",
    "MS2_Effective_Ambiguity": "Cluster-level four-stage shadow separation of raw, positive, formal-tolerance, and formal-match ambiguity.",
    "MS2_Effective_Ambig_Detail": "Physical-peak detail supporting effective ambiguity classification without rematching.",
    "MS2_Effective_Ambig_Summary": "Run-level effective ambiguity counts, definitions, severity, and non-application metadata.",
    "Biological_Prior_Diagnostics": "Run counts for diagnostic biological position prior evaluation.",
    "Biological_Context_Priorities": "Biological context settings used for generic candidate prioritization.",
    "Context_Supported_Candidates": "Ranking candidates receiving a user-configured biological context boost.",
    "P1_Summary": "Summary of P1 observed peak annotation results.",
    "P1_Theoretical_Structures": "P1 monomer and short oligonucleotide theoretical structure candidates.",
    "P1_Peak_Annotations": "Observed P1 peaks matched to theoretical structure candidates, retaining unmatched peaks.",
    "P1_Unmatched_Peaks": "Observed P1 peaks outside tolerance retained for unknown/adduct/phosphate review.",
    "MS2_Summary": "Run-level summary of MS2 c/y ion annotation.",
    "MS2_Spectra": "MS2 spectrum metadata, peak counts, and annotation status.",
    "MS2_Parent_Candidates": "Precursor m/z matches between MS2 spectra and theoretical digestion fragments.",
    "MS2_Theoretical_Ions": "Theoretical c/y RNA fragment ions generated from digestion fragments.",
    "MS2_Ion_Matches": "Matched observed MS2 peaks only; unmatched peaks are reported separately.",
    "MS2_Unmatched_Peaks": "Observed MS2 peaks outside tolerance retained for review.",
    "MS2_Fragment_Evidence": "Spectrum-parent fragment evidence summary from matched MS2 ions.",
    "MS2_Peak_Annotations": "Optional all-peak MS2 annotation sheet, disabled by default.",
    "Warnings": "Warnings and errors recorded during startup, loading, and analysis.",
    "Compound_Modification_Candidates": "Known modification plus an additional simple mass shift (e.g. ncm5s2U + S), not present as a single catalog entry.",
    "Compound_Modification_Summary": "Grouped summary of compound (known modification + extra shift) candidates.",    
}


def _flatten_dict(data: dict[str, Any], prefix: str = "") -> list[dict[str, Any]]:
    rows = []
    for key, value in data.items():
        full_key = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            rows.extend(_flatten_dict(value, full_key))
        else:
            rows.append({"Parameter": full_key, "Value": value})
    return rows


def _autosize_and_freeze(writer: pd.ExcelWriter) -> None:
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2" if worksheet.title == "Index" else "A4"
        for column_cells in worksheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 60))
            worksheet.column_dimensions[column].width = max(10, max_length + 2)


def _sheet_link(sheet_name: str, cell: str = "A1") -> str:
    return f"#'{sheet_name}'!{cell}"


def _coerce_to_frame(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value
    if isinstance(value, list):
        return pd.DataFrame(value)
    if isinstance(value, dict):
        return pd.DataFrame([value])
    return pd.DataFrame([{"Value": value}])


def _json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Enum):
        return _json_safe(value.value)
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, np.generic):
        return _json_safe(value.item())
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if hasattr(value, "to_dict") and callable(value.to_dict):
        return _json_safe(value.to_dict())
    if is_dataclass(value):
        return _json_safe(asdict(value))
    if isinstance(value, Mapping):
        return {
            str(key): _json_safe(item)
            for key, item in sorted(value.items(), key=lambda pair: str(pair[0]))
        }
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, (set, frozenset)):
        normalized = [_json_safe(item) for item in value]
        return sorted(
            normalized,
            key=lambda item: json.dumps(item, ensure_ascii=False, sort_keys=True, default=str),
        )
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _excel_safe_cell(value: Any) -> Any:
    normalized = _json_safe(value)
    if normalized is None:
        return ""
    if isinstance(normalized, (list, dict)):
        return json.dumps(
            normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
    return normalized


def _record_to_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if hasattr(value, "to_dict") and callable(value.to_dict):
        return dict(value.to_dict())
    if is_dataclass(value):
        return asdict(value)
    return dict(value)


def _record_rows(value: Any) -> list[dict[str, Any]]:
    if value is None:
        return []
    if isinstance(value, pd.DataFrame):
        return [dict(row) for row in value.to_dict("records")]
    if isinstance(value, (list, tuple)):
        return [_record_to_dict(item) for item in value]
    return [_record_to_dict(value)]


def _sciex_intact_excel_sheets(value: Any) -> dict[str, pd.DataFrame]:
    if value is None:
        return {}

    wrapper = value if isinstance(value, Mapping) else {}
    result = wrapper.get("result") if "result" in wrapper else value
    source_file = wrapper.get("source_file", "")
    source_file_name = wrapper.get("source_file_name", "")

    if isinstance(result, Mapping) and any(
        key in result for key in ("diagnostics", "peaks", "parameter_provenance", "warnings")
    ):
        diagnostics_value = result.get("diagnostics")
        peaks_value = result.get("peaks")
        provenance_value = result.get("parameter_provenance")
        warnings_value = result.get("warnings")
    else:
        diagnostics_value = (
            result.diagnostics_row() if hasattr(result, "diagnostics_row") else getattr(result, "diagnostics", None)
        )
        peaks_value = result.peak_rows() if hasattr(result, "peak_rows") else getattr(result, "peaks", None)
        provenance_value = (
            result.provenance_rows()
            if hasattr(result, "provenance_rows")
            else getattr(result, "parameter_provenance", None)
        )
        warnings_value = getattr(result, "warnings", None)

    diagnostics_rows = _record_rows(diagnostics_value)
    peak_rows = _record_rows(peaks_value)
    provenance_rows = _record_rows(provenance_value)
    provenance_json = _excel_safe_cell(provenance_rows) if provenance_rows else ""
    warnings_cell = _excel_safe_cell(warnings_value) if warnings_value else ""
    source_text = _excel_safe_cell(source_file) if source_file else ""
    source_name_text = _excel_safe_cell(source_file_name) if source_file_name else ""
    if source_text and not source_name_text:
        source_name_text = Path(str(source_text)).name

    formal_columns = (
        "SCIEX_Intact_Peak_Detection_Applied_To_Formal_Score",
        "SCIEX_Intact_Peak_Detection_Applied_To_Ranking",
        "SCIEX_Intact_Peak_Detection_Applied_To_Candidate_Filtering",
    )
    safe_diagnostics = []
    for source_row in diagnostics_rows:
        row = dict(source_row)
        if source_text:
            row["Source_File"] = source_text
        if source_name_text:
            row["Source_File_Name"] = source_name_text
        elif row.get("Source_File") and not row.get("Source_File_Name"):
            row["Source_File_Name"] = Path(str(row["Source_File"])).name
        row["Parameter_Provenance_JSON"] = provenance_json
        row["Detection_Warnings"] = warnings_cell
        for column in formal_columns:
            row[column] = False
        safe_diagnostics.append({key: _excel_safe_cell(item) for key, item in row.items()})

    safe_peaks = []
    for source_row in peak_rows:
        row = dict(source_row)
        if source_text:
            row["Source_File"] = source_text
        row["Molecular_Identity_Assigned"] = False
        for column in formal_columns:
            row[column] = False
        safe_peaks.append({key: _excel_safe_cell(item) for key, item in row.items()})

    sheets: dict[str, pd.DataFrame] = {}
    if diagnostics_rows:
        sheets[SCIEX_INTACT_DIAGNOSTIC_SHEET] = pd.DataFrame(
            safe_diagnostics, columns=SCIEX_INTACT_DIAGNOSTIC_COLUMNS
        )
        sheets[SCIEX_INTACT_PEAK_SHEET] = pd.DataFrame(
            safe_peaks, columns=SCIEX_INTACT_PEAK_COLUMNS
        )
    elif peak_rows:
        sheets[SCIEX_INTACT_PEAK_SHEET] = pd.DataFrame(
            safe_peaks, columns=SCIEX_INTACT_PEAK_COLUMNS
        )
    return sheets


def _sciex_delta_cluster_excel_sheets(value: Any) -> dict[str, pd.DataFrame]:
    if value is None:
        return {}
    if isinstance(value, Mapping):
        clusters = value.get("cluster_rows", value.get("clusters", []))
        summaries = value.get("summary_rows", value.get("summaries", []))
        relations = value.get("relation_rows", value.get("relations", []))
    else:
        clusters = value.clusters() if hasattr(value, "clusters") else getattr(value, "cluster_rows", [])
        summaries = value.summaries() if hasattr(value, "summaries") else getattr(value, "summary_rows", [])
        relations = value.relations() if hasattr(value, "relations") else getattr(value, "relation_rows", [])
    safe_clusters = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in _record_rows(clusters)
    ]
    safe_summaries = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in _record_rows(summaries)
    ]
    safe_relations = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in _record_rows(relations)
    ]
    sheets = {}
    if safe_summaries:
        sheets[SCIEX_DELTA_CLUSTER_SUMMARY_SHEET] = pd.DataFrame(
            safe_summaries, columns=SCIEX_DELTA_CLUSTER_SUMMARY_COLUMNS,
        )
        sheets[SCIEX_DELTA_CLUSTER_SHEET] = pd.DataFrame(
            safe_clusters, columns=SCIEX_DELTA_CLUSTER_COLUMNS,
        )
        sheets[SCIEX_DELTA_RELATION_SHEET] = pd.DataFrame(
            safe_relations, columns=SCIEX_DELTA_RELATION_COLUMNS,
        )
    return sheets

def _sciex_spacing_resolution_excel_sheets(value: Any) -> dict[str, pd.DataFrame]:
    if value is None:
        return {}
    if isinstance(value, Mapping):
        summaries = value.get("summary_rows", value.get("summaries", []))
        details = value.get("detail_rows", value.get("details", []))
    else:
        summaries = value.summaries() if hasattr(value, "summaries") else getattr(value, "summary_rows", [])
        details = value.details() if hasattr(value, "details") else getattr(value, "detail_rows", [])
    safe_summaries = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in _record_rows(summaries)
    ]
    safe_details = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in _record_rows(details)
    ]
    if not safe_summaries:
        return {}
    return {
        SCIEX_SPACING_RESOLUTION_SUMMARY_SHEET: pd.DataFrame(
            safe_summaries, columns=SCIEX_SPACING_RESOLUTION_SUMMARY_COLUMNS,
        ),
        SCIEX_SPACING_RESOLUTION_DETAIL_SHEET: pd.DataFrame(
            safe_details, columns=SCIEX_SPACING_RESOLUTION_DETAIL_COLUMNS,
        ),
    }


def _sciex_relation_evidence_excel_sheets(value: Any) -> dict[str, pd.DataFrame]:
    if value is None:
        return {}
    if isinstance(value, Mapping):
        details = value.get("detail_rows", value.get("details", []))
        summaries = value.get("summary_rows", value.get("summaries", []))
    else:
        details = value.details() if hasattr(value, "details") else getattr(value, "detail_rows", [])
        summaries = value.summaries() if hasattr(value, "summaries") else getattr(value, "summary_rows", [])
    safe_details = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in _record_rows(details)
    ]
    safe_summaries = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in _record_rows(summaries)
    ]
    if not safe_summaries:
        return {}
    return {
        SCIEX_RELATION_EVIDENCE_DETAIL_SHEET: pd.DataFrame(
            safe_details, columns=SCIEX_RELATION_EVIDENCE_DETAIL_COLUMNS,
        ),
        SCIEX_RELATION_EVIDENCE_SUMMARY_SHEET: pd.DataFrame(
            safe_summaries, columns=SCIEX_RELATION_EVIDENCE_SUMMARY_COLUMNS,
        ),
    }


def _sciex_identity_audit_excel_sheets(value: Any) -> dict[str, pd.DataFrame]:
    if value is None:
        return {}
    if hasattr(value, "row"):
        rows = [value.row()]
    elif isinstance(value, Mapping) and "values" in value:
        rows = _record_rows(value.get("values"))
    else:
        rows = _record_rows(value)
    safe_rows = [
        {key: _excel_safe_cell(item) for key, item in row.items()}
        for row in rows
    ]
    if not safe_rows:
        return {}
    return {
        SCIEX_IDENTITY_AUDIT_SHEET: pd.DataFrame(
            safe_rows, columns=SCIEX_IDENTITY_AUDIT_COLUMNS,
        )
    }


def _sciex_cross_layer_excel_sheets(value: Any | None) -> dict[str, pd.DataFrame]:
    if value is None:
        return {}
    records = audit_cross_layer_optional_result(value)

    def _format(rows: list[dict[str, Any]]) -> pd.DataFrame:
        if not rows:
            return pd.DataFrame()
        safe_rows = [
            {key: _excel_safe_cell(item) for key, item in sorted(row.items())}
            for row in _record_rows(rows)
        ]
        return pd.DataFrame(safe_rows)

    return {
        "XL_Nodes": _format(records.get("node_records", [])),
        "XL_Edges": _format(records.get("edge_records", [])),
        "XL_Hypotheses": _format(records.get("hypothesis_records", [])),
        "XL_Layer_Summary": _format(records.get("layer_summary_records", [])),
        "XL_Consensus": _format(records.get("consensus_records", [])),
        "XL_Next_Evidence": _format(records.get("next_evidence_records", [])),
    }


def _sciex_mass_comparison_excel_sheets(value: Any) -> dict[str, pd.DataFrame]:
    if value is None:
        return {}
    if isinstance(value, Mapping):
        details = value.get("detail_rows", value.get("details", []))
        summaries = value.get("summary_rows", value.get("summaries", []))
    else:
        details = value.details() if hasattr(value, "details") else getattr(value, "detail_rows", [])
        summaries = value.summaries() if hasattr(value, "summaries") else getattr(value, "summary_rows", [])
    safe_details = [{key: _excel_safe_cell(item) for key, item in row.items()} for row in _record_rows(details)]
    safe_summaries = [{key: _excel_safe_cell(item) for key, item in row.items()} for row in _record_rows(summaries)]
    sheets = {}
    if safe_summaries:
        sheets[SCIEX_MASS_COMPARISON_SUMMARY_SHEET] = pd.DataFrame(safe_summaries, columns=SCIEX_MASS_COMPARISON_SUMMARY_COLUMNS)
    if safe_details:
        sheets[SCIEX_MASS_COMPARISON_DETAIL_SHEET] = pd.DataFrame(safe_details, columns=SCIEX_MASS_COMPARISON_DETAIL_COLUMNS)
    return sheets

def _analysis_mode(config) -> str:
    workflow_mode = str((getattr(config, "analysis", {}) or {}).get("mode") or "full")
    if workflow_mode == "intact_only":
        return "intact_only"
    reconstruction_enabled = _as_bool((config.reconstruction or {}).get("enabled"), True)
    digestion_enabled = _as_bool((config.digestion or {}).get("enabled"), True)
    if reconstruction_enabled and digestion_enabled:
        return "Intact + digested fragment analysis"
    if reconstruction_enabled:
        return "Intact reconstruction only"
    if digestion_enabled:
        return "Digested fragment MS1 mapping"
    return "No active mass analysis"


def _add_index_and_backlinks(writer: pd.ExcelWriter, sheet_names: list[str]) -> None:
    workbook = writer.book
    index_sheet = workbook["Index"]
    for row_index, sheet_name in enumerate(sheet_names, start=2):
        link_cell = index_sheet.cell(row=row_index, column=1)
        link_cell.value = sheet_name
        link_cell.hyperlink = _sheet_link(sheet_name, "A1")
        link_cell.style = "Hyperlink"

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        worksheet["A1"] = "← Back to Index"
        worksheet["A1"].hyperlink = _sheet_link("Index", "A1")
        worksheet["A1"].style = "Hyperlink"

_CONCLUSION_CATEGORIES = {FORMAL_CORE, FORMAL_OPTIONAL}
_SEPARATOR_SHEET_NAME = "─── 詳細シート ───"


def _reorder_sheets_conclusion_then_detail(
    sheets: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    """Reorder report sheets so formal/conclusion sheets come first, followed
    by a blank separator sheet, then shadow-audit detail sheets. Run_summary
    is always pinned first. Unclassified sheets are treated as detail
    (conservative default)."""
    conclusion: list[str] = []
    detail: list[str] = []
    for name in sheets:
        if name == "Run_summary":
            continue
        if sheet_category(name) in _CONCLUSION_CATEGORIES:
            conclusion.append(name)
        else:
            detail.append(name)

    ordered: dict[str, pd.DataFrame] = {}
    if "Run_summary" in sheets:
        ordered["Run_summary"] = sheets["Run_summary"]
    for name in conclusion:
        ordered[name] = sheets[name]
    if detail:
        ordered[_SEPARATOR_SHEET_NAME] = pd.DataFrame()
    for name in detail:
        ordered[name] = sheets[name]
    return ordered


def _extract_long_cells_to_word(
    sheets: dict[str, pd.DataFrame],
    collector: WordExportCollector,
) -> dict[str, pd.DataFrame]:
    """Replace any cell spanning more than one sentence with a short bookmark
    label (e.g. "P12"), collecting the original text into `collector` for
    the companion Word document."""
    processed: dict[str, pd.DataFrame] = {}
    for sheet_name, frame in sheets.items():
        if frame.empty or sheet_name in SHEETS_EXCLUDED_FROM_WORD_EXPORT:
            processed[sheet_name] = frame
            continue
        frame = frame.copy()
        for col_position, column in enumerate(frame.columns):
            if frame.iloc[:, col_position].dtype != object:
                continue
            for row_position in range(len(frame)):
                value = frame.iat[row_position, col_position]
                if _is_multi_sentence(value):
                    bookmark = collector.add(
                        sheet_name, str(column), col_position, row_position + 1, value
                    )
                    frame.iat[row_position, col_position] = bookmark
        processed[sheet_name] = frame
    return processed


def _add_word_appendix_hyperlinks(
    writer: pd.ExcelWriter,
    collector: WordExportCollector,
    word_appendix_filename: str,
) -> None:
    workbook = writer.book
    for item in collector.items:
        if item.sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[item.sheet_name]
        excel_row = item.row_index + 3  # header at row 3 (startrow=2); data starts row 4
        excel_col = item.column_position + 1
        cell = worksheet.cell(row=excel_row, column=excel_col)
        cell.hyperlink = f"{word_appendix_filename}#{item.bookmark}"
        cell.style = "Hyperlink"


def _fragment_rows(theoretical_fragments: list[Any]) -> list[dict[str, Any]]:
    rows = []
    for item in theoretical_fragments:
        raw = asdict(item) if is_dataclass(item) else dict(item)
        fragment_warnings = raw.get("warnings", [])
        if isinstance(fragment_warnings, list):
            fragment_warnings = "; ".join(map(str, fragment_warnings))
        rows.append(
            {
                "Fragment_ID": raw.get("fragment_id"),
                "Target_ID": raw.get("target_id"),
                "Sequence": raw.get("sequence"),
                "Length": len(raw.get("sequence") or ""),
                "Start": raw.get("start"),
                "End": raw.get("end"),
                "Enzyme": raw.get("enzyme"),
                "Missed_Cleavages": raw.get("missed_cleavages"),
                "Terminal_Form": raw.get("terminal_form"),
                "Unmodified_Mass": raw.get("unmodified_mass"),
                "Warnings": fragment_warnings,
            }
        )
    return rows


def _match_raw(item: Any) -> dict[str, Any]:
    return asdict(item) if is_dataclass(item) else dict(item)


def _normalize_filter_values(values: Any) -> set[str]:
    if values is None:
        return set()
    if isinstance(values, str):
        return {values.lower()}
    return {str(value).lower() for value in values}


def _fragment_ms1_match_rows(fragment_ms1_matches: list[Any], include_length: bool = False) -> list[dict[str, Any]]:
    rows = []
    for item in fragment_ms1_matches:
        raw = _match_raw(item)
        match_warnings = raw.get("warnings", [])
        if isinstance(match_warnings, list):
            match_warnings = "; ".join(map(str, match_warnings))
        row = {
            "Match_ID": raw.get("match_id"),
            "Fragment_ID": raw.get("fragment_id"),
            "Target_ID": raw.get("target_id"),
            "Sequence": raw.get("sequence"),
            "Start": raw.get("start"),
            "End": raw.get("end"),
            "Enzyme": raw.get("enzyme"),
            "Missed_Cleavages": raw.get("missed_cleavages"),
            "Terminal_Form": raw.get("terminal_form"),
            "Fragment_Mass": raw.get("fragment_mass"),
            "Charge": raw.get("charge"),
            "Theoretical_mz": raw.get("theoretical_mz"),
            "Observed_mz": raw.get("observed_mz"),
            "Mass_Error_Da": raw.get("mass_error_da"),
            "Mass_Error_ppm": raw.get("mass_error_ppm"),
            "Intensity": raw.get("intensity"),
            "RT": raw.get("rt"),
            "Scan_ID": raw.get("scan_id"),
            "Peak_Tier": raw.get("peak_tier"),
            "Confidence": raw.get("confidence"),
            "Warnings": match_warnings,
        }
        if include_length:
            row["Length"] = len(raw.get("sequence") or "")
        rows.append(row)
    return rows


def _filter_fragment_ms1_matches(fragment_ms1_matches: list[Any], mapping_config: dict[str, Any]) -> list[Any]:
    min_length = _as_positive_int(mapping_config.get("min_fragment_length_for_filtered"), 3)
    allowed_tiers = _normalize_filter_values(mapping_config.get("filtered_peak_tiers", ["Major", "Minor"]))
    allowed_confidence = _normalize_filter_values(mapping_config.get("filtered_confidence", ["High", "Medium"]))
    filtered = []
    for item in fragment_ms1_matches:
        raw = _match_raw(item)
        if len(raw.get("sequence") or "") < min_length:
            continue
        if allowed_tiers and str(raw.get("peak_tier") or "").lower() not in allowed_tiers:
            continue
        if allowed_confidence and str(raw.get("confidence") or "").lower() not in allowed_confidence:
            continue
        filtered.append(item)
    return filtered


def _confidence_rank(value: Any) -> int:
    return {"high": 3, "medium": 2, "low": 1}.get(str(value or "").lower(), 0)


def _peak_tier_rank(value: Any) -> int:
    return {"major": 3, "minor": 2, "trace": 1}.get(str(value or "").lower(), 0)


def _best_match_sort_key(item: Any) -> tuple[int, int, float, float]:
    raw = _match_raw(item)
    return (
        -_confidence_rank(raw.get("confidence")),
        -_peak_tier_rank(raw.get("peak_tier")),
        abs(float(raw.get("mass_error_ppm") or 0.0)),
        -float(raw.get("intensity") or 0.0),
    )


def _fragment_ms1_summary_rows(fragment_ms1_matches: list[Any], mapping_config: dict[str, Any]) -> list[dict[str, Any]]:
    group_key = str(mapping_config.get("summary_best_match_by", "fragment_id") or "fragment_id")
    if group_key != "fragment_id":
        group_key = "fragment_id"

    grouped: dict[str, list[Any]] = {}
    for item in fragment_ms1_matches:
        raw = _match_raw(item)
        key = str(raw.get(group_key) or "")
        if not key:
            continue
        grouped.setdefault(key, []).append(item)

    rows = []
    for fragment_id, matches in grouped.items():
        best = min(matches, key=_best_match_sort_key)
        best_raw = _match_raw(best)
        tier_counts = {"Major": 0, "Minor": 0, "Trace": 0}
        confidence_counts = {"High": 0, "Medium": 0, "Low": 0}
        for item in matches:
            raw = _match_raw(item)
            tier = str(raw.get("peak_tier") or "")
            confidence = str(raw.get("confidence") or "")
            if tier in tier_counts:
                tier_counts[tier] += 1
            if confidence in confidence_counts:
                confidence_counts[confidence] += 1
        rows.append(
            {
                "Fragment_ID": fragment_id,
                "Target_ID": best_raw.get("target_id"),
                "Sequence": best_raw.get("sequence"),
                "Length": len(best_raw.get("sequence") or ""),
                "Start": best_raw.get("start"),
                "End": best_raw.get("end"),
                "Enzyme": best_raw.get("enzyme"),
                "Missed_Cleavages": best_raw.get("missed_cleavages"),
                "Terminal_Form": best_raw.get("terminal_form"),
                "Best_Charge": best_raw.get("charge"),
                "Best_Theoretical_mz": best_raw.get("theoretical_mz"),
                "Best_Observed_mz": best_raw.get("observed_mz"),
                "Best_Mass_Error_ppm": best_raw.get("mass_error_ppm"),
                "Best_Intensity": best_raw.get("intensity"),
                "Best_RT": best_raw.get("rt"),
                "Best_Peak_Tier": best_raw.get("peak_tier"),
                "Best_Confidence": best_raw.get("confidence"),
                "Match_Count": len(matches),
                "Major_Count": tier_counts["Major"],
                "Minor_Count": tier_counts["Minor"],
                "Trace_Count": tier_counts["Trace"],
                "High_Count": confidence_counts["High"],
                "Medium_Count": confidence_counts["Medium"],
                "Low_Count": confidence_counts["Low"],
            }
        )
    return sorted(rows, key=lambda row: (row["Start"] or 0, row["End"] or 0, row["Fragment_ID"]))


def _as_bool(value: Any, default: bool = True) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _as_positive_int(value: Any, default: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _append_excel_warning(
    warnings: list[dict[str, Any]],
    sheet_name: str,
    original_rows: int,
    written_rows: int,
) -> None:
    warnings.append(
        {
            "Timestamp": datetime.now().isoformat(timespec="seconds"),
            "Level": "WARNING",
            "Source": "excel_report",
            "Message": "Excel sheet was truncated because it exceeded max_excel_rows_per_sheet.",
            "Context": {"sheet": sheet_name, "original_rows": original_rows, "written_rows": written_rows},
        }
    )


def _truncate_frame_if_needed(
    sheet_name: str,
    frame: pd.DataFrame,
    max_rows: int,
    truncate_large_sheets: bool,
    warnings: list[dict[str, Any]],
    truncations: list[dict[str, Any]],
) -> pd.DataFrame:
    original_rows = len(frame)
    safe_limit = min(max_rows, EXCEL_DATA_ROW_LIMIT)
    if original_rows <= safe_limit:
        return frame

    if truncate_large_sheets:
        written_rows = safe_limit
    else:
        written_rows = EXCEL_DATA_ROW_LIMIT
    written_rows = min(written_rows, original_rows)
    _append_excel_warning(warnings, sheet_name, original_rows, written_rows)
    truncations.append({"sheet": sheet_name, "original_rows": original_rows, "written_rows": written_rows})
    return frame.head(written_rows).copy()


def _truncation_summary(truncations: list[dict[str, Any]]) -> str:
    if not truncations:
        return "None"
    return "; ".join(
        f"{item['sheet']}: {item['original_rows']} -> {item['written_rows']}" for item in truncations
    )


def write_excel_report(
    output_dir: str | Path,
    config,
    diagnostics: dict[str, Any],
    intact_results: list[Any],
    charge_state_peaks: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
    modifications: list[Any] | None = None,
    rule_set: dict[str, Any] | None = None,
    pathways: list[dict[str, Any]] | None = None,
    theoretical_fragments: list[Any] | None = None,
    fragment_ms1_matches: list[Any] | None = None,
    known_modification_candidates: list[dict[str, Any]] | None = None,
    known_modification_summary: list[dict[str, Any]] | None = None,
    unknown_modification_candidates: list[dict[str, Any]] | None = None,
    unknown_modification_summary: list[dict[str, Any]] | None = None,
    compound_modification_candidates: list[dict[str, Any]] | None = None,
    compound_modification_summary: list[dict[str, Any]] | None = None,
    optional_results: dict[str, Any] | None = None,
    audit_policy: AuditPolicy | None = None,
) -> Path:
    audit_policy = audit_policy or AuditPolicy.from_level("full")
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    report_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_path = out_dir / f"RNA_MassHunter_MVP5_{report_timestamp}.xlsx"
    word_appendix_path = out_dir / f"RNA_MassHunter_MVP5_{report_timestamp}_appendix.docx"

    reporting = config.reporting or {}
    max_excel_rows = _as_positive_int(reporting.get("max_excel_rows_per_sheet"), 100000)
    truncate_large_sheets = _as_bool(reporting.get("truncate_large_sheets"), True)
    max_charge_state_peak_rows = _as_positive_int(
        reporting.get("max_charge_state_peak_rows", config.reconstruction.get("max_charge_state_peak_rows")),
        max_excel_rows,
    )
    truncations: list[dict[str, Any]] = []
    reconstruction_enabled = _as_bool(config.reconstruction.get("enabled"), True)
    intact_qc_rows, intact_diagnostic_rows = build_intact_reconstruction_qc(
        intact_results,
        charge_state_peaks,
        config.reconstruction or {},
        reconstruction_enabled=reconstruction_enabled,
    )
    intact_group_rows = build_intact_envelope_group_rows(intact_qc_rows)
    intact_competition_group_rows = build_intact_competition_group_rows(intact_qc_rows)
    intact_competition_score_rows = build_intact_competition_score_rows(intact_qc_rows)
    assignment_dry_run_rows = build_assignment_dry_run_rows(intact_qc_rows)
    assignment_dry_run_summary_rows = build_assignment_dry_run_summary_rows(intact_qc_rows)
    assignment_sensitivity_rows = build_assignment_sensitivity_rows(config.reconstruction or {})
    assignment_stability_rows = build_assignment_stability_rows(intact_qc_rows)
    assignment_candidate_audit_rows = build_assignment_candidate_audit_rows(config.reconstruction or {})
    assignment_ambiguous_rows = build_assignment_ambiguous_rows(intact_qc_rows)
    preassignment_comparison_rows = build_preassignment_comparison_rows(intact_qc_rows)
    intact_comparison_rows = build_intact_comparison_candidate_rows(intact_qc_rows)
    intact_target_review_rows = build_target_review_candidate_rows(intact_qc_rows)
    reconstructed_spectrum_rows = build_reconstructed_mass_spectrum_rows(intact_qc_rows, config.reconstruction or {})
    rt_engine_qc_summary_rows = build_rt_engine_qc_summary_rows(intact_diagnostic_rows)

    intact_rows = []
    for item in intact_results:
        raw = asdict(item) if is_dataclass(item) else dict(item)
        intact_rows.append(
            {
                "Cluster_ID": raw.get("cluster_id"),
                "Reconstructed_Mass": raw.get("observed_mass"),
                "Observed_Mass": raw.get("observed_mass"),
                "In_Neutral_Mass_Search_Range": raw.get("in_neutral_mass_search_range"),
                "Neutral_Mass_Search_Min_Da": raw.get("neutral_mass_search_min_da"),
                "Neutral_Mass_Search_Max_Da": raw.get("neutral_mass_search_max_da"),
                "Neutral_Mass_Range_Status": raw.get("neutral_mass_range_status"),
                "In_Target_Review_Mass_Range": raw.get("in_target_review_mass_range"),
                "Target_Review_Mass_Range_Status": raw.get("target_review_mass_range_status"),
                "Target_Review_Priority": raw.get("target_review_priority"),
                "Envelope_QC_Eligible": raw.get("envelope_qc_eligible"),
                "Intact_Review_Eligible": raw.get("intact_review_eligible"),
                "Intact_Strict_Eligible": raw.get("intact_strict_eligible"),
                "Intact_Envelope_QC_Score": raw.get("intact_envelope_qc_score"),
                "Intact_Envelope_QC_Rank": raw.get("intact_envelope_qc_rank"),
                "Strict_Eligible_Rank": raw.get("strict_eligible_rank"),
                "Review_Eligible_Rank": raw.get("review_eligible_rank"),
                "Dominant_Intact_Envelope_Flag": raw.get("dominant_intact_envelope_flag"),
                "Supporting_Peak_IDs": raw.get("supporting_peak_ids"),
                "Supporting_Peak_Count": raw.get("supporting_peak_count"),
                "Supporting_Scan_IDs": raw.get("supporting_scan_ids"),
                "Supporting_RT_Values": raw.get("supporting_rt_values"),
                "Supporting_Charge_States": raw.get("supporting_charge_states"),
                "Exact_Peak_Set_Key": raw.get("exact_peak_set_key"),
                "Exact_Duplicate_Group_ID": raw.get("exact_duplicate_group_id"),
                "Exact_Duplicate_Count": raw.get("exact_duplicate_count"),
                "Is_Exact_Duplicate_Representative": raw.get("is_exact_duplicate_representative"),
                "Intact_Envelope_Group_ID": raw.get("intact_envelope_group_id"),
                "Envelope_Group_Size": raw.get("envelope_group_size"),
                "Group_Representative": raw.get("group_representative"),
                "Group_Ambiguity_Status": raw.get("group_ambiguity_status"),
                "Comparison_Representative": raw.get("comparison_representative"),
                "Comparison_Representative_Reason": raw.get("comparison_representative_reason"),
                "Comparison_Representative_Rank": raw.get("comparison_representative_rank"),
                "Excluded_From_Comparison_Reason": raw.get("excluded_from_comparison_reason"),
                "Target_Review_Group_Representative": raw.get("target_review_group_representative"),
                "Target_Review_Rank": raw.get("target_review_rank"),
                "Dominant_Target_Review_Eligible_Flag": raw.get("dominant_target_review_eligible_flag"),
                "Reconstruction_Status": raw.get("reconstruction_status"),
                "Reconstruction_Confidence": raw.get("reconstruction_confidence"),
                "Reconstruction_Engine": raw.get("reconstruction_engine"),
                "RT_Window_ID": raw.get("rt_window_id"),
                "RT_Window_Start_Min": raw.get("rt_window_start_min"),
                "RT_Window_End_Min": raw.get("rt_window_end_min"),
                "RT_Window_Center_Min": raw.get("rt_window_center_min"),
                "Num_MS1_Scans_In_Window": raw.get("num_ms1_scans_in_window"),
                "Peak_Aggregation_Method": raw.get("peak_aggregation_method"),
                "Anchor_MZ": raw.get("anchor_mz"),
                "Anchor_Charge": raw.get("anchor_charge"),
                "Predicted_Charge_States": raw.get("predicted_charge_states"),
                "Observed_Charge_States": raw.get("observed_charge_states"),
                "Missing_Charge_States": raw.get("missing_charge_states"),
                "Missing_Charge_Predicted_MZ": raw.get("missing_charge_predicted_mz"),
                "Num_Predicted_Charges": raw.get("num_predicted_charges"),
                "Num_Observed_Charges": raw.get("num_observed_charges"),
                "Charge_Coverage_Fraction": raw.get("charge_coverage_fraction"),
                "Consecutive_Charge_Run_Length": raw.get("consecutive_charge_run_length"),
                "Longest_Consecutive_Charge_Run": raw.get("longest_consecutive_charge_run"),
                "Charge_Gap_Count": raw.get("charge_gap_count"),
                "Charge_Continuity_Fraction": raw.get("charge_continuity_fraction"),
                "Peak_Usage_Count": raw.get("peak_usage_count"),
                "Shared_Peak_Count": raw.get("shared_peak_count"),
                "Shared_Peak_Fraction": raw.get("shared_peak_fraction"),
                "Local_Window_Max_Intensity": raw.get("local_window_max_intensity"),
                "Local_Relative_Peak_Intensity_Percent": raw.get("local_relative_peak_intensity_percent"),
                "Local_Envelope_Relative_Intensity_Percent": raw.get("local_envelope_relative_intensity_percent"),
                "Neutral_Mass_Estimator": raw.get("neutral_mass_estimator"),
                "Neutral_Mass_Unweighted_Mean": raw.get("neutral_mass_unweighted_mean"),
                "Neutral_Mass_Weighted_Mean": raw.get("neutral_mass_weighted_mean"),
                "Neutral_Mass_Median": raw.get("neutral_mass_median"),
                "Envelope_Internal_Error_Max_ppm": raw.get("envelope_internal_error_max_ppm"),
                "Envelope_Internal_Error_Mean_ppm": raw.get("envelope_internal_error_mean_ppm"),
                "Envelope_Internal_Error_Median_ppm": raw.get("envelope_internal_error_median_ppm"),
                "Source_RT_Window_IDs": raw.get("source_rt_window_ids"),
                "Num_Source_RT_Windows": raw.get("num_source_rt_windows"),
                "Merged_Across_RT_Windows": raw.get("merged_across_rt_windows"),
                "Extended_Lower_Charges_Evaluated": raw.get("extended_lower_charges_evaluated"),
                "Extended_Upper_Charges_Evaluated": raw.get("extended_upper_charges_evaluated"),
                "Extended_Charges_Detected": raw.get("extended_charges_detected"),
                "Extended_Weak_Charges_Detected": raw.get("extended_weak_charges_detected"),
                "Extended_Charges_Not_Detected": raw.get("extended_charges_not_detected"),
                "Charge_Extension_Improved_Envelope": raw.get("charge_extension_improved_envelope"),
                "Original_Charge_States": raw.get("original_charge_states"),
                "Final_Charge_States": raw.get("final_charge_states"),
                "Split_Envelope_Group_ID": raw.get("split_envelope_group_id"),
                "Split_Envelope_Member_Count": raw.get("split_envelope_member_count"),
                "Split_Envelope_Merged": raw.get("split_envelope_merged"),
                "Charge_Gaps_Before_Merge": raw.get("charge_gaps_before_merge"),
                "Charge_Gaps_After_Merge": raw.get("charge_gaps_after_merge"),
                "Max_Peak_Usage_Count": raw.get("max_peak_usage_count"),
                "Mean_Peak_Usage_Count": raw.get("mean_peak_usage_count"),
                "Num_Highly_Shared_Peaks": raw.get("num_highly_shared_peaks"),
                "Highly_Shared_Peak_Fraction": raw.get("highly_shared_peak_fraction"),
                "Competing_Candidate_Count": raw.get("competing_candidate_count"),
                "Peak_Sharing_Status": raw.get("peak_sharing_status"),
                "Competing_Envelope_Group_ID": raw.get("competing_envelope_group_id"),
                "Competing_Envelope_Group_Size": raw.get("competing_envelope_group_size"),
                "Shared_Peak_Competitor_Count": raw.get("shared_peak_competitor_count"),
                "Maximum_Shared_Peak_Fraction": raw.get("maximum_shared_peak_fraction"),
                "Mean_Shared_Peak_Fraction": raw.get("mean_shared_peak_fraction"),
                "Competitor_Cluster_IDs": raw.get("competitor_cluster_ids"),
                "Is_Noncompeting_Candidate": raw.get("is_noncompeting_candidate"),
                "Envelope_Evidence_Score": raw.get("envelope_evidence_score"),
                "Evidence_Score_Rank_In_Competition": raw.get("evidence_score_rank_in_competition"),
                "Evidence_Score_Components": raw.get("evidence_score_components"),
                "Evidence_Score_Penalties": raw.get("evidence_score_penalties"),
                "Evidence_Score_Config_Version": raw.get("evidence_score_config_version"),
                "Direct_Competitor_Count": raw.get("direct_competitor_count"),
                "Direct_Competitor_Cluster_IDs": raw.get("direct_competitor_cluster_ids"),
                "Direct_Shared_Peak_Count_Max": raw.get("direct_shared_peak_count_max"),
                "Direct_Shared_Peak_Fraction_Max": raw.get("direct_shared_peak_fraction_max"),
                "Competition_Component_Size": raw.get("competition_component_size"),
                "Dry_Run_Assignment_Status": raw.get("dry_run_assignment_status"),
                "Dry_Run_Selected": raw.get("dry_run_selected"),
                "Dry_Run_Selection_Order": raw.get("dry_run_selection_order"),
                "Supporting_Peak_Count_Before_Assignment": raw.get("supporting_peak_count_before_assignment"),
                "Independent_Supporting_Peak_Count": raw.get("independent_supporting_peak_count"),
                "Independent_Supporting_Peak_Fraction": raw.get("independent_supporting_peak_fraction"),
                "Supporting_Charge_Count_Before_Assignment": raw.get("supporting_charge_count_before_assignment"),
                "Independent_Charge_State_Count": raw.get("independent_charge_state_count"),
                "Peaks_Already_Assigned_Count": raw.get("peaks_already_assigned_count"),
                "Charges_Already_Assigned_Count": raw.get("charges_already_assigned_count"),
                "Excluded_By_Cluster_ID": raw.get("excluded_by_cluster_id"),
                "Dry_Run_Exclusion_Reason": raw.get("dry_run_exclusion_reason"),
                "Score_Margin_To_Excluding_Candidate": raw.get("score_margin_to_excluding_candidate"),
                "Close_Score_Ambiguity": raw.get("close_score_ambiguity"),
                "Assignment_Confidence": raw.get("assignment_confidence"),
                "Shared_Observed_Peak_Count": raw.get("shared_observed_peak_count"),
                "Shared_Peak_Charge_Assignment_Count": raw.get("shared_peak_charge_assignment_count"),
                "Independent_Observed_Peak_Count": raw.get("independent_observed_peak_count"),
                "Pass_Min_Charge_Count": raw.get("pass_min_charge_count"),
                "Pass_Min_Consecutive_Charge_Count": raw.get("pass_min_consecutive_charge_count"),
                "Pass_Charge_Continuity": raw.get("pass_charge_continuity"),
                "Pass_Internal_Error": raw.get("pass_internal_error"),
                "Pass_Neutral_Mass_SD": raw.get("pass_neutral_mass_sd"),
                "Pass_Neutral_Mass_Range": raw.get("pass_neutral_mass_range"),
                "Pass_RT_Consistency": raw.get("pass_rt_consistency"),
                "Pass_Local_Intensity": raw.get("pass_local_intensity"),
                "Pass_Competing_Envelope": raw.get("pass_competing_envelope"),
                "Pass_Peak_Sharing": raw.get("pass_peak_sharing"),
                "Num_Strict_Criteria_Passed": raw.get("num_strict_criteria_passed"),
                "Num_Review_Criteria_Passed": raw.get("num_review_criteria_passed"),
                "Strict_Failure_Reasons": raw.get("strict_failure_reasons"),
                "Review_Failure_Reasons": raw.get("review_failure_reasons"),
                "Intact_Quality_Tier": raw.get("intact_quality_tier"),
                "Quality_Tier_Reason": raw.get("quality_tier_reason"),
                "Quality_Tier_Rank": raw.get("quality_tier_rank"),
                "Comparison_Ready_Strict": raw.get("comparison_ready_strict"),
                "Comparison_Ready_Review": raw.get("comparison_ready_review"),
                "Comparison_Ready": raw.get("comparison_ready"),
                "Comparison_Readiness_Reason": raw.get("comparison_readiness_reason"),
                "Total_Supporting_Intensity": raw.get("total_supporting_intensity"),
                "Mean_Supporting_Intensity": raw.get("mean_supporting_intensity"),
                "Max_Supporting_Intensity": raw.get("max_supporting_intensity"),
                "Reconstructed_Envelope_Intensity": raw.get("reconstructed_envelope_intensity"),
                "Intensity_Method": raw.get("intensity_method"),
                "Relative_Envelope_Intensity_Percent": raw.get("relative_envelope_intensity_percent"),
                "Relative_Overall_Envelope_Intensity_Percent": raw.get("relative_overall_envelope_intensity_percent"),
                "Relative_In_Range_Raw_Intensity_Percent": raw.get("relative_in_range_raw_intensity_percent"),
                "Relative_Intact_Eligible_Intensity_Percent": raw.get("relative_intact_eligible_intensity_percent"),
                "Supporting_Peak_Classes": raw.get("supporting_peak_classes"),
                "Trace_Only_Envelope": raw.get("trace_only_envelope"),
                "Num_Supporting_Charge_States": raw.get("num_supporting_charge_states"),
                "Charge_State_Count": raw.get("charge_state_count"),
                "Charge_States": ",".join(map(str, raw.get("charge_states", []))),
                "Charge_State_Range": raw.get("charge_state_range"),
                "Charge_State_Continuity": raw.get("charge_state_continuity"),
                "Supporting_Peak_Count": raw.get("supporting_peak_count"),
                "RT_Min": raw.get("rt_min"),
                "RT_Max": raw.get("rt_max"),
                "RT_Mean": raw.get("rt_mean"),
                "RT_Range_Min": raw.get("rt_range_min"),
                "Max_RT_Difference_Min": raw.get("max_rt_difference_min"),
                "RT_Consistency": raw.get("rt_consistency"),
                "Neutral_Mass_SD": raw.get("neutral_mass_sd"),
                "Neutral_Mass_Range": raw.get("neutral_mass_range"),
                "Envelope_Internal_Error_ppm": raw.get("envelope_internal_error_ppm"),
                "Max_Mass_Error_ppm": raw.get("max_mass_error_ppm"),
                "Theoretical_Mass": raw.get("theoretical_mass"),
                "Mass_Error_Da": raw.get("mass_error_da"),
                "Mass_Error_ppm": raw.get("mass_error_ppm"),
                "Unmodified_Theory_Delta_Da": raw.get("unmodified_theory_delta_da"),
                "Unmodified_Theory_Delta_ppm": raw.get("unmodified_theory_delta_ppm"),
                "Best_Reference_Label": raw.get("best_reference_label"),
                "Best_Reference_Mass_Da": raw.get("best_reference_mass_da"),
                "Reference_Mass_Error_Da": raw.get("reference_mass_error_da"),
                "Reference_Mass_Error_ppm": raw.get("reference_mass_error_ppm"),
                "Reference_Mass_Matched": raw.get("reference_mass_matched"),
                "Competing_Envelope_Count": raw.get("competing_envelope_count"),
                "Limiting_Factors": raw.get("limiting_factors"),
                "Severe_Limiting_Factors": raw.get("severe_limiting_factors"),
                "Num_Limiting_Factors": raw.get("num_limiting_factors"),
                "Primary_Limiting_Factor": raw.get("primary_limiting_factor"),
                "Total_Intensity": raw.get("total_intensity"),
                "Assignment": raw.get("assignment"),
                "Confidence": raw.get("confidence"),
                "Warnings": raw.get("warnings"),
            }
        )

    charge_state_peak_rows = charge_state_peaks
    if len(charge_state_peaks) > max_charge_state_peak_rows and truncate_large_sheets:
        _append_excel_warning(warnings, "Charge_state_peaks", len(charge_state_peaks), max_charge_state_peak_rows)
        truncations.append(
            {
                "sheet": "Charge_state_peaks",
                "original_rows": len(charge_state_peaks),
                "written_rows": max_charge_state_peak_rows,
            }
        )
        charge_state_peak_rows = charge_state_peaks[:max_charge_state_peak_rows]

    theoretical_fragments = theoretical_fragments or []
    fragment_ms1_matches = fragment_ms1_matches or []
    known_modification_candidates = known_modification_candidates or []
    known_modification_summary = known_modification_summary or []
    unknown_modification_candidates = unknown_modification_candidates or []
    unknown_modification_summary = unknown_modification_summary or []
    compound_modification_candidates = compound_modification_candidates or []
    compound_modification_summary = compound_modification_summary or []
    fragment_ms1_filtered = _filter_fragment_ms1_matches(fragment_ms1_matches, config.fragment_mapping or {})
    fragment_ms1_summary_rows = _fragment_ms1_summary_rows(fragment_ms1_matches, config.fragment_mapping or {})
    input_parameters = {
        "analysis": getattr(config, "analysis", {}),
        "project": config.project,
        "input": config.input,
        "organism": config.organism,
        "sequence": config.sequence,
        "experiment": config.experiment,
        "instrument": config.instrument,
        "sciex_profile": getattr(config, "sciex_profile", {}),
        "reconstruction": config.reconstruction,
        "digestion": config.digestion,
        "alkaline_phosphatase": config.alkaline_phosphatase,
        "fragment_mapping": config.fragment_mapping,
        "modification_search": config.modification_search,
        "peak_filtering": config.peak_filtering,
        "p1_annotation": config.p1_annotation,
        "ms2_annotation": config.ms2_annotation,
        "modification_evidence_ranking": config.modification_evidence_ranking,
        "biological_context": config.biological_context,
        "performance": config.performance,
        "reporting": config.reporting,
    }
    analysis_mode = str((getattr(config, "analysis", {}) or {}).get("mode") or "full")
    optional_results = optional_results or {}
    workflow_summary_rows = optional_results.get("Workflow_Summary") or [
        {
            "Analysis_Mode": analysis_mode,
            "Step_Name": "workflow_summary",
            "Step_Status": "executed",
            "Enabled_By_Config": True,
            "Executed": True,
            "Skip_Reason": "",
            "Output_Sheets": "Workflow_Summary",
            "Notes": "Default summary row generated by report writer.",
        }
    ]

    data_sheets: dict[str, pd.DataFrame] = {
        "Workflow_Summary": pd.DataFrame(workflow_summary_rows, columns=WORKFLOW_SUMMARY_COLUMNS),
        "Input_parameters": pd.DataFrame(_flatten_dict(input_parameters)),
        "mzML_diagnostics": pd.DataFrame([diagnostics] if diagnostics else [{}]),
        "Theoretical_fragments": pd.DataFrame(_fragment_rows(theoretical_fragments), columns=THEORETICAL_FRAGMENT_COLUMNS),
        "Fragment_MS1_matches": pd.DataFrame(_fragment_ms1_match_rows(fragment_ms1_matches), columns=FRAGMENT_MS1_MATCH_COLUMNS),
        "Fragment_MS1_filtered": pd.DataFrame(_fragment_ms1_match_rows(fragment_ms1_filtered, include_length=True), columns=FRAGMENT_MS1_FILTERED_COLUMNS),
        "Fragment_MS1_summary": pd.DataFrame(fragment_ms1_summary_rows, columns=FRAGMENT_MS1_SUMMARY_COLUMNS),
        "Known_Modification_Candidates": pd.DataFrame(known_modification_candidates, columns=KNOWN_MODIFICATION_CANDIDATE_COLUMNS),
        "Known_Modification_Summary": pd.DataFrame(known_modification_summary, columns=KNOWN_MODIFICATION_SUMMARY_COLUMNS),
        "Unknown_Modification_Candidates": pd.DataFrame(unknown_modification_candidates, columns=UNKNOWN_MODIFICATION_CANDIDATE_COLUMNS),
        "Unknown_Modification_Summary": pd.DataFrame(unknown_modification_summary, columns=UNKNOWN_MODIFICATION_SUMMARY_COLUMNS),        
        "Compound_Modification_Candidates": pd.DataFrame(compound_modification_candidates, columns=COMPOUND_MODIFICATION_CANDIDATE_COLUMNS),
        "Compound_Modification_Summary": pd.DataFrame(compound_modification_summary, columns=COMPOUND_MODIFICATION_SUMMARY_COLUMNS),
    }
    intact_qc_sheets = {
        "Intact_Reconstruction_QC": pd.DataFrame(intact_qc_rows, columns=INTACT_QC_COLUMNS),
        "Intact_Reconstruction_Diag": pd.DataFrame(intact_diagnostic_rows, columns=INTACT_DIAGNOSTIC_COLUMNS),
        "Intact_Envelope_Groups": pd.DataFrame(intact_group_rows, columns=INTACT_GROUP_COLUMNS),
        "Intact_Competition_Groups": pd.DataFrame(intact_competition_group_rows, columns=INTACT_COMPETITION_GROUP_COLUMNS),
        "Intact_Competition_Scores": pd.DataFrame(intact_competition_score_rows, columns=INTACT_COMPETITION_SCORE_COLUMNS),
        "Intact_Assignment_Dry_Run": pd.DataFrame(assignment_dry_run_rows, columns=ASSIGNMENT_DRY_RUN_COLUMNS),
        "Competition_Dry_Run_Summary": pd.DataFrame(assignment_dry_run_summary_rows, columns=ASSIGNMENT_DRY_RUN_SUMMARY_COLUMNS),
        "Assignment_Sensitivity": pd.DataFrame(assignment_sensitivity_rows, columns=ASSIGNMENT_SENSITIVITY_COLUMNS),
        "Assignment_Stability": pd.DataFrame(assignment_stability_rows, columns=ASSIGNMENT_STABILITY_COLUMNS),
        "Assignment_Candidate_Audit": pd.DataFrame(assignment_candidate_audit_rows, columns=ASSIGNMENT_CANDIDATE_AUDIT_COLUMNS),
        "Assignment_Ambiguous_Candidates": pd.DataFrame(assignment_ambiguous_rows, columns=ASSIGNMENT_AMBIGUOUS_COLUMNS),
        "Preassignment_Comparison": pd.DataFrame(preassignment_comparison_rows, columns=PREASSIGNMENT_COMPARISON_COLUMNS),
        "Intact_Comparison_Candidates": pd.DataFrame(intact_comparison_rows, columns=INTACT_COMPARISON_CANDIDATE_COLUMNS),
        "Target_Review_Candidates": pd.DataFrame(intact_target_review_rows, columns=INTACT_TARGET_REVIEW_CANDIDATE_COLUMNS),
        "Reconstructed_Mass_Spectrum": pd.DataFrame(reconstructed_spectrum_rows, columns=RECONSTRUCTED_MASS_SPECTRUM_COLUMNS),
        "RT_Engine_QC_Summary": pd.DataFrame(rt_engine_qc_summary_rows, columns=RT_ENGINE_QC_SUMMARY_COLUMNS),
    }
    if reconstruction_enabled:
        data_sheets = {
            "Workflow_Summary": data_sheets["Workflow_Summary"],
            "Input_parameters": data_sheets["Input_parameters"],
            "mzML_diagnostics": data_sheets["mzML_diagnostics"],
            "Intact_mass_reconstruction": pd.DataFrame(intact_rows, columns=INTACT_COLUMNS),
            "Charge_state_peaks": pd.DataFrame(charge_state_peak_rows, columns=CHARGE_COLUMNS),
            **intact_qc_sheets,
            **{key: value for key, value in data_sheets.items() if key not in {"Input_parameters", "mzML_diagnostics"}},
        }
    else:
        data_sheets = {
            "Workflow_Summary": data_sheets["Workflow_Summary"],
            "Input_parameters": data_sheets["Input_parameters"],
            "mzML_diagnostics": data_sheets["mzML_diagnostics"],
            **intact_qc_sheets,
            **{key: value for key, value in data_sheets.items() if key not in {"Input_parameters", "mzML_diagnostics"}},
        }
    if analysis_mode == "intact_only":
        intact_only_sheet_names = {
            "Workflow_Summary",
            "Input_parameters",
            "mzML_diagnostics",
            "Intact_mass_reconstruction",
            "Charge_state_peaks",
            "Intact_Reconstruction_QC",
            "Intact_Reconstruction_Diag",
            "Intact_Envelope_Groups",
            "Intact_Competition_Groups",
            "Intact_Competition_Scores",
            "Intact_Assignment_Dry_Run",
            "Competition_Dry_Run_Summary",
            "Assignment_Sensitivity",
            "Assignment_Stability",
            "Assignment_Candidate_Audit",
            "Assignment_Ambiguous_Candidates",
            "Preassignment_Comparison",
            "Intact_Comparison_Candidates",
            "Target_Review_Candidates",
            "Reconstructed_Mass_Spectrum",
            "RT_Engine_QC_Summary",
            "RT_Envelope_Diagnostics",
            "Missing_Charge_Diagnostics",
            "Intact_Engine_Comparison",
        }
        data_sheets = {key: value for key, value in data_sheets.items() if key in intact_only_sheet_names}

    optional_columns = {
        "P1_Summary": P1_SUMMARY_COLUMNS,
        "P1_Theoretical_Structures": P1_THEORETICAL_COLUMNS,
        "P1_Peak_Annotations": P1_ANNOTATION_COLUMNS,
        "P1_Unmatched_Peaks": P1_UNMATCHED_COLUMNS,
        "P1_SAP_Chemical_State": P1_SAP_CHEMICAL_STATE_COLUMNS,
        "P1_SAP_PT_Family": P1_SAP_PT_FAMILY_COLUMNS,
        "P1_SAP_Terminal_Audit": P1_SAP_TERMINAL_COLUMNS,
        "P1_SAP_Chemistry_Summary": P1_SAP_SUMMARY_COLUMNS,
        "P1_SAP_Features": P1_SAP_FEATURE_COLUMNS,
        "P1_SAP_Competition": P1_SAP_COMPETITION_COLUMNS,
        "Cross_Enzyme_Chemistry": P1_SAP_CROSS_ENZYME_COLUMNS,
        "P1_SAP_MS2_Provenance": P1_SAP_MS2_PROVENANCE_COLUMNS,
        "P1_SAP_Spectrum_Peaks": P1_SAP_SPECTRUM_PEAK_COLUMNS,
        "P1_SAP_Refined_Features": P1_SAP_REFINED_FEATURE_COLUMNS,
        "P1_SAP_Feature_Quality": P1_SAP_FEATURE_QUALITY_COLUMNS,
        "P1_SAP_Isotope_Audit": P1_SAP_ISOTOPE_AUDIT_COLUMNS,
        "P1_SAP_Quality_Summary": P1_SAP_QUALITY_SUMMARY_COLUMNS,
        "P1_SAP_Dinuc_Summary": P1_SAP_DINUC_SUMMARY_COLUMNS,
        "P1_SAP_Dinuc_Groups": P1_SAP_DINUC_GROUP_COLUMNS,
        "P1_SAP_Dinuc_Assignments": P1_SAP_DINUC_ASSIGNMENT_COLUMNS,
        "P1_SAP_Dinuc_SpecPeaks": P1_SAP_DINUC_SPECPEAK_COLUMNS,
        "P1_SAP_Dinuc_Features": P1_SAP_DINUC_FEATURE_COLUMNS,
        "P1_SAP_Dinuc_Isotopes": P1_SAP_DINUC_ISOTOPE_COLUMNS,
        "P1_SAP_Dinuc_Competition": P1_SAP_DINUC_COMPETITION_COLUMNS,
        "P1_SAP_Dinuc_MS2": P1_SAP_DINUC_MS2_COLUMNS,
        "P1_SAP_Dinuc_Targets": P1_SAP_DINUC_TARGET_COLUMNS,
        "P1_SAP_Dinuc_Evidence": P1_SAP_DINUC_EVIDENCE_COLUMNS,
        "P1_SAP_Dinuc_Group_Evidence": P1_SAP_DINUC_GROUP_EVIDENCE_COLUMNS,
        "P1_SAP_Dinuc_Evidence_Summary": P1_SAP_DINUC_EVIDENCE_SUMMARY_COLUMNS,
        "MS2_Summary": MS2_SUMMARY_COLUMNS,
        "MS2_Spectra": MS2_SPECTRA_COLUMNS,
        "MS2_Parent_Candidates": MS2_PARENT_CANDIDATE_COLUMNS,
        "MS2_Modified_Precursor_Candidates": MS2_MODIFIED_PRECURSOR_COLUMNS,
        "MS2_Modified_Theoretical_Ions": MS2_MODIFIED_THEORETICAL_ION_COLUMNS,
        "MS2_Modified_Ion_Matches": MS2_MODIFIED_ION_MATCH_COLUMNS,
        "MS2_Modification_Localization_Evidence": MS2_LOCALIZATION_EVIDENCE_COLUMNS,
        "Modification_Evidence_Summary": SUMMARY_COLUMNS,
        "Modification_Evidence_Ranking": RANKING_COLUMNS + SHADOW_RANKING_COLUMNS + IDENTITY_SHADOW_COLUMNS,
        "Modification_Ambiguity_Groups": AMBIGUITY_GROUP_COLUMNS,
        "Modification_Position_Priors": POSITION_PRIOR_COLUMNS,
        "MS2_Biological_Plausibility": BIOLOGICAL_PLAUSIBILITY_COLUMNS,
        "MS2_Modification_Identity": IDENTITY_COLUMNS,
        "MS2_Identity_Peak_Assignments": PEAK_ASSIGNMENT_COLUMNS,
        "RNase_MS2_Evidence_Summary": RNASE_MS2_EVIDENCE_SUMMARY_COLUMNS,
        "RNase_MS2_Candidate_Evidence": RNASE_MS2_CANDIDATE_EVIDENCE_COLUMNS,
        "RNase_MS2_Peak_Evidence": RNASE_MS2_PEAK_EVIDENCE_COLUMNS,
        "RNase_MS2_Composite_Summary": RNASE_MS2_COMPOSITE_SUMMARY_COLUMNS,
        "RNase_MS2_Composite_Evidence": RNASE_MS2_COMPOSITE_EVIDENCE_COLUMNS,
        "RNase_MS2_Composite_Peak_Evidence": RNASE_MS2_COMPOSITE_PEAK_EVIDENCE_COLUMNS,
        "RNase_MS2_Standard_Composite_Summary": RNASE_MS2_STANDARD_COMPOSITE_SUMMARY_COLUMNS,
        "RNase_MS2_Standard_Composite_Crosswalk": RNASE_MS2_STANDARD_COMPOSITE_CROSSWALK_COLUMNS,
        "RNase_MS2_Consensus_Summary": RNASE_MS2_CONSENSUS_SUMMARY_COLUMNS,
        "RNase_MS2_Consensus_Evidence": RNASE_MS2_CONSENSUS_EVIDENCE_COLUMNS,
        "SCIEX_Profile_Diagnostics": SCIEX_PROFILE_DIAGNOSTIC_COLUMNS,
        "SCIEX_Profile_Input": SCIEX_PROFILE_INPUT_COLUMNS,
        "MS2_Unmatched_Ion_Audit": MS2_UNMATCHED_ION_AUDIT_COLUMNS,
        "MS2_Unmatched_Ion_Summary": MS2_UNMATCHED_ION_SUMMARY_COLUMNS,
        "MS2_Unmatched_Ion_Diagnostics": MS2_UNMATCHED_ION_DIAGNOSTIC_COLUMNS + MS2_AMBIGUOUS_DIAGNOSTIC_COLUMNS + MS2_ZERO_INTENSITY_DIAGNOSTIC_COLUMNS + MS2_EFFECTIVE_AMBIGUITY_DIAGNOSTIC_COLUMNS + MS1_TRUNCATION_DIAGNOSTIC_COLUMNS + MS1_SELECTION_DIAGNOSTIC_COLUMNS + MS1_TOP50_DEDUP_DIAGNOSTIC_COLUMNS + MS1_CROSSFRAG_DIAGNOSTIC_COLUMNS + AUDIT_LEVEL_DIAGNOSTIC_COLUMNS + COMPOSITE_DIAGNOSTIC_COLUMNS,
        "Audit_Status": AUDIT_STATUS_COLUMNS,
        "Composite_Mod_Candidates": COMPOSITE_CANDIDATE_COLUMNS,
        "Composite_Mod_Invalid": COMPOSITE_INVALID_COLUMNS,
        "Composite_Mod_Summary": COMPOSITE_SUMMARY_COLUMNS,
        "Backbone_Mod_Candidates": BACKBONE_COLUMNS,
        "Cleavage_Block_Audit": CLEAVAGE_COLUMNS,
        "Composite_Fragment_Masses": COMPOSITE_FRAGMENT_COLUMNS,
        "Composite_MS1_Matches": COMPOSITE_MS1_COLUMNS,
        "Composite_MS1_Summary": COMPOSITE_MS1_SUMMARY_COLUMNS,
        "Composite_MS2_Ions": COMPOSITE_MS2_ION_COLUMNS,
        "Composite_MS2_Matches": COMPOSITE_MS2_MATCH_COLUMNS,
        "Composite_MS2_Assignment_Competition": COMPOSITE_MS2_ASSIGNMENT_COMPETITION_COLUMNS,
        "Composite_Structure_Position_Map": COMPOSITE_STRUCTURE_POSITION_MAP_COLUMNS,
        "Composite_Structure_Bond_Map": COMPOSITE_STRUCTURE_BOND_MAP_COLUMNS,
        "Composite_Support_Summary": COMPOSITE_SUPPORT_COLUMNS,
        "Blocked_Cleavage_Matches": COMPOSITE_BLOCKED_COLUMNS,
        "Legacy_Composite_Compare": COMPOSITE_COMPARE_COLUMNS,
        "Composite_Shadow_Score": COMPOSITE_SCORE_COLUMNS,
        "Composite_Obs_Summary": COMPOSITE_OBS_SUMMARY_COLUMNS,
        "PT_Paired_Summary": PT_SUMMARY_COLUMNS,
        "PT_Discovery_Candidates": PT_DISCOVERY_COLUMNS,
        "PT_Paired_Evidence": PT_PAIRED_EVIDENCE_COLUMNS,
        "PT_State_Search": PT_STATE_SEARCH_COLUMNS,
        "Composite_Obs_Invalid": COMPOSITE_OBS_INVALID_COLUMNS,
        "MS1_Truncation_Audit": MS1_TRUNCATION_AUDIT_COLUMNS,
        "MS1_Truncation_Detail": MS1_TRUNCATION_DETAIL_COLUMNS,
        "MS1_Truncation_Summary": MS1_TRUNCATION_SUMMARY_COLUMNS,
        "MS1_Selection_Strategy": MS1_SELECTION_STRATEGY_COLUMNS,
        "MS1_Selection_Detail": MS1_SELECTION_DETAIL_COLUMNS,
        "MS1_Selection_Summary": MS1_SELECTION_SUMMARY_COLUMNS,
        "MS1_Top50_Shadow": MS1_TOP50_SHADOW_COLUMNS,
        "MS1_Peak_Dedup_Detail": MS1_PEAK_DEDUP_DETAIL_COLUMNS,
        "MS1_Top50_Dedup_Summary": MS1_TOP50_DEDUP_SUMMARY_COLUMNS,
        "MS1_CrossFrag_Ambiguity": MS1_CROSSFRAG_AMBIGUITY_COLUMNS,
        "MS1_CrossFrag_Detail": MS1_CROSSFRAG_DETAIL_COLUMNS,
        "MS1_CrossFrag_Summary": MS1_CROSSFRAG_SUMMARY_COLUMNS,
        "MS2_Ambiguous_Peak_Clusters": MS2_AMBIGUOUS_CLUSTER_COLUMNS,
        "MS2_Ambiguous_Peak_Detail": MS2_AMBIGUOUS_DETAIL_COLUMNS,
        "MS2_Ambiguity_Summary": MS2_AMBIGUITY_SUMMARY_COLUMNS,
        "MS2_Zero_Intensity_Spectra": MS2_ZERO_INTENSITY_SPECTRA_COLUMNS,
        "MS2_Zero_Intensity_Detail": MS2_ZERO_INTENSITY_DETAIL_COLUMNS,
        "MS2_Zero_Intensity_Summary": MS2_ZERO_INTENSITY_SUMMARY_COLUMNS,
        "MS2_Effective_Ambiguity": MS2_EFFECTIVE_AMBIGUITY_COLUMNS,
        "MS2_Effective_Ambig_Detail": MS2_EFFECTIVE_AMBIGUITY_DETAIL_COLUMNS,
        "MS2_Effective_Ambig_Summary": MS2_EFFECTIVE_AMBIGUITY_SUMMARY_COLUMNS,
        "Biological_Prior_Diagnostics": BIOLOGICAL_PRIOR_DIAGNOSTIC_COLUMNS,
        "Biological_Context_Priorities": CONTEXT_PRIORITY_COLUMNS,
        "Context_Supported_Candidates": RANKING_COLUMNS + SHADOW_RANKING_COLUMNS + IDENTITY_SHADOW_COLUMNS,
        "MS2_Theoretical_Ions": MS2_THEORETICAL_ION_COLUMNS,
        "MS2_Ion_Matches": MS2_ION_MATCH_COLUMNS,
        "MS2_Unmatched_Peaks": MS2_UNMATCHED_COLUMNS,
        "MS2_Fragment_Evidence": MS2_FRAGMENT_EVIDENCE_COLUMNS,
        "MS2_Peak_Annotations": MS2_ION_MATCH_COLUMNS,
        "RT_Envelope_Diagnostics": RT_ENVELOPE_DIAGNOSTIC_COLUMNS,
        "RT_Engine_QC_Summary": RT_ENGINE_QC_SUMMARY_COLUMNS,
        "Missing_Charge_Diagnostics": MISSING_CHARGE_DIAGNOSTIC_COLUMNS,
        "Intact_Engine_Comparison": ENGINE_COMPARISON_COLUMNS,
        "Intact_Competition_Groups": INTACT_COMPETITION_GROUP_COLUMNS,
        "Intact_Competition_Scores": INTACT_COMPETITION_SCORE_COLUMNS,
        "Intact_Assignment_Dry_Run": ASSIGNMENT_DRY_RUN_COLUMNS,
        "Competition_Dry_Run_Summary": ASSIGNMENT_DRY_RUN_SUMMARY_COLUMNS,
        "Assignment_Sensitivity": ASSIGNMENT_SENSITIVITY_COLUMNS,
        "Assignment_Stability": ASSIGNMENT_STABILITY_COLUMNS,
        "Assignment_Candidate_Audit": ASSIGNMENT_CANDIDATE_AUDIT_COLUMNS,
        "Assignment_Ambiguous_Candidates": ASSIGNMENT_AMBIGUOUS_COLUMNS,
        "Preassignment_Comparison": PREASSIGNMENT_COMPARISON_COLUMNS,
    }
    data_sheets.update(
        _sciex_intact_excel_sheets(optional_results.get(SCIEX_INTACT_OPTIONAL_RESULT_KEY))
    )
    data_sheets.update(
        _sciex_identity_audit_excel_sheets(optional_results.get(SCIEX_IDENTITY_AUDIT_RESULT_KEY))
    )
    data_sheets.update(
        _sciex_delta_cluster_excel_sheets(optional_results.get(SCIEX_DELTA_CLUSTER_RESULT_KEY))
    )
    data_sheets.update(
        _sciex_spacing_resolution_excel_sheets(optional_results.get(SCIEX_SPACING_RESOLUTION_RESULT_KEY))
    )
    data_sheets.update(
        _sciex_relation_evidence_excel_sheets(optional_results.get(SCIEX_RELATION_EVIDENCE_RESULT_KEY))
    )
    data_sheets.update(
        _sciex_mass_comparison_excel_sheets(optional_results.get(SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY))
    )
    data_sheets.update(
        _sciex_cross_layer_excel_sheets(optional_results.get(SCIEX_CROSS_LAYER_RESULT_KEY))
    )
    for sheet_name, value in optional_results.items():
        if sheet_name in {
            "Index", "Run_summary", "Warnings", "Workflow_Summary",
            SCIEX_INTACT_OPTIONAL_RESULT_KEY,
            SCIEX_IDENTITY_AUDIT_RESULT_KEY,
            SCIEX_DELTA_CLUSTER_RESULT_KEY,
            SCIEX_SPACING_RESOLUTION_RESULT_KEY,
            SCIEX_RELATION_EVIDENCE_RESULT_KEY,
            SCIEX_MASS_COMPARISON_OPTIONAL_RESULT_KEY,
            SCIEX_CROSS_LAYER_RESULT_KEY,
        }:
            continue
        frame = _coerce_to_frame(value)
        columns = optional_columns.get(sheet_name)
        if columns:
            frame = pd.DataFrame(frame, columns=columns)
        if sheet_name == "Top_Modification_Candidates" and not audit_policy.include_top_shadow_columns:
            frame = frame.drop(columns=[column for column in AUDIT_TOP_SHADOW_COLUMNS if column in frame.columns])
        excel_sheet_name = (
            "Composite_MS2_Assignment_Compe"
            if sheet_name == "Composite_MS2_Assignment_Competition"
            else sheet_name[:31]
        )
        data_sheets[excel_sheet_name] = frame

    included_names, unknown_names = included_sheet_names(data_sheets, audit_policy)
    if unknown_names and audit_policy.level != "full":
        add_warning(
            warnings, "WARNING", "excel_report",
            "Unclassified sheets were omitted for the selected audit level.",
            {"audit_level": audit_policy.level, "sheets": unknown_names},
        )
    data_sheets = {name: frame for name, frame in data_sheets.items() if name in included_names}
    if audit_policy.run_shadow_audits and "MS2_Unmatched_Ion_Diagnostics" in data_sheets:
        actual_shadow_sheet_count = sum(
            1 for name in data_sheets if (sheet_category(name) or "").startswith("AUDIT_")
        )
        diagnostics_frame = data_sheets["MS2_Unmatched_Ion_Diagnostics"].copy()
        diagnostics_frame["Shadow_Audit_Sheet_Count"] = actual_shadow_sheet_count
        data_sheets["MS2_Unmatched_Ion_Diagnostics"] = diagnostics_frame

    truncated_data_sheets = {
        sheet_name: _truncate_frame_if_needed(
            sheet_name,
            frame,
            max_excel_rows,
            truncate_large_sheets,
            warnings,
            truncations,
        )
        for sheet_name, frame in data_sheets.items()
    }

    summary_rows = [
        {"Item": "Project", "Value": config.project.get("name")},
        {"Item": "Generated", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Item": "Analysis mode", "Value": _analysis_mode(config)},
        {"Item": "Modification dictionary entries", "Value": len(modifications or [])},
        {"Item": "Rule set", "Value": config.organism.get("rule_set") or (rule_set or {}).get("id") or (rule_set or {}).get("name")},
        {"Item": "Pathway files", "Value": len(pathways or [])},
        {"Item": "Intact mass candidates", "Value": len(intact_results)},
        {"Item": "Theoretical fragments", "Value": len(theoretical_fragments)},
        {"Item": "Fragment MS1 matches", "Value": len(fragment_ms1_matches)},
        {"Item": "Fragment MS1 filtered", "Value": len(fragment_ms1_filtered)},
        {"Item": "Fragment MS1 summary", "Value": len(fragment_ms1_summary_rows)},
        {"Item": "Known modification candidates", "Value": len(known_modification_candidates)},
        {"Item": "Known modification summary", "Value": len(known_modification_summary)},
        {"Item": "Unknown modification candidates", "Value": len(unknown_modification_candidates)},
        {"Item": "Unknown modification summary", "Value": len(unknown_modification_summary)},
        {"Item": "Compound modification candidates", "Value": len(compound_modification_candidates)},
        {"Item": "Compound modification summary", "Value": len(compound_modification_summary)},
        {"Item": "Truncated sheets", "Value": _truncation_summary(truncations)},
        {"Item": "Warnings", "Value": len(warnings)},
        *audit_policy.run_summary_items(),
    ]

    sheets: dict[str, pd.DataFrame] = {
        "Run_summary": pd.DataFrame(summary_rows),
        **truncated_data_sheets,
        "Warnings": pd.DataFrame(warnings, columns=["Timestamp", "Level", "Source", "Message", "Context"]),
    }
    sheets = {
        sheet_name: _truncate_frame_if_needed(
            sheet_name,
            frame,
            max_excel_rows,
            truncate_large_sheets,
            warnings,
            truncations,
        )
        if sheet_name == "Warnings"
        else frame
        for sheet_name, frame in sheets.items()
    }

    sheets = _reorder_sheets_conclusion_then_detail(sheets)

    word_export_collector = WordExportCollector()
    sheets = _extract_long_cells_to_word(sheets, word_export_collector)

    word_appendix_written_path = None
    if not word_export_collector.is_empty():
        word_appendix_written_path = write_word_appendix(word_export_collector, word_appendix_path)

    index_rows = [
        {
            "Sheet": sheet_name,
            "Description": SHEET_DESCRIPTIONS.get(sheet_name, "Optional result sheet."),
            "Notes": "Data starts at A3.",
        }
        for sheet_name in sheets
    ]

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        pd.DataFrame(index_rows, columns=["Sheet", "Description", "Notes"]).to_excel(writer, sheet_name="Index", index=False)
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        _add_index_and_backlinks(writer, list(sheets))
        if word_appendix_written_path is not None:
            _add_word_appendix_hyperlinks(writer, word_export_collector, word_appendix_written_path.name)
        _autosize_and_freeze(writer)
    return report_path, word_appendix_written_path
    
